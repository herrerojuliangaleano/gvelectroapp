from __future__ import annotations

import csv
import logging
import math
import re
import unicodedata
from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple

import gspread
from google.auth.transport.requests import Request
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from googleapiclient.discovery import build
from googleapiclient.errors import HttpError
from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel


# =========================================================
# CREDENCIALES UNIFICADAS (apuntan a la raíz de la app)
# =========================================================
_APP_ROOT = Path(__file__).resolve().parent.parent.parent
CREDENTIALS_FILE = _APP_ROOT / "credentials.json"
TOKEN_FILE = _APP_ROOT / "token.json"
SCOPES = [
    "https://www.googleapis.com/auth/drive",
    "https://www.googleapis.com/auth/spreadsheets",
]


# =============================================================================
# CONFIGURACIÓN
# =============================================================================

BASE_DIR = Path(__file__).resolve().parent
ARCA_DIR = BASE_DIR / "arca"

SALES_DRIVE_FOLDER_ID = "1aaUJyDOuC5DRX9-qquo_sjste1OL5Vfc"
ACCOUNT_CONFIG_BY_CUIT = {
    "30717199207": {
        "name": "GV",
        "output_root": "1dmpotW8N2g9sDNEORoQHFYi_p_aBZikn",
    },
    "30717985598": {
        "name": "ABC",
        "output_root": "1Eo-_Ak9ZwRbP2ADZVEnow1ZR6pyZFPos",
    },
}

SALES_START_ROW = 4
SALES_END_ROW = 185

# Ventas (1-based)
COL_REMITO = 1      # A
COL_EFECTIVO = 3    # C
COL_TRANSFER = 4    # D
COL_POSNET = 5      # E
COL_USD = 6         # F
COL_VALOR = 10      # J
COL_T = 20          # T
COL_FACTURA_U = 21  # U

# ARCA
ARCA_DATE_COL = 1
ARCA_INVOICE_COL = 4
ARCA_TOTAL_DEFAULT_COL = 28

# POSNET
POSNET_BASE = Decimal("0.085")
POSNET_BANK_MAX = Decimal("0.23")
POSNET_MIN_FACTOR = Decimal("1.085")
POSNET_MAX_FACTOR = Decimal("1.35")
POSNET_SELECTION_FACTOR = (POSNET_MIN_FACTOR + POSNET_MAX_FACTOR) / Decimal("2")

# Tolerancias
ABS_TOLERANCE_NO_POSNET = Decimal("1.00")
ABS_TOLERANCE_POSNET_BOUNDARY = Decimal("1.00")

# Colores
COLOR_GREEN = {"red": 0.74, "green": 0.90, "blue": 0.74}
COLOR_YELLOW = {"red": 1.00, "green": 0.93, "blue": 0.55}
COLOR_ORANGE = {"red": 1.00, "green": 0.82, "blue": 0.49}
COLOR_RED = {"red": 1.00, "green": 0.65, "blue": 0.65}
COLOR_PURPLE = {"red": 0.80, "green": 0.70, "blue": 0.95}

LOG_TO_CONSOLE = True
MARK_BLANK_INVOICE_CELLS = True

# =============================================================================
# LOGGING
# =============================================================================

logger = logging.getLogger("conciliador_facturas")
logger.setLevel(logging.INFO)
logger.handlers.clear()
logger.propagate = False

_file_handler = logging.FileHandler(BASE_DIR / "conciliador_facturas.log", encoding="utf-8")
_file_handler.setLevel(logging.INFO)
_file_handler.setFormatter(logging.Formatter("%(asctime)s - %(levelname)s - %(message)s"))
logger.addHandler(_file_handler)

if LOG_TO_CONSOLE:
    _console_handler = logging.StreamHandler()
    _console_handler.setLevel(logging.ERROR)
    _console_handler.setFormatter(logging.Formatter("[%(levelname)s] %(message)s"))
    logger.addHandler(_console_handler)


def log_console_error_block(title: str, note: str) -> None:
    """Emite un bloque de error fácil de leer en terminal."""
    border = "═" * 90
    logger.error("\n%s\n%s\n%s\n%s", border, title, note, border)


# =============================================================================
# MODELOS
# =============================================================================

@dataclass
class SalesRow:
    row_number: int
    raw_values: List[Any]
    remito_raw: str
    factura_raw: str
    remito: Optional[int]
    factura: Optional[int]
    valor: Decimal
    efectivo: Decimal
    transferencia: Decimal
    posnet: Decimal
    usd: Decimal
    t_flag: str

    def has_payment(self) -> bool:
        return any(
            amount > 0
            for amount in (self.efectivo, self.transferencia, self.posnet, self.usd)
        )

    def payment_tuple(self) -> Tuple[Decimal, Decimal, Decimal, Decimal]:
        return (self.efectivo, self.transferencia, self.posnet, self.usd)


@dataclass
class InvoiceGroup:
    invoice: int
    row_numbers: List[int] = field(default_factory=list)
    remitos: List[int] = field(default_factory=list)
    rows: List[SalesRow] = field(default_factory=list)
    total_j: Decimal = Decimal("0.00")
    payment_rows: List[SalesRow] = field(default_factory=list)

    def add_row(self, row: SalesRow) -> None:
        self.row_numbers.append(row.row_number)
        if row.remito is not None:
            self.remitos.append(row.remito)
        self.rows.append(row)
        self.total_j += row.valor
        if row.has_payment():
            self.payment_rows.append(row)


@dataclass
class ValidationIssue:
    severity: str  # warn | inconsistent | error | missing
    message: str


@dataclass
class ValidationResult:
    status: str  # ok | warn | inconsistent | error | missing
    color: Optional[Dict[str, float]]
    note: str
    arca_record: Optional["ArcaRecord"] = None


@dataclass
class ArcaRecord:
    source_file: str
    source_row: int
    fecha: Optional[date]
    factura_desde: Optional[int]
    importe: Decimal
    output_values: List[Any]


@dataclass
class ArcaTable:
    headers: List[Any]
    records: List[ArcaRecord]
    invoice_col: int
    total_col: Optional[int]
    date_col: int
    output_header: List[Any]


# =============================================================================
# UTILIDADES
# =============================================================================


def remove_accents(text: str) -> str:
    text = unicodedata.normalize("NFKD", text)
    return "".join(ch for ch in text if not unicodedata.combining(ch))



def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    if is_nan(value):
        return ""
    return str(value).strip()



def normalize_tag(value: Any) -> str:
    text = remove_accents(normalize_text(value)).upper()
    text = re.sub(r"\s+", " ", text)
    return text.strip()



def normalize_header(value: Any) -> str:
    return normalize_tag(value).lower()



def is_nan(value: Any) -> bool:
    return isinstance(value, float) and math.isnan(value)



def money_quantize(value: Decimal) -> Decimal:
    return value.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)



def fmt_money(value: Decimal) -> str:
    return f"{money_quantize(value):.2f}"



def normalize_invoice(value: Any) -> Optional[int]:
    """Convierte un valor a entero. Mantiene solo dígitos."""
    if value is None or is_nan(value):
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        if value.is_integer():
            return int(value)
        digits = re.sub(r"\D", "", str(value))
        return int(digits) if digits else None

    text = normalize_text(value)
    if not text:
        return None

    digits = re.sub(r"\D", "", text)
    if not digits:
        return None

    try:
        return int(digits)
    except ValueError:
        return None



def parse_money(value: Any) -> Decimal:
    """
    Parseo robusto de montos con formato local.

    Acepta:
    - 100.000,00
    - 550.000
    - 550,00
    - $ 100.000,00
    - ARS 100.000,00
    - valores vacíos / errores de Sheets / texto residual
    """
    if value is None or value == "" or is_nan(value):
        return Decimal("0.00")
    if isinstance(value, bool):
        return Decimal("0.00")
    if isinstance(value, (int, float, Decimal)):
        try:
            return money_quantize(Decimal(str(value)))
        except (InvalidOperation, ValueError):
            return Decimal("0.00")

    s = str(value).strip()
    if not s:
        return Decimal("0.00")

    if s.upper() in {"#REF!", "#N/A", "#VALUE!", "#ERROR!"}:
        return Decimal("0.00")

    s = (
        s.replace("$", "")
         .replace("ARS", "")
         .replace("AR$", "")
         .replace("USD", "")
         .replace(" ", "")
         .replace("\u00a0", "")
    )

    if re.fullmatch(r"-?\d{1,3}(?:\.\d{3})+(?:,\d+)?", s):
        s = s.replace(".", "").replace(",", ".")
    elif re.fullmatch(r"-?\d{1,3}(?:,\d{3})+(?:\.\d+)?", s):
        s = s.replace(",", "")
    else:
        if "," in s and "." in s:
            if s.rfind(",") > s.rfind("."):
                s = s.replace(".", "").replace(",", ".")
            else:
                s = s.replace(",", "")
        elif "," in s:
            parts = s.split(",")
            if len(parts[-1]) in (1, 2):
                s = s.replace(".", "").replace(",", ".")
            else:
                s = s.replace(",", "")
        elif "." in s:
            parts = s.split(".")
            if len(parts) > 2:
                s = s.replace(".", "")
            elif len(parts[-1]) in (1, 2):
                pass
            else:
                s = s.replace(".", "")

    s = re.sub(r"[^0-9.\-]", "", s)
    if s in ("", ".", "-", "-."):
        return Decimal("0.00")

    try:
        return money_quantize(Decimal(s))
    except (InvalidOperation, ValueError):
        return Decimal("0.00")


DATE_FORMATS = (
    "%d-%m-%Y", "%d/%m/%Y", "%d.%m.%Y",
    "%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d",
    "%d-%m-%y", "%d/%m/%y", "%d.%m.%y",
    "%d %m %Y", "%Y %m %d",
)




def parse_date(value: Any) -> Optional[date]:
    if value is None or value == "" or is_nan(value):
        return None

    if isinstance(value, datetime):
        return value.date()

    if isinstance(value, date):
        return value

    if isinstance(value, (int, float)):
        serial = float(value)
        if 20000 <= serial <= 90000:
            try:
                return from_excel(serial).date()
            except Exception:
                pass

    text_value = str(value).strip()
    if not text_value:
        return None

    text_value = text_value.split(" ", 1)[0].strip()
    normalized_value = re.sub(r"[.\s]+", "/", text_value)

    if re.fullmatch(r"\d+(?:\.\d+)?", text_value):
        try:
            serial = float(text_value)
            if 20000 <= serial <= 90000:
                return from_excel(serial).date()
        except Exception:
            pass

    candidates = {
        text_value,
        normalized_value,
        text_value.replace(".", "-"),
        text_value.replace("/", "-"),
    }

    for candidate in candidates:
        for fmt in DATE_FORMATS:
            try:
                return datetime.strptime(candidate, fmt).date()
            except ValueError:
                continue

    return None



def format_date_ddmmyyyy(d: Optional[date]) -> str:
    return d.strftime("%d-%m-%Y") if d else ""



def column_letter(col_number: int) -> str:
    result = ""
    n = col_number
    while n > 0:
        n, rem = divmod(n - 1, 26)
        result = chr(65 + rem) + result
    return result



def cell_value(row: Sequence[Any], idx1: int) -> Any:
    if idx1 <= 0:
        return None
    idx = idx1 - 1
    if idx >= len(row):
        return None
    return row[idx]



def has_business_content(row: Sequence[Any]) -> bool:
    for idx in (COL_REMITO, COL_EFECTIVO, COL_TRANSFER, COL_POSNET, COL_USD, COL_VALOR, COL_FACTURA_U):
        value = cell_value(row, idx)
        if isinstance(value, (int, float, Decimal)) and not is_nan(value):
            if Decimal(str(value)) != 0:
                return True
        if normalize_text(value):
            return True
    return False


# =============================================================================
# GOOGLE AUTH / CLIENTS
# =============================================================================




def get_credentials() -> Credentials:
    creds: Optional[Credentials] = None

    if TOKEN_FILE.exists():
        creds = Credentials.from_authorized_user_file(str(TOKEN_FILE), SCOPES)

    if creds and creds.expired and creds.refresh_token:
        creds.refresh(Request())
    elif not creds or not creds.valid:
        if not CREDENTIALS_FILE.exists():
            raise FileNotFoundError(f"No encontré {CREDENTIALS_FILE}.")
        flow = InstalledAppFlow.from_client_secrets_file(str(CREDENTIALS_FILE), SCOPES)
        creds = flow.run_local_server(port=0)

    TOKEN_FILE.write_text(creds.to_json(), encoding="utf-8")
    return creds



def get_google_clients():
    creds = get_credentials()
    gc = gspread.authorize(creds)
    drive_service = build("drive", "v3", credentials=creds, cache_discovery=False)
    sheets_service = build("sheets", "v4", credentials=creds, cache_discovery=False)
    return gc, drive_service, sheets_service


# =============================================================================
# DRIVE HELPERS
# =============================================================================


def escape_drive_q(value: str) -> str:
    return value.replace("'", "\\'")



def list_drive_files(
    drive_service,
    q: str,
    fields: str = "files(id, name, mimeType, parents, modifiedTime)",
) -> List[Dict[str, Any]]:
    items: List[Dict[str, Any]] = []
    page_token: Optional[str] = None

    while True:
        resp = (
            drive_service.files()
            .list(
                q=q,
                spaces="drive",
                fields=f"nextPageToken,{fields}",
                pageSize=1000,
                pageToken=page_token,
                includeItemsFromAllDrives=True,
                supportsAllDrives=True,
            )
            .execute()
        )
        items.extend(resp.get("files", []))
        page_token = resp.get("nextPageToken")
        if not page_token:
            break

    return items



def drive_find_exact(
    drive_service,
    name: str,
    parent_id: Optional[str] = None,
    mime_type: Optional[str] = None,
    is_folder: bool = False,
) -> Optional[Dict[str, Any]]:
    q_parts = [f"name = '{escape_drive_q(name)}'", "trashed = false"]
    if parent_id:
        q_parts.append(f"'{parent_id}' in parents")
    if mime_type:
        q_parts.append(f"mimeType = '{mime_type}'")
    if is_folder:
        q_parts.append("mimeType = 'application/vnd.google-apps.folder'")

    files = list_drive_files(drive_service, q=" and ".join(q_parts))
    if not files:
        return None

    files.sort(key=lambda x: x.get("modifiedTime", ""), reverse=True)
    return files[0]



def drive_ensure_folder(drive_service, folder_name: str, parent_id: str) -> str:
    existing = drive_find_exact(drive_service, folder_name, parent_id=parent_id, is_folder=True)
    if existing:
        return existing["id"]

    created = (
        drive_service.files()
        .create(
            body={
                "name": folder_name,
                "mimeType": "application/vnd.google-apps.folder",
                "parents": [parent_id],
            },
            fields="id",
            supportsAllDrives=True,
        )
        .execute()
    )
    return created["id"]



def drive_create_spreadsheet(drive_service, name: str, parent_id: str) -> str:
    created = (
        drive_service.files()
        .create(
            body={
                "name": name,
                "mimeType": "application/vnd.google-apps.spreadsheet",
                "parents": [parent_id],
            },
            fields="id",
            supportsAllDrives=True,
        )
        .execute()
    )
    return created["id"]


def get_sales_worksheet(spreadsheet: gspread.Spreadsheet) -> gspread.Worksheet:
    """Devuelve la hoja principal de ventas.

    Prioriza una pestaña llamada 'Planilla'. Si no existe, cae a la primera hoja.
    """
    for candidate in ("Planilla", "planilla", "PLANILLA"):
        try:
            return spreadsheet.worksheet(candidate)
        except Exception:
            continue
    return spreadsheet.sheet1



def prompt_target_date() -> date:
    """
    Pide al usuario la fecha a analizar y valida hasta recibir una fecha correcta.
    """
    while True:
        raw = input(
            "Ingresá la fecha a analizar (dd/mm/yyyy, dd-mm-yyyy, dd.mm.yyyy, yyyy-mm-dd): "
        ).strip()
        parsed = parse_date(raw)
        if parsed is not None:
            return parsed
        print("Fecha inválida. Probá nuevamente.")


def extract_spreadsheet_id(raw_value: str) -> Optional[str]:
    text = normalize_text(raw_value)
    if not text:
        return None

    match = re.search(r"/spreadsheets/d/([a-zA-Z0-9-_]+)", text)
    if match:
        return match.group(1)

    if re.fullmatch(r"[a-zA-Z0-9-_]{20,}", text):
        return text

    return None


def extract_drive_folder_id(raw_value: str) -> Optional[str]:
    text = normalize_text(raw_value)
    if not text:
        return None

    folder_match = re.search(r"/folders/([a-zA-Z0-9-_]+)", text)
    if folder_match:
        return folder_match.group(1)

    id_match = re.search(r"[?&]id=([a-zA-Z0-9-_]+)", text)
    if id_match:
        return id_match.group(1)

    if re.fullmatch(r"[a-zA-Z0-9-_]{20,}", text):
        return text

    return None


def prompt_sales_spreadsheet(gc: gspread.Client) -> Tuple[str, gspread.Spreadsheet]:
    """
    Pide la URL (o ID) de la planilla de ventas y abre el spreadsheet.
    """
    while True:
        raw = input("IngresÃ¡ la URL de la planilla de ventas: ").strip()
        spreadsheet_id = extract_spreadsheet_id(raw)
        if spreadsheet_id is None:
            print("URL o ID invÃ¡lido. ProbÃ¡ nuevamente.")
            continue

        try:
            return spreadsheet_id, gc.open_by_key(spreadsheet_id)
        except Exception as exc:
            logger.warning("No pude abrir la planilla indicada: %s", exc)
            print("No pude abrir esa planilla. RevisÃ¡ la URL/ID y probÃ¡ nuevamente.")


def get_sales_date_from_worksheet(ws: gspread.Worksheet) -> date:
    """
    Lee la fecha objetivo desde la celda B1 de la hoja principal de ventas.
    """
    b1_raw = ws.acell("B1", value_render_option="UNFORMATTED_VALUE").value
    parsed = parse_date(b1_raw)
    if parsed is None:
        raise RuntimeError("No pude leer una fecha vÃ¡lida en Planilla!B1.")
    return parsed


def list_drive_children(
    drive_service,
    parent_id: str,
    mime_type: Optional[str] = None,
    fields: str = "files(id, name, mimeType, parents, modifiedTime)",
) -> List[Dict[str, Any]]:
    """Lista los hijos directos de una carpeta en Drive."""
    q_parts = [f"'{parent_id}' in parents", "trashed = false"]
    if mime_type:
        q_parts.append(f"mimeType = '{mime_type}'")
    return list_drive_files(drive_service, q=" and ".join(q_parts), fields=fields)


def normalize_folder_name(name: str) -> str:
    """Normaliza nombres de carpetas para comparar año/mes sin depender de mayúsculas o acentos."""
    return normalize_tag(name)


def find_drive_folder_by_candidates(
    drive_service,
    parent_id: str,
    candidates: Sequence[str],
) -> Optional[Dict[str, Any]]:
    """Busca una carpeta hija por nombre exacto o por coincidencia parcial."""
    folders = list_drive_children(
        drive_service,
        parent_id,
        mime_type="application/vnd.google-apps.folder",
    )
    if not folders:
        return None

    normalized_candidates = [normalize_folder_name(c) for c in candidates if normalize_text(c)]
    if not normalized_candidates:
        return None

    for candidate in normalized_candidates:
        for folder in folders:
            if normalize_folder_name(folder.get("name", "")) == candidate:
                return folder

    for candidate in normalized_candidates:
        for folder in folders:
            if candidate in normalize_folder_name(folder.get("name", "")):
                return folder

    return None


SPANISH_MONTH_NAMES = {
    1: "ENERO",
    2: "FEBRERO",
    3: "MARZO",
    4: "ABRIL",
    5: "MAYO",
    6: "JUNIO",
    7: "JULIO",
    8: "AGOSTO",
    9: "SEPTIEMBRE",
    10: "OCTUBRE",
    11: "NOVIEMBRE",
    12: "DICIEMBRE",
}


def year_folder_candidates(target_date: date) -> List[str]:
    return [f"{target_date:%Y}"]


def month_folder_candidates(target_date: date) -> List[str]:
    month_num = f"{target_date:%m}"
    month_plain = str(target_date.month)
    month_name = SPANISH_MONTH_NAMES[target_date.month]
    return [
        month_num,
        month_plain,
        f"{month_num} {month_name}",
        f"{month_num}-{month_name}",
        month_name,
    ]


# =============================================================================
# SALES SHEET
# =============================================================================

DATE_IN_NAME_RE = re.compile(r"(\d{1,2}[-/\.]\d{1,2}[-/\.]\d{4})")




def extract_date_from_name(name: str) -> Optional[date]:
    """
    Intenta extraer una fecha desde el nombre del archivo.
    Acepta formatos como:
    - 31/03/2026
    - 31-03-2026
    - 31.03.2026
    - 2026-03-31
    """
    basename = Path(name).stem
    candidates = re.findall(r"\d{1,4}[-/\.]\d{1,2}[-/\.]\d{1,4}", basename)
    for raw in candidates:
        parsed = parse_date(raw)
        if parsed is not None:
            return parsed
    return None






def find_sales_spreadsheet(
    gc: gspread.Client,
    drive_service,
    target_date: Optional[date] = None,
) -> Tuple[Dict[str, Any], date]:
    """
    Busca la planilla de ventas más adecuada.

    Con target_date, prioriza la carpeta:
    carpeta principal / año / mes
    y luego el archivo cuya fecha coincida.
    """
    root_id = SALES_DRIVE_FOLDER_ID
    scoped_parent_id = root_id

    if target_date is not None:
        year_folder = find_drive_folder_by_candidates(
            drive_service,
            root_id,
            year_folder_candidates(target_date),
        )
        if year_folder is not None:
            scoped_parent_id = year_folder["id"]

            month_folder = find_drive_folder_by_candidates(
                drive_service,
                scoped_parent_id,
                month_folder_candidates(target_date),
            )
            if month_folder is not None:
                scoped_parent_id = month_folder["id"]

    q = (
        f"'{scoped_parent_id}' in parents and "
        f"mimeType = 'application/vnd.google-apps.spreadsheet' and trashed = false"
    )
    candidates = list_drive_files(drive_service, q=q, fields="files(id, name, modifiedTime)")

    # Fallback: si la carpeta año/mes no devolvió nada, buscar en toda la raíz.
    if not candidates and scoped_parent_id != root_id:
        q = (
            f"'{root_id}' in parents and "
            f"mimeType = 'application/vnd.google-apps.spreadsheet' and trashed = false"
        )
        candidates = list_drive_files(drive_service, q=q, fields="files(id, name, modifiedTime)")

    if not candidates:
        raise FileNotFoundError("No encontré planillas de ventas dentro de la carpeta configurada.")

    scored: List[Tuple[int, str, Dict[str, Any], Optional[date], Optional[date]]] = []
    last_error: Optional[Exception] = None

    for file_data in candidates:
        try:
            normalized_name = normalize_tag(file_data["name"])
            if "PLANILLA" not in normalized_name or "VENTAS" not in normalized_name:
                continue

            name_date = extract_date_from_name(file_data["name"])
            sheet_date: Optional[date] = None

            sh = gc.open_by_key(file_data["id"])
            ws = get_sales_worksheet(sh)
            b1_raw = ws.acell("B1", value_render_option="UNFORMATTED_VALUE").value
            sheet_date = parse_date(b1_raw)

            if target_date is not None:
                if sheet_date != target_date and name_date != target_date:
                    continue

                if sheet_date == target_date and name_date == target_date:
                    score = 120
                elif sheet_date == target_date:
                    score = 110
                else:
                    score = 100
            else:
                score = 0
                if name_date and sheet_date and name_date == sheet_date:
                    score = 100
                elif sheet_date is not None:
                    score = 80
                elif name_date is not None:
                    score = 60

            if score > 0:
                scored.append((score, file_data.get("modifiedTime", ""), file_data, sheet_date, name_date))
        except Exception as exc:
            last_error = exc
            continue

    if not scored:
        raise RuntimeError(f"No pude ubicar una planilla de ventas válida. Último error: {last_error!r}")

    scored.sort(key=lambda item: (item[0], item[1]), reverse=True)
    _, _, selected_file, sheet_date, name_date = scored[0]

    selected_date = sheet_date or name_date or target_date
    if not selected_date:
        raise RuntimeError("Encontré una planilla candidata, pero no pude extraer la fecha.")

    logger.info("Planilla candidata seleccionada: %s", selected_file["name"])
    logger.info("Nombre esperado aproximado: Planilla Ventas - %s", format_date_ddmmyyyy(selected_date))

    return selected_file, selected_date


def get_sales_rows(ws: gspread.Worksheet) -> List[SalesRow]:
    values = ws.get_all_values()
    rows: List[SalesRow] = []
    last_row = min(len(values), SALES_END_ROW)

    for row_number in range(SALES_START_ROW, last_row + 1):
        raw = values[row_number - 1]
        t_value = normalize_tag(cell_value(raw, COL_T))
        if t_value == "" or t_value == "SENA":
            continue

        remito_raw = normalize_text(cell_value(raw, COL_REMITO))
        factura_raw = normalize_text(cell_value(raw, COL_FACTURA_U))
        remito = normalize_invoice(remito_raw)
        factura = normalize_invoice(factura_raw)

        valor = parse_money(cell_value(raw, COL_VALOR))
        efectivo = parse_money(cell_value(raw, COL_EFECTIVO))
        transferencia = parse_money(cell_value(raw, COL_TRANSFER))
        posnet = parse_money(cell_value(raw, COL_POSNET))
        usd = parse_money(cell_value(raw, COL_USD))

        row = SalesRow(
            row_number=row_number,
            raw_values=list(raw),
            remito_raw=remito_raw,
            factura_raw=factura_raw,
            remito=remito,
            factura=factura,
            valor=valor,
            efectivo=efectivo,
            transferencia=transferencia,
            posnet=posnet,
            usd=usd,
            t_flag=t_value,
        )

        if has_business_content(raw):
            rows.append(row)

    return rows





def build_invoice_groups(
    sales_rows: List[SalesRow],
) -> Tuple[Dict[int, InvoiceGroup], List[Tuple[int, int, str]]]:
    groups: Dict[int, InvoiceGroup] = {}
    missing_updates: List[Tuple[int, int, str]] = []

    for row in sales_rows:
        if row.factura is None:
            if MARK_BLANK_INVOICE_CELLS:
                missing_updates.append((row.row_number, COL_FACTURA_U, "FALTA FACTURA"))
            continue

        group = groups.get(row.factura)
        if group is None:
            group = InvoiceGroup(invoice=row.factura)
            groups[row.factura] = group

        group.add_row(row)

    return groups, missing_updates


def detect_remito_conflicts(
    groups: Dict[int, InvoiceGroup],
) -> Tuple[List[Tuple[int, int, Dict[str, float]]], List[Tuple[str, List[int], List[int]]]]:
    """
    Detecta conflictos de remito:
    - remito repetido en distintas facturas
    - una factura con más de un remito distinto

    Devuelve acciones de color para la columna A y mensajes para el log.
    """
    color_actions: List[Tuple[int, int, Dict[str, float]]] = []
    conflicts: List[Tuple[str, List[int], List[int]]] = []

    remito_map: Dict[int, List[Tuple[int, int]]] = {}

    for group in groups.values():
        unique_remitos = sorted({r for r in group.remitos if r is not None})

        if len(unique_remitos) > 1:
            conflicts.append(
                (
                    f"Factura {group.invoice} con múltiples remitos {unique_remitos}",
                    list(group.row_numbers),
                    [group.invoice],
                )
            )
            for row_number in group.row_numbers:
                color_actions.append((row_number, COL_REMITO, COLOR_PURPLE))

        for row in group.rows:
            if row.remito is None:
                continue
            remito_map.setdefault(row.remito, []).append((group.invoice, row.row_number))

    for remito, entries in remito_map.items():
        invoices = sorted({invoice for invoice, _ in entries})
        if len(invoices) <= 1:
            continue

        rows = sorted({row_number for _, row_number in entries})
        conflicts.append(
            (
                f"Remito {remito} repetido en facturas {invoices}",
                rows,
                invoices,
            )
        )
        for row_number in rows:
            color_actions.append((row_number, COL_REMITO, COLOR_PURPLE))

    deduped: Dict[Tuple[int, int], Dict[str, float]] = {}
    for row_number, col_number, color in color_actions:
        deduped[(row_number, col_number)] = color

    return [(row, col, color) for (row, col), color in deduped.items()], conflicts



# =============================================================================
# ARCA LOCAL
# =============================================================================


def find_local_arca_files() -> List[Path]:
    if not ARCA_DIR.exists():
        raise FileNotFoundError(f"No existe la carpeta local de ARCA: {ARCA_DIR}")

    files = [p for p in ARCA_DIR.glob("*.xlsx") if not p.name.startswith("~$")]
    files += [p for p in ARCA_DIR.glob("*.csv") if not p.name.startswith("~$")]

    if not files:
        raise FileNotFoundError(f"No encontré ningún .xlsx o .csv en {ARCA_DIR}")

    files.sort(key=lambda p: p.stat().st_mtime, reverse=True)
    return files



def read_csv_rows(path: Path) -> Tuple[List[Any], List[List[Any]]]:
    encodings = ("utf-8-sig", "cp1252", "latin1")
    last_error: Optional[Exception] = None

    for enc in encodings:
        try:
            with path.open("r", encoding=enc, newline="") as f:
                sample = f.read(4096)
                f.seek(0)

                try:
                    dialect = csv.Sniffer().sniff(sample, delimiters=";\t,")
                except Exception:
                    dialect = csv.excel
                    dialect.delimiter = ";"

                rows = list(csv.reader(f, dialect))
                if not rows:
                    return [], []
                return rows[0], rows[1:]
        except Exception as exc:
            last_error = exc
            continue

    raise RuntimeError(f"No pude leer el CSV {path.name}. Último error: {last_error!r}")



def read_xlsx_rows(path: Path) -> Tuple[List[Any], List[List[Any]]]:
    wb = load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], []
    return list(rows[0]), [list(r) for r in rows[1:]]



def find_column_by_candidates(headers: Sequence[Any], candidates: Sequence[str]) -> Optional[int]:
    normalized_headers = [normalize_header(h) for h in headers]
    normalized_candidates = [normalize_header(c) for c in candidates]

    for candidate in normalized_candidates:
        for idx, header in enumerate(normalized_headers, start=1):
            if header == candidate:
                return idx

    for candidate in normalized_candidates:
        for idx, header in enumerate(normalized_headers, start=1):
            if candidate and candidate in header:
                return idx

    return None



def load_arca_table_from_file(path: Path) -> ArcaTable:
    headers, data_rows = read_csv_rows(path) if path.suffix.lower() == ".csv" else read_xlsx_rows(path)
    if not headers:
        raise ValueError(f"El archivo ARCA {path.name} no tiene encabezados.")

    invoice_col = find_column_by_candidates(
        headers,
        ["cbte desde", "comprobante desde", "desde", "factura", "numero de comprobante", "nro comprobante"],
    ) or ARCA_INVOICE_COL

    date_col = find_column_by_candidates(headers, ["fecha", "fecha comprobante", "fecha desde"]) or ARCA_DATE_COL

    total_col = find_column_by_candidates(
        headers,
        ["imp. total", "imp total", "importe total", "total", "importe", "monto total", "neto", "importe neto"],
    )
    if total_col is None and len(headers) >= ARCA_TOTAL_DEFAULT_COL:
        total_col = ARCA_TOTAL_DEFAULT_COL

    output_indices = [1, 2, 4, 9, 28]
    output_header: List[Any] = []
    for idx, fallback_name in zip(output_indices, ["A", "B", "D", "I", "AB"]):
        if idx <= len(headers) and normalize_text(headers[idx - 1]):
            output_header.append(headers[idx - 1])
        else:
            output_header.append(fallback_name)

    records: List[ArcaRecord] = []

    for i, row in enumerate(data_rows, start=2):
        if not any(normalize_text(v) for v in row):
            continue

        raw_invoice = cell_value(row, invoice_col)
        factura_desde = normalize_invoice(raw_invoice)
        if factura_desde is None:
            continue

        raw_date = cell_value(row, date_col)
        raw_total = cell_value(row, total_col) if total_col else None

        records.append(
            ArcaRecord(
                source_file=path.name,
                source_row=i,
                fecha=parse_date(raw_date),
                factura_desde=factura_desde,
                importe=parse_money(raw_total),
                output_values=[cell_value(row, idx) for idx in output_indices],
            )
        )

    logger.info("ARCA cargado: %s | registros=%s", path.name, len(records))
    return ArcaTable(list(headers), records, invoice_col, total_col, date_col, output_header)



def load_all_arca_records() -> ArcaTable:
    files = find_local_arca_files()
    all_records: List[ArcaRecord] = []
    table_template: Optional[ArcaTable] = None

    for path in files:
        table = load_arca_table_from_file(path)
        all_records.extend(table.records)
        if table_template is None:
            table_template = table

    if table_template is None:
        raise RuntimeError("No pude construir la tabla ARCA.")

    return ArcaTable(
        headers=table_template.headers,
        records=all_records,
        invoice_col=table_template.invoice_col,
        total_col=table_template.total_col,
        date_col=table_template.date_col,
        output_header=table_template.output_header,
    )



def build_arca_lookup(records: Sequence[ArcaRecord], target_date: Optional[date]) -> Dict[int, List[ArcaRecord]]:
    lookup: Dict[int, List[ArcaRecord]] = {}
    for rec in records:
        if target_date is not None and rec.fecha != target_date:
            continue
        if rec.factura_desde is None:
            continue
        lookup.setdefault(rec.factura_desde, []).append(rec)
    return lookup



def select_best_arca_record(candidates: List[ArcaRecord], expected_amount: Decimal) -> Optional[ArcaRecord]:
    if not candidates:
        return None

    if len(candidates) == 1:
        return candidates[0]

    ranked = sorted(
        candidates,
        key=lambda r: (
            abs(r.importe - expected_amount),
            r.source_file,
            r.source_row,
        ),
    )
    return ranked[0]


# =============================================================================
# VALIDACIÓN
# =============================================================================


def calculate_usd_adjustment(total_j: Decimal, efectivo: Decimal, transferencia: Decimal, usd_pesos: Decimal) -> Tuple[Decimal, Decimal, Decimal]:
    """
    Corrige el excedente generado por USD pagado de a 1000.
    Devuelve: usd_ajustado, exceso_absorbido, exceso_restante
    """
    base_without_posnet = efectivo + transferencia + usd_pesos
    excess = base_without_posnet - total_j
    if excess <= 0:
        return usd_pesos, Decimal("0.00"), Decimal("0.00")

    reduction = min(usd_pesos, excess)
    adjusted = usd_pesos - reduction
    leftover = excess - reduction
    return adjusted, reduction, leftover


def pick_severity(issues: List[ValidationIssue]) -> str:
    if not issues:
        return "ok"

    order = {"warn": 1, "inconsistent": 2, "missing": 3, "error": 4}
    highest = max(issues, key=lambda i: order.get(i.severity, 0)).severity
    return highest


def severity_to_color(status: str) -> Optional[Dict[str, float]]:
    if status == "ok":
        return COLOR_GREEN
    if status == "warn":
        return COLOR_YELLOW
    if status == "inconsistent":
        return COLOR_ORANGE
    if status in ("error", "missing"):
        return COLOR_RED
    return None


def validate_invoice_group(
    group: InvoiceGroup,
    arca_lookup_date: Dict[int, List[ArcaRecord]],
    arca_lookup_all: Dict[int, List[ArcaRecord]],
    cotizacion_e1: Decimal,
) -> ValidationResult:
    """
    Valida una factura completa sumando los medios de pago de todas las filas.

    Estrategia:
    - recorrer cada fila con la misma factura;
    - sumar las columnas de pago, tratando vacíos como cero;
    - convertir USD a pesos con E1;
    - absorber el excedente de USD automáticamente;
    - comparar con ARCA.
    """
    issues: List[ValidationIssue] = []

    total_j = money_quantize(group.total_j)

    total_efectivo = money_quantize(sum((row.efectivo for row in group.rows if row.efectivo > 0), Decimal("0.00")))
    total_transferencia = money_quantize(sum((row.transferencia for row in group.rows if row.transferencia > 0), Decimal("0.00")))
    total_posnet = money_quantize(sum((row.posnet for row in group.rows if row.posnet > 0), Decimal("0.00")))
    total_usd = money_quantize(sum((row.usd for row in group.rows if row.usd > 0), Decimal("0.00")))

    payment_rows = [row for row in group.rows if row.has_payment()]

    unique_remitos = sorted({r for r in group.remitos if r is not None})
    if len(unique_remitos) > 1:
        issues.append(
            ValidationIssue(
                "error",
                f"FACTURA CON MÚLTIPLES REMITOS | remitos={unique_remitos}",
            )
        )

    for row in payment_rows:
        logger.info(
            "Factura %s | fila %s | C=%s | D=%s | E=%s | F=%s | J=%s",
            group.invoice,
            row.row_number,
            fmt_money(row.efectivo),
            fmt_money(row.transferencia),
            fmt_money(row.posnet),
            fmt_money(row.usd),
            fmt_money(row.valor),
        )

    if total_j <= 0:
        issues.append(ValidationIssue("error", "TOTAL J INVÁLIDO O CERO"))

    if not payment_rows:
        issues.append(ValidationIssue("error", "SIN MEDIO DE PAGO"))

    if total_usd > 0 and cotizacion_e1 <= 0:
        issues.append(ValidationIssue("error", "E1 INVÁLIDA CON USD PRESENTE"))

    usd_pesos = money_quantize(total_usd * cotizacion_e1) if total_usd > 0 else Decimal("0.00")
    usd_ajustado = usd_pesos
    usd_absorbido = Decimal("0.00")
    usd_restante = Decimal("0.00")

    if total_usd > 0:
        usd_ajustado, usd_absorbido, usd_restante = calculate_usd_adjustment(
            total_j,
            total_efectivo,
            total_transferencia,
            usd_pesos,
        )
        if usd_restante > 0:
            issues.append(
                ValidationIssue(
                    "error",
                    f"EXCESO NO ABSORBIDO POR USD | exceso_restante={fmt_money(usd_restante)}",
                )
            )

    base_real = money_quantize(total_efectivo + total_transferencia + usd_ajustado)

    arca_candidates = arca_lookup_date.get(group.invoice, [])
    used_fallback_date = False
    if not arca_candidates:
        arca_candidates = arca_lookup_all.get(group.invoice, [])
        if arca_candidates:
            used_fallback_date = True
            issues.append(ValidationIssue("warn", "ARCA ENCONTRADO FUERA DE LA FECHA OBJETIVO"))

    if not arca_candidates:
        issues.append(ValidationIssue("error", f"FACTURA NO ENCONTRADA EN ARCA | factura={group.invoice}"))
        status = pick_severity(issues)
        note = build_group_note(
            group,
            total_j,
            total_efectivo,
            total_transferencia,
            total_posnet,
            total_usd,
            usd_pesos,
            usd_ajustado,
            base_real,
            None,
            issues,
            used_fallback_date=used_fallback_date,
        )
        return ValidationResult(status=status, color=severity_to_color(status), note=note, arca_record=None)

    expected_for_selection = base_real
    if total_posnet > 0:
        expected_for_selection = money_quantize(base_real + total_posnet * POSNET_SELECTION_FACTOR)

    arca_record = select_best_arca_record(arca_candidates, expected_for_selection)
    if arca_record is None:
        issues.append(ValidationIssue("error", f"FACTURA NO ENCONTRADA EN ARCA | factura={group.invoice}"))
        status = pick_severity(issues)
        note = build_group_note(
            group,
            total_j,
            total_efectivo,
            total_transferencia,
            total_posnet,
            total_usd,
            usd_pesos,
            usd_ajustado,
            base_real,
            None,
            issues,
            used_fallback_date=used_fallback_date,
        )
        return ValidationResult(status=status, color=severity_to_color(status), note=note, arca_record=None)

    arca_value = money_quantize(arca_record.importe)

    if total_posnet > 0:
        min_expected = money_quantize(base_real + total_posnet * POSNET_MIN_FACTOR)
        max_expected = money_quantize(base_real + total_posnet * POSNET_MAX_FACTOR)
        if not (min_expected - ABS_TOLERANCE_POSNET_BOUNDARY <= arca_value <= max_expected + ABS_TOLERANCE_POSNET_BOUNDARY):
            issues.append(
                ValidationIssue(
                    "error",
                    f"ARCA FUERA DE RANGO POSNET | esperado={fmt_money(min_expected)}-{fmt_money(max_expected)} | arca={fmt_money(arca_value)}",
                )
            )
    else:
        tolerance = max(ABS_TOLERANCE_NO_POSNET, money_quantize(total_j * Decimal("0.01")))
        if abs(arca_value - base_real) > tolerance:
            issues.append(
                ValidationIssue(
                    "error",
                    f"ARCA DIF | esperado={fmt_money(base_real)} ± {fmt_money(tolerance)} | arca={fmt_money(arca_value)}",
                )
            )

    status = pick_severity(issues)
    note = build_group_note(
        group,
        total_j,
        total_efectivo,
        total_transferencia,
        total_posnet,
        total_usd,
        usd_pesos,
        usd_ajustado,
        base_real,
        arca_record,
        issues,
        used_fallback_date=used_fallback_date,
    )
    return ValidationResult(status=status, color=severity_to_color(status), note=note, arca_record=arca_record)



def build_group_note(
    group: InvoiceGroup,
    total_j: Decimal,
    efectivo: Decimal,
    transferencia: Decimal,
    posnet: Decimal,
    usd: Decimal,
    usd_pesos: Decimal,
    usd_ajustado: Decimal,
    base_real: Decimal,
    arca_record: Optional[ArcaRecord],
    issues: List[ValidationIssue],
    used_fallback_date: bool = False,
) -> str:
    parts = [
        f"Factura {group.invoice}",
        f"filas={group.row_numbers}",
        f"remitos={sorted(set(group.remitos))}" if group.remitos else "remitos=[]",
        f"J_total={fmt_money(total_j)}",
        f"efectivo={fmt_money(efectivo)}",
        f"transferencia={fmt_money(transferencia)}",
        f"posnet={fmt_money(posnet)}",
        f"usd={fmt_money(usd)}",
        f"usd_pesos={fmt_money(usd_pesos)}",
        f"usd_ajustado={fmt_money(usd_ajustado)}",
        f"base_real={fmt_money(base_real)}",
    ]

    if arca_record is not None:
        parts.append(f"ARCA={arca_record.source_file}:{arca_record.source_row}")
        parts.append(f"arca={fmt_money(arca_record.importe)}")

    if used_fallback_date:
        parts.append("ARCA_FUERA_DE_FECHA")

    for issue in issues:
        parts.append(issue.message)

    return " | ".join(parts)


# =============================================================================
# GOOGLE SHEETS: COLORES Y VALORES
# =============================================================================


def build_repeat_cell_request(sheet_id: int, row_number: int, col_number: int, color: Dict[str, float]) -> Dict[str, Any]:
    return {
        "repeatCell": {
            "range": {
                "sheetId": sheet_id,
                "startRowIndex": row_number - 1,
                "endRowIndex": row_number,
                "startColumnIndex": col_number - 1,
                "endColumnIndex": col_number,
            },
            "cell": {"userEnteredFormat": {"backgroundColor": color}},
            "fields": "userEnteredFormat.backgroundColor",
        }
    }



def apply_cell_colors(
    sheets_service,
    spreadsheet_id: str,
    worksheet_id: int,
    actions: List[Tuple[int, int, Dict[str, float]]],
) -> None:
    if not actions:
        return

    requests = [build_repeat_cell_request(worksheet_id, row, col, color) for row, col, color in actions]
    sheets_service.spreadsheets().batchUpdate(spreadsheetId=spreadsheet_id, body={"requests": requests}).execute()


def get_protected_ranges(
    sheets_service,
    spreadsheet_id: str,
    worksheet_id: int,
) -> List[Dict[str, Any]]:
    response = sheets_service.spreadsheets().get(
        spreadsheetId=spreadsheet_id,
        fields="sheets(properties(sheetId),protectedRanges(range,warningOnly))",
    ).execute()

    for sheet in response.get("sheets", []):
        properties = sheet.get("properties", {})
        if properties.get("sheetId") != worksheet_id:
            continue

        protected_ranges: List[Dict[str, Any]] = []
        for protected_range in sheet.get("protectedRanges", []):
            if protected_range.get("warningOnly"):
                continue

            grid_range = protected_range.get("range") or {}
            if grid_range.get("sheetId", worksheet_id) != worksheet_id:
                continue
            protected_ranges.append(grid_range)

        return protected_ranges

    return []


def is_cell_in_grid_range(row_number: int, col_number: int, grid_range: Dict[str, Any]) -> bool:
    row_index = row_number - 1
    col_index = col_number - 1

    start_row = grid_range.get("startRowIndex", 0)
    end_row = grid_range.get("endRowIndex")
    start_col = grid_range.get("startColumnIndex", 0)
    end_col = grid_range.get("endColumnIndex")

    if row_index < start_row:
        return False
    if end_row is not None and row_index >= end_row:
        return False
    if col_index < start_col:
        return False
    if end_col is not None and col_index >= end_col:
        return False
    return True


def is_protected_cell(row_number: int, col_number: int, protected_ranges: Sequence[Dict[str, Any]]) -> bool:
    return any(is_cell_in_grid_range(row_number, col_number, grid_range) for grid_range in protected_ranges)


def filter_missing_updates_for_protection(
    missing_updates: List[Tuple[int, int, str]],
    protected_ranges: Sequence[Dict[str, Any]],
) -> Tuple[List[Tuple[int, int, str]], int]:
    allowed: List[Tuple[int, int, str]] = []
    skipped = 0

    for row_number, col_number, value in missing_updates:
        if is_protected_cell(row_number, col_number, protected_ranges):
            skipped += 1
            continue
        allowed.append((row_number, col_number, value))

    return allowed, skipped


def filter_color_actions_for_protection(
    color_actions: List[Tuple[int, int, Dict[str, float]]],
    protected_ranges: Sequence[Dict[str, Any]],
) -> Tuple[List[Tuple[int, int, Dict[str, float]]], int]:
    allowed: List[Tuple[int, int, Dict[str, float]]] = []
    skipped = 0

    for row_number, col_number, color in color_actions:
        if is_protected_cell(row_number, col_number, protected_ranges):
            skipped += 1
            continue
        allowed.append((row_number, col_number, color))

    return allowed, skipped



def apply_missing_values(ws: gspread.Worksheet, missing_updates: List[Tuple[int, int, str]]) -> None:
    if not missing_updates:
        return

    data = []
    for row_number, col_number, value in missing_updates:
        data.append({"range": f"{column_letter(col_number)}{row_number}", "values": [[value]]})

    ws.batch_update(data, value_input_option="USER_ENTERED")


# =============================================================================
# SALIDA LIMPIA EN DRIVE
# =============================================================================


def build_clean_rows_for_output(table: ArcaTable, target_date: Optional[date]) -> List[List[Any]]:
    rows: List[List[Any]] = [table.output_header]
    for rec in table.records:
        if target_date is None or rec.fecha == target_date:
            rows.append(rec.output_values)
    return rows


def detect_account_cuit_from_source_file(source_file: str) -> Optional[str]:
    normalized_name = normalize_text(source_file)
    for cuit in ACCOUNT_CONFIG_BY_CUIT:
        if cuit in normalized_name:
            return cuit
    return None


def build_clean_rows_for_output_by_account(
    table: ArcaTable,
    target_date: Optional[date],
) -> Tuple[Dict[str, List[List[Any]]], List[str]]:
    rows_by_account: Dict[str, List[List[Any]]] = {}
    unknown_sources: set[str] = set()

    for rec in table.records:
        if target_date is not None and rec.fecha != target_date:
            continue

        account_cuit = detect_account_cuit_from_source_file(rec.source_file)
        if account_cuit is None:
            unknown_sources.add(rec.source_file)
            continue

        if account_cuit not in rows_by_account:
            rows_by_account[account_cuit] = [table.output_header]
        rows_by_account[account_cuit].append(rec.output_values)

    return rows_by_account, sorted(unknown_sources)



def create_or_update_output_sheet(
    gc: gspread.Client,
    drive_service,
    output_root_folder_id: str,
    target_date: date,
    rows: List[List[Any]],
) -> Tuple[str, str, str, str]:
    year_name = f"{target_date:%Y}"
    month_name = f"{target_date:%m}"
    day_name = format_date_ddmmyyyy(target_date)

    year_folder_id = drive_ensure_folder(drive_service, year_name, output_root_folder_id)
    month_folder_id = drive_ensure_folder(drive_service, month_name, year_folder_id)

    spreadsheet_title = f"Facturas_{day_name}"
    existing = drive_find_exact(
        drive_service,
        spreadsheet_title,
        parent_id=month_folder_id,
        mime_type="application/vnd.google-apps.spreadsheet",
    )
    spreadsheet_id = existing["id"] if existing else drive_create_spreadsheet(drive_service, spreadsheet_title, month_folder_id)

    sh = gc.open_by_key(spreadsheet_id)
    ws = sh.sheet1
    ws.clear()
    ws.update(values=rows, range_name="A1")

    try:
        ws.freeze(rows=1)
    except Exception:
        pass

    return spreadsheet_id, spreadsheet_title, year_folder_id, month_folder_id

# =============================================================================
# FLUJO PRINCIPAL
# =============================================================================



def process() -> None:
    gc, drive_service, sheets_service = get_google_clients()

    sales_spreadsheet_id, sales_sh = prompt_sales_spreadsheet(gc)
    sales_ws = get_sales_worksheet(sales_sh)
    sales_date = get_sales_date_from_worksheet(sales_ws)

    logger.info("Planilla de ventas indicada por URL: %s", sales_sh.title)
    logger.info("Fecha detectada en Planilla!B1: %s", format_date_ddmmyyyy(sales_date))

    cotizacion_raw = sales_ws.acell("E1", value_render_option="UNFORMATTED_VALUE").value
    cotizacion_e1 = parse_money(cotizacion_raw)
    if cotizacion_e1 <= 0:
        raise RuntimeError("Cotización E1 inválida o vacía.")
    logger.info("Cotización E1 detectada: %s", fmt_money(cotizacion_e1))

    sales_rows = get_sales_rows(sales_ws)
    logger.info("Filas candidatas de ventas: %s", len(sales_rows))

    groups, missing_updates = build_invoice_groups(sales_rows)
    logger.info("Facturas detectadas: %s", len(groups))

    arca_table = load_all_arca_records()
    logger.info("ARCA total cargado | registros=%s", len(arca_table.records))

    arca_lookup_date = build_arca_lookup(arca_table.records, sales_date)
    arca_lookup_all: Dict[int, List[ArcaRecord]] = {}
    for rec in arca_table.records:
        if rec.factura_desde is None:
            continue
        arca_lookup_all.setdefault(rec.factura_desde, []).append(rec)

    if not arca_lookup_date:
        logger.warning(
            "No encontré registros de ARCA para la fecha %s. Se usará búsqueda global por factura.",
            format_date_ddmmyyyy(sales_date),
        )

    # Detectar conflictos de remito:
    # - remito repetido en distintas facturas
    # - factura con más de un remito distinto
    remito_color_actions, remito_conflicts = detect_remito_conflicts(groups)
    for message, rows, invoices in remito_conflicts:
        log_console_error_block(
            "REMITO EN CONFLICTO",
            f"{message}. Filas involucradas: {rows}. Facturas: {invoices}. "
            f"Se marcó la columna A en violeta.",
        )

    color_actions: List[Tuple[int, int, Dict[str, float]]] = list(remito_color_actions)
    summary = {"ok": 0, "warn": 0, "inconsistent": 0, "error": 0, "missing": 0}

    for invoice in sorted(groups.keys()):
        group = groups[invoice]
        result = validate_invoice_group(group, arca_lookup_date, arca_lookup_all, cotizacion_e1)
        summary[result.status] = summary.get(result.status, 0) + 1

        if result.color is not None:
            for row_number in group.row_numbers:
                color_actions.append((row_number, COL_FACTURA_U, result.color))

        if result.status == "warn":
            logger.warning(result.note)
        elif result.status != "ok":
            log_console_error_block(
                f"FACTURA {group.invoice} [{result.status.upper()}]",
                result.note,
            )

    protected_ranges: List[Dict[str, Any]] = []
    try:
        protected_ranges = get_protected_ranges(sheets_service, sales_spreadsheet_id, sales_ws.id)
    except Exception as exc:
        logger.warning("No pude leer los rangos protegidos de la hoja: %s", exc)

    if protected_ranges:
        missing_updates, skipped_missing_updates = filter_missing_updates_for_protection(
            missing_updates,
            protected_ranges,
        )
        color_actions, skipped_color_actions = filter_color_actions_for_protection(
            color_actions,
            protected_ranges,
        )

        if skipped_missing_updates:
            logger.info("Se omitieron %s escrituras en celdas protegidas.", skipped_missing_updates)
        if skipped_color_actions:
            logger.info("Se omitieron %s colores en celdas protegidas.", skipped_color_actions)

    try:
        apply_missing_values(sales_ws, missing_updates)
    except Exception as exc:
        logger.error("No pude escribir las celdas faltantes porque están protegidas: %s", exc)

    try:
        apply_cell_colors(sheets_service, sales_spreadsheet_id, sales_ws.id, color_actions)
    except Exception as exc:
        logger.error("No pude aplicar colores en la hoja: %s", exc)

    output_rows_by_account, unknown_output_sources = build_clean_rows_for_output_by_account(arca_table, sales_date)
    created_outputs: List[Tuple[str, str, str, str, str, str]] = []

    for account_cuit, rows in sorted(output_rows_by_account.items()):
        if len(rows) <= 1:
            continue

        account_config = ACCOUNT_CONFIG_BY_CUIT.get(account_cuit)
        if account_config is None:
            logger.warning("No encontrÃ© configuraciÃ³n de salida para el CUIT %s.", account_cuit)
            continue

        account_name = normalize_text(account_config.get("name")) or account_cuit
        output_root_raw = normalize_text(account_config.get("output_root"))
        output_root_folder_id = extract_drive_folder_id(output_root_raw)
        if output_root_folder_id is None:
            logger.warning(
                "La carpeta de salida de %s (%s) no tiene una URL o ID vÃ¡lido: %s",
                account_name,
                account_cuit,
                output_root_raw or "<vacÃ­o>",
            )
            continue

        output_sheet_id, output_title, year_folder_id, month_folder_id = create_or_update_output_sheet(
            gc=gc,
            drive_service=drive_service,
            output_root_folder_id=output_root_folder_id,
            target_date=sales_date,
            rows=rows,
        )
        created_outputs.append(
            (account_name, output_sheet_id, output_title, output_root_folder_id, year_folder_id, month_folder_id)
        )

        logger.info("Sheet limpia creada/actualizada [%s]: %s", account_name, output_title)
        logger.info("Output ID [%s]: %s", account_name, output_sheet_id)
        logger.info(
            "Carpetas [%s]: root=%s | year=%s | month=%s",
            account_name,
            output_root_folder_id,
            year_folder_id,
            month_folder_id,
        )

    if unknown_output_sources:
        logger.warning(
            "No pude clasificar por cuenta estos archivos ARCA para exportar: %s",
            ", ".join(unknown_output_sources),
        )

    if not created_outputs:
        logger.warning(
            "No encontrÃ© registros ARCA exportables para la fecha %s y las cuentas configuradas.",
            format_date_ddmmyyyy(sales_date),
        )

    print(
        f"\nResumen final | OK={summary.get('ok', 0)} | WARN={summary.get('warn', 0)} | "
        f"INCONSISTENT={summary.get('inconsistent', 0)} | ERROR={summary.get('error', 0)} | "
        f"MISSING={summary.get('missing', 0)}"
    )

    if created_outputs:
        export_summary = " | ".join(
            f"{account_name}={output_title}"
            for account_name, _, output_title, _, _, _ in created_outputs
        )
        print(f"Exports creados | {export_summary}")


# =============================================================================
# MAIN
# =============================================================================

# =============================================================================


def main() -> None:
    try:
        process()
        print("Proceso terminado correctamente.")
    except Exception as exc:
        logger.exception("Error ejecutando el analizador")
        print(f"\nError: {exc}")
        raise


if __name__ == "__main__":
    main()
