from __future__ import annotations

import io
import re
import unicodedata
import uuid
from collections import defaultdict
from datetime import date, datetime, timedelta, timezone
from decimal import Decimal
from pathlib import Path
from typing import Any

import openpyxl
from sqlalchemy import and_, case, func, or_, select
from sqlalchemy.orm import Session, selectinload

from .commercial.matching import build_product_indexes, normalize_descripcion, resolve_product
from .db import db_session
from .google_sheets import sheets_service
from .models.auth import User
from .models.org import Branch
from .models.products import Product
from .models.sales_bi import SalesBalance, SalesBIProductAlias, SalesImport, SalesRecord
from .operational_config import extract_spreadsheet_id
from .product_catalog import sku_key


# ── helpers ──────────────────────────────────────────────────────────────────


def _norm(v: Any) -> str:
    if v is None:
        return ""
    s = unicodedata.normalize("NFKD", str(v))
    s = "".join(c for c in s if not unicodedata.combining(c))
    s = s.upper().strip()
    s = re.sub(r"[^A-Z0-9]+", " ", s)
    return re.sub(r"\s+", " ", s).strip()


def _parse_num(v: Any) -> float:
    if v is None:
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    s = re.sub(r"[$\s]", "", str(v).strip())
    if not s or s == "-":
        return 0.0
    if "," in s and "." in s:
        s = s.replace(".", "").replace(",", ".")
    elif "," in s:
        s = s.replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return 0.0


def _is_date_like(v: Any) -> bool:
    if isinstance(v, (datetime, date)):
        return True
    # Google Sheets serial date (days since 1899-12-30)
    if isinstance(v, (int, float)) and 40_000 < v < 60_000:
        return True
    if isinstance(v, str) and re.search(r"\d{1,2}[/\-]\d{1,2}[/\-]\d{2,4}", v):
        return True
    return False


def _parse_date(v: Any) -> str:
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, date):
        return v.isoformat()
    if isinstance(v, (int, float)) and 40_000 < v < 60_000:
        return (date(1899, 12, 30) + timedelta(days=int(v))).isoformat()
    s = str(v).strip()
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%d/%m/%y", "%d-%m-%y"):
        try:
            return datetime.strptime(s, fmt).date().isoformat()
        except ValueError:
            pass
    return s


def utc_now() -> str:
    return datetime.now(timezone.utc).isoformat()


def utc_now_dt() -> datetime:
    return datetime.now(timezone.utc)


def _parse_date_value(value: Any) -> date | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value or "").strip()
    if not text:
        return None
    try:
        return date.fromisoformat(text[:10])
    except ValueError:
        return None


def _fmt_date(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return str(value)


def _fmt_dt(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.isoformat()
    return str(value)


def _decimal(value: Any) -> Decimal:
    try:
        if value is None or value == "":
            return Decimal("0")
        return Decimal(str(value))
    except Exception:
        return Decimal("0")


def _num(value: Any) -> float:
    if value is None:
        return 0.0
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, (int, float)):
        return float(value)
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0.0


# ── column aliases ────────────────────────────────────────────────────────────

_DATE_LABELS = frozenset(["DIA", "FECHA", "DATE", "DIA DE VENTA", "FECHA DE VENTA"])
_RATE_LABELS = frozenset(["TIPO DE CAMBIO", "COTIZACION", "COTIZACION DOLAR", "TC", "T C", "DOLAR"])
_SKIP_LABELS = frozenset(["$", "USD", "PESOS"])
_NON_SUCURSAL_LABELS = frozenset([
    "TOTAL", "TOTALES", "SUBTOTAL", "RESUMEN", "SHARE",
    "LINEAS", "LINEAS TOTALES", "CANTIDAD", "MONTO",
])

# Aliases written in the spreadsheets that map to canonical branch names used in the DB
_SUCURSAL_ALIASES: dict[str, str] = {
    "NORTE": "Norcenter",
    "SUR": "Lanus",
}


def _normalize_sucursal(name: str) -> str:
    """Map spreadsheet aliases (NORTE → Norcenter, SUR → Lanus) to canonical names."""
    return _SUCURSAL_ALIASES.get(name.upper().strip(), name.strip())

def _is_placeholder_sucursal(value: str, sheet_name: str) -> bool:
    normalized = _norm(value)
    if not normalized:
        return True
    return (
        normalized == _norm(sheet_name)
        or normalized in (_SHEET_NAME_LOCAL | _SHEET_NAME_ONLINE)
        or normalized in _NON_SUCURSAL_LABELS
    )


def _infer_sucursal_from_source_name(source_name: str) -> str:
    source_norm = _norm(Path(str(source_name or "")).stem)
    if not source_norm:
        return ""
    words = set(source_norm.split())
    for alias, canonical in _SUCURSAL_ALIASES.items():
        if _norm(alias) in words:
            return canonical
    for canonical in set(_SUCURSAL_ALIASES.values()):
        if _norm(canonical) in source_norm:
            return canonical
    return ""


def _infer_workbook_sucursal(sheets: dict[str, list[list]], source_name: str = "") -> str:
    for name, rows in sheets.items():
        if _classify_sheet_name(name) is None:
            continue
        meta = _detect_metadata(rows)
        sucursal = _normalize_sucursal(str(meta.get("sucursal") or "").strip())
        if sucursal and not _is_placeholder_sucursal(sucursal, name):
            return sucursal
    return _infer_sucursal_from_source_name(source_name)


_COLUMN_ALIASES: dict[str, list[str]] = {
    "remito": ["REMITO", "NRO REMITO", "N REMITO", "NUMERO REMITO", "N REMITO", "REM"],
    "pedido": ["PEDIDO", "NUMERO PEDIDO", "N PEDIDO", "NRO PEDIDO", "N ORDEN", "ORDEN"],
    "vendedor": ["VENDEDOR", "VEND"],
    "producto": ["PRODUCTO", "DESCRIPCION", "ARTICULO", "DESCRIPCION ARTICULO"],
    "sku": ["SKU", "CODIGO", "COD", "COD ARTICULO", "CODIGO ARTICULO"],
    "marca": ["MARCA"],
    "tipo": ["TIPO", "TIPO PRODUCTO", "RUBRO"],
    "condicion": ["CONDICION", "COND"],
    "cantidad": ["CANTIDAD", "CANT", "QTY", "CTD"],
    "pvp": ["PVP", "PRECIO", "PRECIO VENTA", "P VENTA", "PRECIO UNITARIO", "P UNITARIO", "IMPORTE", "VALOR", "MONTO"],
    "costo": ["COSTO", "COSTO VIGENTE", "COSTO VIG", "P COSTO"],
    "monto_ingresado": ["MONTO INGRESADO", "MONTO ING.", "MONTO ING", "INGRESADO", "PAGO INGRESADO"],
    "efectivo": ["EFECTIVO", "EFECT", "EFECT.", "EFT"],
    "transferencia": ["TRANSFERENCIA", "TRANSFER", "TRANSFER.", "TRANSF", "TRANSF.", "TRF"],
    "tarjeta": ["TARJETA", "TAR", "TAR.", "TC", "CREDITO", "DEBITO", "POSNET", "POS"],
    "usd": ["USD", "DOLARES", "U$S", "U S D"],
    "cuenta_corriente": ["CUENTA CORRIENTE", "CTA CTE", "CTA. CTE.", "CC", "CTA CORRIENTE"],
    "otros": ["OTROS", "OTRO", "OTROS MEDIOS"],
    "total": ["TOTAL", "IMPORTE TOTAL", "TOTAL VENTA", "TOTAL COBRADO", "MONTO"],
}

_ALIAS_TO_FIELD: dict[str, str] = {}
for _f, _al in _COLUMN_ALIASES.items():
    for _a in _al:
        _k = _norm(_a)
        if _k not in _ALIAS_TO_FIELD:
            _ALIAS_TO_FIELD[_k] = _f

# ── classifiers ───────────────────────────────────────────────────────────────

_PAYMENT_SUBHEADER_ALIASES: dict[str, str] = {
    "EFECTIVO": "efectivo",
    "EFECT": "efectivo",
    "EFECT.": "efectivo",
    "EFEC": "efectivo",
    "EFEC.": "efectivo",
    "EFT": "efectivo",
    "TRANSFERENCIA": "transferencia",
    "TRANSFER": "transferencia",
    "TRANSFER.": "transferencia",
    "TRANSF": "transferencia",
    "TRANSF.": "transferencia",
    "TRAN": "transferencia",
    "TRF": "transferencia",
    "TARJETA": "tarjeta",
    "TAR": "tarjeta",
    "TC": "tarjeta",
    "POSNET": "tarjeta",
    "POS": "tarjeta",
    "USD": "usd",
    "DOLARES": "usd",
    "U$S": "usd",
}
_PAYMENT_SUBHEADER_TO_FIELD = {_norm(k): v for k, v in _PAYMENT_SUBHEADER_ALIASES.items()}


def _parent_header_for_column(normed_header: list[str], column: int) -> str:
    for idx in range(min(column, len(normed_header) - 1), -1, -1):
        if normed_header[idx]:
            return normed_header[idx]
    return ""


def _is_sena_parent(parent: str) -> bool:
    compact = parent.replace(" ", "")
    return compact in {"SENA", "SENIA", "SEA"} or parent.startswith("SEN")


def _field_for_subheader(normed_header: list[str], column: int, normed_subheader: str) -> str | None:
    parent = _parent_header_for_column(normed_header, column)
    payment_field = _PAYMENT_SUBHEADER_TO_FIELD.get(normed_subheader)
    if payment_field and _is_sena_parent(parent):
        return f"sena_{payment_field}"
    return _ALIAS_TO_FIELD.get(normed_subheader) or payment_field


# ── Taxonomía de categorías (5 buckets) ────────────────────────────────────
# Reemplaza el esquema viejo de 3 niveles (Gran / Medio / Pequeño electro)
# por las 5 líneas comerciales reales que usa el negocio. Tipos no listados
# caen en "OTROS" (en vez de PEQUEÑO ELECTRO por descarte).
#
# Si agregás un tipo nuevo:
#   1. Sumarlo al frozenset correspondiente.
#   2. Si la planilla puede traer descripciones libres (no solo el tipo
#      canónico), agregar también el keyword en `_CATEGORIA_KEYWORDS`
#      para el fallback de texto suelto.
_CAT_LINEA_BLANCA = frozenset([
    "FREEZER", "HELADERA", "LAVARROPAS", "LAVASECARROPAS", "LAVAVAJILLAS",
    "SECARROPAS", "TORRE DE LAVADO",
])
_CAT_COCINA = frozenset([
    "ANAFE", "CAMPANA", "COCINA", "HORNO", "MICROONDAS",
])
_CAT_CLIMATIZACION = frozenset([
    "AIRE ACONDICIONADO", "CALEFON", "CALOVENTOR", "CONVECTOR", "PANEL",
    "PURIFICADOR", "TERMOTANQUE", "VENTILADOR",
])
_CAT_TV_AUDIO = frozenset([
    "MINICOMPONENTE", "MONITOR", "PARLANTE", "TV",
])
_CAT_PEQUENOS = frozenset([
    "ARROCERA", "ASPIRADORA", "BATIDORA", "CAFETERA", "CERVECERA",
    "CHOPPER", "ESPUMADOR", "EXPRIMIDOR", "EXTRACTOR", "FREIDORA",
    "JARRA", "LICUADORA", "LIMPIADOR ZAP", "MIXER", "MOLINO", "MOLINILLO",
    "MULTIOLLA", "MULTIPROCESADORA", "PAVA", "PICADORA", "PLANCHA",
    "PROCESADORA", "QUITAPELUSAS", "SANDWICHERA", "SOPERA", "TOSTADORA",
    "VAPORIZADOR", "YOGURTERA",
])

# Etiquetas finales que se guardan en `sales_records.categoria` y que el
# frontend usa para mostrar/filtrar. Mantener constantes y en MAYÚSCULAS.
CATEGORIA_LINEA_BLANCA = "LINEA BLANCA"
CATEGORIA_COCINA = "COCINA"
CATEGORIA_CLIMATIZACION = "CLIMATIZACION"
CATEGORIA_TV_AUDIO = "TV / AUDIO"
CATEGORIA_PEQUENOS = "PEQUENOS"
CATEGORIA_OTROS = "OTROS"

# Fallback por substring para descripciones libres (cuando la planilla no
# trae el tipo canónico exacto). El orden importa: el primero que matchee
# gana, así que las palabras más específicas van primero.
_CATEGORIA_KEYWORDS: tuple[tuple[str, str], ...] = (
    ("AIRE ACONDICIONADO", CATEGORIA_CLIMATIZACION),
    ("HELADERA", CATEGORIA_LINEA_BLANCA),
    ("LAVARROPAS", CATEGORIA_LINEA_BLANCA),
    ("LAVASECARROPAS", CATEGORIA_LINEA_BLANCA),
    ("LAVAVAJILLAS", CATEGORIA_LINEA_BLANCA),
    ("SECARROPAS", CATEGORIA_LINEA_BLANCA),
    ("FREEZER", CATEGORIA_LINEA_BLANCA),
    ("MICROONDAS", CATEGORIA_COCINA),
    ("COCINA", CATEGORIA_COCINA),
    ("HORNO", CATEGORIA_COCINA),
    ("ANAFE", CATEGORIA_COCINA),
    ("CAMPANA", CATEGORIA_COCINA),
    ("TERMOTANQUE", CATEGORIA_CLIMATIZACION),
    ("CALEFON", CATEGORIA_CLIMATIZACION),
    ("VENTILADOR", CATEGORIA_CLIMATIZACION),
    ("CALOVENTOR", CATEGORIA_CLIMATIZACION),
    ("CONVECTOR", CATEGORIA_CLIMATIZACION),
    ("PURIFICADOR", CATEGORIA_CLIMATIZACION),
    ("MONITOR", CATEGORIA_TV_AUDIO),
    ("PARLANTE", CATEGORIA_TV_AUDIO),
    ("MINICOMPONENTE", CATEGORIA_TV_AUDIO),
    (" TV ", CATEGORIA_TV_AUDIO),  # con espacios para evitar match contra "TVS" en producto
)


def _classify(tipo: str) -> tuple[str, str]:
    """Clasifica un tipo de producto en una de las 5 categorías comerciales.

    Returns:
        (categoria, linea) — por contrato histórico devolvemos un tuple.
        Hoy ambos campos contienen el mismo valor (la categoría de 5 buckets).
        El campo `linea` queda como alias por compatibilidad con código viejo
        que lo lee aparte. Podría dropear en un futuro refactor.
    """
    t = _norm(tipo)
    if t in _CAT_LINEA_BLANCA:
        categoria = CATEGORIA_LINEA_BLANCA
    elif t in _CAT_COCINA:
        categoria = CATEGORIA_COCINA
    elif t in _CAT_CLIMATIZACION:
        categoria = CATEGORIA_CLIMATIZACION
    elif t in _CAT_TV_AUDIO:
        categoria = CATEGORIA_TV_AUDIO
    elif t in _CAT_PEQUENOS:
        categoria = CATEGORIA_PEQUENOS
    elif not t:
        categoria = ""
    else:
        # Fallback: descripción libre. Buscar keywords (con padding para
        # que " TV " no matchee dentro de "TVS" u otros nombres comerciales).
        padded = f" {t} "
        categoria = CATEGORIA_OTROS
        for kw, target in _CATEGORIA_KEYWORDS:
            if kw.startswith(" "):
                if kw in padded:
                    categoria = target
                    break
            elif kw in t:
                categoria = target
                break

    # `linea` queda igual a `categoria` por compatibilidad — antes diferían
    # (categoria=GRAN/MEDIO/PEQUEÑO, linea=LINEA BLANCA/COCINA/...) pero
    # ahora la 5-bucket es la única taxonomía y la duplicamos para no
    # romper código que lee `record.linea`.
    return categoria, categoria


# ── Set público para validación y migraciones ──────────────────────────────
CATEGORIAS_VALIDAS: frozenset[str] = frozenset([
    CATEGORIA_LINEA_BLANCA,
    CATEGORIA_COCINA,
    CATEGORIA_CLIMATIZACION,
    CATEGORIA_TV_AUDIO,
    CATEGORIA_PEQUENOS,
    CATEGORIA_OTROS,
])


def reclassify_existing_records(*, dry_run: bool = False) -> dict[str, int]:
    """Re-clasifica `sales_records.categoria` y `.linea` aplicando la
    taxonomía actual sobre el `tipo_producto` ya almacenado.

    Pensada para correr una sola vez después de cambiar la taxonomía,
    sin tener que pedirle al usuario que re-importe planillas.

    Args:
        dry_run: si True solo cuenta cuántos cambiarían sin tocar la DB.

    Returns:
        {categoria_nueva: cantidad_de_records} + claves "updated" y "scanned".
    """
    from .models.sales_bi import SalesRecord  # local: evita ciclo con el __init__

    counts: dict[str, int] = {"scanned": 0, "updated": 0}
    with db_session() as session:
        for record in session.scalars(select(SalesRecord)).all():
            counts["scanned"] += 1
            new_cat, new_linea = _classify(str(record.tipo_producto or ""))
            current_cat = str(record.categoria or "")
            current_linea = str(record.linea or "")
            if current_cat == new_cat and current_linea == new_linea:
                continue
            counts["updated"] += 1
            counts[new_cat] = counts.get(new_cat, 0) + 1
            if not dry_run:
                record.categoria = new_cat
                record.linea = new_linea
        if not dry_run:
            session.commit()
    return counts


def _detect_condicion(sku: str, producto: str) -> str:
    if re.search(r"\(o\)", sku, re.IGNORECASE) or re.search(r"\(o\)", producto, re.IGNORECASE):
        return "OUTLET"
    return "PRIMERA"


def _normalize_sku(sku: str) -> str:
    """Normalize SKU with the same key used by the product catalog."""
    return sku_key(sku)


def _normalize_seller(value: Any) -> str:
    return _norm(value) or "SIN VENDEDOR"


def _load_product_match_context(session: Session) -> tuple[dict[str, dict], dict[str, SalesBIProductAlias], dict[str, SalesBIProductAlias]]:
    products = session.scalars(select(Product).where(Product.is_active.is_(True))).all()
    aliases = session.scalars(select(SalesBIProductAlias)).all()
    indexes = build_product_indexes(products, aliases=aliases)
    aliases_by_sku = {str(a.alias_sku_norm or "").strip(): a for a in aliases if str(a.alias_sku_norm or "").strip()}
    aliases_by_desc = {str(a.alias_desc_norm or "").strip(): a for a in aliases if str(a.alias_desc_norm or "").strip()}
    return indexes, aliases_by_sku, aliases_by_desc


def _resolve_record_match(
    rec: dict,
    indexes: dict[str, dict],
    aliases_by_sku: dict[str, SalesBIProductAlias],
    aliases_by_desc: dict[str, SalesBIProductAlias],
) -> tuple[Product | None, str, str, int | None]:
    raw_sku = str(rec.get("sku") or "").strip()
    raw_desc = str(rec.get("producto") or "").strip()
    sku_norm = _normalize_sku(raw_sku)
    desc_norm = normalize_descripcion(raw_desc)
    product = resolve_product(sku_normalized=sku_norm, descripcion=raw_desc, indexes=indexes)
    if product is None:
        return None, sku_norm, "unmatched", None

    alias_id: int | None = None
    status = "matched"
    exact_desc = indexes.get("by_desc", {}).get(desc_norm) if desc_norm else None
    if exact_desc is None and desc_norm:
        alias = aliases_by_desc.get(desc_norm)
        if alias is not None and int(alias.product_id) == int(product.id):
            alias_id = int(alias.id)
            status = "matched_by_alias"

    if alias_id is None and sku_norm:
        exact_sku = indexes.get("by_sku", {}).get(sku_norm)
        if exact_sku is None:
            alias = aliases_by_sku.get(sku_norm)
            if alias is not None and int(alias.product_id) == int(product.id):
                alias_id = int(alias.id)
                status = "matched_by_alias"

    return product, sku_norm, status, alias_id


def _apply_product_match_to_record(
    rec: dict,
    product: Product | None,
    sku_norm: str,
    status: str,
    alias_id: int | None,
) -> dict:
    rec["sku_normalized"] = sku_norm
    rec["product_id"] = int(product.id) if product is not None else None
    rec["product_match_status"] = status
    rec["product_alias_id"] = alias_id

    if product is not None:
        if not rec.get("marca") and product.marca:
            rec["marca"] = product.marca
        if not rec.get("tipo_producto") and product.tipo:
            rec["tipo_producto"] = product.tipo
        if not rec.get("costo") and product.costo_vigente:
            rec["costo"] = float(product.costo_vigente)

    rec["categoria"], rec["linea"] = _classify(rec.get("tipo_producto", ""))
    costo = rec.get("costo", 0.0)
    total_cobrado = rec.get("total_cobrado", 0.0)
    cantidad = rec.get("cantidad", 1)
    if costo:
        rec["diferencia"] = total_cobrado - costo * cantidad
        rec["margen_porcentaje"] = round(rec["diferencia"] / total_cobrado * 100, 2) if total_cobrado else 0.0
    return rec


def enrich_from_catalog(records: list[dict]) -> list[dict]:
    """
    Look up each record's SKU in the products catalog.
    Fills in marca, tipo_producto, costo (if missing or zero), then
    recomputes categoria, linea, diferencia, margen_porcentaje.
    Does NOT overwrite existing non-zero values — same policy as the AppScript.
    """
    with db_session() as session:
        indexes, aliases_by_sku, aliases_by_desc = _load_product_match_context(session)

        for rec in records:
            product, sku_norm, status, alias_id = _resolve_record_match(rec, indexes, aliases_by_sku, aliases_by_desc)
            _apply_product_match_to_record(rec, product, sku_norm, status, alias_id)

    return records


# ── DB ───────────────────────────────────────────────────────────────────────


# ── temp file storage ─────────────────────────────────────────────────────────

_TEMP: dict[str, Path] = {}
_TEMP_DIR = Path(__file__).parent.parent / "storage" / "tmp"


def save_temp_file(content: bytes) -> str:
    _TEMP_DIR.mkdir(parents=True, exist_ok=True)
    key = uuid.uuid4().hex
    p = _TEMP_DIR / f"sbi_{key}.xlsx"
    p.write_bytes(content)
    _TEMP[key] = p
    return key


def load_temp_file(key: str) -> bytes | None:
    p = _TEMP.get(key) or _TEMP_DIR / f"sbi_{key}.xlsx"
    if p.exists():
        _TEMP[key] = p
        return p.read_bytes()
    return None


def delete_temp_file(key: str) -> None:
    p = _TEMP.pop(key, None) or _TEMP_DIR / f"sbi_{key}.xlsx"
    try:
        if p and p.exists():
            p.unlink()
    except Exception:
        pass


# ── spreadsheet reading ───────────────────────────────────────────────────────


def read_excel(content: bytes) -> dict[str, list[list]]:
    wb = openpyxl.load_workbook(filename=io.BytesIO(content), data_only=True)
    result: dict[str, list[list]] = {}
    for name in wb.sheetnames:
        ws = wb[name]
        rows = [list(row) for row in ws.iter_rows(values_only=True)]
        result[name] = rows
    return result


def read_google_sheet(url: str) -> dict[str, list[list]]:
    spreadsheet_id = extract_spreadsheet_id(url)
    if not spreadsheet_id:
        raise ValueError(f"URL de Google Sheets inválida: {url}")
    svc = sheets_service()
    meta = svc.spreadsheets().get(spreadsheetId=spreadsheet_id).execute()
    result: dict[str, list[list]] = {}
    for sh in meta.get("sheets", []):
        title = sh["properties"]["title"]
        resp = (
            svc.spreadsheets()
            .values()
            .get(spreadsheetId=spreadsheet_id, range=title, valueRenderOption="UNFORMATTED_VALUE")
            .execute()
        )
        result[title] = resp.get("values", [])
    return result


# ── metadata detection ────────────────────────────────────────────────────────


def _detect_metadata(rows: list[list]) -> dict:
    """
    Detect date, branch, and exchange rate from the top rows.

    Handles the common format:
        Row 1:  Dia | 11/05/2026 | SUR | Tipo de cambio | $ | 1.375,00
    """
    meta: dict[str, Any] = {"fecha": "", "sucursal": "", "cotizacion_dolar": None}

    for row in rows[:15]:
        non_empty = [(i, c) for i, c in enumerate(row) if c is not None and str(c).strip()]

        for idx, (col_i, cell) in enumerate(non_empty):
            cn = _norm(cell)
            rest = non_empty[idx + 1 :]

            # ── Date by label ────────────────────────────────────────────────
            if not meta["fecha"] and cn in _DATE_LABELS:
                for _, nxt in rest[:4]:
                    if _is_date_like(nxt):
                        try:
                            meta["fecha"] = _parse_date(nxt)
                        except Exception:
                            pass
                        break

                # Sucursal: next non-numeric, non-label text after the date value
                if meta["fecha"]:
                    skip_one = True  # skip the date value itself
                    for _, nxt in rest:
                        if skip_one and _is_date_like(nxt):
                            skip_one = False
                            continue
                        skip_one = False
                        nxt_n = _norm(nxt)
                        if (
                            not meta["sucursal"]
                            and nxt_n
                            and nxt_n not in _DATE_LABELS
                            and nxt_n not in _NON_SUCURSAL_LABELS
                            and not any(lbl in nxt_n for lbl in _RATE_LABELS | _SKIP_LABELS)
                        ):
                            try:
                                float(str(nxt).replace(",", ".").replace(".", "", 1))
                            except ValueError:
                                meta["sucursal"] = str(nxt).strip()
                                break

            # ── Sucursal by label ─────────────────────────────────────────────
            if not meta["sucursal"] and "SUCURSAL" in cn:
                for _, nxt in rest[:2]:
                    s = str(nxt).strip()
                    if s:
                        meta["sucursal"] = s
                        break

            # ── Exchange rate ─────────────────────────────────────────────────
            if meta["cotizacion_dolar"] is None and any(lbl in cn for lbl in _RATE_LABELS):
                for _, nxt in rest[:5]:
                    v = _parse_num(nxt)
                    if v > 0:
                        meta["cotizacion_dolar"] = v
                        break

    return meta


# ── SALDOS table detection (anywhere in the grid) ────────────────────────────


def _find_saldos_table(rows: list[list]) -> list[dict]:
    """
    Find the SALDOS section anywhere in the 2D grid (it can be to the right of
    the main table, not necessarily below it).
    Returns a list of balance dicts keyed by payment method.
    """
    # Step 1: find "SALDOS" cell
    saldos_ri = -1
    saldos_ci = -1
    for ri, row in enumerate(rows):
        for ci, cell in enumerate(row):
            if _norm(cell) == "SALDOS":
                saldos_ri = ri
                saldos_ci = ci
                break
        if saldos_ri >= 0:
            break

    if saldos_ri < 0:
        return []

    # Step 2: find sub-header row (REMITO, EFECT., TRANSFER., etc.)
    # Search within ±3 rows of the SALDOS header
    header_col_map: dict[str, int] = {}
    data_start_row = -1

    for ri in range(max(0, saldos_ri - 1), min(len(rows), saldos_ri + 4)):
        row = rows[ri]
        tmp: dict[str, int] = {}
        for ci in range(saldos_ci, min(saldos_ci + 20, len(row))):
            if ci >= len(row):
                break
            n = _norm(row[ci])
            field = _ALIAS_TO_FIELD.get(n)
            if field in ("remito", "efectivo", "transferencia", "tarjeta", "usd", "otros", "total", "cuenta_corriente"):
                tmp[field] = ci
        if tmp:
            header_col_map = tmp
            data_start_row = ri + 1
            break

    if not header_col_map or data_start_row < 0:
        return []

    # Step 3: read data rows within the SALDOS column range
    saldos: list[dict] = []
    remito_col = header_col_map.get("remito")

    for ri in range(data_start_row, len(rows)):
        row = rows[ri]
        # Check if there's any non-empty value in the SALDOS column range
        has_data = any(
            col < len(row) and row[col] is not None and str(row[col]).strip()
            for col in header_col_map.values()
        )
        if not has_data:
            continue

        remito_val = (row[remito_col] if remito_col is not None and remito_col < len(row) else None)
        if not remito_val or not str(remito_val).strip():
            continue

        entry: dict[str, Any] = {
            "remito": str(remito_val).strip(),
            "efectivo": 0.0,
            "transferencia": 0.0,
            "tarjeta": 0.0,
            "usd": 0.0,
            "otros": 0.0,
        }
        for field, col in header_col_map.items():
            if field != "remito" and col < len(row):
                entry[field] = _parse_num(row[col])
        entry["total"] = (
            entry["efectivo"] + entry["transferencia"] + entry["tarjeta"]
            + entry["usd"] + entry["otros"]
        )
        saldos.append(entry)

    return saldos


# ── header and data parsing ───────────────────────────────────────────────────


def _find_header_row(rows: list[list]) -> tuple[int, dict[str, int]] | None:
    """
    Find the main sales table header. Handles:
    - Single-row headers (REMITO | VENDEDOR | EFECT. | ...)
    - Two-row headers with group label on row N and sub-columns on row N+1
      (REMITO | VENDEDOR | MEDIOS DE PAGO) + (blank | blank | EFECT. | TRANSFER. | ...)
    """
    for i, row in enumerate(rows):
        if i > 80:
            break
        normed = [_norm(c) for c in row]

        has_remito = any(_ALIAS_TO_FIELD.get(n) == "remito" for n in normed)
        has_vendedor = any(_ALIAS_TO_FIELD.get(n) == "vendedor" for n in normed)
        has_product = any(_ALIAS_TO_FIELD.get(n) == "producto" for n in normed)
        has_payment = any(
            _ALIAS_TO_FIELD.get(n) in ("efectivo", "transferencia", "tarjeta", "usd")
            for n in normed
        )

        is_header = (
            (has_product and (has_remito or has_vendedor or has_payment))
            or (has_remito and (has_vendedor or has_payment))
            or (has_remito and has_vendedor)
        )
        if not is_header:
            continue

        col_map: dict[str, int] = {}
        for j, n in enumerate(normed):
            field = _ALIAS_TO_FIELD.get(n)
            if field and field not in col_map:
                col_map[field] = j

        # Check next 1-2 rows for sub-headers (payment sub-columns)
        sub_header_rows = 0
        for offset in range(1, 3):
            if i + offset >= len(rows):
                break
            sub_normed = [_norm(c) for c in rows[i + offset]]
            added = False
            for j, n in enumerate(sub_normed):
                field = _field_for_subheader(normed, j, n)
                if field and field not in col_map:
                    col_map[field] = j
                    added = True
            if added:
                sub_header_rows += 1
            else:
                break

        if col_map:
            return i, col_map, sub_header_rows  # type: ignore[return-value]

    return None


_STOP_KEYWORDS = frozenset([
    "TOTAL", "TOTALES", "SUBTOTAL",
    "VENTA TOTAL", "TOTAL VENTA", "TOTAL VENTAS",
    "VENTA DIARIAS", "VENTAS DIARIAS",
    "ADMINISTRACION", "ADMINISTRACIÓN",
])

# Primeras palabras que indican una sección de cierre. Permite cazar
# "TOTAL GV", "SUBTOTAL DEL MES", "RETIRO A CARGO DE:", "DEVOLUCIONES" etc.
# Estos textos suelen aparecer en filas footer que reaprovechan las columnas
# REMITO/VENDEDOR/SKU/PRODUCTO para mostrar resúmenes — y antes los leíamos
# como si fueran ventas.
_STOP_PREFIXES = frozenset([
    "TOTAL", "TOTALES", "SUBTOTAL", "SUBTOTALES",
    "RETIRO", "RETIROS", "DEVOLUCION", "DEVOLUCIONES",
])


def _is_stop_marker(text: Any) -> bool:
    """¿Esta celda parece un marcador de fin de tabla (label de resumen)?"""
    n = _norm(text)
    if not n:
        return False
    if n in _STOP_KEYWORDS:
        return True
    # Primera palabra normalizada (ej. "TOTAL GV" → "TOTAL").
    return n.split(" ", 1)[0] in _STOP_PREFIXES


def _is_stop_row(row: list, remito_col: int | None) -> bool:
    """Detecta filas de cierre/resumen que marcan el final de la tabla de ventas.

    Reglas:
      1. Si la celda REMITO tiene un marcador de cierre (TOTAL, RETIRO, ...)
         → parar. Caso clásico de fila de totales debajo de la tabla.
      2. Si REMITO tiene un valor con dígitos (ej. "76", "3546", "RET-2026-1")
         es una fila de DATOS legítima — no parar aunque otra columna haga
         match con un keyword. Evita falsos positivos como un vendedor
         llamado "Administracion" o un producto cuya descripción mencione
         "TOTAL".
      3. Si REMITO está vacío o sin dígitos, miramos el resto de las celdas
         como fallback (fila de resumen tipo "Administracion: $1.770.000",
         "TOTAL GV | SUCURSAL + ON LINE", "Retiro a cargo de:" etc.).
    """
    if remito_col is not None and remito_col < len(row):
        val = row[remito_col]
        n = _norm(val) if val is not None else ""
        if _is_stop_marker(val):
            return True
        # Hay número de remito real → es una fila de datos, no de cierre.
        if n and re.search(r"\d", n):
            return False
    # Fallback: REMITO vacío/sin dígitos. Mirar todas las celdas por si la
    # fila es un resumen suelto (ej. label en una columna que no es REMITO).
    for cell in row:
        if _is_stop_marker(cell):
            return True
    return False


def _parse_data_rows(
    rows: list[list],
    header_idx: int,
    sub_header_rows: int,
    col_map: dict[str, int],
    is_online: bool,
    saldos_col_start: int,
) -> tuple[list[dict], int]:
    data_start = header_idx + 1 + sub_header_rows
    records: list[dict] = []
    last_data_row = data_start
    remito_col = col_map.get("remito")

    for i, row in enumerate(rows[data_start:], start=data_start):
        def get(field: str, default: Any = None) -> Any:
            idx = col_map.get(field)
            if idx is None or idx >= min(len(row), saldos_col_start):
                return default
            v = row[idx]
            return v if v is not None else default

        remito = str(get("remito", "")).strip()
        producto = str(get("producto", "")).strip()
        sku = str(get("sku", "")).strip()

        # Need at least a remito or a producto to consider this a data row
        if not remito and not producto:
            continue

        if _is_stop_row(row, remito_col):
            break

        # Una venta legítima necesita al menos un identificador de producto
        # (SKU o descripción). Si una fila trae REMITO + VENDEDOR pero los
        # dos identificadores de producto están vacíos, es casi seguro una
        # fila de footer/resumen de otra tabla que la planilla pega abajo
        # de las ventas reales (ej. "Retiros del día", "Ale Paulo: total"
        # con la columna REMITO reusada para enumerar). La saltamos.
        if not producto and not sku:
            continue

        # Skip rows where remito is a dash/empty and producto is purely numeric
        # (these are usually totals rows like "927500")
        remito_clean = remito.lstrip("-").strip()
        if not remito_clean and re.fullmatch(r"[\d\s.,]+", producto):
            continue

        vendedor = str(get("vendedor", "")).strip()
        marca = str(get("marca", "")).strip()
        tipo_produto = str(get("tipo", "")).strip()

        raw_cond = str(get("condicion", "")).strip()
        condicion = _detect_condicion(sku, producto)
        if raw_cond and "OUTLET" in _norm(raw_cond):
            condicion = "OUTLET"

        cantidad = max(1, int(_parse_num(get("cantidad", 1)) or 1))
        pvp = _parse_num(get("pvp", 0))
        # En planillas ONLINE la columna VALOR/MONTO es el TOTAL de la línea
        # (precio unitario × cantidad), no el precio unitario. Si cantidad > 1
        # y la dejamos como está, el cálculo posterior `pvp × cantidad` infla
        # el total. Lo normalizamos a precio unitario para que el resto del
        # pipeline (que asume "pvp = precio unitario") dé números correctos.
        # Ejemplo real: HY11INV cantidad=2 valor=1.900.000 → pvp_unit=950.000,
        # total=950.000×2=1.900.000 (en vez de 1.900.000×2=3.800.000).
        if is_online and cantidad > 1 and pvp > 0:
            pvp = pvp / cantidad
        costo = _parse_num(get("costo", 0))
        monto_ingresado = _parse_num(get("monto_ingresado", 0))
        efectivo = _parse_num(get("efectivo", 0))
        transferencia = _parse_num(get("transferencia", 0))
        tarjeta = _parse_num(get("tarjeta", 0))
        usd = _parse_num(get("usd", 0))
        cuenta_corriente = _parse_num(get("cuenta_corriente", 0))
        otros = _parse_num(get("otros", 0))
        efectivo += _parse_num(get("sena_efectivo", 0))
        transferencia += _parse_num(get("sena_transferencia", 0))
        tarjeta += _parse_num(get("sena_tarjeta", 0))
        usd += _parse_num(get("sena_usd", 0))

        if is_online:
            # Venta online ingresa como transferencia. Si el monto ingresado es
            # menor al total del remito, el saldo queda como seña pendiente.
            total_cobrado = monto_ingresado
            transferencia = monto_ingresado
            efectivo = tarjeta = usd = cuenta_corriente = otros = 0.0
        else:
            total_cobrado = efectivo + transferencia + tarjeta + usd + cuenta_corriente + otros

        diferencia = total_cobrado - costo * cantidad if costo else 0.0
        margen = (diferencia / total_cobrado * 100) if total_cobrado else 0.0

        categoria, linea = _classify(tipo_produto)

        records.append({
            "remito": remito,
            "vendedor": vendedor,
            "vendedor_normalized": _normalize_seller(vendedor),
            "producto": producto,
            "sku": sku,
            "sku_normalized": _normalize_sku(sku),
            "product_id": None,
            "product_alias_id": None,
            "product_match_status": "unmatched",
            "marca": marca,
            "tipo_producto": tipo_produto,
            "condicion": condicion,
            "categoria": categoria,
            "linea": linea,
            "cantidad": cantidad,
            "pvp": pvp,
            "costo": costo,
            "diferencia": diferencia,
            "margen_porcentaje": round(margen, 2),
            "efectivo": efectivo,
            "transferencia": transferencia,
            "tarjeta": tarjeta,
            "usd": usd,
            "cuenta_corriente": cuenta_corriente,
            "otros": otros,
            "total_cobrado": total_cobrado,
            "saldo": 0.0,  # computed after payment distribution
        })
        last_data_row = i

    return records, last_data_row


_PAYMENT_FIELDS = ("efectivo", "transferencia", "tarjeta", "usd", "cuenta_corriente", "otros")


def _raw_cobrado(rec: dict) -> float:
    return sum(rec.get(f, 0.0) for f in _PAYMENT_FIELDS)


def _recompute_financials(rec: dict) -> None:
    pvp_total = rec["pvp"] * rec["cantidad"]
    total_cobrado = float(rec.get("total_cobrado") or 0.0)
    rec["saldo"] = round(max(0.0, pvp_total - total_cobrado), 2)
    if rec.get("costo"):
        diferencia = total_cobrado - rec["costo"] * rec["cantidad"]
        rec["diferencia"] = round(diferencia, 2)
        rec["margen_porcentaje"] = round((diferencia / total_cobrado * 100) if total_cobrado else 0.0, 2)
    else:
        rec["diferencia"] = 0.0
        rec["margen_porcentaje"] = 0.0


def _distribute_remito_payments(records: list[dict]) -> list[dict]:
    """
    The remito is the accounting key for an imported sale.
    When it appears on multiple rows, sellers often write the payment only once.
    This compares the sum of payment-method fields against the sum of products
    for the remito and redistributes the real collected amount proportionally.
    """
    from collections import defaultdict

    groups: dict[str, list[int]] = defaultdict(list)
    for i, rec in enumerate(records):
        remito = rec.get("remito", "").strip().lstrip("-").strip()
        if remito:
            groups[remito].append(i)

    for remito, indices in groups.items():
        if len(indices) < 2:
            continue
        group_recs = [records[i] for i in indices]
        actual_cobrado = sum(_raw_cobrado(r) for r in group_recs)
        total_pvp = sum(r["pvp"] * r["cantidad"] for r in group_recs)
        if actual_cobrado == 0 or total_pvp == 0:
            continue

        # Distribute proportionally; last record absorbs rounding delta
        remaining = actual_cobrado
        for idx in indices[:-1]:
            pvp_i = records[idx]["pvp"] * records[idx]["cantidad"]
            share = round(pvp_i / total_pvp * actual_cobrado, 2)
            records[idx]["total_cobrado"] = share
            remaining = round(remaining - share, 2)
        records[indices[-1]]["total_cobrado"] = remaining

    for rec in records:
        _recompute_financials(rec)

    return records


_SHEET_NAME_LOCAL = frozenset(["PLANILLA"])
_SHEET_NAME_ONLINE = frozenset(["ONLINE", "ON LINE"])


def _classify_sheet_name(name: str) -> str | None:
    """
    Returns 'local', 'online', or None (skip).
    Only the exact sheet names 'Planilla' (local) and 'On Line'/'Online' (online) are imported.
    """
    nn = _norm(name)
    if nn in _SHEET_NAME_ONLINE:
        return "online"
    if nn in _SHEET_NAME_LOCAL:
        return "local"
    return None


def _is_online_sheet(name: str, col_map: dict[str, int]) -> bool:
    tipo = _classify_sheet_name(name)
    if tipo is not None:
        return tipo == "online"
    # fallback: infer from columns
    return "pedido" in col_map and "remito" not in col_map


def _find_saldos_col_start(rows: list[list]) -> int:
    """Return the column index where a separate SALDOS *section* starts.
    Only the plural 'SALDOS' triggers this — the singular 'SALDO' is a regular
    column within the main table and must not cut off the data read.
    """
    for row in rows:
        for ci, cell in enumerate(row):
            if _norm(cell) == "SALDOS":
                return ci
    return 9999


def _parse_sheet(name: str, rows: list[list], sucursal_override: str = "") -> dict:
    warnings: list[str] = []
    meta = _detect_metadata(rows)

    if not meta["fecha"]:
        warnings.append("No se detectó la fecha en la planilla.")
    if sucursal_override:
        meta["sucursal"] = _normalize_sucursal(sucursal_override)
    elif meta["sucursal"]:
        meta["sucursal"] = _normalize_sucursal(meta["sucursal"])
    else:
        meta["sucursal"] = name

    result = _find_header_row(rows)
    if result is None:
        return {
            "sheet_name": name,
            "fecha": meta["fecha"],
            "sucursal": meta["sucursal"],
            "tipo": "local",
            "cotizacion_dolar": meta["cotizacion_dolar"],
            "records": [],
            "balances": [],
            "warnings": warnings + ["No se encontró la tabla de ventas en esta hoja."],
            "ok": False,
        }

    header_idx, col_map, sub_header_rows = result  # type: ignore[misc]

    is_online = _is_online_sheet(name, col_map)
    saldos_col = _find_saldos_col_start(rows)

    records, _ = _parse_data_rows(rows, header_idx, sub_header_rows, col_map, is_online, saldos_col)

    records = _distribute_remito_payments(records)

    # Enrich with product catalog data (marca, tipo, costo, categoria, linea)
    try:
        records = enrich_from_catalog(records)
    except Exception:
        pass  # catalog enrichment is best-effort

    if not records:
        warnings.append("La tabla de ventas está vacía o no se pudieron leer los datos.")

    balances: list[dict] = []
    if not is_online:
        balances = _find_saldos_table(rows)

    return {
        "sheet_name": name,
        "fecha": meta["fecha"],
        "sucursal": meta["sucursal"],
        "tipo": "online" if is_online else "local",
        "cotizacion_dolar": meta["cotizacion_dolar"],
        "records": records,
        "balances": balances,
        "warnings": warnings,
        "ok": True,
    }


def _sheet_totals(records: list[dict]) -> dict:
    return {
        "total_records": len(records),
        "matched_products": sum(1 for r in records if r.get("product_match_status") == "matched"),
        "matched_by_alias": sum(1 for r in records if r.get("product_match_status") == "matched_by_alias"),
        "unmatched_products": sum(1 for r in records if r.get("product_match_status") == "unmatched"),
        "total_pvp": sum(r["pvp"] * r["cantidad"] for r in records),
        "total_costo": sum(r["costo"] * r["cantidad"] for r in records),
        "total_efectivo": sum(r["efectivo"] for r in records),
        "total_transferencia": sum(r["transferencia"] for r in records),
        "total_tarjeta": sum(r["tarjeta"] for r in records),
        "total_usd": sum(r["usd"] for r in records),
        "total_cuenta_corriente": sum(r["cuenta_corriente"] for r in records),
        "total_otros": sum(r["otros"] for r in records),
    }


# ── public API ────────────────────────────────────────────────────────────────


def analyze_sheets(sheets: dict[str, list[list]], sucursal_override: str = "", source_name: str = "") -> list[dict]:
    results = []
    workbook_sucursal = "" if sucursal_override else _infer_workbook_sucursal(sheets, source_name)
    for name, rows in sheets.items():
        tipo_hoja = _classify_sheet_name(name)
        if tipo_hoja is None:
            # Skip sheets that are not 'Planilla' or 'On Line'
            continue
        parsed = _parse_sheet(name, rows, sucursal_override=sucursal_override)
        if workbook_sucursal and _is_placeholder_sucursal(parsed.get("sucursal", ""), parsed.get("sheet_name", name)):
            parsed["sucursal"] = workbook_sucursal
        totals = _sheet_totals(parsed["records"])
        parsed.update(totals)
        results.append(parsed)
    return results


# ── DB operations ─────────────────────────────────────────────────────────────


def _branch_to_dict(branch: Branch | None) -> dict | None:
    if not branch:
        return None
    return {
        "id": str(branch.id),
        "name": str(branch.name or ""),
        "code": str(branch.code or ""),
        "type": str(branch.type or ""),
    }


def _find_branch_in_session(session: Session, sucursal: str, tipo: str) -> dict | None:
    code_guess = str(sucursal or "").upper().strip().replace(" ", "_")
    if not code_guess:
        return None
    if tipo == "online":
        candidates = [code_guess + "_WEB", code_guess]
        preferred_type = "web"
    else:
        candidates = [code_guess]
        preferred_type = "physical"

    for code in candidates:
        branch = session.scalar(
            select(Branch).where(Branch.code == code, Branch.is_active.is_(True)).limit(1)
        )
        if branch:
            return _branch_to_dict(branch)

    branch = session.scalar(
        select(Branch)
        .where(
            func.upper(Branch.name).like(f"%{str(sucursal or '').upper()}%"),
            Branch.is_active.is_(True),
        )
        .order_by(case((Branch.type == preferred_type, 1), else_=0).desc())
        .limit(1)
    )
    return _branch_to_dict(branch)


def find_branch(sucursal: str, tipo: str) -> dict | None:
    with db_session() as session:
        return _find_branch_in_session(session, sucursal, tipo)


def _user_id_from_username(session: Session, username: str) -> int | None:
    uname = str(username or "").strip().lower()
    if not uname:
        return None
    return session.scalar(select(User.id).where(User.username == uname))


def _user_identity(session: Session, user_id: int | None) -> tuple[str, str]:
    if user_id is None:
        return "", ""
    user = session.get(User, user_id)
    if not user:
        return "", ""
    return str(user.username or ""), str(user.display_name or "")


def _record_to_dict(record: SalesRecord, imp: SalesImport | None = None) -> dict:
    data = {
        "id": int(record.id),
        "import_id": int(record.import_id),
        "nro_linea": int(record.nro_linea or 0),
        "remito": str(record.remito or ""),
        "vendedor": str(record.vendedor or ""),
        "vendedor_normalized": str(record.vendedor_normalized or ""),
        "seller_user_id": int(record.seller_user_id) if record.seller_user_id else None,
        "producto": str(record.producto or ""),
        "sku": str(record.sku or ""),
        "sku_normalized": str(record.sku_normalized or ""),
        "product_id": int(record.product_id) if record.product_id else None,
        "product_alias_id": int(record.product_alias_id) if record.product_alias_id else None,
        "product_match_status": str(record.product_match_status or "unmatched"),
        "marca": str(record.marca or ""),
        "tipo_producto": str(record.tipo_producto or ""),
        "condicion": str(record.condicion or ""),
        "categoria": str(record.categoria or ""),
        "linea": str(record.linea or ""),
        "cantidad": int(record.cantidad or 0),
        "pvp": _num(record.pvp),
        "costo": _num(record.costo),
        "diferencia": _num(record.diferencia),
        "margen_porcentaje": _num(record.margen_porcentaje),
        "efectivo": _num(record.efectivo),
        "transferencia": _num(record.transferencia),
        "tarjeta": _num(record.tarjeta),
        "usd": _num(record.usd),
        "cuenta_corriente": _num(record.cuenta_corriente),
        "otros": _num(record.otros),
        "total_cobrado": _num(record.total_cobrado),
        "saldo": _num(record.saldo),
    }
    if imp is not None:
        data.update({
            "fecha": _fmt_date(imp.fecha),
            "sucursal": str(imp.sucursal or ""),
            "tipo": str(imp.tipo or ""),
        })
    return data


def _balance_to_dict(balance: SalesBalance, imp: SalesImport | None = None) -> dict:
    data = {
        "id": int(balance.id),
        "import_id": int(balance.import_id),
        "remito": str(balance.remito or ""),
        "efectivo": _num(balance.efectivo),
        "transferencia": _num(balance.transferencia),
        "tarjeta": _num(balance.tarjeta),
        "usd": _num(balance.usd),
        "otros": _num(balance.otros),
        "total": _num(balance.total),
    }
    if imp is not None:
        data.update({
            "fecha": _fmt_date(imp.fecha),
            "sucursal": str(imp.sucursal or ""),
        })
    return data


def _import_to_dict(imp: SalesImport, session: Session, *, include_children: bool = False) -> dict:
    imported_by, imported_by_name = _user_identity(session, imp.imported_by_user_id)
    voided_by, _voided_by_name = _user_identity(session, imp.voided_by_user_id)
    branch = session.get(Branch, imp.branch_id) if imp.branch_id else None
    data = {
        "id": int(imp.id),
        "fecha": _fmt_date(imp.fecha),
        "sucursal": str(imp.sucursal or ""),
        "tipo": str(imp.tipo or ""),
        "branch_id": str(imp.branch_id) if imp.branch_id else None,
        "branch_name": str(branch.name or "") if branch else None,
        "branch_type": str(branch.type or "") if branch else None,
        "fuente": str(imp.fuente or ""),
        "fuente_url": str(imp.fuente_url or ""),
        "fuente_nombre": str(imp.fuente_nombre or ""),
        "status": str(imp.status or "activo"),
        "total_records": int(imp.total_records or 0),
        "total_pvp": _num(imp.total_pvp),
        "total_costo": _num(imp.total_costo),
        "total_efectivo": _num(imp.total_efectivo),
        "total_transferencia": _num(imp.total_transferencia),
        "total_tarjeta": _num(imp.total_tarjeta),
        "total_usd": _num(imp.total_usd),
        "total_cuenta_corriente": _num(imp.total_cuenta_corriente),
        "total_otros": _num(imp.total_otros),
        "cotizacion_dolar": None if imp.cotizacion_dolar is None else _num(imp.cotizacion_dolar),
        "imported_by": imported_by,
        "imported_by_name": imported_by_name,
        "created_at": _fmt_dt(imp.created_at),
        "voided_at": _fmt_dt(imp.voided_at),
        "voided_by": voided_by,
        "void_reason": str(imp.void_reason or ""),
        "warnings": list(imp.warnings or []),
    }
    if include_children:
        records = sorted(list(imp.records or []), key=lambda r: int(r.nro_linea or 0))
        balances = sorted(list(imp.balances or []), key=lambda b: int(b.id or 0))
        data["records"] = [_record_to_dict(record) for record in records]
        data["balances"] = [_balance_to_dict(balance) for balance in balances]
    return data


def get_active_import(fecha: str, sucursal: str, tipo: str) -> dict | None:
    fecha_value = _parse_date_value(fecha)
    if not fecha_value:
        return None
    with db_session() as session:
        imp = session.scalar(
            select(SalesImport).where(
                SalesImport.fecha == fecha_value,
                SalesImport.sucursal == str(sucursal or ""),
                SalesImport.tipo == str(tipo or ""),
                SalesImport.status == "activo",
            )
        )
        return _import_to_dict(imp, session) if imp else None


def save_import(
    sheet: dict,
    fuente: str,
    fuente_url: str,
    fuente_nombre: str,
    username: str,
    display_name: str,
    branch_id: str | None = None,
) -> int:
    now = utc_now_dt()
    totals = _sheet_totals(sheet["records"])

    with db_session() as session:
        fecha = _parse_date_value(sheet.get("fecha"))
        if fecha is None:
            raise ValueError("La importacion no tiene fecha valida.")

        if not branch_id:
            matched = _find_branch_in_session(session, sheet["sucursal"], sheet["tipo"])
            branch_id = matched["id"] if matched else None
        elif not session.get(Branch, branch_id):
            branch_id = None

        imp = SalesImport(
            fecha=fecha,
            sucursal=str(sheet["sucursal"]),
            tipo=str(sheet["tipo"]),
            fuente=fuente,
            fuente_url=fuente_url,
            fuente_nombre=fuente_nombre,
            status="activo",
            total_records=int(totals["total_records"]),
            total_pvp=_decimal(totals["total_pvp"]),
            total_costo=_decimal(totals["total_costo"]),
            total_efectivo=_decimal(totals["total_efectivo"]),
            total_transferencia=_decimal(totals["total_transferencia"]),
            total_tarjeta=_decimal(totals["total_tarjeta"]),
            total_usd=_decimal(totals["total_usd"]),
            total_cuenta_corriente=_decimal(totals["total_cuenta_corriente"]),
            total_otros=_decimal(totals["total_otros"]),
            cotizacion_dolar=None if sheet.get("cotizacion_dolar") is None else _decimal(sheet.get("cotizacion_dolar")),
            imported_by_user_id=_user_id_from_username(session, username),
            created_at=now,
            warnings=list(sheet.get("warnings", [])),
            branch_id=branch_id,
        )
        session.add(imp)
        session.flush()

        for i, rec in enumerate(sheet["records"], start=1):
            session.add(
                SalesRecord(
                    import_=imp,
                    nro_linea=i,
                    remito=str(rec.get("remito") or ""),
                    vendedor=str(rec.get("vendedor") or ""),
                    vendedor_normalized=str(rec.get("vendedor_normalized") or _normalize_seller(rec.get("vendedor"))),
                    seller_user_id=None,
                    producto=str(rec.get("producto") or ""),
                    sku=str(rec.get("sku") or ""),
                    sku_normalized=str(rec.get("sku_normalized") or _normalize_sku(rec.get("sku", ""))),
                    product_id=rec.get("product_id"),
                    product_alias_id=rec.get("product_alias_id"),
                    product_match_status=str(rec.get("product_match_status") or "unmatched"),
                    marca=str(rec.get("marca") or ""),
                    tipo_producto=str(rec.get("tipo_producto") or ""),
                    condicion=str(rec.get("condicion") or ""),
                    categoria=str(rec.get("categoria") or ""),
                    linea=str(rec.get("linea") or ""),
                    cantidad=int(rec.get("cantidad") or 1),
                    pvp=_decimal(rec.get("pvp")),
                    costo=_decimal(rec.get("costo")),
                    diferencia=_decimal(rec.get("diferencia")),
                    margen_porcentaje=_decimal(rec.get("margen_porcentaje")),
                    efectivo=_decimal(rec.get("efectivo")),
                    transferencia=_decimal(rec.get("transferencia")),
                    tarjeta=_decimal(rec.get("tarjeta")),
                    usd=_decimal(rec.get("usd")),
                    cuenta_corriente=_decimal(rec.get("cuenta_corriente")),
                    otros=_decimal(rec.get("otros")),
                    total_cobrado=_decimal(rec.get("total_cobrado")),
                    saldo=_decimal(rec.get("saldo", 0.0)),
                )
            )

        for bal in sheet.get("balances", []):
            session.add(
                SalesBalance(
                    import_=imp,
                    remito=str(bal.get("remito") or ""),
                    efectivo=_decimal(bal.get("efectivo")),
                    transferencia=_decimal(bal.get("transferencia")),
                    tarjeta=_decimal(bal.get("tarjeta")),
                    usd=_decimal(bal.get("usd")),
                    otros=_decimal(bal.get("otros")),
                    total=_decimal(bal.get("total")),
                )
            )

        import_id = int(imp.id)
        session.commit()
    return import_id


def void_import(import_id: int, username: str, reason: str) -> None:
    now = utc_now_dt()
    with db_session() as session:
        imp = session.get(SalesImport, import_id)
        if not imp:
            return
        imp.status = "anulado"
        imp.voided_at = now
        imp.voided_by_user_id = _user_id_from_username(session, username)
        imp.void_reason = str(reason or "")
        session.commit()


def get_import_detail(import_id: int) -> dict | None:
    with db_session() as session:
        imp = session.scalar(
            select(SalesImport)
            .options(selectinload(SalesImport.records), selectinload(SalesImport.balances))
            .where(SalesImport.id == import_id)
        )
        if not imp:
            return None
        return _import_to_dict(imp, session, include_children=True)


def list_imports(
    fecha_desde: str | None = None,
    fecha_hasta: str | None = None,
    sucursal: str | None = None,
    tipo: str | None = None,
    status: str | None = None,
    limit: int = 100,
    offset: int = 0,
) -> tuple[list[dict], int]:
    filters: list[Any] = []
    if fecha_desde:
        value = _parse_date_value(fecha_desde)
        if value:
            filters.append(SalesImport.fecha >= value)
    if fecha_hasta:
        value = _parse_date_value(fecha_hasta)
        if value:
            filters.append(SalesImport.fecha <= value)
    if sucursal:
        filters.append(SalesImport.sucursal == sucursal)
    if tipo:
        filters.append(SalesImport.tipo == tipo)
    if status:
        filters.append(SalesImport.status == status)
    with db_session() as session:
        total = int(session.scalar(select(func.count()).select_from(SalesImport).where(*filters)) or 0)
        rows = session.scalars(
            select(SalesImport)
            .where(*filters)
            .order_by(SalesImport.fecha.desc(), SalesImport.id.desc())
            .limit(limit)
            .offset(offset)
        ).all()
        return [_import_to_dict(row, session) for row in rows], total


def list_records(
    import_id: int | None = None,
    fecha_desde: str | None = None,
    fecha_hasta: str | None = None,
    sucursal: str | None = None,
    tipo: str | None = None,
    vendedor: str | None = None,
    categoria: str | None = None,
    condicion: str | None = None,
    q: str | None = None,
    limit: int = 200,
    offset: int = 0,
) -> tuple[list[dict], int]:
    filters: list[Any] = [SalesImport.status == "activo"]
    if import_id is not None:
        filters.append(SalesRecord.import_id == import_id)
    if fecha_desde:
        value = _parse_date_value(fecha_desde)
        if value:
            filters.append(SalesImport.fecha >= value)
    if fecha_hasta:
        value = _parse_date_value(fecha_hasta)
        if value:
            filters.append(SalesImport.fecha <= value)
    if sucursal:
        filters.append(SalesImport.sucursal == sucursal)
    if tipo:
        filters.append(SalesImport.tipo == tipo)
    if vendedor:
        filters.append(SalesRecord.vendedor.ilike(f"%{vendedor}%"))
    if categoria:
        filters.append(SalesRecord.categoria == categoria)
    if condicion:
        filters.append(SalesRecord.condicion == condicion)
    if q:
        text = f"%{q}%"
        filters.append(or_(
            SalesRecord.producto.ilike(text),
            SalesRecord.sku.ilike(text),
            SalesRecord.marca.ilike(text),
            SalesRecord.remito.ilike(text),
        ))

    with db_session() as session:
        total = int(session.scalar(
            select(func.count())
            .select_from(SalesRecord)
            .join(SalesImport, SalesImport.id == SalesRecord.import_id)
            .where(*filters)
        ) or 0)
        rows = session.execute(
            select(SalesRecord, SalesImport)
            .join(SalesImport, SalesImport.id == SalesRecord.import_id)
            .where(*filters)
            .order_by(SalesImport.fecha.desc(), SalesRecord.id.desc())
            .limit(limit)
            .offset(offset)
        ).all()
        return [_record_to_dict(record, imp) for record, imp in rows], total


def list_balances(
    import_id: int | None = None,
    fecha_desde: str | None = None,
    fecha_hasta: str | None = None,
    sucursal: str | None = None,
    limit: int = 200,
    offset: int = 0,
) -> tuple[list[dict], int]:
    filters: list[Any] = [SalesImport.status == "activo"]
    if import_id is not None:
        filters.append(SalesBalance.import_id == import_id)
    if fecha_desde:
        value = _parse_date_value(fecha_desde)
        if value:
            filters.append(SalesImport.fecha >= value)
    if fecha_hasta:
        value = _parse_date_value(fecha_hasta)
        if value:
            filters.append(SalesImport.fecha <= value)
    if sucursal:
        filters.append(SalesImport.sucursal == sucursal)

    with db_session() as session:
        total = int(session.scalar(
            select(func.count())
            .select_from(SalesBalance)
            .join(SalesImport, SalesImport.id == SalesBalance.import_id)
            .where(*filters)
        ) or 0)
        rows = session.execute(
            select(SalesBalance, SalesImport)
            .join(SalesImport, SalesImport.id == SalesBalance.import_id)
            .where(*filters)
            .order_by(SalesImport.fecha.desc(), SalesBalance.id.desc())
            .limit(limit)
            .offset(offset)
        ).all()
        return [_balance_to_dict(balance, imp) for balance, imp in rows], total


def _product_snapshot(product: Product | None) -> dict[str, Any] | None:
    if product is None:
        return None
    return {
        "id": int(product.id),
        "sku": str(product.sku or ""),
        "descripcion": str(product.descripcion or ""),
        "marca": str(product.marca or ""),
        "tipo": str(product.tipo or ""),
    }


def _alias_to_dict(alias: SalesBIProductAlias, product: Product | None = None) -> dict[str, Any]:
    return {
        "id": int(alias.id),
        "product_id": int(alias.product_id),
        "alias_sku_norm": str(alias.alias_sku_norm or ""),
        "alias_desc_norm": str(alias.alias_desc_norm or ""),
        "alias_sku_raw": str(alias.alias_sku_raw or ""),
        "alias_desc_raw": str(alias.alias_desc_raw or ""),
        "created_at": _fmt_dt(alias.created_at),
        "product": _product_snapshot(product),
    }


def create_product_alias(product_id: int, alias_sku: str, alias_desc: str, username: str) -> dict:
    sku_raw = str(alias_sku or "").strip()
    desc_raw = str(alias_desc or "").strip()
    sku_norm = _normalize_sku(sku_raw) if sku_raw else ""
    desc_norm = normalize_descripcion(desc_raw) if desc_raw else ""
    if not sku_norm and not desc_norm:
        raise ValueError("Debe enviar un SKU o una descripcion para vincular.")

    with db_session() as session:
        product = session.get(Product, int(product_id))
        if not product or not product.is_active:
            raise ValueError("Producto de catalogo inexistente o inactivo.")

        duplicate_filters = []
        if sku_norm:
            duplicate_filters.append(SalesBIProductAlias.alias_sku_norm == sku_norm)
        if desc_norm:
            duplicate_filters.append(SalesBIProductAlias.alias_desc_norm == desc_norm)
        existing = session.scalar(select(SalesBIProductAlias).where(or_(*duplicate_filters)).limit(1))
        if existing:
            if int(existing.product_id) != int(product_id):
                raise ValueError("Ese alias ya esta vinculado a otro producto.")
            return _alias_to_dict(existing, product)

        alias = SalesBIProductAlias(
            product_id=int(product_id),
            alias_sku_norm=sku_norm or None,
            alias_desc_norm=desc_norm or None,
            alias_sku_raw=sku_raw,
            alias_desc_raw=desc_raw,
            created_by_user_id=_user_id_from_username(session, username),
            created_at=utc_now_dt(),
        )
        session.add(alias)
        session.flush()
        out = _alias_to_dict(alias, product)
        session.commit()
        return out


def delete_product_alias(alias_id: int) -> bool:
    with db_session() as session:
        alias = session.get(SalesBIProductAlias, int(alias_id))
        if not alias:
            return False
        session.delete(alias)
        session.commit()
        return True


def list_unmatched_products(
    fecha_desde: str | None = None,
    fecha_hasta: str | None = None,
    sucursal: str | None = None,
    tipo: str | None = None,
    q: str | None = None,
    limit: int = 100,
) -> list[dict]:
    filters: list[Any] = [SalesImport.status == "activo"]
    if fecha_desde:
        value = _parse_date_value(fecha_desde)
        if value:
            filters.append(SalesImport.fecha >= value)
    if fecha_hasta:
        value = _parse_date_value(fecha_hasta)
        if value:
            filters.append(SalesImport.fecha <= value)
    if sucursal:
        filters.append(SalesImport.sucursal == sucursal)
    if tipo:
        filters.append(SalesImport.tipo == tipo)
    if q:
        text = f"%{q}%"
        filters.append(or_(SalesRecord.sku.ilike(text), SalesRecord.producto.ilike(text), SalesRecord.marca.ilike(text)))

    unmatched_filter = or_(
        SalesRecord.product_match_status == "unmatched",
        and_(SalesRecord.product_id.is_(None), SalesRecord.product_match_status.is_(None)),
    )

    with db_session() as session:
        rows = session.execute(
            select(SalesRecord, SalesImport)
            .join(SalesImport, SalesImport.id == SalesRecord.import_id)
            .where(*filters, unmatched_filter)
            .order_by(SalesImport.fecha.desc(), SalesRecord.id.desc())
            .limit(max(1, min(limit * 10, 2000)))
        ).all()

    grouped: dict[tuple[str, str], dict[str, Any]] = {}
    for record, imp in rows:
        sku_norm = str(record.sku_normalized or _normalize_sku(record.sku))
        desc_norm = normalize_descripcion(record.producto)
        key = (sku_norm, desc_norm)
        bucket = grouped.setdefault(key, {
            "sku": str(record.sku or ""),
            "sku_normalized": sku_norm,
            "producto": str(record.producto or ""),
            "descripcion_normalized": desc_norm,
            "marca": str(record.marca or ""),
            "lineas": 0,
            "unidades": 0,
            "total_cobrado": 0.0,
            "import_ids": set(),
            "sucursales": set(),
            "first_fecha": _fmt_date(imp.fecha),
            "last_fecha": _fmt_date(imp.fecha),
        })
        bucket["lineas"] += 1
        bucket["unidades"] += int(record.cantidad or 0)
        bucket["total_cobrado"] += _num(record.total_cobrado)
        bucket["import_ids"].add(int(record.import_id))
        bucket["sucursales"].add(str(imp.sucursal or ""))
        fecha_text = _fmt_date(imp.fecha)
        bucket["first_fecha"] = min(bucket["first_fecha"], fecha_text)
        bucket["last_fecha"] = max(bucket["last_fecha"], fecha_text)

    out = []
    for item in grouped.values():
        item["import_ids"] = sorted(item["import_ids"])
        item["sucursales"] = sorted(s for s in item["sucursales"] if s)
        out.append(item)
    out.sort(key=lambda r: (float(r["total_cobrado"]), int(r["unidades"])), reverse=True)
    return out[:limit]


def rematch_import_products(import_id: int) -> dict:
    with db_session() as session:
        imp = session.scalar(
            select(SalesImport)
            .options(selectinload(SalesImport.records))
            .where(SalesImport.id == int(import_id))
        )
        if not imp:
            return {"ok": False, "message": "Importacion no encontrada.", "import_id": import_id}
        indexes, aliases_by_sku, aliases_by_desc = _load_product_match_context(session)
        counts = {"matched": 0, "matched_by_alias": 0, "unmatched": 0}
        for record in imp.records:
            rec = _record_to_dict(record)
            product, sku_norm, status, alias_id = _resolve_record_match(rec, indexes, aliases_by_sku, aliases_by_desc)
            rec = _apply_product_match_to_record(rec, product, sku_norm, status, alias_id)
            record.sku_normalized = str(rec.get("sku_normalized") or "")
            record.product_id = rec.get("product_id")
            record.product_alias_id = rec.get("product_alias_id")
            record.product_match_status = str(rec.get("product_match_status") or "unmatched")
            record.vendedor_normalized = str(record.vendedor_normalized or _normalize_seller(record.vendedor))
            record.marca = str(rec.get("marca") or "")
            record.tipo_producto = str(rec.get("tipo_producto") or "")
            record.categoria = str(rec.get("categoria") or "")
            record.linea = str(rec.get("linea") or "")
            record.costo = _decimal(rec.get("costo"))
            record.diferencia = _decimal(rec.get("diferencia"))
            record.margen_porcentaje = _decimal(rec.get("margen_porcentaje"))
            counts[record.product_match_status] = counts.get(record.product_match_status, 0) + 1
        session.commit()
        return {"ok": True, "import_id": int(import_id), **counts}


def _parse_vendor_filter(vendedores: list[str] | str | None) -> set[str]:
    if not vendedores:
        return set()
    if isinstance(vendedores, str):
        raw = [part for part in vendedores.split(",")]
    else:
        raw = vendedores
    return {_normalize_seller(v) for v in raw if str(v or "").strip()}


def _parse_csv_list(value: list[str] | str | None) -> list[str]:
    """Normaliza un parámetro que puede venir como lista o CSV string."""
    if not value:
        return []
    if isinstance(value, str):
        return [v.strip() for v in value.split(",") if v.strip()]
    return [str(v).strip() for v in value if str(v or "").strip()]


def _sales_rows_for_report(
    fecha_desde: str | None,
    fecha_hasta: str | None,
    sucursal: str | None = None,
    tipo: str | None = None,
    vendedores: list[str] | str | None = None,
    *,
    empresa: str | None = None,
    sucursales: list[str] | str | None = None,
) -> list[tuple[SalesRecord, SalesImport]]:
    """Trae los SalesRecord+SalesImport del rango aplicando filtros.

    Filtros:
        sucursal: nombre legacy (Caseros, Canning, ...) — single. Backward
            compat con el endpoint original.
        sucursales: lista (o CSV) de nombres de sucursal. Si viene junto a
            `sucursal`, los dos se aplican (ambos como `IN`).
        empresa: slug de Company (ej. `electro_gv`, `electro_abc_srl`).
            Filtra los imports que estén linkeados a un branch de esa
            empresa (`sales_imports.branch_id -> branches.company_id`).
        tipo: 'local' | 'online'.
        vendedores: lista o CSV de vendedor_normalized.
    """
    fd = _parse_date_value(fecha_desde) if fecha_desde else None
    fh = _parse_date_value(fecha_hasta) if fecha_hasta else None
    if fd is None or fh is None:
        today = date.today()
        fh = fh or today
        fd = fd or (fh - timedelta(days=29))
    if fh < fd:
        fd, fh = fh, fd

    filters: list[Any] = [SalesImport.status == "activo", SalesImport.fecha >= fd, SalesImport.fecha <= fh]

    # Sucursal(es) — soporta el `sucursal=Caseros` legacy y el nuevo
    # `sucursales=Caseros,Canning`. Si vienen los dos, ambos se aplican.
    sucursales_list = _parse_csv_list(sucursales)
    if sucursal and not sucursales_list:
        filters.append(SalesImport.sucursal == sucursal)
    elif sucursales_list:
        filters.append(SalesImport.sucursal.in_(sucursales_list))

    if tipo:
        filters.append(SalesImport.tipo == tipo)

    # Empresa — slug de la tabla `companies`. Filtra contra los branches
    # que pertenezcan a esa empresa, vía subquery sobre branch_id.
    if empresa:
        from .models.org import Branch  # local import — evita ciclo
        filters.append(
            SalesImport.branch_id.in_(
                select(Branch.id).where(Branch.company_id == empresa)
            )
        )

    vendor_filter = _parse_vendor_filter(vendedores)
    with db_session() as session:
        rows = session.execute(
            select(SalesRecord, SalesImport)
            .join(SalesImport, SalesImport.id == SalesRecord.import_id)
            .where(*filters)
            .order_by(SalesImport.fecha.asc(), SalesRecord.id.asc())
        ).all()

    if vendor_filter:
        rows = [
            (record, imp)
            for record, imp in rows
            if (str(record.vendedor_normalized or "") or _normalize_seller(record.vendedor)) in vendor_filter
        ]
    return rows


def _metric_bucket() -> dict[str, Any]:
    return {
        "total_vendido": 0.0,
        "total_cobrado": 0.0,
        "saldo": 0.0,
        "unidades": 0,
        "lineas": 0,
        "diferencia": 0.0,
        "_tickets": set(),
        "_ticket_totals": {},
    }


def _branch_report_key(imp: SalesImport) -> str:
    return str(imp.sucursal or imp.branch_id or "Sin sucursal").strip() or "Sin sucursal"


def _add_metric(bucket: dict[str, Any], record: SalesRecord) -> None:
    cantidad = int(record.cantidad or 0)
    total_vendido = _num(record.pvp) * cantidad
    total_cobrado = _num(record.total_cobrado)
    saldo = _num(record.saldo)
    bucket["total_vendido"] += total_vendido
    bucket["total_cobrado"] += total_cobrado
    bucket["saldo"] += saldo
    bucket["unidades"] += cantidad
    bucket["lineas"] += 1
    bucket["diferencia"] += _num(record.diferencia)
    ticket_key = str(record.remito or "").strip() or f"linea-{record.id}"
    bucket["_tickets"].add(ticket_key)
    ticket_totals = bucket.setdefault("_ticket_totals", {})
    ticket_bucket = ticket_totals.setdefault(ticket_key, {"total_vendido": 0.0, "total_cobrado": 0.0, "saldo": 0.0})
    ticket_bucket["total_vendido"] += total_vendido
    ticket_bucket["total_cobrado"] += total_cobrado
    ticket_bucket["saldo"] += saldo


def _finalize_metric(bucket: dict[str, Any], *, include_margin: bool, total_reference: float | None = None) -> dict[str, Any]:
    tickets = len(bucket.get("_tickets") or set())
    total_cobrado = float(bucket.get("total_cobrado") or 0.0)
    diferencia = float(bucket.get("diferencia") or 0.0)
    sena_tickets = 0
    sena_total_vendido = 0.0
    sena_monto_cobrado = 0.0
    sena_saldo_pendiente = 0.0
    for ticket in (bucket.get("_ticket_totals") or {}).values():
        ticket_vendido = float(ticket.get("total_vendido") or 0.0)
        ticket_cobrado = float(ticket.get("total_cobrado") or 0.0)
        ticket_saldo = max(float(ticket.get("saldo") or 0.0), ticket_vendido - ticket_cobrado)
        # Sena = remito con cobro parcial. Un remito sin pago cargado queda
        # como saldo, pero no se cuenta como sena comercial.
        if ticket_cobrado > 0.01 and ticket_saldo > 0.01:
            sena_tickets += 1
            sena_total_vendido += ticket_vendido
            sena_monto_cobrado += ticket_cobrado
            sena_saldo_pendiente += ticket_saldo
    out = {
        "total_vendido": round(float(bucket.get("total_vendido") or 0.0), 2),
        "total_cobrado": round(total_cobrado, 2),
        "saldo": round(float(bucket.get("saldo") or 0.0), 2),
        "unidades": int(bucket.get("unidades") or 0),
        "lineas": int(bucket.get("lineas") or 0),
        "tickets": tickets,
        "ticket_promedio": round(total_cobrado / tickets, 2) if tickets else 0.0,
        "participacion_pct": round(total_cobrado / total_reference * 100, 2) if total_reference else 0.0,
        "sena_tickets": sena_tickets,
        "sena_total_vendido": round(sena_total_vendido, 2),
        "sena_monto_cobrado": round(sena_monto_cobrado, 2),
        "sena_saldo_pendiente": round(sena_saldo_pendiente, 2),
        "sena_pct_tickets": round(sena_tickets / tickets * 100, 2) if tickets else 0.0,
        "sena_ticket_promedio": round(sena_monto_cobrado / sena_tickets, 2) if sena_tickets else 0.0,
    }
    if include_margin:
        out["diferencia"] = round(diferencia, 2)
        out["margen_porcentaje"] = round(diferencia / total_cobrado * 100, 2) if total_cobrado else 0.0
    return out


def _delta_metric(base: dict[str, Any], compare: dict[str, Any]) -> dict[str, Any]:
    out: dict[str, Any] = {}
    for key in (
        "total_vendido", "total_cobrado", "saldo", "unidades", "tickets", "ticket_promedio",
        "sena_tickets", "sena_total_vendido", "sena_monto_cobrado", "sena_saldo_pendiente",
        "sena_pct_tickets", "sena_ticket_promedio", "diferencia", "margen_porcentaje",
    ):
        if key not in base and key not in compare:
            continue
        a = float(base.get(key) or 0)
        b = float(compare.get(key) or 0)
        delta = a - b
        out[key] = {
            "actual": a,
            "comparado": b,
            "delta": round(delta, 2),
            "delta_pct": round(delta / b * 100, 2) if b else None,
        }
    return out


def build_sellers_report(
    fecha_desde: str | None = None,
    fecha_hasta: str | None = None,
    sucursal: str | None = None,
    tipo: str | None = None,
    vendedores: list[str] | str | None = None,
    *,
    empresa: str | None = None,
    sucursales: list[str] | str | None = None,
    include_costs: bool = False,
    include_margin: bool = False,
) -> dict[str, Any]:
    rows = _sales_rows_for_report(
        fecha_desde, fecha_hasta, sucursal, tipo, vendedores,
        empresa=empresa, sucursales=sucursales,
    )
    fd = _parse_date_value(fecha_desde) if fecha_desde else None
    fh = _parse_date_value(fecha_hasta) if fecha_hasta else None
    if fd is None or fh is None:
        today = date.today()
        fh = fh or today
        fd = fd or (fh - timedelta(days=29))
    if fh < fd:
        fd, fh = fh, fd

    overall = _metric_bucket()
    sellers: dict[str, dict[str, Any]] = {}
    seller_branches: dict[str, dict[str, dict[str, Any]]] = defaultdict(lambda: defaultdict(_metric_bucket))
    branch_sellers: dict[str, dict[str, dict[str, Any]]] = defaultdict(lambda: defaultdict(_metric_bucket))
    branch_totals: dict[str, dict[str, Any]] = defaultdict(_metric_bucket)
    daily: dict[str, dict[str, Any]] = {}
    payments = {key: 0.0 for key in ("efectivo", "transferencia", "tarjeta", "usd", "cuenta_corriente", "otros")}
    categories: dict[str, dict[str, Any]] = defaultdict(_metric_bucket)
    brands: dict[str, dict[str, Any]] = defaultdict(_metric_bucket)
    products: dict[tuple[str, str], dict[str, Any]] = {}
    unmatched_count = 0

    for record, imp in rows:
        seller_key = str(record.vendedor_normalized or "") or _normalize_seller(record.vendedor)
        seller_label = str(record.vendedor or "").strip() or "Sin vendedor"
        if seller_key not in sellers:
            sellers[seller_key] = {"vendedor": seller_label, "vendedor_normalized": seller_key, **_metric_bucket()}
        _add_metric(sellers[seller_key], record)
        _add_metric(overall, record)
        branch_key = _branch_report_key(imp)
        _add_metric(seller_branches[seller_key][branch_key], record)
        _add_metric(branch_sellers[branch_key][seller_key], record)
        _add_metric(branch_totals[branch_key], record)

        day_key = _fmt_date(imp.fecha)
        if day_key not in daily:
            daily[day_key] = {"fecha": day_key, **_metric_bucket()}
        _add_metric(daily[day_key], record)

        for field in payments:
            payments[field] += _num(getattr(record, field))

        category_key = str(record.categoria or "Sin categoria")
        brand_key = str(record.marca or "Sin marca")
        _add_metric(categories[category_key], record)
        _add_metric(brands[brand_key], record)

        product_key = (str(record.sku or ""), str(record.producto or ""))
        if product_key not in products:
            products[product_key] = {
                "sku": product_key[0],
                "producto": product_key[1],
                "marca": str(record.marca or ""),
                **_metric_bucket(),
            }
        _add_metric(products[product_key], record)

        if str(record.product_match_status or "unmatched") == "unmatched":
            unmatched_count += 1

    totals = _finalize_metric(overall, include_margin=include_margin)
    total_reference = float(totals["total_cobrado"] or 0)

    branch_reference_totals = {
        key: float(_finalize_metric(value, include_margin=False).get("total_cobrado") or 0.0)
        for key, value in branch_totals.items()
    }
    branch_rankings: dict[str, dict[str, int]] = {}
    for branch_key, branch_seller_map in branch_sellers.items():
        ranked = sorted(
            branch_seller_map.items(),
            key=lambda pair: (float(pair[1].get("total_cobrado") or 0.0), int(pair[1].get("unidades") or 0)),
            reverse=True,
        )
        branch_rankings[branch_key] = {seller_key: idx + 1 for idx, (seller_key, _) in enumerate(ranked)}

    seller_items = []
    for item in sellers.values():
        seller_key = str(item["vendedor_normalized"])
        metric = _finalize_metric(item, include_margin=include_margin, total_reference=total_reference)
        seller_branch_map = seller_branches.get(seller_key, {})
        primary_branch_key = ""
        if seller_branch_map:
            primary_branch_key = max(
                seller_branch_map,
                key=lambda key: (
                    float(seller_branch_map[key].get("total_cobrado") or 0.0),
                    int(seller_branch_map[key].get("unidades") or 0),
                ),
            )
        branch_reference = branch_reference_totals.get(primary_branch_key, 0.0)
        branch_metric = (
            _finalize_metric(seller_branch_map[primary_branch_key], include_margin=include_margin, total_reference=branch_reference)
            if primary_branch_key else {}
        )
        seller_items.append({
            "vendedor": item["vendedor"],
            "vendedor_normalized": seller_key,
            "sucursal": primary_branch_key,
            "sucursales": sorted(seller_branch_map.keys()),
            "empresa_total_cobrado": total_reference,
            "sucursal_total_cobrado": round(branch_reference, 2),
            "empresa_participacion_pct": metric["participacion_pct"],
            "sucursal_participacion_pct": float(branch_metric.get("participacion_pct") or 0.0),
            "rank_sucursal": branch_rankings.get(primary_branch_key, {}).get(seller_key, 0),
            "sellers_en_sucursal": len(branch_sellers.get(primary_branch_key, {})) if primary_branch_key else 0,
            "sellers_en_empresa": len(sellers),
            **metric,
        })
    seller_items.sort(key=lambda r: (float(r["total_cobrado"]), int(r["unidades"])), reverse=True)
    for idx, seller in enumerate(seller_items, start=1):
        seller["rank_empresa"] = idx

    daily_series = [
        {"fecha": key, **_finalize_metric(value, include_margin=include_margin)}
        for key, value in sorted(daily.items())
    ]

    def ranked_mix(source: dict[str, dict[str, Any]], limit: int = 12) -> list[dict[str, Any]]:
        out = [{"name": k, **_finalize_metric(v, include_margin=include_margin, total_reference=total_reference)} for k, v in source.items()]
        out.sort(key=lambda r: float(r["total_cobrado"]), reverse=True)
        return out[:limit]

    top_products = [
        {
            "sku": value["sku"],
            "producto": value["producto"],
            "marca": value["marca"],
            **_finalize_metric(value, include_margin=include_margin, total_reference=total_reference),
        }
        for value in products.values()
    ]
    top_products.sort(key=lambda r: float(r["total_cobrado"]), reverse=True)

    detail = [_record_to_dict(record, imp) for record, imp in rows]
    if not include_costs:
        for rec in detail:
            rec.pop("costo", None)
            rec.pop("diferencia", None)
    if not include_margin:
        for rec in detail:
            rec.pop("margen_porcentaje", None)

    return {
        "filters": {
            "fecha_desde": fd.isoformat(),
            "fecha_hasta": fh.isoformat(),
            "sucursal": sucursal or "",
            "tipo": tipo or "",
            "vendedores": sorted(_parse_vendor_filter(vendedores)),
        },
        "totals": totals,
        "sellers": seller_items,
        "daily_series": daily_series,
        "payment_mix": [{"name": k, "value": round(v, 2)} for k, v in payments.items() if v],
        "category_mix": ranked_mix(categories),
        "brand_mix": ranked_mix(brands),
        "top_products": top_products[:20],
        "unmatched_count": unmatched_count,
        "detail": detail,
    }


def compare_sellers_report(
    base_desde: str,
    base_hasta: str,
    compare_desde: str,
    compare_hasta: str,
    sucursal: str | None = None,
    tipo: str | None = None,
    vendedores: list[str] | str | None = None,
    *,
    empresa: str | None = None,
    sucursales: list[str] | str | None = None,
    include_costs: bool = False,
    include_margin: bool = False,
) -> dict[str, Any]:
    base = build_sellers_report(
        base_desde, base_hasta, sucursal, tipo, vendedores,
        empresa=empresa, sucursales=sucursales,
        include_costs=include_costs, include_margin=include_margin,
    )
    compare = build_sellers_report(
        compare_desde, compare_hasta, sucursal, tipo, vendedores,
        empresa=empresa, sucursales=sucursales,
        include_costs=include_costs, include_margin=include_margin,
    )
    base_by_seller = {s["vendedor_normalized"]: s for s in base["sellers"]}
    compare_by_seller = {s["vendedor_normalized"]: s for s in compare["sellers"]}
    sellers = []
    for key in sorted(set(base_by_seller) | set(compare_by_seller)):
        b = base_by_seller.get(key, {"vendedor": key, "vendedor_normalized": key})
        c = compare_by_seller.get(key, {"vendedor": b.get("vendedor", key), "vendedor_normalized": key})
        sellers.append({
            "vendedor": b.get("vendedor") or c.get("vendedor") or key,
            "vendedor_normalized": key,
            "delta": _delta_metric(b, c),
        })
    sellers.sort(key=lambda r: float(r["delta"].get("total_cobrado", {}).get("actual") or 0), reverse=True)
    return {
        "base": {k: v for k, v in base.items() if k != "detail"},
        "compare": {k: v for k, v in compare.items() if k != "detail"},
        "delta": _delta_metric(base["totals"], compare["totals"]),
        "sellers": sellers,
    }


def get_sellers_filter_options() -> dict[str, Any]:
    """Devuelve empresas + sucursales disponibles para los filtros del dashboard.

    Las sucursales se sirven con el TEXTO que aparece en `sales_imports.sucursal`
    (legacy: "Caseros", "Canning", "Norcenter", "Lanus") agrupado por empresa
    a través del branch al que están linkeadas.

    Estructura:
        {
          "empresas": [{"id": "electro_gv", "name": "Electro GV"}, ...],
          "sucursales": [
              {"name": "Caseros", "empresa_id": "electro_gv",
               "branch_ids": ["caseros", "caseros_web"]},
              ...
          ],
        }

    Nota operativa: en las planillas ABC, Lanús aparece como "Sur" y
    Norcenter como "Norte" (codename interno). La importación las
    mapea al texto user-friendly ("Lanus"/"Norcenter") antes de
    guardar en `sales_imports.sucursal`, así que el filtro del
    dashboard ya las muestra con el nombre que la gente usa.
    """
    from .models.org import Company, Branch  # local: evita ciclo

    empresas: list[dict[str, str]] = []
    sucursales_by_name: dict[str, dict[str, Any]] = {}

    with db_session() as session:
        for company in session.scalars(
            select(Company).where(Company.is_active.is_(True)).order_by(Company.name)
        ).all():
            empresas.append({"id": str(company.id), "name": str(company.name)})

        # Junto branches con los nombres reales que aparecen en sales_imports
        # (puede que un branch nuevo aún no tenga imports — lo incluimos igual
        # para que el usuario pueda preseleccionarlo y filtrar sin resultados).
        used_sucursales: set[tuple[str, str]] = set()  # (sucursal_text, company_id)
        for row in session.execute(
            select(SalesImport.sucursal, Branch.company_id)
            .join(Branch, Branch.id == SalesImport.branch_id, isouter=True)
            .where(SalesImport.status == "activo")
            .distinct()
        ).all():
            sucursal_text = str(row[0] or "").strip()
            company_id = str(row[1] or "")
            if sucursal_text:
                used_sucursales.add((sucursal_text, company_id))

        # Agrupo por nombre de sucursal: cada uno trae todos los branch_ids
        # que matchean (física + web cuando existen).
        for sucursal_text, company_id in used_sucursales:
            entry = sucursales_by_name.setdefault(
                sucursal_text,
                {"name": sucursal_text, "empresa_id": company_id, "branch_ids": []},
            )
            entry["empresa_id"] = entry["empresa_id"] or company_id

        for branch in session.scalars(
            select(Branch).where(Branch.is_active.is_(True))
        ).all():
            for entry in sucursales_by_name.values():
                # Branch matchea si su código empieza con el slug de la sucursal
                # (`caseros` → "Caseros"; `caseros_web` también).
                slug = str(branch.id or "")
                base = slug.replace("_web", "")
                sucursal_slug = entry["name"].lower().replace(" ", "_")
                # Caso especial: ABC usa codename "Sur" → branch `sur` (Lanús),
                # "Norte" → branch `norte` (Norcenter). Lo mapeamos a mano.
                aliases = {"lanus": ["sur"], "norcenter": ["norte"]}
                slugs_to_check = [sucursal_slug] + aliases.get(sucursal_slug, [])
                if base in slugs_to_check:
                    if slug not in entry["branch_ids"]:
                        entry["branch_ids"].append(slug)
                    if not entry["empresa_id"]:
                        entry["empresa_id"] = str(branch.company_id or "")

    sucursales = sorted(sucursales_by_name.values(), key=lambda r: str(r["name"]).lower())
    return {"empresas": empresas, "sucursales": sucursales}


def get_stats() -> dict:
    with db_session() as session:
        total_imports = int(session.scalar(
            select(func.count()).select_from(SalesImport).where(SalesImport.status == "activo")
        ) or 0)
        total_records = int(session.scalar(
            select(func.count())
            .select_from(SalesRecord)
            .join(SalesImport, SalesImport.id == SalesRecord.import_id)
            .where(SalesImport.status == "activo")
        ) or 0)
        sum_pvp = _num(session.scalar(
            select(func.coalesce(func.sum(SalesRecord.pvp * SalesRecord.cantidad), 0))
            .select_from(SalesRecord)
            .join(SalesImport, SalesImport.id == SalesRecord.import_id)
            .where(SalesImport.status == "activo")
        ))
        last_import = session.scalar(
            select(SalesImport)
            .where(SalesImport.status == "activo")
            .order_by(SalesImport.id.desc())
            .limit(1)
        )
    return {
        "total_imports": total_imports,
        "total_records": total_records,
        "total_pvp": sum_pvp,
        "last_import": {
            "fecha": _fmt_date(last_import.fecha),
            "sucursal": str(last_import.sucursal or ""),
            "created_at": _fmt_dt(last_import.created_at),
        } if last_import else None,
    }
