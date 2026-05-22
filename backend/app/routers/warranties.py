from __future__ import annotations

import json
import re
import sqlite3
import threading
import time
import unicodedata
from datetime import date, datetime, timezone
from pathlib import Path
from uuid import uuid4
from typing import Annotated, Any
from zoneinfo import ZoneInfo

from fastapi import APIRouter, Depends, HTTPException, Query, status
from fastapi.responses import FileResponse
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.units import mm
from reportlab.pdfgen import canvas
from reportlab.platypus import Table, TableStyle, Paragraph
from reportlab.lib.styles import ParagraphStyle
from pydantic import BaseModel, Field, model_validator
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table as XLTable, TableStyleInfo

from ..audit import audit
from ..auth import require_current_user, require_permission
from ..config import get_settings
from ..google_sheets import quote_sheet_name, sheets_service
from ..operational_config import runtime_warranty_config, load_operational_config, save_operational_config
from ..permissions import has_permission
from ..product_catalog import search_products as search_local_products, get_provider_for_brand, ensure_product_catalog_tables, runtime_product_catalog_config
from ..users import load_roles, load_users
from .notifications import notify_many

router = APIRouter(prefix="/api/warranties", tags=["warranties"])

# La app pasa a operar Garantías desde DB propia.
# Google Sheet queda como fuente auxiliar para productos/opciones y espejo futuro.

DEFAULT_STATUSES = [
    "1 - INGRESO",
    "2 - PENDIENTE",
    "3 - LISTO PARA ENVIAR",
    "4 - ENVIADO AL PROVEEDOR",
    "5 - EN EL PROVEEDOR",
    "6 - RESPONDIDO POR PROVEEDOR",
    "7 - RESUELTO",
    "8 - RECHAZADO",
    "9 - ANULADA",
    "10 - FINALIZADO",
]
DEFAULT_SUCURSALES = ["1 - CASEROS", "2 - LANUS", "3 - CANNING", "4 - NORCENTER"]
DEFAULT_DEPOSITOS = ["6 - CHICLANA", "7 - CORRALES", "8 - CACHI"]
DEFAULT_FINAL_STATUSES = ["10 - FINALIZADO"]

# Sub-tipos de resolución cuando estado = "7 - RESUELTO".
# Fase 12: solo se consideran resoluciones finales reales del proveedor.
# RESUELTO no equivale a FINALIZADO; el cierre se hace aparte con estado 10.
RESOLUTION_OPTIONS = {
    "nota_credito": "Nota de crédito",
    "reparacion": "Reparación",
    "cambio_equipo": "Cambio de equipo",
}
RESOLUTION_ALIASES = {
    "nc": "nota_credito",
    "nota de credito": "nota_credito",
    "nota de crédito": "nota_credito",
    "cambio": "cambio_equipo",
    "cambio_aprobado": "cambio_equipo",
    "cambio aprobado": "cambio_equipo",
    "reparado": "reparacion",
    "reparación": "reparacion",
}

def normalize_resolution_result(value: Any) -> str:
    raw = str(value or "").strip().lower()
    if not raw:
        return ""
    key = raw.replace("-", "_").replace(" ", "_")
    if key in RESOLUTION_OPTIONS:
        return key
    text_key = raw.replace("_", " ")
    return RESOLUTION_ALIASES.get(key) or RESOLUTION_ALIASES.get(text_key) or key

# Estado logístico del retiro del proveedor.
# No reemplaza al estado principal: indica si el proveedor ya pidió retirar y
# si la mercadería está lista físicamente para entregar.
PROVIDER_PICKUP_STATUSES = {
    "sin_solicitud": "Sin solicitud",
    "retiro_solicitado": "Retiro solicitado",
    "listo_para_retiro": "Listo para retiro",
    "retirado": "Retirado por proveedor",
}

# Fase 26 — roles/acciones: un usuario operativo de depósito NO es gestor.
# Aunque el catálogo de permisos haya quedado viejo o se le hayan colado permisos,
# si su único rol operativo es DEPOSITO no debe poder revisar, exportar, gestionar proveedor
# ni resolver garantías. Debe poder cargar cliente en depósito, recibir remitos y mover depósito→depósito.
WARRANTY_PRIVILEGED_ROLES = {"SUPERADMIN", "GERENTE", "ADMINISTRADOR", "ADMIN", "GESTOR", "GESTOR_GARANTIAS", "JEFE_POSVENTA"}

def _user_role_keys(user: Any) -> set[str]:
    values = [getattr(user, "role", ""), *(getattr(user, "roles", []) or [])]
    return {str(v or "").strip().upper() for v in values if str(v or "").strip()}

def is_plain_deposit_operator(user: Any) -> bool:
    roles = _user_role_keys(user)
    if "DEPOSITO" not in roles:
        return False
    return not bool(roles & WARRANTY_PRIVILEGED_ROLES)

def deny_plain_deposit_operator(user: Any, action: str = "esta acción") -> None:
    if is_plain_deposit_operator(user):
        raise HTTPException(status_code=403, detail=f"El rol Depósito operativo no puede realizar {action}.")

def normalize_provider_pickup_status(value: Any) -> str:
    raw = str(value or "").strip().lower().replace(" ", "_").replace("-", "_")
    return raw if raw in PROVIDER_PICKUP_STATUSES else "sin_solicitud"

DEFAULT_DELAY_RANGES = [3, 7, 14, 30]
DEFAULT_REQUIRED_REVIEW_FIELDS = ["producto", "sku", "marca", "serie", "falla", "sucursal", "deposito"]
CANCELLED_STATUS = "ANULADA"
REVIEW_PENDING = "pendiente_revision"
REVIEW_IN_PROGRESS = "en_revision"
REVIEW_INCOMPLETE = "requiere_correccion"
REVIEW_APPROVED = "revisada"
REVIEW_LABELS = {
    REVIEW_PENDING: "Pendiente de revisión",
    REVIEW_IN_PROGRESS: "En revisión interna",
    REVIEW_INCOMPLETE: "Requiere corrección",
    REVIEW_APPROVED: "Revisada",
}

# ── Tipos de ingreso ──────────────────────────────────────────────────────────
# Cómo llegó la garantía al sistema. Determina origen_ingreso y ubicacion_actual.
VALID_TIPO_INGRESO = [
    "cliente_sucursal",           # cliente dejó el producto en una sucursal
    "cliente_deposito",           # cliente trajo el producto directamente al depósito
    "falla_recepcion_mercaderia", # falla detectada al recibir mercadería en depósito
    "stock_interno",              # problema detectado en stock propio
    "otro",                       # cualquier otro caso
]
# Tipos permitidos para un usuario VENDEDOR (rol sin warranties.manage).
# Todo ingreso desde sucursal debe ser cliente_sucursal.
TIPOS_INGRESO_VENDEDOR = {"cliente_sucursal"}
TIPO_INGRESO_LABELS: dict[str, str] = {
    "cliente_sucursal":           "Cliente en sucursal",
    "cliente_deposito":           "Cliente en depósito",
    "falla_recepcion_mercaderia": "Falla al recibir mercadería",
    "stock_interno":              "Stock interno",
    "otro":                       "Otro",
}
UBICACION_LABELS: dict[str, str] = {
    "sucursal":              "En sucursal",
    "en_transito":           "En tránsito",
    "deposito":              "En depósito",
    "en_transito_proveedor": "En tránsito al proveedor",
    "proveedor":             "En el proveedor",
    "devuelto":              "Devuelto",
    "entregado":             "Entregado al cliente",
    "desconocida":           "Ubicación desconocida",
}

DEFAULT_RAW_HEADERS = [
    "ID GARANTIA",
    "RESPONSABLE",
    "INGRESO",
    "PRODUCTO",
    "SKU",
    "MARCA",
    "SERIE",
    "FALLA",
    "SUCURSAL",
    "DEPOSITO",
    "ESTADO",
    "DIAS PENDIENTE",
    "FECHA DE INICIO DE GESTION",
    "ID DE CASO",
    "DIAS SIN RESPUESTA",
    "FECHA DE RETIRO",
    "FECHA DE RESOLUCION",
    "OBSERVACIONES",
    "VUELVE A",
    "FINALIZACION",
    "TIPO",
    "LUGAR LLEGADA",
    "USUARIO",
    "FECHA ULTIMA ACTUALIZACION",
    "ACTUALIZADO POR",
]

# Fase 27 — Espejo Google Sheets.
# La app sigue siendo fuente principal; estas pestañas son espejo/reporting.
# 00_RAW_GARANTIAS se mantiene por compatibilidad histórica/importación.
MIRROR_SHEETS: dict[str, list[str]] = {
    "GARANTIAS": [
        "ID GARANTIA", "FECHA INGRESO", "EMPRESA", "SUCURSAL CARGA", "SUCURSAL RESPONSABLE",
        "ORIGEN INGRESO", "TIPO INGRESO", "UBICACION ACTUAL", "DEPOSITO DESTINO",
        "ESTADO", "REVISION", "REMITO INTERNO", "ENV", "PROVEEDOR",
        "FECHA ENVIO PROVEEDOR", "FECHA ULTIMO MAIL", "DIAS SIN RESPUESTA",
        "RETIRO PROVEEDOR", "FECHA RETIRO PROVEEDOR", "RESPUESTA PROVEEDOR",
        "RESULTADO RESOLUCION", "FECHA RESOLUCION", "FECHA FINALIZACION",
        "CLIENTE", "TELEFONO", "EMAIL", "FACTURA", "FECHA COMPRA",
        "RESPONSABLE", "CREADO POR", "ACTUALIZADO", "ACTUALIZADO POR", "OBSERVACIONES",
    ],
    "GARANTIA_ITEMS": [
        "ID GARANTIA", "ITEM", "PRODUCTO", "SKU", "MARCA", "TIPO", "SERIE", "FALLA",
        "PROVEEDOR", "OBSERVACIONES", "UPDATED_AT",
    ],
    "REMITOS": [
        "CODIGO REMITO", "TIPO REMITO", "EMPRESA", "ORIGEN", "DESTINO", "ESTADO",
        "FECHA CREACION", "FECHA DESPACHO", "FECHA LLEGADA", "CREADO POR", "DESPACHADO POR",
        "RECIBIDO POR", "CANTIDAD ITEMS", "PDF", "NOTA",
    ],
    "REMITO_ITEMS": [
        "CODIGO REMITO", "ID GARANTIA", "PRODUCTO", "SKU", "SERIE", "SUCURSAL RESPONSABLE",
        "ORIGEN", "DESTINO", "ESTADO REMITO", "UPDATED_AT",
    ],
    "LOTES_ENV": [
        "ENV", "PROVEEDOR", "MARCA", "ESTADO LOTE", "FECHA CREACION", "CREADO POR",
        "ARCHIVO EXCEL", "CANTIDAD ITEMS", "FECHA MAIL", "FECHA ULTIMO MAIL", "OBSERVACIONES",
    ],
    "LOTE_ITEMS": [
        "ENV", "ID GARANTIA", "PRODUCTO", "SKU", "SERIE", "PROVEEDOR", "ESTADO GARANTIA",
        "RESPUESTA PROVEEDOR", "RESULTADO RESOLUCION", "UPDATED_AT",
    ],
    "EVENTOS": [
        "FECHA", "ID GARANTIA", "USUARIO", "NOMBRE", "ACCION", "ESTADO ANTERIOR", "ESTADO NUEVO",
        "REVISION ANTERIOR", "REVISION NUEVA", "DETALLE", "METADATA",
    ],
}

PRODUCT_CACHE: dict[str, Any] = {"loaded_at": 0.0, "items": []}
COUNTER_LOCK = threading.RLock()


class WarrantyItemIn(BaseModel):
    # ── Datos del artículo ─────────────────────────────────────────────────────
    producto: str = Field(min_length=1)
    sku: str | None = None
    marca: str | None = None
    tipo: str | None = None
    serie: str | None = None
    falla: str = Field(min_length=1)
    observaciones: str | None = None
    # ── Origen y ubicación ─────────────────────────────────────────────────────
    tipo_ingreso: str = Field(min_length=1)     # obligatorio — determina origen y ubicación inicial
    sucursal: str = ""                          # obligatorio solo si tipo_ingreso = "cliente_sucursal"
    deposito: str = Field(min_length=1)
    lugar_llegada: str | None = None
    # ── Sucursal comercialmente responsable (depósito/gestor solo) ───────────
    # sucursal_responsable_id: branch_id real del sistema (preferido).
    # sucursal_responsable: texto de display (se deriva del ID si se conoce; fallback legado).
    # Los VENDEDORES no envían ninguno — el backend los deriva de su asignación.
    sucursal_responsable_id: str = ""   # branch_id de la sucursal responsable
    sucursal_responsable: str = ""      # nombre de display (derivado del ID o texto libre)
    # ── Proveedor (sugerido por catálogo, opcional) ────────────────────────────
    proveedor: str | None = None
    # ── Datos del cliente (opcionales) ────────────────────────────────────────
    cliente_nombre: str | None = None
    cliente_telefono: str | None = None
    cliente_email: str | None = None
    numero_factura: str | None = None
    fecha_compra: str | None = None
    # Fecha real en la que ingresó físicamente/operativamente la garantía.
    # Si no viene, se usa la fecha/hora actual.
    fecha_ingreso: str | None = None

    @model_validator(mode="after")
    def _validate_tipo_and_sucursal(self) -> "WarrantyItemIn":
        tipo = (self.tipo_ingreso or "").strip()
        if tipo not in VALID_TIPO_INGRESO:
            raise ValueError(
                f"tipo_ingreso inválido: '{tipo}'. "
                f"Valores permitidos: {', '.join(VALID_TIPO_INGRESO)}"
            )
        if tipo == "cliente_sucursal" and not (self.sucursal or "").strip():
            raise ValueError(
                "sucursal es obligatoria cuando tipo_ingreso es 'cliente_sucursal'"
            )
        return self


class WarrantyCreateRequest(BaseModel):
    items: list[WarrantyItemIn] = Field(min_length=1, max_length=100)
    group_under_one_id: bool = False


class WarrantyCreatedItem(BaseModel):
    id_garantia: str
    producto: str
    sku: str | None = None
    parent_warranty_code: str = ""
    parent_item_index: int | None = None


class WarrantyCreateResponse(BaseModel):
    ok: bool
    count: int
    ids: list[str]
    items: list[WarrantyCreatedItem]


class WarrantyRow(BaseModel):
    row_number: int
    id_garantia: str
    responsable: str = ""
    usuario: str = ""
    ingreso: str = ""
    producto: str = ""
    sku: str = ""
    marca: str = ""
    tipo: str = ""
    serie: str = ""
    falla: str = ""
    sucursal: str = ""
    deposito: str = ""
    lugar_llegada: str = ""
    estado: str = ""
    observaciones: str = ""
    actualizado_por: str = ""
    fecha_ultima_actualizacion: str = ""
    cancelled: bool = False
    cancel_reason: str = ""
    cancelled_by: str = ""
    cancelled_at: str = ""


class WarrantySummary(BaseModel):
    id_garantia: str
    parent_warranty_code: str = ""
    parent_item_index: int | None = None
    grouped_item_label: str = ""
    ingreso: str = ""
    ingreso_iso: str = ""
    responsible_username: str = ""
    responsable: str = ""
    usuario: str = ""
    producto_principal: str = ""
    productos: list[str] = []
    cantidad_items: int = 0
    marca: str = ""
    sku: str = ""
    serie: str = ""
    falla: str = ""
    sucursal: str = ""
    sucursal_code: str = ""
    branch_id: str = ""
    company_id: str = ""
    sucursal_responsable: str = ""
    sucursal_responsable_id: str = ""
    deposito: str = ""
    lugar_llegada: str = ""
    estado: str = ""
    review_status: str = "pendiente_revision"
    review_status_label: str = "Pendiente de revisión"
    reviewed_by: str = ""
    reviewed_by_name: str = ""
    reviewed_at: str = ""
    review_note: str = ""
    observaciones: str = ""
    photos_reference: str = ""
    # ── Origen / tipo / ubicación física (Fase 1) ──────────────────────────────
    tipo_ingreso: str = ""
    tipo_ingreso_label: str = ""
    origen_ingreso: str = ""
    ubicacion_actual: str = ""
    ubicacion_actual_label: str = ""
    # ── Datos del cliente (Fase 1) ─────────────────────────────────────────────
    cliente_nombre: str = ""
    cliente_telefono: str = ""
    cliente_email: str = ""
    numero_factura: str = ""
    fecha_compra: str = ""
    # ── Proveedor / gestión ────────────────────────────────────────────────────
    provider_name: str = ""
    id_de_caso: str = ""
    fecha_envio_proveedor: str = ""
    fecha_ultima_respuesta: str = ""
    fecha_ultimo_reclamo: str = ""
    estado_retiro_proveedor: str = "sin_solicitud"
    estado_retiro_proveedor_label: str = "Sin solicitud"
    fecha_solicitud_retiro_proveedor: str = ""
    fecha_retiro_proveedor: str = ""
    dias_pendiente: int = 0
    dias_sin_respuesta: int | None = None
    shipment_code: str = ""
    shipment_file_name: str = ""
    resolution_note: str = ""
    resolution_reference: str = ""
    resultado_resolucion: str = ""
    resultado_resolucion_label: str = ""
    numero_nota_credito: str = ""
    importe_nota_credito: str = ""
    fecha_nota_credito: str = ""
    detalle_reparacion: str = ""
    fecha_reparacion: str = ""
    producto_reemplazo: str = ""
    sku_reemplazo: str = ""
    serie_reemplazo: str = ""
    fecha_recepcion_reemplazo: str = ""
    fecha_finalizacion: str = ""
    finalizacion: str = ""
    remito_interno: str = ""
    remito_proveedor: str = ""
    transit_status: str = ""  # '' | 'en_transito' | 'en_deposito'
    synced_to_google_sheet: bool = False
    fecha_ultima_sincronizacion: str = ""
    actualizado_por: str = ""
    fecha_ultima_actualizacion: str = ""
    cancelled: bool = False
    cancel_reason: str = ""
    cancelled_by: str = ""
    cancelled_at: str = ""


class WarrantyListResponse(BaseModel):
    items: list[WarrantySummary]
    total: int
    limit: int


class WarrantyDetailResponse(BaseModel):
    summary: WarrantySummary
    rows: list[WarrantyRow]
    history: list[dict[str, Any]]


class WarrantyItemUpdateRequest(BaseModel):
    row_number: int
    producto: str | None = None
    sku: str | None = None
    marca: str | None = None
    tipo: str | None = None
    serie: str | None = None
    falla: str | None = None
    observaciones: str | None = None


class WarrantyUpdateRequest(BaseModel):
    estado: str | None = None
    sucursal: str | None = None
    deposito: str | None = None
    lugar_llegada: str | None = None
    ubicacion_actual: str | None = None
    sucursal_responsable: str | None = None
    observaciones: str | None = None
    photos_reference: str | None = None
    append_observation: str | None = None
    items: list[WarrantyItemUpdateRequest] | None = None


class WarrantyEntryBaseUpdateRequest(BaseModel):
    """Edición controlada de la base del ingreso.

    Pensado para corregir datos recién cargados antes de que avancen a revisión/gestión.
    No permite cambiar estado, remitos, ENV ni proveedor operativo.
    """
    fecha_ingreso: str | None = None
    observaciones: str | None = None
    photos_reference: str | None = None
    proveedor: str | None = None
    cliente_nombre: str | None = None
    cliente_telefono: str | None = None
    cliente_email: str | None = None
    numero_factura: str | None = None
    fecha_compra: str | None = None
    items: list[WarrantyItemUpdateRequest] | None = None


class WarrantyReviewRequest(BaseModel):
    note: str | None = None


class WarrantyProviderSendRequest(BaseModel):
    provider_name: str = Field(min_length=1)
    provider_case_id: str | None = None
    note: str | None = None


class WarrantyProviderResponseRequest(BaseModel):
    note: str | None = None
    provider_case_id: str | None = None
    estado: str | None = None


class WarrantyClaimRequest(BaseModel):
    note: str = Field(min_length=1)


class WarrantyResendMailRequest(BaseModel):
    note: str | None = None


class WarrantyStatusChangeRequest(BaseModel):
    estado: str = Field(min_length=1)
    note: str | None = None
    resolution_note: str | None = None       # Detalle de la resolución (motivo rechazo, descripción reparación, etc.)
    resolution_reference: str | None = None  # Referencia numérica (N° NC, remito cambio, etc.)
    resultado_resolucion: str | None = None  # nota_credito | reparacion | cambio_equipo (requerido si estado = "7 - RESUELTO")
    numero_nota_credito: str | None = None
    importe_nota_credito: str | None = None
    fecha_nota_credito: str | None = None
    detalle_reparacion: str | None = None
    fecha_reparacion: str | None = None
    producto_reemplazo: str | None = None
    sku_reemplazo: str | None = None
    serie_reemplazo: str | None = None
    fecha_recepcion_reemplazo: str | None = None
    finalizacion: str | None = None


class WarrantyExportRequest(BaseModel):
    marca: str | None = None
    proveedor: str | None = None
    estado: str | None = None
    sucursal: str | None = None
    deposito: str | None = None
    fecha_desde: str | None = None
    fecha_hasta: str | None = None


class WarrantyExportInfo(BaseModel):
    id: int
    created_at: str
    created_by: str = ""
    provider_name: str = ""
    marca: str = ""
    filters: dict[str, Any] = {}
    file_name: str
    row_count: int
    download_url: str
    shipment_code: str = ""
    file_format: str = "excel"
    logo_brand: str = "gv_electro"


class WarrantyExportListResponse(BaseModel):
    items: list[WarrantyExportInfo]


class WarrantyBatchExportRequest(BaseModel):
    warranty_ids: list[str] = Field(min_length=1)
    proveedor: str | None = None
    nota: str | None = None
    formato: str | None = "excel"
    logo_brand: str | None = "gv_electro"


def normalize_export_format(value: Any) -> str:
    raw = str(value or "excel").strip().lower()
    return "pdf" if raw == "pdf" else "excel"


def normalize_export_logo(value: Any) -> str:
    raw = str(value or "gv_electro").strip().lower().replace("-", "_")
    if raw in {"abc", "abc_electro", "electro_abc"}:
        return "abc_electro"
    return "gv_electro"


class ConfirmShipmentRequest(BaseModel):
    shipment_code: str = Field(min_length=1)
    provider_name: str | None = None
    nota: str | None = None

class ProviderPickupRequest(BaseModel):
    note: str | None = None
    provider_case_id: str | None = None
    fecha_retiro_acordada: str | None = None


class WarrantySyncStatus(BaseModel):
    last_sync_at: str = ""
    last_sync_type: str = ""
    last_sync_status: str = ""
    last_sync_user: str = ""
    pending_to_sheet: int = 0
    total_guarantees: int = 0
    errors: list[str] = []


class WarrantySyncResult(BaseModel):
    ok: bool
    sync_type: str
    status: str
    started_at: str
    finished_at: str
    rows_processed: int = 0
    rows_created: int = 0
    rows_updated: int = 0
    rows_skipped: int = 0
    errors: list[str] = []


class WarrantySyncLogInfo(BaseModel):
    id: int
    sync_type: str
    status: str
    started_at: str
    finished_at: str
    actor_username: str = ""
    actor_name: str = ""
    rows_processed: int = 0
    rows_created: int = 0
    rows_updated: int = 0
    rows_skipped: int = 0
    errors: list[str] = []


class WarrantySyncLogsResponse(BaseModel):
    items: list[WarrantySyncLogInfo]




class WarrantyConfigCatalog(BaseModel):
    statuses: list[str] = []
    final_statuses: list[str] = []
    sucursales: list[str] = []
    depositos: list[str] = []
    delay_ranges: list[int] = []
    required_review_fields: list[str] = []
    sheet_raw: str = ""
    spreadsheet_url: str = ""
    products_source_label: str = "Catálogo local"


class WarrantyConfigResponse(BaseModel):
    config: WarrantyConfigCatalog
    providers_count: int = 0
    brands_count: int = 0
    mapped_brands_count: int = 0
    unmapped_brands_count: int = 0
    pending_review_count: int = 0
    active_count: int = 0


class WarrantyConfigSaveRequest(BaseModel):
    statuses: list[str] | None = None
    final_statuses: list[str] | None = None
    sucursales: list[str] | None = None
    depositos: list[str] | None = None
    delay_ranges: list[int] | None = None
    required_review_fields: list[str] | None = None
    raw_sheet: str | None = None
    spreadsheet_url: str | None = None


class WarrantyCancelRequest(BaseModel):
    reason: str = Field(min_length=3)


class WarrantyDashboardPoint(BaseModel):
    label: str
    value: float
    extra: dict[str, Any] = {}


class WarrantyDashboardMetrics(BaseModel):
    total: int = 0
    ingreso: int = 0
    pendientes_revision: int = 0
    pendientes_proveedor: int = 0
    enviadas_proveedor: int = 0
    en_revision: int = 0
    resueltas: int = 0
    rechazadas: int = 0
    demoradas_7: int = 0
    demoradas_15: int = 0
    promedio_dias_pendiente: float = 0
    promedio_resolucion: float = 0
    promedio_dias_sin_respuesta: float = 0


class WarrantyDashboardResponse(BaseModel):
    metrics: WarrantyDashboardMetrics
    by_status: list[WarrantyDashboardPoint] = []
    by_brand: list[WarrantyDashboardPoint] = []
    by_provider: list[WarrantyDashboardPoint] = []
    by_branch: list[WarrantyDashboardPoint] = []
    by_deposit: list[WarrantyDashboardPoint] = []
    by_delay_range: list[WarrantyDashboardPoint] = []
    monthly_entries: list[WarrantyDashboardPoint] = []
    avg_resolution_by_provider: list[WarrantyDashboardPoint] = []
    final_resolutions: list[WarrantyDashboardPoint] = []
    critical: list[WarrantySummary] = []
    filters: dict[str, Any] = {}




class WarrantyDiagnosticItem(BaseModel):
    key: str
    label: str
    status: str = "ok"
    detail: str = ""
    count: int = 0


class WarrantyDiagnosticsResponse(BaseModel):
    status: str = "ok"
    generated_at: str
    items: list[WarrantyDiagnosticItem] = []
    next_actions: list[str] = []

class WarrantyCounterInfo(BaseModel):
    year: int
    sucursal: str
    last_number: int


class WarrantyCountersResponse(BaseModel):
    counters: list[WarrantyCounterInfo]


class WarrantyResetSummary(BaseModel):
    guarantees: int = 0
    guarantee_items: int = 0
    guarantee_history: int = 0
    remitos: int = 0
    exports: int = 0
    sync_logs: int = 0
    counters: int = 0
    generated_export_files: int = 0


class WarrantyResetPreviewResponse(BaseModel):
    ok: bool = True
    generated_at: str
    summary: WarrantyResetSummary
    preserved: list[str]
    warning: str
    confirmation_phrase: str


class WarrantyResetRequest(BaseModel):
    confirmation: str = Field(min_length=1)
    reset_generated_files: bool = True


class WarrantyResetResponse(BaseModel):
    ok: bool
    reset_at: str
    summary_before: WarrantyResetSummary
    backup_file: str
    deleted_generated_files: int = 0
    message: str


RESET_CONFIRMATION_PHRASE = "RESET GARANTIAS PRODUCCION"


# =========================================================
# Utilidades generales
# =========================================================

def db_connect() -> sqlite3.Connection:
    settings = get_settings()
    settings.ensure_dirs()
    conn = sqlite3.connect(settings.database_path, check_same_thread=False)
    conn.row_factory = sqlite3.Row
    return conn


def utc_now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def now_ar() -> datetime:
    return datetime.now(ZoneInfo("America/Argentina/Buenos_Aires"))


def normalize_text(value: Any) -> str:
    text = "" if value is None else str(value)
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = text.upper().strip()
    text = re.sub(r"[^A-Z0-9]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def canonical_status_key(value: Any) -> str:
    """Clave estable para comparar estados aunque vengan con número o texto.

    La normalización separa revisión interna (review_status) de estados operativos.
    Compatibiliza estados viejos sin volver a usarlos como fuente de verdad.
    """
    original = normalize_text(value)
    if not original:
        return ""
    # Casos numerados antiguos: antes el proveedor usaba "5 - EN REVISION".
    if re.match(r"^5\s+EN\s+REVISION$", original):
        return "EN EL PROVEEDOR"
    text = re.sub(r"^\d+\s+", "", original).strip()
    aliases = {
        "INGRESADO": "INGRESO",
        "PENDIENTE REVISION": "INGRESO",
        "PENDIENTE DE REVISION": "INGRESO",
        # Estados internos escritos en status por versiones anteriores. Ahora viven en review_status.
        "EN REVISION": "INGRESO",
        "EN REVISION INTERNA": "INGRESO",
        "CORRECCION PENDIENTE": "INGRESO",
        # Variantes del flujo proveedor.
        "EN REVISION PROVEEDOR": "EN EL PROVEEDOR",
        "EN REVISION DEL PROVEEDOR": "EN EL PROVEEDOR",
        "PROVEEDOR": "EN EL PROVEEDOR",
        "EN PROVEEDOR": "EN EL PROVEEDOR",
        "RESPONDIDO": "RESPONDIDO POR PROVEEDOR",
        "RESPUESTA PROVEEDOR": "RESPONDIDO POR PROVEEDOR",
        "RESPONDIDO PROVEEDOR": "RESPONDIDO POR PROVEEDOR",
        "LISTO RETIRO": "LISTO PARA RETIRO",
        "NC": "NOTA DE CREDITO",
        "NOTA CREDITO": "NOTA DE CREDITO",
        "ANULADO": "ANULADA",
        "CANCELADO": "ANULADA",
    }
    return aliases.get(text, text)


def normalize_status(value: Any) -> str:
    """Devuelve la etiqueta canónica visible de un estado operativo.

    Fase 16 hotfix: algunos filtros del listado usan normalize_status(), pero
    versiones anteriores sólo dejaron canonical_status_key(). Esta función mantiene
    compatibilidad con estados viejos y devuelve siempre una etiqueta oficial
    cuando puede mapearla.
    """
    key = canonical_status_key(value)
    if not key:
        return ""
    for label in DEFAULT_STATUSES:
        if canonical_status_key(label) == key:
            return label
    # Compatibilidad con valores viejos que no existen en DEFAULT_STATUSES.
    if key == canonical_status_key(CANCELLED_STATUS):
        return "9 - ANULADA"
    return str(value or "").strip()


def _assigned_deposit_branch_from_user(user: Any) -> dict[str, Any] | None:
    branches = getattr(user, "branches", []) or []
    ordered = sorted(
        [b for b in branches if isinstance(b, dict)],
        key=lambda b: 0 if b.get("is_primary") else 1,
    )
    for branch in ordered:
        b_type = normalize_text(branch.get("type", ""))
        b_name = normalize_text(branch.get("name", ""))
        if b_type in {"DEPOSIT", "DEPOSITO"} or b_name.startswith("DEPOSITO ") or b_name == "DEPOSITO":
            return branch
    return None


def is_deposit_operator_user(user: Any) -> bool:
    branch_type_key = normalize_text(getattr(user, "branch_type", "") or "")
    branch_name_key = normalize_text(getattr(user, "branch_name", "") or getattr(user, "sucursal", "") or "")
    role_key = normalize_text(getattr(user, "role", "") or "")
    roles_keys = {normalize_text(r) for r in (getattr(user, "roles", []) or [])}
    return (
        branch_type_key in {"DEPOSIT", "DEPOSITO"}
        or role_key == "DEPOSITO"
        or "DEPOSITO" in roles_keys
        or branch_name_key.startswith("DEPOSITO ")
        or branch_name_key == "DEPOSITO"
        or _assigned_deposit_branch_from_user(user) is not None
    )


def ensure_warranty_intake_access(user: Any) -> None:
    if getattr(user, "must_change_password", False):
        raise HTTPException(status_code=403, detail="Tenés que crear tu contraseña antes de continuar")
    if user.has("warranties.view") or user.has("warranties.create") or is_deposit_operator_user(user):
        return
    raise HTTPException(status_code=403, detail="No tenés permiso para realizar esta acción")


def status_matches(value: Any, expected: Any) -> bool:
    return canonical_status_key(value) == canonical_status_key(expected)


def _canonical_review_status(value: Any) -> str:
    """Normaliza un valor de review_status a su constante canónica.

    NO usa canonical_status_key() porque esa función mapea 'EN REVISION' y
    'PENDIENTE REVISION' a 'INGRESO' (para compatibilidad con estados operativos
    viejos), lo que hace que pendiente_revision y en_revision sean indistinguibles.
    """
    s = normalize_text(value)
    if not s:
        return REVIEW_PENDING
    # Quitar prefijo numérico si existiera
    s = re.sub(r"^\d+\s+", "", s).strip()
    aliases: dict[str, str] = {
        # pendiente_revision
        "PENDIENTE REVISION":          REVIEW_PENDING,
        "PENDIENTE DE REVISION":       REVIEW_PENDING,
        "PENDING":                     REVIEW_PENDING,
        # en_revision
        "EN REVISION":                 REVIEW_IN_PROGRESS,
        "EN REVISION INTERNA":         REVIEW_IN_PROGRESS,
        # requiere_correccion
        "REQUIERE CORRECCION":         REVIEW_INCOMPLETE,
        "CORRECCION PENDIENTE":        REVIEW_INCOMPLETE,
        "INCOMPLETA":                  REVIEW_INCOMPLETE,
        # revisada
        "REVISADA":                    REVIEW_APPROVED,
    }
    return aliases.get(s, s)


def review_status_matches(value: Any, expected: str) -> bool:
    """Compara review_status con normalización propia (sin canonical_status_key)."""
    return _canonical_review_status(value) == _canonical_review_status(expected)


def header_key(value: Any) -> str:
    return normalize_text(value).replace(" ", "")


def format_date_ar(dt: datetime | date | None = None) -> str:
    if dt is None:
        dt = now_ar()
    if isinstance(dt, datetime):
        return dt.astimezone(ZoneInfo("America/Argentina/Buenos_Aires")).strftime("%d/%m/%Y")
    return dt.strftime("%d/%m/%Y")


def format_datetime_ar(dt: datetime | None = None) -> str:
    if dt is None:
        dt = now_ar()
    return dt.astimezone(ZoneInfo("America/Argentina/Buenos_Aires")).strftime("%d/%m/%Y %H:%M")


def parse_iso_datetime(value: Any) -> datetime | None:
    text = str(value or "").strip()
    if not text:
        return None
    try:
        dt = datetime.fromisoformat(text.replace("Z", "+00:00"))
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=timezone.utc)
        return dt
    except Exception:
        return None


def parse_date_filter(value: Any) -> date | None:
    text = str(value or "").strip()
    if not text:
        return None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%d/%m/%y"):
        try:
            return datetime.strptime(text[:10], fmt).date()
        except Exception:
            pass
    return None


def ingreso_at_from_input(value: Any, *, fallback_now: bool = True) -> str:
    """Normaliza una fecha de ingreso a ISO.

    Acepta YYYY-MM-DD, DD/MM/YYYY o un ISO existente. Para fechas sin hora,
    guardamos mediodía de Argentina para evitar corrimientos visuales por zona horaria.
    """
    text = str(value or "").strip()
    if not text:
        return utc_now_iso() if fallback_now else ""
    dt = parse_iso_datetime(text)
    if dt:
        return dt.isoformat()
    d = parse_date_filter(text)
    if d:
        return datetime(d.year, d.month, d.day, 12, 0, tzinfo=ZoneInfo("America/Argentina/Buenos_Aires")).isoformat()
    raise HTTPException(status_code=400, detail="Fecha de ingreso inválida. Usá formato AAAA-MM-DD o DD/MM/AAAA.")


def date_input_from_iso(value: Any) -> str:
    dt = parse_iso_datetime(value)
    if not dt:
        return ""
    return dt.astimezone(ZoneInfo("America/Argentina/Buenos_Aires")).date().isoformat()


def clean_option_value(value: Any) -> str:
    text = str(value or "").strip()
    return re.sub(r"\s+", " ", text)


def strip_numeric_prefix(value: Any) -> str:
    return re.sub(r"^\s*\d+\s*[-.)]\s*", "", str(value or "").strip()).strip()


def canonical_deposit_display(value: Any) -> str:
    clean = strip_numeric_prefix(value)
    key = normalize_text(clean)
    if key in {"CHICLANA", "DEPOSITO CHICLANA"}:
        return "Depósito Chiclana"
    if key in {"CORRALES", "DEPOSITO CORRALES"}:
        return "Depósito Corrales"
    if key in {"CACHI", "DEPOSITO CACHI"}:
        return "Depósito Cachi"
    return clean


def clean_select_options(values: list[Any], *, deposit: bool = False) -> list[str]:
    cleaned: list[str] = []
    for value in values or []:
        text = canonical_deposit_display(value) if deposit else strip_numeric_prefix(value)
        text = clean_option_value(text)
        if text:
            cleaned.append(text)
    return unique_keep_order(cleaned)


def unique_keep_order(values: list[str]) -> list[str]:
    seen: set[str] = set()
    out: list[str] = []
    for value in values:
        clean = clean_option_value(value)
        if not clean:
            continue
        key = header_key(clean)
        if key in seen:
            continue
        seen.add(key)
        out.append(clean)
    return out


def sucursal_code(value: Any) -> str:
    clean = normalize_text(strip_numeric_prefix(value))
    compact = clean.replace(" ", "")
    mapping = {
        "CASEROS": "CAS",
        "LANUS": "LAN",
        "LANUSOESTE": "LAN",
        "CANNING": "CAN",
        "NORCENTER": "NOR",
        "NORTE": "NOR",
        "NORTH": "NOR",
        "SUR": "SUR",
    }
    if compact in mapping:
        return mapping[compact]
    for key, code in mapping.items():
        if key in compact:
            return code
    compact = re.sub(r"[^A-Z0-9]+", "", compact)
    return (compact[:3] or "GEN").upper()


def _origen_from_tipo(tipo: str) -> str:
    """Deriva origen_ingreso a partir del tipo_ingreso."""
    return "sucursal" if tipo == "cliente_sucursal" else "deposito"


def _ubicacion_from_tipo(tipo: str) -> str:
    """Deriva ubicación genérica legacy a partir del tipo_ingreso."""
    if tipo == "cliente_sucursal":
        return "sucursal"
    if tipo in ("cliente_deposito", "falla_recepcion_mercaderia", "stock_interno"):
        return "deposito"
    return "desconocida"


def _initial_ubicacion_actual(tipo: str, sucursal_carga: str = "", deposito_carga: str = "") -> str:
    """Ubicación física real al crear la garantía.

    Regla operativa fase 36:
    - Cliente en sucursal: el equipo queda físicamente en esa sucursal
      (Caseros, Canning, etc.), no en el depósito destino.
    - Cliente en depósito / falla recepción / stock interno: el equipo queda
      físicamente en el depósito de carga (Chiclana, Corrales, Cachi).

    Los valores genéricos 'sucursal'/'deposito' quedan solo como fallback legacy.
    """
    clean_tipo = (tipo or "").strip()
    if clean_tipo == "cliente_sucursal":
        return (sucursal_carga or "").strip() or "sucursal"
    if clean_tipo in ("cliente_deposito", "falla_recepcion_mercaderia", "stock_interno"):
        return (deposito_carga or "").strip() or "deposito"
    return "desconocida"


def _location_is_deposit_value(value: str) -> bool:
    key = normalize_text(value)
    return key == "DEPOSITO" or key.startswith("DEPOSITO ")


def _code_source_for_tipo(item_sucursal: str, item_deposito: str, tipo: str) -> str:
    """Devuelve la fuente de código (sucursal o depósito) para next_warranty_code."""
    if tipo == "cliente_sucursal" and item_sucursal.strip():
        return item_sucursal
    return item_deposito


def _fetch_branch_info(branch_id: str) -> dict[str, str] | None:
    """Busca en la tabla branches el nombre, empresa y tipo de una branch por su ID."""
    if not branch_id.strip():
        return None
    try:
        with db_connect() as conn:
            row = conn.execute(
                """
                SELECT b.id, b.name, b.code, b.type, b.company_id, c.name AS company_name
                FROM branches b
                LEFT JOIN companies c ON c.id = b.company_id
                WHERE b.id = ?
                """,
                (branch_id.strip(),),
            ).fetchone()
            if not row:
                return None
            return {
                "id": str(row["id"]),
                "name": str(row["name"] or ""),
                "code": str(row["code"] or ""),
                "type": str(row["type"] or ""),
                "company_id": str(row["company_id"] or ""),
                "company_name": str(row["company_name"] or ""),
            }
    except Exception:
        return None


def _branch_key(value: Any) -> str:
    return normalize_text(strip_numeric_prefix(value))


def _branch_by_name(branches: list[dict[str, str]], name: str, branch_type: str | None = None) -> dict[str, str] | None:
    key = _branch_key(name)
    if not key:
        return None
    for branch in branches:
        if branch_type and str(branch.get("type") or "") != branch_type:
            continue
        if _branch_key(branch.get("name", "")) == key or _branch_key(branch.get("code", "")) == key:
            return branch
    return None


def _fetch_branches_operativas() -> list[dict[str, str]]:
    """Devuelve las branches físicas y de depósito activas del sistema (physical + deposit).
    Usadas en el selector de sucursal_responsable y para validar IDs del frontend.
    """
    try:
        with db_connect() as conn:
            rows = conn.execute(
                """
                SELECT b.id, b.name, b.code, b.type, b.company_id, c.name AS company_name
                FROM branches b
                LEFT JOIN companies c ON c.id = b.company_id
                WHERE b.is_active = 1 AND b.type IN ('physical', 'deposit', 'admin')
                ORDER BY
                    CASE b.type WHEN 'physical' THEN 1 WHEN 'deposit' THEN 2 ELSE 3 END,
                    b.name COLLATE NOCASE
                """,
            ).fetchall()
            return [
                {
                    "id": str(r["id"]),
                    "name": str(r["name"] or ""),
                    "code": str(r["code"] or ""),
                    "type": str(r["type"] or ""),
                    "company_id": str(r["company_id"] or ""),
                    "company_name": str(r["company_name"] or ""),
                }
                for r in rows
            ]
    except Exception:
        return []

def _warranty_central_deposit_from_branches(branches: list[dict[str, str]]) -> dict[str, str] | None:
    """Depósito operativo principal de Garantías.

    Regla actual del negocio: Chiclana es el depósito destino principal.
    Corrales y Cachi son depósitos de guarda, por lo que nunca deben ser
    fallback automático para remitos/ingresos de sucursal.
    """
    deposits = [b for b in branches if str(b.get("type") or "") == "deposit"]
    for b in deposits:
        key = _branch_key(f"{b.get('code','')} {b.get('name','')}")
        if "chiclana" in key:
            return b
    return None

def _warranty_central_deposit_name(branches: list[dict[str, str]] | None = None) -> str:
    if branches is None:
        branches = _fetch_branches_operativas()
    branch = _warranty_central_deposit_from_branches(branches)
    if branch and str(branch.get("name") or "").strip():
        return str(branch.get("name") or "").strip()
    for value in DEFAULT_DEPOSITOS:
        if "chiclana" in normalize_text(value):
            return strip_numeric_prefix(value)
    return "Depósito Chiclana"



# ──────────────────────────────────────────────────────────────────────────────
# Fuente de verdad organizativa para Garantías
# ──────────────────────────────────────────────────────────────────────────────
# Estos campos son los que deben conducir la lógica nueva:
#   company_id                 -> empresa imputada/relacionada
#   branch_id                  -> unidad de carga real (sucursal o depósito asignado)
#   sucursal_responsable_id    -> sucursal comercial responsable, cuando aplica
#   origen_ingreso             -> sucursal | deposito
#   tipo_ingreso               -> cliente_sucursal | cliente_deposito | falla_recepcion_mercaderia | ...
#   ubicacion_actual           -> sucursal | deposito | en_transito | proveedor | ...
# Los campos texto heredados (sucursal, deposito, lugar_llegada, transit_status)
# se mantienen para compatibilidad/display, pero no deberían ser la única fuente
# de permisos ni de decisiones nuevas.
ORG_TRUTH_FIELDS = [
    "company_id", "branch_id", "sucursal_responsable_id",
    "origen_ingreso", "tipo_ingreso", "ubicacion_actual",
]
LEGACY_DISPLAY_FIELDS = ["sucursal", "deposito", "lugar_llegada", "transit_status"]


def _branches_by_id(conn: sqlite3.Connection) -> dict[str, dict[str, str]]:
    try:
        rows = conn.execute(
            """
            SELECT b.id, b.name, b.code, b.type, b.company_id, c.name AS company_name
            FROM branches b
            LEFT JOIN companies c ON c.id = b.company_id
            WHERE b.is_active = 1
            """
        ).fetchall()
    except Exception:
        return {}
    return {
        str(r["id"]): {
            "id": str(r["id"]),
            "name": str(r["name"] or ""),
            "code": str(r["code"] or ""),
            "type": str(r["type"] or ""),
            "company_id": str(r["company_id"] or ""),
            "company_name": str(r["company_name"] or ""),
        }
        for r in rows
    }


def _branches_by_key(branches: dict[str, dict[str, str]]) -> dict[str, dict[str, str]]:
    out: dict[str, dict[str, str]] = {}
    for b in branches.values():
        for value in (b.get("id", ""), b.get("name", ""), b.get("code", "")):
            key = _branch_key(value)
            if key and key not in out:
                out[key] = b
    return out


def _infer_tipo_ingreso_from_legacy(row: sqlite3.Row, branch: dict[str, str] | None) -> str:
    current = str(row["tipo_ingreso"] or "") if "tipo_ingreso" in row.keys() else ""
    if current.strip():
        return current.strip()
    branch_type = str((branch or {}).get("type") or "").lower()
    sucursal = str(row["sucursal"] or "") if "sucursal" in row.keys() else ""
    deposito = str(row["deposito"] or "") if "deposito" in row.keys() else ""
    if branch_type == "deposit" or (not sucursal.strip() and deposito.strip()):
        return "cliente_deposito"
    return "cliente_sucursal"


def _backfill_warranty_org_fields(conn: sqlite3.Connection) -> int:
    """Completa campos organizativos nuevos en garantías heredadas sin borrar datos.

    Es deliberadamente conservador: solo rellena blancos. No corrige manualmente
    valores existentes porque eso podría alterar datos reales ya usados.
    """
    try:
        rows = conn.execute("SELECT * FROM guarantees").fetchall()
    except Exception:
        return 0
    branches = _branches_by_id(conn)
    by_key = _branches_by_key(branches)
    updated = 0
    for row in rows:
        row_id = int(row["id"])
        branch_id = str(row["branch_id"] or "") if "branch_id" in row.keys() else ""
        branch = branches.get(branch_id) if branch_id else None
        sucursal_text = str(row["sucursal"] or "") if "sucursal" in row.keys() else ""
        deposito_text = str(row["deposito"] or "") if "deposito" in row.keys() else ""
        if not branch:
            # Preferimos resolver por sucursal de carga; si no existe, por depósito.
            branch = by_key.get(_branch_key(sucursal_text)) or by_key.get(_branch_key(deposito_text))
            branch_id = branch.get("id", "") if branch else ""

        tipo = _infer_tipo_ingreso_from_legacy(row, branch)
        origen = str(row["origen_ingreso"] or "") if "origen_ingreso" in row.keys() else ""
        ubicacion = str(row["ubicacion_actual"] or "") if "ubicacion_actual" in row.keys() else ""
        if not origen:
            origen = _origen_from_tipo(tipo)
        if not ubicacion:
            ubicacion = _ubicacion_from_tipo(tipo)

        suc_resp_id = str(row["sucursal_responsable_id"] or "") if "sucursal_responsable_id" in row.keys() else ""
        suc_resp = str(row["sucursal_responsable"] or "") if "sucursal_responsable" in row.keys() else ""
        resp_branch = branches.get(suc_resp_id) if suc_resp_id else None
        if not resp_branch:
            # Para ingresos de sucursal, la responsable suele ser la misma sucursal.
            # Para ingresos de depósito heredados, no inventamos sucursal responsable si no se sabe.
            source_for_resp = suc_resp or (sucursal_text if tipo == "cliente_sucursal" else "")
            resp_branch = by_key.get(_branch_key(source_for_resp)) if source_for_resp else None
            if resp_branch:
                suc_resp_id = resp_branch.get("id", "")
                suc_resp = resp_branch.get("name", "")
        elif not suc_resp:
            suc_resp = resp_branch.get("name", "")

        company_id = str(row["company_id"] or "") if "company_id" in row.keys() else ""
        if not company_id:
            company_id = (resp_branch or {}).get("company_id") or (branch or {}).get("company_id") or ""

        updates: dict[str, str] = {}
        if branch_id and not (str(row["branch_id"] or "") if "branch_id" in row.keys() else ""):
            updates["branch_id"] = branch_id
        if tipo and not (str(row["tipo_ingreso"] or "") if "tipo_ingreso" in row.keys() else ""):
            updates["tipo_ingreso"] = tipo
        if origen and not (str(row["origen_ingreso"] or "") if "origen_ingreso" in row.keys() else ""):
            updates["origen_ingreso"] = origen
        if ubicacion and not (str(row["ubicacion_actual"] or "") if "ubicacion_actual" in row.keys() else ""):
            updates["ubicacion_actual"] = ubicacion
        if suc_resp and not (str(row["sucursal_responsable"] or "") if "sucursal_responsable" in row.keys() else ""):
            updates["sucursal_responsable"] = suc_resp
        if suc_resp_id and not (str(row["sucursal_responsable_id"] or "") if "sucursal_responsable_id" in row.keys() else ""):
            updates["sucursal_responsable_id"] = suc_resp_id
        if company_id and not (str(row["company_id"] or "") if "company_id" in row.keys() else ""):
            updates["company_id"] = company_id
        if not updates:
            continue
        sets = ", ".join(f"{key} = ?" for key in updates)
        conn.execute(f"UPDATE guarantees SET {sets} WHERE id = ?", (*updates.values(), row_id))
        updated += 1
    return updated



def _derive_sucursal_fields(item: "WarrantyItemIn", user: Any, is_vendedor_sucursal: bool) -> tuple[str, str]:
    """Devuelve (sucursal_carga, sucursal_responsable).

    sucursal_carga       → origen físico del ingreso (se guarda en guarantees.sucursal).
    sucursal_responsable → rama comercialmente responsable (nueva columna).

    Reglas:
    · VENDEDOR (is_vendedor_sucursal=True): ambos campos = sucursal del usuario.
      El tipo de ingreso siempre es cliente_sucursal; no puede escalar ni sobre-escribir.
    · GESTOR/DEPOSITO — cliente_sucursal:
        carga = item.sucursal; responsable = item.sucursal_responsable o item.sucursal.
    · GESTOR/DEPOSITO — cliente_deposito:
        carga = sucursal del usuario (depósito); responsable = item.sucursal_responsable (obligatorio).
    · GESTOR/DEPOSITO — otros (falla_recepcion_mercaderia, stock_interno, otro):
        carga = sucursal del usuario (depósito); responsable = item.sucursal_responsable o user_branch.
    """
    tipo = (item.tipo_ingreso or "").strip()
    user_branch = str(getattr(user, "sucursal", "") or getattr(user, "branch_name", "") or "").strip()

    if is_vendedor_sucursal:
        return user_branch, user_branch

    suc_item = (item.sucursal or "").strip()
    suc_resp = (item.sucursal_responsable or "").strip()

    if tipo == "cliente_sucursal":
        return suc_item, suc_resp or suc_item

    if tipo == "cliente_deposito":
        # suc_resp ya fue validado como obligatorio en create_warranty_entries
        return user_branch, suc_resp

    # falla_recepcion_mercaderia, stock_interno, otro → depósito es carga y responsable
    return user_branch, suc_resp or user_branch


def _notify_gestor_garantias_pickup(title: str, message: str) -> None:
    """Notifica a usuarios con warranties.remitos.provider_delivery cuando se necesita acción urgente de logística."""
    try:
        roles = load_roles()
        users_map = load_users()
        targets: list[str] = []
        for u in users_map.values():
            if not getattr(u, "is_active", True):
                continue
            perms: list[str] = []
            for role_name in (getattr(u, "roles", None) or [getattr(u, "role", "")]):
                role = roles.get(role_name)
                if role:
                    perms.extend(getattr(role, "permissions", []))
            if has_permission(perms, "warranties.remitos.provider_delivery"):
                targets.append(u.username)
        if targets:
            notify_many(targets, title, message, type_="warning")
    except Exception:
        pass  # notificaciones no son críticas


def is_provider_waiting_closed_status(status_value: str) -> bool:
    """Estados que cortan el contador de días sin respuesta del proveedor."""
    s = canonical_status_key(status_value)
    return any(token in s for token in [
        "RESPONDIDO POR PROVEEDOR", "RESUELTO", "RECHAZADO", "ANULADA", "FINALIZADO",
        # Tokens antiguos (mantener para historial)
        "LISTO PARA RETIRO", "APROBADO CAMBIO", "NOTA DE CREDITO", "REPARADO",
    ])


def is_resolved_status(status_value: str) -> bool:
    """Estados donde el proveedor ya definió o el caso quedó cerrado administrativamente."""
    s = canonical_status_key(status_value)
    return any(token in s for token in [
        "RESUELTO", "RECHAZADO", "ANULADA", "FINALIZADO",
        # Tokens antiguos (mantener para historial)
        "LISTO PARA RETIRO", "APROBADO CAMBIO", "NOTA DE CREDITO", "REPARADO",
    ])



def internal_logistics_ready_for_provider(row: sqlite3.Row) -> bool:
    """Indica si la garantía puede avanzar a flujo proveedor.

    Para garantías nacidas en sucursal, primero deben llegar físicamente a
    Chiclana/depósito mediante remito interno. Esto evita estados incoherentes
    como "ENVIADO AL PROVEEDOR" mientras el producto sigue en tránsito interno.
    Las garantías nacidas directamente en depósito se consideran listas.
    """
    keys = set(row.keys())
    origin = str(row["origen_ingreso"] or "").strip().lower() if "origen_ingreso" in keys else ""
    transit = str(row["transit_status"] or "").strip().lower() if "transit_status" in keys else ""
    location = str(row["ubicacion_actual"] or "").strip().lower() if "ubicacion_actual" in keys else ""
    if origin == "sucursal":
        return transit == "en_deposito" or location == "deposito" or _location_is_deposit_value(location)
    return True


def assert_internal_logistics_ready_for_provider(row: sqlite3.Row) -> None:
    if internal_logistics_ready_for_provider(row):
        return
    remito = str(row["remito_interno"] or "").strip() if "remito_interno" in row.keys() else ""
    if remito:
        raise HTTPException(
            status_code=400,
            detail=f"La garantía todavía no llegó a Depósito Chiclana. Confirmá la llegada del remito {remito} antes de avanzar con proveedor.",
        )
    raise HTTPException(
        status_code=400,
        detail="La garantía todavía está en sucursal. Generá/confirmá el remito interno a Depósito Chiclana antes de avanzar con proveedor.",
    )


def provider_flow_started(row: sqlite3.Row) -> bool:
    status_key = canonical_status_key(str(row["status"] or ""))
    return bool(str(row["sent_to_provider_at"] or "").strip()) or status_key in {
        "ENVIADO AL PROVEEDOR",
        "EN EL PROVEEDOR",
        "RESPONDIDO POR PROVEEDOR",
        "RESUELTO",
        "RECHAZADO",
        "ANULADA",
        "FINALIZADO",
    }


def provider_has_physical_product(row: sqlite3.Row) -> bool:
    """True cuando ya se registró el retiro físico del proveedor."""
    keys = set(row.keys())
    status_key = canonical_status_key(str(row["status"] or ""))
    location = str(row["ubicacion_actual"] or "").strip().lower() if "ubicacion_actual" in keys else ""
    pickup = normalize_provider_pickup_status(row["estado_retiro_proveedor"]) if "estado_retiro_proveedor" in keys else "sin_solicitud"
    return status_key == "EN EL PROVEEDOR" or location == "proveedor" or pickup == "retirado"


def assert_provider_has_physical_product(row: sqlite3.Row) -> None:
    if provider_has_physical_product(row):
        return
    raise HTTPException(
        status_code=400,
        detail="El proveedor todavía no tiene físicamente el producto. Primero registrá el retiro desde Depósito Chiclana.",
    )

def days_between(start_iso: str, end_iso: str | None = None) -> int:
    start = parse_iso_datetime(start_iso)
    if not start:
        return 0
    end = parse_iso_datetime(end_iso or "") or datetime.now(timezone.utc)
    return max(0, (end.date() - start.date()).days)


def compute_pending_days(row: sqlite3.Row) -> int:
    resolution = row["fecha_resolucion"] or row["cancelled_at"] or ""
    return days_between(row["ingreso_at"] or row["created_at"], resolution or None)


def compute_no_response_days(row: sqlite3.Row) -> int | None:
    if not row["sent_to_provider_at"]:
        return None
    if is_provider_waiting_closed_status(row["status"] or ""):
        return 0
    keys = set(row.keys())
    # El contador debe arrancar desde el último mail enviado/re-enviado al proveedor.
    # Las respuestas cierran la espera mediante is_provider_waiting_closed_status.
    last_mail = str(row["fecha_ultimo_mail_proveedor"] or "").strip() if "fecha_ultimo_mail_proveedor" in keys else ""
    base = last_mail or row["sent_to_provider_at"]
    return days_between(base)


def ensure_column(conn: sqlite3.Connection, table: str, column: str, ddl: str) -> None:
    existing = {str(row["name"]) for row in conn.execute(f"PRAGMA table_info({table})").fetchall()}
    if column not in existing:
        conn.execute(f"ALTER TABLE {table} ADD COLUMN {column} {ddl}")


def ensure_warranty_tables(conn: sqlite3.Connection) -> None:
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS guarantees (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            warranty_code TEXT UNIQUE NOT NULL,
            status TEXT NOT NULL DEFAULT '1 - INGRESO',
            review_status TEXT NOT NULL DEFAULT 'pendiente_revision',
            reviewed_by TEXT NOT NULL DEFAULT '',
            reviewed_by_name TEXT NOT NULL DEFAULT '',
            reviewed_at TEXT NOT NULL DEFAULT '',
            review_note TEXT NOT NULL DEFAULT '',
            responsible_username TEXT NOT NULL DEFAULT '',
            responsible_name TEXT NOT NULL DEFAULT '',
            created_by TEXT NOT NULL DEFAULT '',
            created_by_name TEXT NOT NULL DEFAULT '',
            created_at TEXT NOT NULL,
            ingreso_at TEXT NOT NULL,
            updated_at TEXT NOT NULL,
            updated_by TEXT NOT NULL DEFAULT '',
            updated_by_name TEXT NOT NULL DEFAULT '',
            sucursal TEXT NOT NULL DEFAULT '',
            sucursal_code TEXT NOT NULL DEFAULT '',
            branch_id TEXT NOT NULL DEFAULT '',
            deposito TEXT NOT NULL DEFAULT '',
            lugar_llegada TEXT NOT NULL DEFAULT '',
            provider_name TEXT NOT NULL DEFAULT '',
            provider_case_id TEXT NOT NULL DEFAULT '',
            sent_to_provider_at TEXT NOT NULL DEFAULT '',
            last_provider_response_at TEXT NOT NULL DEFAULT '',
            last_claim_at TEXT NOT NULL DEFAULT '',
            fecha_retiro TEXT NOT NULL DEFAULT '',
            fecha_resolucion TEXT NOT NULL DEFAULT '',
            finalizacion TEXT NOT NULL DEFAULT '',
            vuelve_a TEXT NOT NULL DEFAULT '',
            observations TEXT NOT NULL DEFAULT '',
            photos_reference TEXT NOT NULL DEFAULT '',
            cancelled INTEGER NOT NULL DEFAULT 0,
            cancel_reason TEXT NOT NULL DEFAULT '',
            cancelled_by TEXT NOT NULL DEFAULT '',
            cancelled_at TEXT NOT NULL DEFAULT '',
            synced_to_google_sheet INTEGER NOT NULL DEFAULT 0,
            last_google_sync_at TEXT NOT NULL DEFAULT '',
            google_sheet_row_id TEXT NOT NULL DEFAULT '',
            google_sheet_updated_at TEXT NOT NULL DEFAULT ''
        )
        """
    )
    ensure_column(conn, "guarantees", "reviewed_by", "TEXT NOT NULL DEFAULT ''")
    ensure_column(conn, "guarantees", "reviewed_by_name", "TEXT NOT NULL DEFAULT ''")
    ensure_column(conn, "guarantees", "reviewed_at", "TEXT NOT NULL DEFAULT ''")
    ensure_column(conn, "guarantees", "review_note", "TEXT NOT NULL DEFAULT ''")
    # Fase 9 — trazabilidad específica de revisión. No reemplaza guarantee_history;
    # estos campos permiten filtrar/reportar rápido sin parsear JSON histórico.
    ensure_column(conn, "guarantees", "review_started_at", "TEXT NOT NULL DEFAULT ''")
    ensure_column(conn, "guarantees", "review_started_by", "TEXT NOT NULL DEFAULT ''")
    ensure_column(conn, "guarantees", "correction_requested_at", "TEXT NOT NULL DEFAULT ''")
    ensure_column(conn, "guarantees", "correction_requested_by", "TEXT NOT NULL DEFAULT ''")
    ensure_column(conn, "guarantees", "correction_resubmitted_at", "TEXT NOT NULL DEFAULT ''")
    ensure_column(conn, "guarantees", "correction_resubmitted_by", "TEXT NOT NULL DEFAULT ''")
    ensure_column(conn, "guarantees", "provider_name", "TEXT NOT NULL DEFAULT ''")
    ensure_column(conn, "guarantees", "provider_case_id", "TEXT NOT NULL DEFAULT ''")
    ensure_column(conn, "guarantees", "sent_to_provider_at", "TEXT NOT NULL DEFAULT ''")
    ensure_column(conn, "guarantees", "last_provider_response_at", "TEXT NOT NULL DEFAULT ''")
    ensure_column(conn, "guarantees", "last_claim_at", "TEXT NOT NULL DEFAULT ''")
    ensure_column(conn, "guarantees", "synced_to_google_sheet", "INTEGER NOT NULL DEFAULT 0")
    ensure_column(conn, "guarantees", "last_google_sync_at", "TEXT NOT NULL DEFAULT ''")
    ensure_column(conn, "guarantees", "google_sheet_row_id", "TEXT NOT NULL DEFAULT ''")
    ensure_column(conn, "guarantees", "google_sheet_updated_at", "TEXT NOT NULL DEFAULT ''")
    ensure_column(conn, "guarantees", "sync_error", "TEXT NOT NULL DEFAULT ''")
    ensure_column(conn, "guarantees", "shipment_code", "TEXT NOT NULL DEFAULT ''")
    ensure_column(conn, "guarantees", "shipment_file_name", "TEXT NOT NULL DEFAULT ''")
    ensure_column(conn, "guarantees", "resolution_note", "TEXT NOT NULL DEFAULT ''")
    ensure_column(conn, "guarantees", "resolution_reference", "TEXT NOT NULL DEFAULT ''")
    # Tránsito interno
    ensure_column(conn, "guarantees", "remito_interno", "TEXT NOT NULL DEFAULT ''")
    ensure_column(conn, "guarantees", "transit_status", "TEXT NOT NULL DEFAULT ''")
    ensure_column(conn, "guarantees", "fecha_salida_transito", "TEXT NOT NULL DEFAULT ''")
    ensure_column(conn, "guarantees", "fecha_llegada_transito", "TEXT NOT NULL DEFAULT ''")
    ensure_column(conn, "guarantees", "lugar_salida_transito", "TEXT NOT NULL DEFAULT ''")
    # Fase 1 — origen, tipo de ingreso y ubicación física
    ensure_column(conn, "guarantees", "tipo_ingreso",   "TEXT NOT NULL DEFAULT ''")
    ensure_column(conn, "guarantees", "origen_ingreso", "TEXT NOT NULL DEFAULT ''")
    ensure_column(conn, "guarantees", "ubicacion_actual", "TEXT NOT NULL DEFAULT ''")
    # Sucursal comercialmente responsable (puede diferir de sucursal_carga en flujo depósito)
    ensure_column(conn, "guarantees", "sucursal_responsable", "TEXT NOT NULL DEFAULT ''")
    # Vínculos con el sistema real de organización (companies + branches)
    ensure_column(conn, "guarantees", "company_id",              "TEXT NOT NULL DEFAULT ''")
    ensure_column(conn, "guarantees", "sucursal_responsable_id", "TEXT NOT NULL DEFAULT ''")
    # Datos opcionales del cliente / compra
    ensure_column(conn, "guarantees", "cliente_nombre",   "TEXT NOT NULL DEFAULT ''")
    ensure_column(conn, "guarantees", "cliente_telefono", "TEXT NOT NULL DEFAULT ''")
    ensure_column(conn, "guarantees", "cliente_email", "TEXT NOT NULL DEFAULT ''")
    ensure_column(conn, "guarantees", "numero_factura",   "TEXT NOT NULL DEFAULT ''")
    ensure_column(conn, "guarantees", "fecha_compra",     "TEXT NOT NULL DEFAULT ''")
    # Fase 34 — garantías agrupadas: ID madre + ítem operativo.
    # Ej: GAR-2026-CAS-0005-01 con parent_warranty_code = GAR-2026-CAS-0005.
    ensure_column(conn, "guarantees", "parent_warranty_code", "TEXT NOT NULL DEFAULT ''")
    ensure_column(conn, "guarantees", "parent_item_index", "INTEGER NOT NULL DEFAULT 0")
    # Campos para fases futuras (se agregan ahora para no migrar después)
    ensure_column(conn, "guarantees", "resultado_resolucion",       "TEXT NOT NULL DEFAULT ''")
    # Fase 12 — resolución normalizada: NC / reparación / cambio de equipo.
    ensure_column(conn, "guarantees", "numero_nota_credito", "TEXT NOT NULL DEFAULT ''")
    ensure_column(conn, "guarantees", "importe_nota_credito", "TEXT NOT NULL DEFAULT ''")
    ensure_column(conn, "guarantees", "fecha_nota_credito", "TEXT NOT NULL DEFAULT ''")
    ensure_column(conn, "guarantees", "detalle_reparacion", "TEXT NOT NULL DEFAULT ''")
    ensure_column(conn, "guarantees", "fecha_reparacion", "TEXT NOT NULL DEFAULT ''")
    ensure_column(conn, "guarantees", "producto_reemplazo", "TEXT NOT NULL DEFAULT ''")
    ensure_column(conn, "guarantees", "sku_reemplazo", "TEXT NOT NULL DEFAULT ''")
    ensure_column(conn, "guarantees", "serie_reemplazo", "TEXT NOT NULL DEFAULT ''")
    ensure_column(conn, "guarantees", "fecha_recepcion_reemplazo", "TEXT NOT NULL DEFAULT ''")
    ensure_column(conn, "guarantees", "fecha_finalizacion", "TEXT NOT NULL DEFAULT ''")
    ensure_column(conn, "guarantees", "fecha_ultimo_mail_proveedor", "TEXT NOT NULL DEFAULT ''")
    ensure_column(conn, "guarantees", "estado_retiro_proveedor", "TEXT NOT NULL DEFAULT 'sin_solicitud'")
    ensure_column(conn, "guarantees", "fecha_solicitud_retiro_proveedor", "TEXT NOT NULL DEFAULT ''")
    ensure_column(conn, "guarantees", "fecha_retiro_proveedor", "TEXT NOT NULL DEFAULT ''")
    ensure_column(conn, "guarantees", "nota_retiro_proveedor", "TEXT NOT NULL DEFAULT ''")
    ensure_column(conn, "guarantees", "remito_proveedor",      "TEXT NOT NULL DEFAULT ''")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_guarantees_code ON guarantees(warranty_code)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_guarantees_status ON guarantees(status)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_guarantees_review_status ON guarantees(review_status)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_guarantees_review_started ON guarantees(review_started_at)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_guarantees_provider ON guarantees(provider_name)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_guarantees_sucursal ON guarantees(sucursal)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_guarantees_branch ON guarantees(branch_id)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_guarantees_company ON guarantees(company_id)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_guarantees_sucursal_resp ON guarantees(sucursal_responsable_id)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_guarantees_tipo_ingreso ON guarantees(tipo_ingreso)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_guarantees_origen_ingreso ON guarantees(origen_ingreso)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_guarantees_ubicacion ON guarantees(ubicacion_actual)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_guarantees_ingreso ON guarantees(ingreso_at)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_guarantees_updated ON guarantees(updated_at)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_guarantees_parent ON guarantees(parent_warranty_code, parent_item_index)")
    _backfill_warranty_org_fields(conn)
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS guarantee_items (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            guarantee_id INTEGER NOT NULL,
            producto TEXT NOT NULL DEFAULT '',
            sku TEXT NOT NULL DEFAULT '',
            marca TEXT NOT NULL DEFAULT '',
            tipo TEXT NOT NULL DEFAULT '',
            serie TEXT NOT NULL DEFAULT '',
            falla TEXT NOT NULL DEFAULT '',
            observaciones TEXT NOT NULL DEFAULT '',
            created_at TEXT NOT NULL,
            updated_at TEXT NOT NULL,
            FOREIGN KEY(guarantee_id) REFERENCES guarantees(id) ON DELETE CASCADE
        )
        """
    )
    ensure_column(conn, "guarantee_items", "item_index", "INTEGER NOT NULL DEFAULT 1")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_guarantee_items_gid ON guarantee_items(guarantee_id)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_guarantee_items_sku ON guarantee_items(sku)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_guarantee_items_marca ON guarantee_items(marca)")
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS guarantee_history (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            guarantee_id INTEGER NOT NULL,
            warranty_code TEXT NOT NULL DEFAULT '',
            created_at TEXT NOT NULL,
            actor_username TEXT NOT NULL DEFAULT '',
            actor_name TEXT NOT NULL DEFAULT '',
            action TEXT NOT NULL,
            old_status TEXT NOT NULL DEFAULT '',
            new_status TEXT NOT NULL DEFAULT '',
            field_name TEXT NOT NULL DEFAULT '',
            old_value TEXT NOT NULL DEFAULT '',
            new_value TEXT NOT NULL DEFAULT '',
            note TEXT NOT NULL DEFAULT '',
            details_json TEXT NOT NULL DEFAULT '{}',
            FOREIGN KEY(guarantee_id) REFERENCES guarantees(id) ON DELETE CASCADE
        )
        """
    )
    conn.execute("CREATE INDEX IF NOT EXISTS idx_guarantee_history_gid ON guarantee_history(guarantee_id)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_guarantee_history_code ON guarantee_history(warranty_code)")
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS guarantee_exports (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            created_at TEXT NOT NULL,
            created_by TEXT NOT NULL DEFAULT '',
            created_by_name TEXT NOT NULL DEFAULT '',
            provider_name TEXT NOT NULL DEFAULT '',
            marca TEXT NOT NULL DEFAULT '',
            filters_json TEXT NOT NULL DEFAULT '{}',
            file_path TEXT NOT NULL,
            file_name TEXT NOT NULL,
            row_count INTEGER NOT NULL DEFAULT 0
        )
        """
    )
    conn.execute("CREATE INDEX IF NOT EXISTS idx_guarantee_exports_created ON guarantee_exports(created_at)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_guarantee_exports_provider ON guarantee_exports(provider_name)")
    ensure_column(conn, "guarantee_exports", "shipment_code", "TEXT NOT NULL DEFAULT ''")
    ensure_column(conn, "guarantee_exports", "warranty_ids_json", "TEXT NOT NULL DEFAULT '[]'")
    ensure_column(conn, "guarantee_exports", "file_format", "TEXT NOT NULL DEFAULT 'excel'")
    ensure_column(conn, "guarantee_exports", "logo_brand", "TEXT NOT NULL DEFAULT 'gv_electro'")
    # Confirmación de retiro del proveedor
    ensure_column(conn, "guarantee_exports", "punto_retiro", "TEXT NOT NULL DEFAULT ''")
    ensure_column(conn, "guarantee_exports", "tipo_retiro", "TEXT NOT NULL DEFAULT ''")
    ensure_column(conn, "guarantee_exports", "fecha_retiro_acordada", "TEXT NOT NULL DEFAULT ''")
    ensure_column(conn, "guarantee_exports", "respuesta_proveedor_pickup", "TEXT NOT NULL DEFAULT ''")
    ensure_column(conn, "guarantee_exports", "pickup_alert_sent", "INTEGER NOT NULL DEFAULT 0")
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS guarantee_sync_logs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            sync_type TEXT NOT NULL,
            status TEXT NOT NULL,
            started_at TEXT NOT NULL,
            finished_at TEXT NOT NULL DEFAULT '',
            actor_username TEXT NOT NULL DEFAULT '',
            actor_name TEXT NOT NULL DEFAULT '',
            rows_processed INTEGER NOT NULL DEFAULT 0,
            rows_created INTEGER NOT NULL DEFAULT 0,
            rows_updated INTEGER NOT NULL DEFAULT 0,
            rows_skipped INTEGER NOT NULL DEFAULT 0,
            errors_json TEXT NOT NULL DEFAULT '[]'
        )
        """
    )
    conn.execute("CREATE INDEX IF NOT EXISTS idx_guarantee_sync_logs_created ON guarantee_sync_logs(started_at)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_guarantee_sync_logs_type ON guarantee_sync_logs(sync_type)")
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS guarantee_counters (
            year INTEGER NOT NULL,
            sucursal_code TEXT NOT NULL,
            last_number INTEGER NOT NULL DEFAULT 0,
            updated_at TEXT NOT NULL,
            PRIMARY KEY(year, sucursal_code)
        )
        """
    )
    # Normalización compatible de estados viejos/internos. No borra datos: solo evita
    # que el flujo visual mezcle revisión interna, proveedor y cierre.
    now_norm = utc_now_iso()
    conn.execute(
        """
        UPDATE guarantees
        SET review_status = 'en_revision'
        WHERE UPPER(TRIM(review_status)) IN ('', 'PENDIENTE_REVISION', 'PENDIENTE REVISION', 'PENDIENTE DE REVISION')
          AND UPPER(TRIM(status)) IN ('EN REVISIÓN', 'EN REVISION')
        """
    )
    conn.execute(
        """
        UPDATE guarantees
        SET review_status = 'requiere_correccion'
        WHERE UPPER(TRIM(status)) IN ('CORRECCIÓN PENDIENTE', 'CORRECCION PENDIENTE')
        """
    )
    conn.execute(
        """
        UPDATE guarantees
        SET status = '1 - INGRESO', updated_at = CASE WHEN updated_at = '' THEN ? ELSE updated_at END
        WHERE UPPER(TRIM(status)) IN ('EN REVISIÓN', 'EN REVISION', 'CORRECCIÓN PENDIENTE', 'CORRECCION PENDIENTE')
        """,
        (now_norm,),
    )
    conn.execute(
        """
        UPDATE guarantees
        SET status = '6 - RESPONDIDO POR PROVEEDOR'
        WHERE UPPER(TRIM(status)) IN ('5 - EN REVISION', '5 - EN REVISIÓN')
          AND TRIM(COALESCE(last_provider_response_at, '')) <> ''
        """
    )
    conn.execute(
        """
        UPDATE guarantees
        SET status = '5 - EN EL PROVEEDOR'
        WHERE UPPER(TRIM(status)) IN ('5 - EN REVISION', '5 - EN REVISIÓN')
          AND TRIM(COALESCE(last_provider_response_at, '')) = ''
        """
    )
    conn.execute("UPDATE guarantees SET status = '7 - RESUELTO' WHERE UPPER(TRIM(status)) = '6 - RESUELTO'")
    conn.execute("UPDATE guarantees SET status = '8 - RECHAZADO' WHERE UPPER(TRIM(status)) = '7 - RECHAZADO'")
    conn.execute("UPDATE guarantees SET status = '9 - ANULADA' WHERE UPPER(TRIM(status)) IN ('ANULADA', 'ANULADO')")


def add_history(
    conn: sqlite3.Connection,
    guarantee_id: int,
    warranty_code: str,
    user: Any,
    action: str,
    *,
    old_status: str = "",
    new_status: str = "",
    field_name: str = "",
    old_value: str = "",
    new_value: str = "",
    note: str = "",
    details: dict[str, Any] | None = None,
) -> None:
    conn.execute(
        """
        INSERT INTO guarantee_history (
            guarantee_id, warranty_code, created_at, actor_username, actor_name, action,
            old_status, new_status, field_name, old_value, new_value, note, details_json
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            guarantee_id,
            warranty_code,
            utc_now_iso(),
            getattr(user, "username", "") or "",
            getattr(user, "display_name", "") or "",
            action,
            old_status or "",
            new_status or "",
            field_name or "",
            old_value or "",
            new_value or "",
            note or "",
            json.dumps(details or {}, ensure_ascii=False),
        ),
    )


# =========================================================
# Google Sheets: opciones/productos auxiliares
# =========================================================

def require_spreadsheet_id() -> str:
    spreadsheet_id = runtime_warranty_config().get("spreadsheet_id")
    if not spreadsheet_id:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Falta configurar la planilla de Garantías en Configuración operativa.",
        )
    return str(spreadsheet_id)


def get_values(sheet_name: str, a1: str) -> list[list[Any]]:
    service = sheets_service()
    spreadsheet_id = require_spreadsheet_id()
    result = service.spreadsheets().values().get(
        spreadsheetId=spreadsheet_id,
        range=f"{quote_sheet_name(sheet_name)}!{a1}",
    ).execute()
    return result.get("values", [])


def find_column(headers: list[str], candidates: list[str], fallback_index: int | None = None) -> int | None:
    keys = {header_key(c) for c in candidates}
    for index, header in enumerate(headers):
        if header_key(header) in keys:
            return index
    if fallback_index is not None and fallback_index < len(headers):
        return fallback_index
    return None


def read_options_from_sheet(sheet_name: str, max_rows: int = 1000) -> dict[str, list[str]]:
    try:
        values = get_values(sheet_name, f"A1:ZZ{max_rows}")
    except Exception:
        return {"sucursales": [], "depositos": [], "estados": []}
    if not values:
        return {"sucursales": [], "depositos": [], "estados": []}
    headers = [str(x).strip() for x in values[0]]
    sucursal_col = find_column(headers, ["SUCURSAL", "LOCAL", "ORIGEN", "SUCURSAL ORIGEN"])
    deposito_col = find_column(headers, ["DEPOSITO", "DEPÓSITO", "DESTINO", "DEPOSITO DESTINO", "LUGAR LLEGADA", "LUGAR DONDE LLEGA"])
    estado_col = find_column(headers, ["ESTADO", "STATUS", "ESTADO GARANTIA", "ESTADO GARANTÍA"])

    def collect(col: int | None) -> list[str]:
        if col is None:
            return []
        out: list[str] = []
        for row in values[1:]:
            if col >= len(row):
                continue
            value = clean_option_value(row[col])
            if not value:
                continue
            if header_key(value) in {"SUCURSAL", "DEPOSITO", "ESTADO", "STATUS"}:
                continue
            out.append(value)
        return unique_keep_order(out)

    return {"sucursales": collect(sucursal_col), "depositos": collect(deposito_col), "estados": collect(estado_col)}


def merge_options(*groups: list[str]) -> list[str]:
    merged: list[str] = []
    for group in groups:
        merged.extend(group or [])
    return unique_keep_order(merged)


def warranty_config_values() -> dict[str, Any]:
    cfg = runtime_warranty_config()
    statuses = list(cfg.get("statuses") or cfg.get("estados") or [])
    if not statuses:
        statuses = list(DEFAULT_STATUSES)
    # Estados canónicos. La config vieja puede traer estados desactualizados; se dejan
    # fuera para que el flujo visual y operativo sea único.
    statuses = unique_keep_order(DEFAULT_STATUSES)
    sucursales = list(cfg.get("sucursales") or DEFAULT_SUCURSALES)
    depositos = list(cfg.get("depositos") or DEFAULT_DEPOSITOS)
    # Finalizado es el cierre real. RESUELTO/RECHAZADO no significan cierre automático.
    final_statuses = list(DEFAULT_FINAL_STATUSES)
    delay_ranges_raw = cfg.get("delay_ranges") or DEFAULT_DELAY_RANGES
    delay_ranges: list[int] = []
    for value in delay_ranges_raw:
        try:
            number = int(value)
            if number > 0 and number not in delay_ranges:
                delay_ranges.append(number)
        except Exception:
            continue
    delay_ranges = sorted(delay_ranges or DEFAULT_DELAY_RANGES)
    required_review_fields = list(cfg.get("required_review_fields") or DEFAULT_REQUIRED_REVIEW_FIELDS)
    return {
        "statuses": statuses,
        "sucursales": sucursales,
        "depositos": depositos,
        "final_statuses": final_statuses,
        "delay_ranges": delay_ranges,
        "required_review_fields": required_review_fields,
        "raw_sheet": str(cfg.get("raw_sheet") or "00_RAW_GARANTIAS"),
        "spreadsheet_url": str(cfg.get("spreadsheet_url") or ""),
    }


def runtime_warranty_options() -> dict[str, Any]:
    cfg = runtime_warranty_config()
    values = warranty_config_values()

    # ── Branches reales del sistema (fuente de verdad) ────────────────────────
    branches_operativas = _fetch_branches_operativas()
    branches_physical = [b for b in branches_operativas if b["type"] == "physical"]
    branches_deposit  = [b for b in branches_operativas if b["type"] == "deposit"]

    # Sucursales: derivadas de branches físicas del sistema.
    # Si no hay branches configuradas todavía, caer a la config operativa como fallback.
    if branches_physical:
        sucursales = clean_select_options([b["name"] for b in branches_physical])
    else:
        sucursales = clean_select_options(list(values["sucursales"] or DEFAULT_SUCURSALES))

    # Depósitos: derivados de branches de tipo deposit del sistema.
    if branches_deposit:
        depositos = clean_select_options([b["name"] for b in branches_deposit], deposit=True)
    else:
        depositos = clean_select_options(list(values["depositos"] or DEFAULT_DEPOSITOS), deposit=True)

    warranty_central_deposit = _warranty_central_deposit_from_branches(branches_operativas)
    warranty_central_deposit_name = _warranty_central_deposit_name(branches_operativas)

    # Estados: forzar DEFAULT_STATUSES como única fuente de verdad visual/operativa.
    estados = unique_keep_order(DEFAULT_STATUSES)

    return {
        "sucursales": sucursales,
        "depositos": depositos,
        "warranty_central_deposit": warranty_central_deposit or {"id": "", "name": warranty_central_deposit_name, "code": "CHICLANA", "type": "deposit", "company_id": "", "company_name": ""},
        "estados": estados,
        "estado_default": str(cfg.get("estado_default") or DEFAULT_STATUSES[0]),
        "review_statuses": [{"value": key, "label": value} for key, value in REVIEW_LABELS.items()],
        "tipos_ingreso": [{"value": k, "label": v} for k, v in TIPO_INGRESO_LABELS.items()],
        "ubicacion_labels": UBICACION_LABELS,
        "resolution_options": [{"value": k, "label": v} for k, v in RESOLUTION_OPTIONS.items()],
        "final_statuses": values["final_statuses"],
        "delay_ranges": values["delay_ranges"],
        "required_review_fields": values["required_review_fields"],
        # Branches con IDs reales — el frontend las usa para selectores con ID.
        "branches_operativas": branches_operativas,
        "source": {"raw_sheet": values["raw_sheet"], "product_sheet": "Catálogo local", "mode": "database_primary"},
    }


def load_product_catalog() -> list[dict[str, str]]:
    cfg = runtime_warranty_config()
    now = time.time()
    cache_seconds = int(cfg.get("product_cache_seconds", 300) or 300)
    if PRODUCT_CACHE["items"] and now - float(PRODUCT_CACHE["loaded_at"]) < cache_seconds:
        return PRODUCT_CACHE["items"]
    try:
        values = get_values(str(cfg.get("product_sheet") or "Productos PVP"), "A:Z")
    except Exception:
        PRODUCT_CACHE["loaded_at"] = now
        PRODUCT_CACHE["items"] = []
        return []
    if not values:
        PRODUCT_CACHE["loaded_at"] = now
        PRODUCT_CACHE["items"] = []
        return []
    headers = [str(x).strip() for x in values[0]]
    producto_col = find_column(headers, ["PRODUCTO", "DESCRIPCION", "DESCRIPCIÓN", "ARTICULO", "ARTÍCULO", "NOMBRE"], fallback_index=2)
    sku_col = find_column(headers, ["SKU", "CODIGO", "CÓDIGO"], fallback_index=3)
    marca_col = find_column(headers, ["MARCA"], fallback_index=0)
    tipo_col = find_column(headers, ["TIPO", "RUBRO", "TIPO PRODUCTO"], fallback_index=1)
    items: list[dict[str, str]] = []
    for raw in values[1:]:
        def get(col: int | None) -> str:
            if col is None or col >= len(raw):
                return ""
            return str(raw[col]).strip()
        producto = get(producto_col)
        sku = get(sku_col)
        marca = get(marca_col)
        tipo = get(tipo_col)
        if not producto and not sku:
            continue
        label = " — ".join(part for part in [producto, sku] if part)
        items.append({"producto": producto, "sku": sku, "marca": marca, "tipo": tipo, "label": label, "search": normalize_text(" ".join([producto, sku, marca, tipo]))})
    PRODUCT_CACHE["loaded_at"] = now
    PRODUCT_CACHE["items"] = items
    return items


# =========================================================
# DB mappers
# =========================================================

def next_warranty_code(conn: sqlite3.Connection, sucursal: str) -> str:
    code = sucursal_code(sucursal)
    year = now_ar().year
    with COUNTER_LOCK:
        ensure_warranty_tables(conn)
        row = conn.execute("SELECT last_number FROM guarantee_counters WHERE year = ? AND sucursal_code = ?", (year, code)).fetchone()
        last = int(row["last_number"] if row else 0)
        # Seguridad adicional por si se importaron IDs sin contador.
        like = f"GAR-{year}-{code}-%"
        rows = conn.execute("SELECT warranty_code FROM guarantees WHERE warranty_code LIKE ?", (like,)).fetchall()
        for existing in rows:
            # Contar también IDs hijos de garantías agrupadas: GAR-YYYY-COD-0005-01
            # para no reutilizar el número madre si no existe una fila con el ID sin sufijo.
            m = re.fullmatch(rf"GAR-{year}-{re.escape(code)}-(\d+)(?:-\d+)?", str(existing["warranty_code"] or ""))
            if m:
                last = max(last, int(m.group(1)))
        next_number = last + 1
        conn.execute(
            "INSERT OR REPLACE INTO guarantee_counters (year, sucursal_code, last_number, updated_at) VALUES (?, ?, ?, ?)",
            (year, code, next_number, utc_now_iso()),
        )
        return f"GAR-{year}-{code}-{next_number:04d}"


def next_shipment_code(conn: sqlite3.Connection) -> str:
    year = now_ar().year
    prefix = f"ENV-{year}-"
    rows = conn.execute(
        "SELECT shipment_code FROM guarantee_exports WHERE shipment_code LIKE ? ORDER BY id DESC",
        (f"{prefix}%",),
    ).fetchall()
    last = 0
    for row in rows:
        m = re.fullmatch(rf"ENV-{year}-(\d+)", str(row["shipment_code"] or ""))
        if m:
            last = max(last, int(m.group(1)))
    return f"ENV-{year}-{(last + 1):04d}"


def collect_export_rows_by_ids(conn: sqlite3.Connection, warranty_codes: list[str]) -> list[dict[str, Any]]:
    """Fetch export rows for a specific list of warranty codes."""
    if not warranty_codes:
        return []
    placeholders = ",".join("?" * len(warranty_codes))
    rows = conn.execute(
        f"""
        SELECT
            g.id AS guarantee_id,
            g.warranty_code,
            g.status,
            g.ingreso_at,
            g.created_at,
            g.fecha_resolucion,
            g.cancelled_at,
            g.sucursal,
            g.deposito,
            g.lugar_llegada,
            g.provider_name,
            g.provider_case_id,
            g.sent_to_provider_at,
            g.last_provider_response_at,
            g.observations,
            i.producto,
            i.sku,
            i.marca,
            i.serie,
            i.falla,
            i.observaciones AS item_observaciones
        FROM guarantees g
        JOIN guarantee_items i ON i.guarantee_id = g.id
        WHERE g.warranty_code IN ({placeholders})
        ORDER BY g.ingreso_at ASC, g.warranty_code ASC, i.id ASC
        """,
        warranty_codes,
    ).fetchall()
    return [{key: row[key] for key in row.keys()} for row in rows]


def insert_guarantee(
    conn: sqlite3.Connection,
    warranty_code: str,
    item: "WarrantyItemIn",
    user: Any,
    *,
    sucursal_carga: str | None = None,
    sucursal_carga_branch_id: str | None = None,
    sucursal_responsable_override: str | None = None,
    sucursal_responsable_id_override: str | None = None,
    company_id_override: str | None = None,
    parent_warranty_code: str | None = None,
    parent_item_index: int | None = None,
) -> int:
    """Inserta una garantía en la DB.

    Parámetros de override (todos keyword-only):
    · sucursal_carga              → nombre de la sucursal de carga (display).
    · sucursal_carga_branch_id   → branch_id real de la sucursal de carga.
    · sucursal_responsable_override    → nombre de la sucursal responsable (display).
    · sucursal_responsable_id_override → branch_id real de la sucursal responsable.
    · company_id_override         → ID de empresa (si no se pasa, se lee del usuario).

    Si no se pasan, se usan los valores de item o del usuario.
    """
    now      = utc_now_iso()
    ingreso_at = ingreso_at_from_input(item.fecha_ingreso)
    lugar    = (item.lugar_llegada or item.deposito or "").strip()
    tipo     = (item.tipo_ingreso or "").strip()
    suc      = sucursal_carga if sucursal_carga is not None else item.sucursal.strip()
    suc_bid  = sucursal_carga_branch_id if sucursal_carga_branch_id is not None else (getattr(user, "branch_id", "") or "")
    suc_resp = sucursal_responsable_override if sucursal_responsable_override is not None else (item.sucursal_responsable or "").strip()
    suc_resp_id = sucursal_responsable_id_override if sucursal_responsable_id_override is not None else (item.sucursal_responsable_id or "").strip()
    company_id = company_id_override if company_id_override is not None else (getattr(user, "company_id", "") or "")
    # Si la garantía nace en depósito, ya está físicamente en depósito (no necesita remito).
    origen = _origen_from_tipo(tipo)
    transit = "en_deposito" if origen == "deposito" else ""
    conn.execute(
        """
        INSERT INTO guarantees (
            warranty_code, parent_warranty_code, parent_item_index, status, review_status,
            responsible_username, responsible_name,
            created_by, created_by_name,
            created_at, ingreso_at, updated_at,
            updated_by, updated_by_name,
            sucursal, sucursal_code, branch_id,
            company_id,
            deposito, lugar_llegada,
            observations, synced_to_google_sheet,
            tipo_ingreso, origen_ingreso, ubicacion_actual,
            transit_status,
            sucursal_responsable, sucursal_responsable_id,
            provider_name,
            cliente_nombre, cliente_telefono, cliente_email, numero_factura, fecha_compra
        ) VALUES (
            ?, ?, ?, ?, ?,
            ?, ?,
            ?, ?,
            ?, ?, ?,
            ?, ?,
            ?, ?, ?,
            ?,
            ?, ?,
            ?, 0,
            ?, ?, ?,
            ?,
            ?, ?,
            ?,
            ?, ?, ?, ?, ?
        )
        """,
        (
            warranty_code,
            (parent_warranty_code or "").strip(),
            int(parent_item_index or 0),
            DEFAULT_STATUSES[0],
            REVIEW_PENDING,
            getattr(user, "username", "") or "",
            getattr(user, "display_name", "") or getattr(user, "username", "") or "",
            getattr(user, "username", "") or "",
            getattr(user, "display_name", "") or "",
            now, ingreso_at, now,
            getattr(user, "username", "") or "",
            getattr(user, "display_name", "") or "",
            suc,
            sucursal_code(suc or item.deposito),
            suc_bid,
            company_id,
            item.deposito.strip(),
            lugar,
            (item.observaciones or "").strip(),
            tipo,
            origen,
            _initial_ubicacion_actual(tipo, suc, item.deposito.strip() or lugar),
            transit,
            suc_resp,
            suc_resp_id,
            (item.proveedor or "").strip(),
            (item.cliente_nombre or "").strip(),
            (item.cliente_telefono or "").strip(),
            (item.cliente_email or "").strip(),
            (item.numero_factura or "").strip(),
            (item.fecha_compra or "").strip(),
        ),
    )
    return int(conn.execute("SELECT last_insert_rowid() AS id").fetchone()["id"])


def insert_item(conn: sqlite3.Connection, guarantee_id: int, item: WarrantyItemIn, item_index: int = 1) -> None:
    now = utc_now_iso()
    conn.execute(
        """
        INSERT INTO guarantee_items (guarantee_id, item_index, producto, sku, marca, tipo, serie, falla, observaciones, created_at, updated_at)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            guarantee_id,
            int(item_index or 1),
            item.producto.strip(),
            (item.sku or "").strip(),
            (item.marca or "").strip(),
            (item.tipo or "").strip(),
            (item.serie or "").strip(),
            item.falla.strip(),
            (item.observaciones or "").strip(),
            now,
            now,
        ),
    )


def fetch_guarantee_with_items(conn: sqlite3.Connection, warranty_code: str) -> tuple[sqlite3.Row, list[sqlite3.Row]] | None:
    ensure_warranty_tables(conn)
    row = conn.execute("SELECT * FROM guarantees WHERE warranty_code = ?", (warranty_code,)).fetchone()
    if not row:
        return None
    items = conn.execute("SELECT * FROM guarantee_items WHERE guarantee_id = ? ORDER BY id", (row["id"],)).fetchall()
    return row, items


def row_to_summary(row: sqlite3.Row, items: list[sqlite3.Row]) -> WarrantySummary:
    products = [str(item["producto"] or "") for item in items if str(item["producto"] or "").strip()]
    first = items[0] if items else None
    ingreso = format_date_ar(parse_iso_datetime(row["ingreso_at"]) or now_ar())
    updated = format_datetime_ar(parse_iso_datetime(row["updated_at"]) or now_ar())
    return WarrantySummary(
        id_garantia=str(row["warranty_code"] or ""),
        parent_warranty_code=str(row["parent_warranty_code"] or "") if "parent_warranty_code" in row.keys() else "",
        parent_item_index=int(row["parent_item_index"] or 0) if "parent_item_index" in row.keys() and row["parent_item_index"] else None,
        grouped_item_label=(f"Ítem {int(row['parent_item_index']):02d} de {row['parent_warranty_code']}" if "parent_warranty_code" in row.keys() and row["parent_warranty_code"] and "parent_item_index" in row.keys() and row["parent_item_index"] else ""),
        ingreso=ingreso,
        ingreso_iso=date_input_from_iso(row["ingreso_at"]),
        responsible_username=str(row["responsible_username"] or ""),
        responsable=str(row["responsible_name"] or ""),
        usuario=str(row["created_by"] or ""),
        producto_principal=(products[0] if products else (str(first["producto"] or "") if first else "")),
        productos=products[:12],
        cantidad_items=len(items),
        marca=str(first["marca"] or "") if first else "",
        sku=str(first["sku"] or "") if first else "",
        serie=str(first["serie"] or "") if first else "",
        falla=str(first["falla"] or "") if first else "",
        sucursal=str(row["sucursal"] or ""),
        sucursal_code=str(row["sucursal_code"] or ""),
        branch_id=str(row["branch_id"] or "") if "branch_id" in row.keys() else "",
        company_id=str(row["company_id"] or "") if "company_id" in row.keys() else "",
        sucursal_responsable=str(row["sucursal_responsable"] or "") if "sucursal_responsable" in row.keys() else "",
        sucursal_responsable_id=str(row["sucursal_responsable_id"] or "") if "sucursal_responsable_id" in row.keys() else "",
        deposito=str(row["deposito"] or ""),
        lugar_llegada=str(row["lugar_llegada"] or row["deposito"] or ""),
        estado=str(row["status"] or ""),
        tipo_ingreso=str(row["tipo_ingreso"] or "") if "tipo_ingreso" in row.keys() else "",
        tipo_ingreso_label=TIPO_INGRESO_LABELS.get(str(row["tipo_ingreso"] or ""), str(row["tipo_ingreso"] or "")) if "tipo_ingreso" in row.keys() else "",
        origen_ingreso=str(row["origen_ingreso"] or "") if "origen_ingreso" in row.keys() else "",
        ubicacion_actual=str(row["ubicacion_actual"] or "") if "ubicacion_actual" in row.keys() else "",
        ubicacion_actual_label=UBICACION_LABELS.get(str(row["ubicacion_actual"] or ""), str(row["ubicacion_actual"] or "")) if "ubicacion_actual" in row.keys() else "",
        cliente_nombre=str(row["cliente_nombre"] or "") if "cliente_nombre" in row.keys() else "",
        cliente_telefono=str(row["cliente_telefono"] or "") if "cliente_telefono" in row.keys() else "",
        cliente_email=str(row["cliente_email"] or "") if "cliente_email" in row.keys() else "",
        numero_factura=str(row["numero_factura"] or "") if "numero_factura" in row.keys() else "",
        fecha_compra=str(row["fecha_compra"] or "") if "fecha_compra" in row.keys() else "",
        review_status=str(row["review_status"] or REVIEW_PENDING),
        review_status_label=REVIEW_LABELS.get(str(row["review_status"] or REVIEW_PENDING), str(row["review_status"] or REVIEW_PENDING)),
        reviewed_by=str(row["reviewed_by"] or ""),
        reviewed_by_name=str(row["reviewed_by_name"] or ""),
        reviewed_at=format_datetime_ar(parse_iso_datetime(row["reviewed_at"])) if row["reviewed_at"] else "",
        review_note=str(row["review_note"] or ""),
        observaciones=str(row["observations"] or ""),
        photos_reference=str(row["photos_reference"] or ""),
        provider_name=str(row["provider_name"] or ""),
        id_de_caso=str(row["provider_case_id"] or ""),
        fecha_envio_proveedor=format_datetime_ar(parse_iso_datetime(row["sent_to_provider_at"])) if row["sent_to_provider_at"] else "",
        fecha_ultima_respuesta=format_datetime_ar(parse_iso_datetime(row["last_provider_response_at"])) if row["last_provider_response_at"] else "",
        fecha_ultimo_reclamo=format_datetime_ar(parse_iso_datetime(row["last_claim_at"])) if "last_claim_at" in row.keys() and row["last_claim_at"] else "",
        estado_retiro_proveedor=normalize_provider_pickup_status(row["estado_retiro_proveedor"]) if "estado_retiro_proveedor" in row.keys() else "sin_solicitud",
        estado_retiro_proveedor_label=PROVIDER_PICKUP_STATUSES.get(normalize_provider_pickup_status(row["estado_retiro_proveedor"]) if "estado_retiro_proveedor" in row.keys() else "sin_solicitud", "Sin solicitud"),
        fecha_solicitud_retiro_proveedor=format_datetime_ar(parse_iso_datetime(row["fecha_solicitud_retiro_proveedor"])) if "fecha_solicitud_retiro_proveedor" in row.keys() and row["fecha_solicitud_retiro_proveedor"] else "",
        fecha_retiro_proveedor=format_datetime_ar(parse_iso_datetime(row["fecha_retiro_proveedor"])) if "fecha_retiro_proveedor" in row.keys() and row["fecha_retiro_proveedor"] else (format_datetime_ar(parse_iso_datetime(row["fecha_retiro"])) if "fecha_retiro" in row.keys() and row["fecha_retiro"] else ""),
        dias_pendiente=compute_pending_days(row),
        dias_sin_respuesta=compute_no_response_days(row),
        shipment_code=str(row["shipment_code"] or "") if "shipment_code" in row.keys() else "",
        shipment_file_name=str(row["shipment_file_name"] or "") if "shipment_file_name" in row.keys() else "",
        resolution_note=str(row["resolution_note"] or "") if "resolution_note" in row.keys() else "",
        resolution_reference=str(row["resolution_reference"] or "") if "resolution_reference" in row.keys() else "",
        resultado_resolucion=normalize_resolution_result(row["resultado_resolucion"]) if "resultado_resolucion" in row.keys() else "",
        resultado_resolucion_label=RESOLUTION_OPTIONS.get(normalize_resolution_result(row["resultado_resolucion"]) if "resultado_resolucion" in row.keys() else "", ""),
        numero_nota_credito=str(row["numero_nota_credito"] or "") if "numero_nota_credito" in row.keys() else "",
        importe_nota_credito=str(row["importe_nota_credito"] or "") if "importe_nota_credito" in row.keys() else "",
        fecha_nota_credito=str(row["fecha_nota_credito"] or "") if "fecha_nota_credito" in row.keys() else "",
        detalle_reparacion=str(row["detalle_reparacion"] or "") if "detalle_reparacion" in row.keys() else "",
        fecha_reparacion=str(row["fecha_reparacion"] or "") if "fecha_reparacion" in row.keys() else "",
        producto_reemplazo=str(row["producto_reemplazo"] or "") if "producto_reemplazo" in row.keys() else "",
        sku_reemplazo=str(row["sku_reemplazo"] or "") if "sku_reemplazo" in row.keys() else "",
        serie_reemplazo=str(row["serie_reemplazo"] or "") if "serie_reemplazo" in row.keys() else "",
        fecha_recepcion_reemplazo=str(row["fecha_recepcion_reemplazo"] or "") if "fecha_recepcion_reemplazo" in row.keys() else "",
        fecha_finalizacion=format_datetime_ar(parse_iso_datetime(row["fecha_finalizacion"])) if "fecha_finalizacion" in row.keys() and row["fecha_finalizacion"] else "",
        finalizacion=str(row["finalizacion"] or "") if "finalizacion" in row.keys() else "",
        remito_interno=str(row["remito_interno"] or "") if "remito_interno" in row.keys() else "",
        remito_proveedor=str(row["remito_proveedor"] or "") if "remito_proveedor" in row.keys() else "",
        transit_status=str(row["transit_status"] or "") if "transit_status" in row.keys() else "",
        synced_to_google_sheet=bool(row["synced_to_google_sheet"]),
        fecha_ultima_sincronizacion=format_datetime_ar(parse_iso_datetime(row["last_google_sync_at"])) if row["last_google_sync_at"] else "",
        actualizado_por=str(row["updated_by_name"] or row["updated_by"] or ""),
        fecha_ultima_actualizacion=updated,
        cancelled=bool(row["cancelled"]),
        cancel_reason=str(row["cancel_reason"] or ""),
        cancelled_by=str(row["cancelled_by"] or ""),
        cancelled_at=format_datetime_ar(parse_iso_datetime(row["cancelled_at"])) if row["cancelled_at"] else "",
    )


def item_to_row(warranty_row: sqlite3.Row, item: sqlite3.Row, index: int) -> WarrantyRow:
    return WarrantyRow(
        row_number=int(item["id"] or index),
        id_garantia=str(warranty_row["warranty_code"] or ""),
        responsable=str(warranty_row["responsible_name"] or ""),
        usuario=str(warranty_row["created_by"] or ""),
        ingreso=format_date_ar(parse_iso_datetime(warranty_row["ingreso_at"]) or now_ar()),
        producto=str(item["producto"] or ""),
        sku=str(item["sku"] or ""),
        marca=str(item["marca"] or ""),
        tipo=str(item["tipo"] or ""),
        serie=str(item["serie"] or ""),
        falla=str(item["falla"] or ""),
        sucursal=str(warranty_row["sucursal"] or ""),
        deposito=str(warranty_row["deposito"] or ""),
        lugar_llegada=str(warranty_row["lugar_llegada"] or warranty_row["deposito"] or ""),
        estado=str(warranty_row["status"] or ""),
        observaciones=str(item["observaciones"] or warranty_row["observations"] or ""),
        actualizado_por=str(warranty_row["updated_by_name"] or warranty_row["updated_by"] or ""),
        fecha_ultima_actualizacion=format_datetime_ar(parse_iso_datetime(warranty_row["updated_at"]) or now_ar()),
    )


def history_for_guarantee(conn: sqlite3.Connection, guarantee_id: int, limit: int = 200) -> list[dict[str, Any]]:
    rows = conn.execute(
        "SELECT * FROM guarantee_history WHERE guarantee_id = ? ORDER BY id DESC LIMIT ?",
        (guarantee_id, limit),
    ).fetchall()
    events: list[dict[str, Any]] = []
    for row in rows:
        details: dict[str, Any] = {}
        try:
            details = json.loads(row["details_json"] or "{}")
        except Exception:
            details = {}
        events.append({
            "id": row["id"],
            "created_at": format_datetime_ar(parse_iso_datetime(row["created_at"]) or now_ar()),
            "event_type": row["action"],
            "actor_username": row["actor_username"],
            "actor_display_name": row["actor_name"],
            "actor_role": None,
            "resource_type": "warranty",
            "resource_id": row["warranty_code"],
            "status": "ok",
            "message": row["note"] or "",
            "details": {
                **details,
                "old_status": row["old_status"],
                "new_status": row["new_status"],
                "field_name": row["field_name"],
                "old_value": row["old_value"],
                "new_value": row["new_value"],
            },
        })
    return events


def fetch_all_guarantee_summaries() -> list[WarrantySummary]:
    with db_connect() as conn:
        ensure_warranty_tables(conn)
        rows = conn.execute("SELECT * FROM guarantees ORDER BY ingreso_at DESC, id DESC").fetchall()
        all_items = conn.execute("SELECT * FROM guarantee_items ORDER BY id").fetchall()
    by_gid: dict[int, list[sqlite3.Row]] = {}
    for item in all_items:
        by_gid.setdefault(int(item["guarantee_id"]), []).append(item)
    return [row_to_summary(row, by_gid.get(int(row["id"]), [])) for row in rows]


def validate_status_or_400(estado: str) -> str:
    clean = str(estado or "").strip()
    allowed = {canonical_status_key(x): x for x in DEFAULT_STATUSES}
    key = canonical_status_key(clean)
    if key not in allowed:
        raise HTTPException(status_code=400, detail=f"Estado inválido para garantías: {clean}")
    return allowed[key]


def update_guarantee_provider_fields(
    conn: sqlite3.Connection,
    *,
    row: sqlite3.Row,
    user: Any,
    updates: dict[str, str],
    action: str,
    note: str,
    old_status: str = "",
    new_status: str = "",
    details: dict[str, Any] | None = None,
) -> None:
    updates["updated_at"] = utc_now_iso()
    updates["updated_by"] = getattr(user, "username", "") or ""
    updates["updated_by_name"] = getattr(user, "display_name", "") or ""
    updates["synced_to_google_sheet"] = "0"
    assignments = ", ".join([f"{key} = ?" for key in updates])
    conn.execute(f"UPDATE guarantees SET {assignments} WHERE id = ?", [*updates.values(), int(row["id"])])
    add_history(
        conn,
        int(row["id"]),
        str(row["warranty_code"]),
        user,
        action,
        old_status=old_status,
        new_status=new_status,
        note=note,
        details=details or {},
    )




def warranty_exports_dir() -> Path:
    settings = get_settings()
    settings.ensure_dirs()
    path = settings.outputs_dir / "warranties" / "exports"
    path.mkdir(parents=True, exist_ok=True)
    return path


def safe_filename_part(value: Any, fallback: str = "general") -> str:
    text = normalize_text(value).lower().replace(" ", "-")
    text = re.sub(r"[^a-z0-9-]+", "", text).strip("-")
    return text or fallback


def export_row_matches(row: dict[str, Any], filters: WarrantyExportRequest) -> bool:
    if filters.marca and normalize_text(row.get("marca")) != normalize_text(filters.marca):
        return False
    if filters.proveedor and normalize_text(row.get("provider_name")) != normalize_text(filters.proveedor):
        return False
    if filters.estado and normalize_text(row.get("status")) != normalize_text(filters.estado):
        return False
    if filters.sucursal and normalize_text(row.get("sucursal")) != normalize_text(filters.sucursal):
        return False
    if filters.deposito:
        dep = normalize_text(row.get("deposito"))
        llegada = normalize_text(row.get("lugar_llegada"))
        wanted = normalize_text(filters.deposito)
        if wanted not in {dep, llegada}:
            return False
    date_from = parse_date_filter(filters.fecha_desde)
    date_to = parse_date_filter(filters.fecha_hasta)
    ingreso_dt = parse_iso_datetime(row.get("ingreso_at"))
    ingreso_date = ingreso_dt.date() if ingreso_dt else None
    if date_from and ingreso_date and ingreso_date < date_from:
        return False
    if date_to and ingreso_date and ingreso_date > date_to:
        return False
    return True


def collect_export_rows(conn: sqlite3.Connection, filters: WarrantyExportRequest) -> list[dict[str, Any]]:
    ensure_warranty_tables(conn)
    rows = conn.execute(
        """
        SELECT
            g.id AS guarantee_id,
            g.warranty_code,
            g.status,
            g.ingreso_at,
            g.created_at,
            g.fecha_resolucion,
            g.cancelled_at,
            g.sucursal,
            g.deposito,
            g.lugar_llegada,
            g.provider_name,
            g.provider_case_id,
            g.sent_to_provider_at,
            g.last_provider_response_at,
            g.observations,
            i.producto,
            i.sku,
            i.marca,
            i.serie,
            i.falla,
            i.observaciones AS item_observaciones
        FROM guarantees g
        JOIN guarantee_items i ON i.guarantee_id = g.id
        WHERE COALESCE(g.cancelled, 0) = 0
        ORDER BY g.ingreso_at ASC, g.warranty_code ASC, i.id ASC
        """
    ).fetchall()
    output: list[dict[str, Any]] = []
    for raw in rows:
        item = {key: raw[key] for key in raw.keys()}
        if export_row_matches(item, filters):
            output.append(item)
    return output


def export_info_from_row(row: sqlite3.Row) -> WarrantyExportInfo:
    filters: dict[str, Any] = {}
    try:
        parsed = json.loads(row["filters_json"] or "{}")
        if isinstance(parsed, dict):
            filters = parsed
    except Exception:
        filters = {}
    return WarrantyExportInfo(
        id=int(row["id"]),
        created_at=format_datetime_ar(parse_iso_datetime(row["created_at"]) or now_ar()),
        created_by=str(row["created_by_name"] or row["created_by"] or ""),
        provider_name=str(row["provider_name"] or ""),
        marca=str(row["marca"] or ""),
        filters=filters,
        file_name=str(row["file_name"] or ""),
        row_count=int(row["row_count"] or 0),
        download_url=f"/api/warranties/exports/{int(row['id'])}/download",
        shipment_code=str(row["shipment_code"] or "") if "shipment_code" in row.keys() else "",
        file_format=str(row["file_format"] or "excel") if "file_format" in row.keys() else "excel",
        logo_brand=str(row["logo_brand"] or "gv_electro") if "logo_brand" in row.keys() else "gv_electro",
    )


def _export_brand_info(logo_brand: str) -> dict[str, str]:
    brand = normalize_export_logo(logo_brand)
    if brand == "abc_electro":
        return {"key": "abc_electro", "label": "ABC Electro", "logo_file": "abc_electro.png", "accent": "2563EB"}
    return {"key": "gv_electro", "label": "GV Electro", "logo_file": "gv_electro.png", "accent": "1E293B"}


def _export_logo_path(logo_brand: str) -> Path | None:
    info = _export_brand_info(logo_brand)
    cfg = get_settings()
    candidates = [
        cfg.storage_dir / "logos" / info["logo_file"],
        cfg.project_dir / "storage" / "logos" / info["logo_file"],
        cfg.project_dir / "backend" / "storage" / "logos" / info["logo_file"],
        Path("storage") / "logos" / info["logo_file"],
        Path("backend") / "storage" / "logos" / info["logo_file"],
    ]
    for candidate in candidates:
        try:
            if candidate.exists():
                return candidate
        except Exception:
            continue
    return None


def build_provider_excel(rows: list[dict[str, Any]], file_path: Path, *, provider_name: str = "", shipment_code: str = "", logo_brand: str = "gv_electro") -> None:
    """Excel externo profesional para proveedor.

    La cabecera puede incluir logo/título, pero la tabla mantiene solo los campos
    que el proveedor necesita: ID, producto, SKU, serie y falla.
    """
    brand = _export_brand_info(logo_brand)
    accent = brand["accent"]
    dark = "0F172A"
    muted = "64748B"
    soft = "F8FAFC"
    row_alt = "F1F5F9"

    wb = Workbook()
    ws = wb.active
    ws.title = "Garantías"
    ws.sheet_view.showGridLines = False

    # Configuración de impresión: que el archivo se vea presentable al abrirlo
    # y también al exportarlo/imprimirlo desde Excel.
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins.left = 0.35
    ws.page_margins.right = 0.35
    ws.page_margins.top = 0.45
    ws.page_margins.bottom = 0.45

    # Anchos pensados para lectura real, no solo para que entren los datos.
    widths = [24, 64, 20, 22, 56]
    for index, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(index)].width = width

    # Bloque de encabezado profesional.
    for row_num in range(1, 7):
        ws.row_dimensions[row_num].height = 24
        for col_num in range(1, 6):
            ws.cell(row=row_num, column=col_num).fill = PatternFill("solid", fgColor=soft)

    ws.row_dimensions[1].height = 32
    ws.row_dimensions[2].height = 24
    ws.row_dimensions[3].height = 22
    ws.row_dimensions[4].height = 22
    ws.row_dimensions[5].height = 18
    ws.row_dimensions[6].height = 8

    logo_path = _export_logo_path(logo_brand)
    if logo_path:
        try:
            logo = XLImage(str(logo_path))
            logo.width = 86
            logo.height = 86
            ws.add_image(logo, "A1")
        except Exception:
            pass

    # Encabezado textual. Se deja espacio al logo en A1:A4.
    ws.merge_cells("B1:E1")
    ws["B1"] = "LOTE DE GARANTÍAS PARA PROVEEDOR"
    ws["B1"].font = Font(bold=True, size=18, color=dark)
    ws["B1"].alignment = Alignment(horizontal="left", vertical="center")

    ws.merge_cells("B2:E2")
    ws["B2"] = f"N° {shipment_code or 'ENV'}"
    ws["B2"].font = Font(bold=True, size=11, color=accent)
    ws["B2"].alignment = Alignment(horizontal="left", vertical="center")

    ws.merge_cells("B3:E3")
    ws["B3"] = f"Proveedor: {provider_name or '—'}"
    ws["B3"].font = Font(bold=True, size=11, color=dark)
    ws["B3"].alignment = Alignment(horizontal="left", vertical="center")

    ws.merge_cells("B4:E4")
    ws["B4"] = f"Fecha de emisión: {format_datetime_ar(now_ar())} · Cantidad: {len(rows)} garantía(s)"
    ws["B4"].font = Font(size=9, color=muted)
    ws["B4"].alignment = Alignment(horizontal="left", vertical="center")

    ws.merge_cells("A6:E6")
    ws["A6"] = ""
    ws["A6"].fill = PatternFill("solid", fgColor=accent)

    headers = ["ID GARANTÍA", "PRODUCTO", "SKU", "N° SERIE", "FALLA"]
    header_row = 8
    for col, value in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col, value=value)
        cell.fill = PatternFill("solid", fgColor=dark)
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[header_row].height = 26

    thin = Side(style="thin", color="CBD5E1")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for row_idx, item in enumerate(rows, start=header_row + 1):
        values = [
            item.get("warranty_code") or "",
            item.get("producto") or "",
            item.get("sku") or "",
            item.get("serie") or "",
            item.get("falla") or "",
        ]
        fill = PatternFill("solid", fgColor=row_alt if (row_idx - header_row) % 2 == 0 else "FFFFFF")
        # Altura dinámica simple para que fallas o productos largos no queden aplastados.
        product_len = len(str(values[1] or ""))
        falla_len = len(str(values[4] or ""))
        ws.row_dimensions[row_idx].height = max(28, min(64, 18 + 10 * max(product_len // 55, falla_len // 45)))
        for col, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.fill = fill
            cell.border = border
            cell.font = Font(size=10, color="111827")
            if col == 1:
                cell.font = Font(size=10, bold=True, color=dark)
            if col in (3, 4):
                # Texto explícito para evitar que Excel convierta series/SKU largos en números.
                cell.number_format = "@"
            cell.alignment = Alignment(vertical="top", wrap_text=True, horizontal="left")

    # Bordes de encabezado al final para que queden por encima de los estilos.
    for cell in ws[header_row]:
        cell.border = border

    last_row = header_row + max(len(rows), 1)
    if not rows:
        ws.cell(row=header_row + 1, column=1, value="Sin garantías seleccionadas")
        ws.merge_cells(start_row=header_row + 1, start_column=1, end_row=header_row + 1, end_column=5)
        ws.cell(row=header_row + 1, column=1).alignment = Alignment(horizontal="center")
        ws.cell(row=header_row + 1, column=1).font = Font(italic=True, color=muted)
        last_row = header_row + 1

    try:
        table_ref = f"A{header_row}:E{last_row}"
        tab = XLTable(displayName="TablaGarantiasProveedor", ref=table_ref)
        style = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False, showRowStripes=False, showColumnStripes=False)
        tab.tableStyleInfo = style
        ws.add_table(tab)
    except Exception:
        ws.auto_filter.ref = f"A{header_row}:E{last_row}"

    ws.freeze_panes = f"A{header_row + 1}"
    ws.print_title_rows = f"{header_row}:{header_row}"

    # Sin leyendas internas ni datos de empresa: el logo identifica la marca emisora.

    file_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(file_path)

def build_provider_pdf(rows: list[dict[str, Any]], file_path: Path, *, provider_name: str = "", shipment_code: str = "", logo_brand: str = "gv_electro") -> None:
    """PDF externo profesional para proveedor con tabla mínima.

    Fase 44: se genera en horizontal y con celdas envueltas para evitar
    superposición en productos/SKU/series largos.
    """
    brand = _export_brand_info(logo_brand)
    file_path.parent.mkdir(parents=True, exist_ok=True)

    page_size = landscape(A4)
    c = canvas.Canvas(str(file_path), pagesize=page_size)
    width, height = page_size
    margin_x = 12 * mm
    top_y = height - 12 * mm
    logo_path = _export_logo_path(logo_brand)

    # Encabezado horizontal: más aire y más ancho útil para la tabla.
    if logo_path:
        try:
            c.drawImage(str(logo_path), margin_x, top_y - 24 * mm, width=24 * mm, height=24 * mm, mask="auto")
        except Exception:
            pass

    c.setFillColor(colors.HexColor("#111827"))
    c.setFont("Helvetica-Bold", 18)
    c.drawRightString(width - margin_x, top_y - 6 * mm, "LOTE DE GARANTÍAS")
    c.setFont("Helvetica-Bold", 11)
    c.drawRightString(width - margin_x, top_y - 13 * mm, "PARA PROVEEDOR")
    c.setFont("Helvetica-Bold", 10)
    c.drawRightString(width - margin_x, top_y - 20 * mm, f"N° {shipment_code or 'ENV'}")
    c.setFont("Helvetica", 8)
    c.drawRightString(width - margin_x, top_y - 25 * mm, format_datetime_ar(now_ar()))

    y = top_y - 32 * mm
    c.setStrokeColor(colors.HexColor("#334155"))
    c.setLineWidth(0.8)
    c.line(margin_x, y, width - margin_x, y)

    y -= 8 * mm
    c.setFont("Helvetica-Bold", 10)
    c.setFillColor(colors.HexColor("#0F172A"))
    c.drawString(margin_x, y, f"Proveedor: {provider_name or '—'}")
    y -= 9 * mm

    header_style = ParagraphStyle(
        "ProviderHeaderCell",
        fontName="Helvetica-Bold",
        fontSize=7.5,
        leading=9,
        textColor=colors.white,
        alignment=0,
    )
    cell_style = ParagraphStyle(
        "ProviderBodyCell",
        fontName="Helvetica",
        fontSize=7.4,
        leading=9.2,
        textColor=colors.HexColor("#111827"),
        wordWrap="CJK",
        splitLongWords=True,
    )

    def pcell(value: Any, style: ParagraphStyle = cell_style) -> Paragraph:
        text = str(value or "")
        text = (
            text.replace("&", "&amp;")
            .replace("<", "&lt;")
            .replace(">", "&gt;")
            .replace("\n", "<br/>")
        )
        return Paragraph(text, style)

    headers = ["ID GARANTIA", "PRODUCTO", "SKU", "N° SERIE", "FALLA"]
    data: list[list[Any]] = [[pcell(h, header_style) for h in headers]]
    for item in rows:
        data.append([
            pcell(item.get("warranty_code") or ""),
            pcell(item.get("producto") or ""),
            pcell(item.get("sku") or ""),
            pcell(item.get("serie") or ""),
            pcell(item.get("falla") or ""),
        ])

    # Anchos para A4 apaisado. El producto y falla tienen prioridad; SKU/serie
    # también envuelven para no invadir columnas vecinas.
    col_widths = [38 * mm, 96 * mm, 42 * mm, 38 * mm, 58 * mm]
    table = Table(data, colWidths=col_widths, repeatRows=1, hAlign="LEFT")
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#" + brand["accent"])),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#CBD5E1")),
        ("LEFTPADDING", (0, 0), (-1, -1), 4),
        ("RIGHTPADDING", (0, 0), (-1, -1), 4),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8FAFC")]),
    ]))

    # Permitir múltiples páginas si hay muchos ítems.
    available_height = y - 14 * mm
    parts = table.splitOn(c, width - 2 * margin_x, available_height) or [table]
    for idx, part in enumerate(parts):
        if idx > 0:
            c.showPage()
            y = height - 16 * mm
            available_height = y - 14 * mm
        tw, th = part.wrapOn(c, width - 2 * margin_x, available_height)
        part.drawOn(c, margin_x, y - th)

    c.setFont("Helvetica", 6.5)
    c.setFillColor(colors.HexColor("#94A3B8"))
    c.drawCentredString(width / 2, 8 * mm, f"{shipment_code or ''}")
    c.save()


# =========================================================
# Sincronización controlada con Google Sheets
# =========================================================

def sheet_raw_name() -> str:
    cfg = runtime_warranty_config()
    return str(cfg.get("raw_sheet") or "00_RAW_GARANTIAS")


def sheet_date_to_iso(value: Any) -> str:
    text = str(value or "").strip()
    if not text:
        return utc_now_iso()
    for fmt in ("%d/%m/%Y %H:%M", "%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%d/%m/%y"):
        try:
            dt = datetime.strptime(text[:16] if "%H" in fmt else text[:10], fmt)
            return dt.replace(tzinfo=ZoneInfo("America/Argentina/Buenos_Aires")).astimezone(timezone.utc).isoformat()
        except Exception:
            pass
    parsed = parse_iso_datetime(text)
    return parsed.isoformat() if parsed else utc_now_iso()


def read_raw_sheet_rows() -> tuple[list[str], list[list[Any]]]:
    values = get_values(sheet_raw_name(), "A1:ZZ50000")
    if not values:
        return DEFAULT_RAW_HEADERS[:], []
    headers = [str(x).strip() for x in values[0]]
    return headers, values[1:]


def ensure_raw_sheet_tab_exists() -> bool:
    """Crea la pestaña raw en el spreadsheet si no existe. Devuelve True si la tuvo que crear."""
    service = sheets_service()
    spreadsheet_id = require_spreadsheet_id()
    raw_sheet = sheet_raw_name()
    spreadsheet = service.spreadsheets().get(spreadsheetId=spreadsheet_id).execute()
    existing = [s["properties"]["title"] for s in spreadsheet.get("sheets", [])]
    if raw_sheet in existing:
        return False
    service.spreadsheets().batchUpdate(
        spreadsheetId=spreadsheet_id,
        body={"requests": [{"addSheet": {"properties": {"title": raw_sheet}}}]},
    ).execute()
    return True


def ensure_raw_sheet_headers() -> None:
    service = sheets_service()
    spreadsheet_id = require_spreadsheet_id()
    raw_sheet = sheet_raw_name()
    # Crea la pestaña si no existe antes de escribir los headers
    ensure_raw_sheet_tab_exists()
    service.spreadsheets().values().update(
        spreadsheetId=spreadsheet_id,
        range=f"{quote_sheet_name(raw_sheet)}!A1:{get_column_letter(len(DEFAULT_RAW_HEADERS))}1",
        valueInputOption="USER_ENTERED",
        body={"values": [DEFAULT_RAW_HEADERS]},
    ).execute()


def _spreadsheet_titles(service: Any, spreadsheet_id: str) -> set[str]:
    meta = service.spreadsheets().get(spreadsheetId=spreadsheet_id, fields="sheets(properties(title))").execute()
    return {str(sheet.get("properties", {}).get("title", "")) for sheet in meta.get("sheets", [])}


def ensure_sheet_with_headers(service: Any, spreadsheet_id: str, sheet_name: str, headers: list[str], existing_titles: set[str] | None = None) -> bool:
    """Crea/verifica una pestaña y escribe headers. Devuelve True si la creó."""
    created = False
    titles = existing_titles if existing_titles is not None else _spreadsheet_titles(service, spreadsheet_id)
    if sheet_name not in titles:
        service.spreadsheets().batchUpdate(
            spreadsheetId=spreadsheet_id,
            body={"requests": [{"addSheet": {"properties": {"title": sheet_name}}}]},
        ).execute()
        titles.add(sheet_name)
        created = True
    service.spreadsheets().values().update(
        spreadsheetId=spreadsheet_id,
        range=f"{quote_sheet_name(sheet_name)}!A1:{get_column_letter(len(headers))}1",
        valueInputOption="USER_ENTERED",
        body={"values": [headers]},
    ).execute()
    return created


def ensure_mirror_sheets() -> dict[str, bool]:
    service = sheets_service()
    spreadsheet_id = require_spreadsheet_id()
    existing = _spreadsheet_titles(service, spreadsheet_id)
    created: dict[str, bool] = {}
    # Raw histórico + pestañas espejo nuevas.
    created[sheet_raw_name()] = ensure_sheet_with_headers(service, spreadsheet_id, sheet_raw_name(), DEFAULT_RAW_HEADERS, existing)
    for name, headers in MIRROR_SHEETS.items():
        created[name] = ensure_sheet_with_headers(service, spreadsheet_id, name, headers, existing)
    return created


def _table_exists(conn: sqlite3.Connection, table_name: str) -> bool:
    row = conn.execute("SELECT name FROM sqlite_master WHERE type='table' AND name = ?", (table_name,)).fetchone()
    return bool(row)


def _row_value(row: sqlite3.Row | dict[str, Any] | None, key: str, default: Any = "") -> Any:
    if row is None:
        return default
    try:
        if hasattr(row, "keys") and key in row.keys():
            return row[key]
        if isinstance(row, dict):
            return row.get(key, default)
    except Exception:
        pass
    return default


def _fmt_sheet_dt(value: Any) -> str:
    if not value:
        return ""
    return format_datetime_ar(parse_iso_datetime(value)) if parse_iso_datetime(value) else str(value)


def _fmt_sheet_date(value: Any) -> str:
    if not value:
        return ""
    return format_date_ar(parse_iso_datetime(value)) if parse_iso_datetime(value) else str(value)


def _json_list(value: Any) -> list[Any]:
    try:
        parsed = json.loads(str(value or "[]"))
        return parsed if isinstance(parsed, list) else []
    except Exception:
        return []


def _warranty_item_summary(conn: sqlite3.Connection, guarantee_id: int) -> tuple[int, str, str, str, str]:
    rows = conn.execute("SELECT * FROM guarantee_items WHERE guarantee_id = ? ORDER BY id", (guarantee_id,)).fetchall()
    if not rows:
        return 0, "", "", "", ""
    first = rows[0]
    return len(rows), str(first["producto"] or ""), str(first["sku"] or ""), str(first["serie"] or ""), str(first["falla"] or "")


def mirror_rows_for_sheet(conn: sqlite3.Connection, sheet_name: str) -> list[list[Any]]:
    ensure_warranty_tables(conn)
    sheet = sheet_name.upper().strip()
    if sheet == "GARANTIAS":
        rows = conn.execute("SELECT * FROM guarantees ORDER BY ingreso_at ASC, warranty_code ASC, id ASC").fetchall()
        values: list[list[Any]] = []
        for row in rows:
            item_count, producto, sku, serie, falla = _warranty_item_summary(conn, int(row["id"]))
            days = compute_no_response_days(row)
            values.append([
                row["warranty_code"],
                _fmt_sheet_dt(row["ingreso_at"]),
                row["company_id"] if "company_id" in row.keys() else "",
                row["sucursal"] or row["branch_name"] if "branch_name" in row.keys() else row["sucursal"],
                row["sucursal_responsable"] if "sucursal_responsable" in row.keys() else row["sucursal"],
                row["origen_ingreso"] if "origen_ingreso" in row.keys() else "",
                row["tipo_ingreso"] if "tipo_ingreso" in row.keys() else "",
                row["ubicacion_actual"] if "ubicacion_actual" in row.keys() else row["lugar_llegada"],
                row["deposito"],
                normalize_status(row["status"]),
                row["review_status"] if "review_status" in row.keys() else "",
                row["remito_interno"] if "remito_interno" in row.keys() else "",
                row["shipment_code"] if "shipment_code" in row.keys() else "",
                row["provider_name"] if "provider_name" in row.keys() else "",
                _fmt_sheet_dt(row["sent_to_provider_at"] if "sent_to_provider_at" in row.keys() else ""),
                _fmt_sheet_dt(row["fecha_ultimo_mail_proveedor"] if "fecha_ultimo_mail_proveedor" in row.keys() else ""),
                "" if days is None else days,
                row["estado_retiro_proveedor"] if "estado_retiro_proveedor" in row.keys() else "",
                _fmt_sheet_dt(row["fecha_retiro"] if "fecha_retiro" in row.keys() else ""),
                row["provider_response"] if "provider_response" in row.keys() else "",
                row["resultado_resolucion"] if "resultado_resolucion" in row.keys() else "",
                _fmt_sheet_dt(row["fecha_resolucion"] if "fecha_resolucion" in row.keys() else ""),
                _fmt_sheet_dt(row["fecha_finalizacion"] if "fecha_finalizacion" in row.keys() else ""),
                row["cliente_nombre"] if "cliente_nombre" in row.keys() else "",
                row["cliente_telefono"] if "cliente_telefono" in row.keys() else "",
                row["cliente_email"] if "cliente_email" in row.keys() else "",
                row["numero_factura"] if "numero_factura" in row.keys() else "",
                row["fecha_compra"] if "fecha_compra" in row.keys() else "",
                row["responsible_name"] or row["responsible_username"],
                row["created_by_name"] or row["created_by"],
                _fmt_sheet_dt(row["updated_at"]),
                row["updated_by_name"] or row["updated_by"],
                row["observations"] if "observations" in row.keys() else "",
            ])
        return values
    if sheet == "GARANTIA_ITEMS":
        rows = conn.execute(
            """
            SELECT g.warranty_code, g.updated_at, i.*
            FROM guarantee_items i
            JOIN guarantees g ON g.id = i.guarantee_id
            ORDER BY g.ingreso_at ASC, g.warranty_code ASC, i.id ASC
            """
        ).fetchall()
        return [[
            row["warranty_code"], row["item_index"] if "item_index" in row.keys() else row["id"],
            row["producto"], row["sku"], row["marca"], row["tipo"], row["serie"], row["falla"],
            row["proveedor"] if "proveedor" in row.keys() else "", row["observaciones"] if "observaciones" in row.keys() else "",
            _fmt_sheet_dt(row["updated_at"]),
        ] for row in rows]
    if sheet in {"REMITOS", "REMITO_ITEMS"}:
        if not _table_exists(conn, "warranty_remitos"):
            return []
        remitos = conn.execute("SELECT * FROM warranty_remitos ORDER BY created_at ASC, remito_code ASC").fetchall()
        if sheet == "REMITOS":
            return [[
                r["remito_code"], r["tipo_remito"] if "tipo_remito" in r.keys() else "sucursal_a_deposito",
                r["company_brand"], r["origen_sucursal"], r["destino_deposito"], r["status"],
                _fmt_sheet_dt(r["created_at"]), _fmt_sheet_dt(r["fecha_despacho"]), _fmt_sheet_dt(r["fecha_llegada"]),
                r["created_by_name"] or r["created_by"], r["despachado_por_name"] or r["despachado_por"],
                r["recibido_por_name"] or r["recibido_por"], len(_json_list(r["warranty_ids_json"])),
                f"/api/warranties/remitos/pdf/{r['remito_code']}", r["nota"],
            ] for r in remitos]
        values: list[list[Any]] = []
        for r in remitos:
            for code in _json_list(r["warranty_ids_json"]):
                g = conn.execute("SELECT * FROM guarantees WHERE warranty_code = ?", (str(code),)).fetchone()
                item_count, producto, sku, serie, falla = _warranty_item_summary(conn, int(g["id"])) if g else (0, "", "", "", "")
                values.append([
                    r["remito_code"], code, producto, sku, serie,
                    _row_value(g, "sucursal_responsable", _row_value(g, "sucursal", "")),
                    r["origen_sucursal"], r["destino_deposito"], r["status"], _fmt_sheet_dt(_row_value(g, "updated_at", r["created_at"])),
                ])
        return values
    if sheet in {"LOTES_ENV", "LOTE_ITEMS"}:
        if not _table_exists(conn, "guarantee_exports"):
            return []
        exports = conn.execute("SELECT * FROM guarantee_exports ORDER BY created_at ASC, shipment_code ASC").fetchall()
        if sheet == "LOTES_ENV":
            return [[
                r["shipment_code"], r["provider_name"], r["marca"],
                "mail_enviado" if any(str(_row_value(conn.execute("SELECT status FROM guarantees WHERE shipment_code = ? LIMIT 1", (r["shipment_code"],)).fetchone(), "status", "")).startswith("4 -") for _ in [0]) else "excel_generado",
                _fmt_sheet_dt(r["created_at"]), r["created_by_name"] or r["created_by"], r["file_name"],
                r["row_count"], "", "", r["filters_json"],
            ] for r in exports]
        values: list[list[Any]] = []
        for r in exports:
            codes = _json_list(r["warranty_ids_json"])
            for code in codes:
                g = conn.execute("SELECT * FROM guarantees WHERE warranty_code = ?", (str(code),)).fetchone()
                item_count, producto, sku, serie, falla = _warranty_item_summary(conn, int(g["id"])) if g else (0, "", "", "", "")
                values.append([
                    r["shipment_code"], code, producto, sku, serie,
                    _row_value(g, "provider_name", r["provider_name"]), normalize_status(_row_value(g, "status", "")),
                    _row_value(g, "provider_response", ""), _row_value(g, "resultado_resolucion", ""), _fmt_sheet_dt(_row_value(g, "updated_at", r["created_at"])),
                ])
        return values
    if sheet == "EVENTOS":
        if not _table_exists(conn, "guarantee_history"):
            return []
        rows = conn.execute("SELECT * FROM guarantee_history ORDER BY created_at ASC, id ASC").fetchall()
        values = []
        for r in rows:
            details = r["details_json"] if "details_json" in r.keys() else ""
            values.append([
                _fmt_sheet_dt(r["created_at"]), r["warranty_code"], r["actor_username"], r["actor_name"], r["action"],
                r["previous_status"] if "previous_status" in r.keys() else "",
                r["new_status"] if "new_status" in r.keys() else "",
                r["previous_review_status"] if "previous_review_status" in r.keys() else "",
                r["new_review_status"] if "new_review_status" in r.keys() else "",
                r["note"] if "note" in r.keys() else "",
                details,
            ])
        return values
    return []


def write_sheet_values(service: Any, spreadsheet_id: str, sheet_name: str, headers: list[str], values: list[list[Any]]) -> None:
    ensure_sheet_with_headers(service, spreadsheet_id, sheet_name, headers)
    last_col = get_column_letter(len(headers))
    service.spreadsheets().values().clear(
        spreadsheetId=spreadsheet_id,
        range=f"{quote_sheet_name(sheet_name)}!A2:{last_col}50000",
        body={},
    ).execute()
    if values:
        service.spreadsheets().values().update(
            spreadsheetId=spreadsheet_id,
            range=f"{quote_sheet_name(sheet_name)}!A2:{last_col}{len(values) + 1}",
            valueInputOption="USER_ENTERED",
            body={"values": values},
        ).execute()


def make_header_index(headers: list[str]) -> dict[str, int]:
    index: dict[str, int] = {}
    for i, header in enumerate(headers):
        index[header_key(header)] = i
    return index


def get_sheet_cell(row: list[Any], index: dict[str, int], header: str) -> str:
    idx = index.get(header_key(header))
    if idx is None or idx >= len(row):
        return ""
    return str(row[idx]).strip()


def warranty_sheet_row(guarantee_row: sqlite3.Row | dict[str, Any], item_row: sqlite3.Row | dict[str, Any]) -> list[Any]:
    def g(key: str, default: Any = "") -> Any:
        try:
            if hasattr(guarantee_row, "keys") and key in guarantee_row.keys():
                return guarantee_row[key]
            if isinstance(guarantee_row, dict):
                return guarantee_row.get(key, default)
        except Exception:
            pass
        return default

    def i(key: str, default: Any = "") -> Any:
        try:
            if hasattr(item_row, "keys") and key in item_row.keys():
                return item_row[key]
            if isinstance(item_row, dict):
                return item_row.get(key, default)
        except Exception:
            pass
        return default

    pending_days = compute_pending_days(guarantee_row)
    no_response_days = compute_no_response_days(guarantee_row)
    fecha_inicio_gestion = g("reviewed_at") or g("sent_to_provider_at") or ""
    observations = str(i("observaciones") or g("observations") or "")
    return [
        g("warranty_code"),
        g("responsible_name"),
        format_date_ar(parse_iso_datetime(g("ingreso_at")) or now_ar()),
        i("producto"),
        i("sku"),
        i("marca"),
        i("serie"),
        i("falla"),
        g("sucursal"),
        g("deposito"),
        g("status"),
        pending_days,
        format_datetime_ar(parse_iso_datetime(fecha_inicio_gestion)) if fecha_inicio_gestion else "",
        g("provider_case_id"),
        "" if no_response_days is None else no_response_days,
        format_datetime_ar(parse_iso_datetime(g("fecha_retiro"))) if g("fecha_retiro") else "",
        format_datetime_ar(parse_iso_datetime(g("fecha_resolucion"))) if g("fecha_resolucion") else "",
        observations,
        g("vuelve_a"),
        g("finalizacion"),
        i("tipo"),
        g("lugar_llegada"),
        g("created_by"),
        format_datetime_ar(parse_iso_datetime(g("updated_at"))) if g("updated_at") else "",
        g("updated_by_name") or g("updated_by"),
    ]


def warranty_rows_for_sheet(conn: sqlite3.Connection) -> tuple[list[list[Any]], dict[int, tuple[int, int]]]:
    ensure_warranty_tables(conn)
    rows = conn.execute(
        """
        SELECT * FROM guarantees
        WHERE COALESCE(cancelled, 0) = 0
        ORDER BY ingreso_at ASC, warranty_code ASC, id ASC
        """
    ).fetchall()
    values: list[list[Any]] = []
    row_ranges: dict[int, tuple[int, int]] = {}
    next_row = 2
    for guarantee in rows:
        items = conn.execute("SELECT * FROM guarantee_items WHERE guarantee_id = ? ORDER BY id", (int(guarantee["id"]),)).fetchall()
        if not items:
            continue
        start = next_row
        for item in items:
            values.append(warranty_sheet_row(guarantee, item))
            next_row += 1
        row_ranges[int(guarantee["id"])] = (start, next_row - 1)
    return values, row_ranges


def insert_sync_log(
    conn: sqlite3.Connection,
    *,
    sync_type: str,
    status_value: str,
    started_at: str,
    finished_at: str,
    user: Any,
    rows_processed: int,
    rows_created: int,
    rows_updated: int,
    rows_skipped: int,
    errors: list[str],
) -> int:
    cur = conn.execute(
        """
        INSERT INTO guarantee_sync_logs (
            sync_type, status, started_at, finished_at, actor_username, actor_name,
            rows_processed, rows_created, rows_updated, rows_skipped, errors_json
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            sync_type,
            status_value,
            started_at,
            finished_at,
            getattr(user, "username", "") or "",
            getattr(user, "display_name", "") or "",
            rows_processed,
            rows_created,
            rows_updated,
            rows_skipped,
            json.dumps(errors, ensure_ascii=False),
        ),
    )
    return int(cur.lastrowid)


def sync_log_info(row: sqlite3.Row) -> WarrantySyncLogInfo:
    errors: list[str] = []
    try:
        parsed = json.loads(row["errors_json"] or "[]")
        if isinstance(parsed, list):
            errors = [str(x) for x in parsed]
    except Exception:
        errors = []
    return WarrantySyncLogInfo(
        id=int(row["id"]),
        sync_type=str(row["sync_type"] or ""),
        status=str(row["status"] or ""),
        started_at=format_datetime_ar(parse_iso_datetime(row["started_at"]) or now_ar()),
        finished_at=format_datetime_ar(parse_iso_datetime(row["finished_at"])) if row["finished_at"] else "",
        actor_username=str(row["actor_username"] or ""),
        actor_name=str(row["actor_name"] or ""),
        rows_processed=int(row["rows_processed"] or 0),
        rows_created=int(row["rows_created"] or 0),
        rows_updated=int(row["rows_updated"] or 0),
        rows_skipped=int(row["rows_skipped"] or 0),
        errors=errors,
    )


def import_sheet_group(conn: sqlite3.Connection, warranty_code: str, raw_rows: list[dict[str, str]], user: Any) -> tuple[str, str]:
    existing = conn.execute("SELECT * FROM guarantees WHERE warranty_code = ?", (warranty_code,)).fetchone()
    if existing:
        # MVP seguro: la app es fuente principal. No se pisa automáticamente una garantía ya existente.
        incoming_status = raw_rows[0].get("ESTADO", "")
        incoming_case = raw_rows[0].get("ID DE CASO", "")
        current_status = str(existing["status"] or "")
        current_case = str(existing["provider_case_id"] or "")
        if normalize_text(incoming_status) != normalize_text(current_status) or str(incoming_case).strip() != current_case.strip():
            return "conflict", f"{warranty_code}: existe en app y Sheet con diferencias. No se pisó automáticamente."
        return "skipped", f"{warranty_code}: ya existe en la app."

    first = raw_rows[0]
    now = utc_now_iso()
    ingreso = sheet_date_to_iso(first.get("INGRESO"))
    status_value = first.get("ESTADO") or DEFAULT_STATUSES[0]
    suc = first.get("SUCURSAL") or ""
    cur = conn.execute(
        """
        INSERT INTO guarantees (
            warranty_code, status, review_status, responsible_username, responsible_name,
            created_by, created_by_name, created_at, ingreso_at, updated_at, updated_by, updated_by_name,
            sucursal, sucursal_code, deposito, lugar_llegada, provider_name, provider_case_id,
            observations, vuelve_a, finalizacion, synced_to_google_sheet, last_google_sync_at, google_sheet_updated_at
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 1, ?, ?)
        """,
        (
            warranty_code,
            status_value,
            REVIEW_PENDING if normalize_text(status_value) == normalize_text(DEFAULT_STATUSES[0]) else REVIEW_APPROVED,
            first.get("USUARIO") or first.get("RESPONSABLE") or "",
            first.get("RESPONSABLE") or "",
            first.get("USUARIO") or getattr(user, "username", "") or "",
            first.get("RESPONSABLE") or getattr(user, "display_name", "") or "",
            now,
            ingreso,
            sheet_date_to_iso(first.get("FECHA ULTIMA ACTUALIZACION")) if first.get("FECHA ULTIMA ACTUALIZACION") else now,
            first.get("ACTUALIZADO POR") or getattr(user, "username", "") or "",
            first.get("ACTUALIZADO POR") or getattr(user, "display_name", "") or "",
            suc,
            sucursal_code(suc),
            first.get("DEPOSITO") or "",
            first.get("LUGAR LLEGADA") or first.get("DEPOSITO") or "",
            "",
            first.get("ID DE CASO") or "",
            first.get("OBSERVACIONES") or "",
            first.get("VUELVE A") or "",
            first.get("FINALIZACION") or "",
            now,
            first.get("FECHA ULTIMA ACTUALIZACION") or "",
        ),
    )
    gid = int(cur.lastrowid)
    for item in raw_rows:
        conn.execute(
            """
            INSERT INTO guarantee_items (guarantee_id, producto, sku, marca, tipo, serie, falla, observaciones, created_at, updated_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                gid,
                item.get("PRODUCTO") or "",
                item.get("SKU") or "",
                item.get("MARCA") or "",
                item.get("TIPO") or "",
                item.get("SERIE") or "",
                item.get("FALLA") or "",
                item.get("OBSERVACIONES") or "",
                now,
                now,
            ),
        )
    add_history(conn, gid, warranty_code, user, "sheet_sync_pull", new_status=status_value, note="Garantía importada desde Google Sheet", details={"items": len(raw_rows)})
    return "created", f"{warranty_code}: importada."

# =========================================================
# Endpoints
# =========================================================

@router.get("/options")
def warranty_options(user: Annotated[Any, Depends(require_current_user)]):
    # Depósito operativo necesita opciones para cargar Cliente en depósito,
    # aunque no tenga permiso de listado/gestión global.
    ensure_warranty_intake_access(user)
    return runtime_warranty_options()


@router.get("/products")
def warranty_products(
    user: Annotated[Any, Depends(require_current_user)],
    q: str = Query(default="", min_length=0),
    limit: int = Query(default=20, ge=1, le=50),
):
    ensure_warranty_intake_access(user)
    query = normalize_text(q)
    if len(query) < 2:
        return []

    # Fase 7: Garantías busca primero en el catálogo local sincronizado desde Planilla Madre.
    local = search_local_products(q, limit=limit)
    if local:
        out = []
        for item in local:
            provider = get_provider_for_brand(str(item.get("marca") or ""))
            out.append({
                "producto": item.get("producto") or item.get("descripcion") or item.get("sku") or "",
                "sku": item.get("sku") or "",
                "marca": item.get("marca") or "",
                "tipo": item.get("tipo") or "",
                "pvp_texto": item.get("pvp_text") or item.get("precio_texto") or "",
                "costo_texto": item.get("costo_text") or "",
                "provider_name": provider.get("name") if provider else "",
                "label": item.get("label") or "",
            })
        return out

    # Fallback de compatibilidad: lectura anterior desde Google Sheets si todavía no hay catálogo local.
    tokens = query.split()
    matches = []
    for item in load_product_catalog():
        haystack = item.get("search", "")
        if all(token in haystack for token in tokens):
            score = 0
            if haystack.startswith(query):
                score += 10
            if item.get("sku") and normalize_text(item["sku"]).startswith(query):
                score += 20
            matches.append((score, item))
    matches.sort(key=lambda pair: pair[0], reverse=True)
    return [{k: v for k, v in item.items() if k != "search"} for _, item in matches[:limit]]


@router.post("/entries", response_model=WarrantyCreateResponse)
def create_warranty_entries(data: WarrantyCreateRequest, user: Annotated[Any, Depends(require_current_user)]):
    settings = get_settings()
    if not settings.app_enabled:
        raise HTTPException(status_code=403, detail="La aplicación está deshabilitada por el administrador.")
    if getattr(user, "must_change_password", False):
        raise HTTPException(status_code=403, detail="Tenés que crear tu contraseña antes de continuar")

    # El rol DEPOSITO operativo debe poder cargar garantías de cliente en depósito
    # aunque el catálogo de roles local todavía no se haya resincronizado.
    # Para vendedores/gestores/admin seguimos usando warranties.create como permiso normal.
    #
    # Hotfix Fase 19:
    # En instalaciones existentes puede venir como branch_type="deposit", "deposito",
    # rol="DEPOSITO", o directamente por la branch asignada con nombre "Depósito ...".
    # Normalizamos sin acentos para no bloquear la carga operativa por un detalle de catálogo.
    def _assigned_deposit_branch(u: Any) -> dict[str, Any] | None:
        """Devuelve la branch depósito asignada aunque el token legacy no traiga branch_name.

        En instalaciones existentes el usuario DEPOSITO puede venir con role=DEPOSITO
        y la unidad real dentro de user.branches, pero branch_name/branch_type vacíos.
        La carga operativa no debe fallar por ese desfasaje.
        """
        branches = getattr(u, "branches", []) or []
        ordered = sorted(
            [b for b in branches if isinstance(b, dict)],
            key=lambda b: 0 if b.get("is_primary") else 1,
        )
        for branch in ordered:
            b_type = normalize_text(branch.get("type", ""))
            b_name = normalize_text(branch.get("name", ""))
            if b_type in {"DEPOSIT", "DEPOSITO"} or b_name.startswith("DEPOSITO ") or b_name == "DEPOSITO":
                return branch
        return None

    def _is_deposit_operator_user(u: Any) -> bool:
        branch_type_key = normalize_text(getattr(u, "branch_type", "") or "")
        branch_name_key = normalize_text(getattr(u, "branch_name", "") or getattr(u, "sucursal", "") or "")
        role_key = normalize_text(getattr(u, "role", "") or "")
        roles_keys = {normalize_text(r) for r in (getattr(u, "roles", []) or [])}
        return (
            branch_type_key in {"DEPOSIT", "DEPOSITO"}
            or role_key == "DEPOSITO"
            or "DEPOSITO" in roles_keys
            or branch_name_key.startswith("DEPOSITO ")
            or branch_name_key == "DEPOSITO"
            or _assigned_deposit_branch(u) is not None
        )

    is_deposit_operator_for_permission = _is_deposit_operator_user(user)
    if not user.has("warranties.create") and not is_deposit_operator_for_permission:
        raise HTTPException(status_code=403, detail="No tenés permiso para realizar esta acción")

    # ── Perfil del usuario ─────────────────────────────────────────────────────
    # Fuente de verdad organizativa: usuario -> empresa -> branch asignada.
    assigned_deposit_branch = _assigned_deposit_branch(user)
    branch_type = str(getattr(user, "branch_type", "") or (assigned_deposit_branch or {}).get("type", "") or "").lower().strip()
    branch_type_key = normalize_text(branch_type)
    user_branch_id = str(getattr(user, "branch_id", "") or (assigned_deposit_branch or {}).get("id", "") or "").strip()
    user_branch_name = str(getattr(user, "branch_name", "") or getattr(user, "sucursal", "") or (assigned_deposit_branch or {}).get("name", "") or "").strip()
    user_branch_name_key = normalize_text(user_branch_name)
    user_company_id = str(getattr(user, "company_id", "") or (assigned_deposit_branch or {}).get("company_id", "") or "").strip()
    can_manage = user.has("warranties.manage") or user.has("warranties.manage_provider")

    if branch_type_key == "WEB":
        raise HTTPException(
            status_code=403,
            detail=(
                "Los usuarios de sucursal web no pueden registrar garantías directamente. "
                "Las garantías deben ingresarse desde la sucursal física o el depósito correspondiente."
            ),
        )

    is_sucursal_fisica = branch_type_key in {"PHYSICAL", "FISICA", "SUCURSAL", "SUCURSAL FISICA"} and not can_manage
    # Importante: tener una branch/rol de depósito NO debe encerrar a gestores/admin.
    # Para usuarios con permisos de gestión, la unidad asignada es un default visual;
    # la carga puede ser desde sucursal o depósito según el tipo elegido.
    is_deposito = is_deposit_operator_for_permission and not can_manage
    if not branch_type_key and not can_manage and user_branch_name:
        # Compatibilidad con usuarios legacy: sin branch_type pero con sucursal asignada.
        # Si el nombre asignado es Depósito..., tratarlo como depósito; si no, como sucursal.
        if user_branch_name_key.startswith("DEPOSITO ") or user_branch_name_key == "DEPOSITO":
            is_deposito = True
        else:
            is_sucursal_fisica = True

    options = runtime_warranty_options()
    branches_operativas: list[dict[str, str]] = options.get("branches_operativas") or []
    branches_map: dict[str, dict[str, str]] = {b["id"]: b for b in branches_operativas if b.get("id")}
    physical_branches = [b for b in branches_operativas if b.get("type") == "physical"]
    deposit_branches = [b for b in branches_operativas if b.get("type") == "deposit"]
    central_deposit_name = _warranty_central_deposit_name(branches_operativas)

    allowed_sucursales = {normalize_text(x) for x in ([b["name"] for b in physical_branches] or options.get("sucursales", []))}
    allowed_depositos = {normalize_text(x) for x in ([b["name"] for b in deposit_branches] or options.get("depositos", []))}

    if (is_sucursal_fisica or is_deposito) and not user_branch_name:
        raise HTTPException(
            status_code=403,
            detail="Tu usuario no tiene sucursal/depósito asignado. Pedile a un administrador que revise tu configuración.",
        )

    user_branch_info = branches_map.get(user_branch_id) if user_branch_id else None
    if user_branch_id and not user_branch_info:
        user_branch_info = _fetch_branch_info(user_branch_id)
        if user_branch_info:
            branches_map[user_branch_id] = user_branch_info
    if is_deposito and user_branch_info and normalize_text(user_branch_info.get("type", "")) not in {"DEPOSIT", "DEPOSITO"}:
        # Permitir compatibilidad si el nombre de la unidad es claramente un depósito.
        if not normalize_text(user_branch_info.get("name", "")).startswith("DEPOSITO "):
            raise HTTPException(status_code=403, detail="Tu usuario no está asignado a un depósito válido.")
    if is_sucursal_fisica and user_branch_info and normalize_text(user_branch_info.get("type", "")) not in {"PHYSICAL", "FISICA", "SUCURSAL", "SUCURSAL FISICA"}:
        raise HTTPException(status_code=403, detail="Tu usuario no está asignado a una sucursal física válida.")

    def _resolve_branch_by_id(branch_id: str, *, expected_type: str | None = None, label: str = "branch") -> dict[str, str]:
        clean_id = (branch_id or "").strip()
        if not clean_id:
            raise HTTPException(status_code=400, detail=f"Falta indicar {label}.")
        branch = branches_map.get(clean_id) or _fetch_branch_info(clean_id)
        if not branch:
            raise HTTPException(status_code=400, detail=f"{label} inexistente en el sistema (ID: {clean_id}).")
        branches_map[clean_id] = branch
        if expected_type and branch.get("type") != expected_type:
            raise HTTPException(status_code=400, detail=f"{label} debe ser de tipo {expected_type}.")
        return branch

    def _resolve_physical_by_name(name: str) -> dict[str, str] | None:
        return _branch_by_name(branches_operativas, name, "physical")

    def _resolve_deposit_by_name(name: str) -> dict[str, str] | None:
        return _branch_by_name(branches_operativas, name, "deposit")

    for item in data.items:
        tipo = (item.tipo_ingreso or "").strip()

        # Todo ingreso hecho desde sucursal tiene como destino operativo de garantías
        # el depósito Chiclana. Corrales/Cachi quedan para movimientos internos
        # posteriores hechos por usuarios de depósito.
        if tipo == "cliente_sucursal":
            object.__setattr__(item, "deposito", central_deposit_name)
            object.__setattr__(item, "lugar_llegada", central_deposit_name)

        # ── Seguridad por perfil ───────────────────────────────────────────────
        if is_sucursal_fisica:
            if tipo not in TIPOS_INGRESO_VENDEDOR:
                raise HTTPException(
                    status_code=403,
                    detail="Los usuarios de sucursal solo pueden cargar garantías como cliente_sucursal.",
                )
            # El frontend puede mostrar un dato, pero el backend fuerza lo real.
            object.__setattr__(item, "sucursal", user_branch_name)
            object.__setattr__(item, "deposito", (item.deposito or "").strip())
        elif is_deposito:
            # Personal operativo de depósito: carga acotada al cliente que llega al depósito.
            # Gestores/Admin con branch de depósito NO entran acá; caen en el bloque amplio de abajo.
            if tipo == "cliente_sucursal":
                raise HTTPException(status_code=403, detail="Los usuarios de depósito no pueden cargar como cliente_sucursal.")
            if tipo != "cliente_deposito":
                raise HTTPException(
                    status_code=403,
                    detail="Los usuarios de depósito solo pueden cargar garantías como cliente en depósito. Otras opciones quedan para gestión/admin.",
                )
            # El depósito de carga también es fuente de verdad backend.
            object.__setattr__(item, "deposito", user_branch_name)
        else:
            # Gestor/admin/encargado puede operar más amplio, pero siempre validado.
            # Si carga desde sucursal, debe poder elegir la sucursal responsable/carga.
            # Si carga desde depósito y no eligió uno, sugerimos la unidad asignada si es depósito.
            if tipo != "cliente_sucursal" and not (item.deposito or "").strip() and user_branch_name:
                object.__setattr__(item, "deposito", user_branch_name)

        if tipo == "cliente_sucursal":
            if allowed_sucursales and normalize_text(item.sucursal) not in allowed_sucursales:
                raise HTTPException(status_code=400, detail=f"Sucursal inválida: {item.sucursal}")
        if allowed_depositos and (item.deposito or "").strip() and normalize_text(item.deposito) not in allowed_depositos:
            raise HTTPException(status_code=400, detail=f"Depósito inválido: {item.deposito}")

        # cliente_deposito requiere sucursal responsable real o fallback textual.
        suc_resp_id = (item.sucursal_responsable_id or "").strip()
        if suc_resp_id:
            branch = _resolve_branch_by_id(suc_resp_id, expected_type="physical", label="Sucursal responsable")
            object.__setattr__(item, "sucursal_responsable", branch.get("name", ""))
        elif (item.sucursal_responsable or "").strip():
            branch = _resolve_physical_by_name(item.sucursal_responsable)
            if branch:
                object.__setattr__(item, "sucursal_responsable_id", branch.get("id", ""))
                object.__setattr__(item, "sucursal_responsable", branch.get("name", ""))

        if tipo == "cliente_deposito":
            if not (item.sucursal_responsable_id or item.sucursal_responsable or "").strip():
                raise HTTPException(
                    status_code=400,
                    detail="Cuando el cliente deja el producto en depósito, indicá la sucursal donde compró.",
                )

        if tipo in {"cliente_sucursal", "cliente_deposito"}:
            if not (item.cliente_nombre or "").strip():
                raise HTTPException(status_code=400, detail="Indicá el nombre del cliente.")
            if not (item.cliente_telefono or "").strip():
                raise HTTPException(status_code=400, detail="Indicá el teléfono del cliente.")
            if not (item.numero_factura or "").strip():
                raise HTTPException(status_code=400, detail="Indicá el número de factura o ticket.")
            if not (item.fecha_compra or "").strip():
                raise HTTPException(status_code=400, detail="Indicá la fecha de compra.")

    def _derive_org_fields(item: WarrantyItemIn) -> dict[str, str]:
        """Campos organizativos definitivos para guardar la garantía.

        Fuente de verdad:
        - branch_id guarda la unidad de carga: sucursal física o depósito asignado/usado.
        - sucursal_responsable_id guarda la sucursal comercial responsable.
        - company_id se deriva de la sucursal responsable cuando corresponde.
        """
        tipo = (item.tipo_ingreso or "").strip()
        s_resp_id = (item.sucursal_responsable_id or "").strip()
        s_resp_name = (item.sucursal_responsable or "").strip()
        s_resp_branch = branches_map.get(s_resp_id) if s_resp_id else None

        if is_sucursal_fisica:
            branch = user_branch_info or {"id": user_branch_id, "name": user_branch_name, "company_id": user_company_id}
            return {
                "sucursal_carga": branch.get("name") or user_branch_name,
                "sucursal_carga_bid": branch.get("id") or user_branch_id,
                "sucursal_responsable": branch.get("name") or user_branch_name,
                "sucursal_responsable_id": branch.get("id") or user_branch_id,
                "company_id": branch.get("company_id") or user_company_id,
            }

        if is_deposito:
            deposit_branch = user_branch_info or {"id": user_branch_id, "name": user_branch_name, "company_id": user_company_id}
            if tipo == "cliente_deposito":
                # La garantía se imputa a la empresa de la sucursal donde compró el cliente.
                company_id = (s_resp_branch or {}).get("company_id") or user_company_id
                return {
                    "sucursal_carga": deposit_branch.get("name") or user_branch_name,
                    "sucursal_carga_bid": deposit_branch.get("id") or user_branch_id,
                    "sucursal_responsable": s_resp_name,
                    "sucursal_responsable_id": s_resp_id,
                    "company_id": company_id,
                }
            # Falla recepción / stock interno: nace y se imputa operativamente al depósito.
            return {
                "sucursal_carga": deposit_branch.get("name") or user_branch_name,
                "sucursal_carga_bid": deposit_branch.get("id") or user_branch_id,
                "sucursal_responsable": s_resp_name or (deposit_branch.get("name") or user_branch_name),
                "sucursal_responsable_id": s_resp_id,
                "company_id": (deposit_branch.get("company_id") or user_company_id),
            }

        # Gestor/admin con libertad, pero derivando IDs/empresa cuando hay branches reales.
        if tipo == "cliente_sucursal":
            branch = _resolve_physical_by_name(item.sucursal) or {}
            company_id = branch.get("company_id") or (s_resp_branch or {}).get("company_id") or user_company_id
            return {
                "sucursal_carga": branch.get("name") or (item.sucursal or "").strip(),
                "sucursal_carga_bid": branch.get("id", ""),
                "sucursal_responsable": s_resp_name or branch.get("name") or (item.sucursal or "").strip(),
                "sucursal_responsable_id": s_resp_id or branch.get("id", ""),
                "company_id": company_id,
            }

        deposit_branch = _resolve_deposit_by_name(item.deposito) or user_branch_info or {}
        company_id = (s_resp_branch or {}).get("company_id") or deposit_branch.get("company_id") or user_company_id
        return {
            "sucursal_carga": deposit_branch.get("name") or user_branch_name or (item.deposito or "").strip(),
            "sucursal_carga_bid": deposit_branch.get("id") or user_branch_id,
            "sucursal_responsable": s_resp_name or (deposit_branch.get("name") or user_branch_name or ""),
            "sucursal_responsable_id": s_resp_id,
            "company_id": company_id,
        }

    derived = [_derive_org_fields(item) for item in data.items]

    if data.group_under_one_id:
        first_d = derived[0]
        first_item = data.items[0]
        first_cs = normalize_text(_code_source_for_tipo(first_d["sucursal_carga"], first_item.deposito, first_item.tipo_ingreso))
        for d, it in zip(derived[1:], data.items[1:]):
            cs = normalize_text(_code_source_for_tipo(d["sucursal_carga"], it.deposito, it.tipo_ingreso))
            if cs != first_cs:
                raise HTTPException(
                    status_code=400,
                    detail="Para usar un mismo ID, todas las filas deben tener la misma sucursal/depósito de origen.",
                )

    created: list[WarrantyCreatedItem] = []
    ids: list[str] = []
    with db_connect() as conn:
        ensure_warranty_tables(conn)
        if data.group_under_one_id:
            # Fase 34: “todo pertenece a una sola garantía” significa mismo caso madre,
            # pero cada producto nace como garantía operativa independiente.
            # Ej: GAR-2026-CAS-0005-01, GAR-2026-CAS-0005-02.
            # Así revisión, remitos, ENV/proveedor y resolución pueden trabajar ítem por ítem.
            first_item = data.items[0]
            d0 = derived[0]
            code_source = _code_source_for_tipo(d0["sucursal_carga"], first_item.deposito, first_item.tipo_ingreso)
            parent_code = next_warranty_code(conn, code_source)
            total_items = len(data.items)
            for idx, (item, d) in enumerate(zip(data.items, derived), start=1):
                warranty_code = f"{parent_code}-{idx:02d}"
                guarantee_id = insert_guarantee(
                    conn, warranty_code, item, user,
                    sucursal_carga=d["sucursal_carga"],
                    sucursal_carga_branch_id=d["sucursal_carga_bid"],
                    sucursal_responsable_override=d["sucursal_responsable"],
                    sucursal_responsable_id_override=d["sucursal_responsable_id"],
                    company_id_override=d["company_id"],
                    parent_warranty_code=parent_code,
                    parent_item_index=idx,
                )
                insert_item(conn, guarantee_id, item, item_index=idx)
                add_history(
                    conn, guarantee_id, warranty_code, user, "created",
                    new_status=DEFAULT_STATUSES[0],
                    note="Ítem de garantía agrupada creado",
                    details={
                        "items": 1,
                        "grouped": True,
                        "parent_warranty_code": parent_code,
                        "parent_item_index": idx,
                        "parent_items_total": total_items,
                        "fecha_ingreso": date_input_from_iso(ingreso_at_from_input(item.fecha_ingreso)),
                        "tipo_ingreso": item.tipo_ingreso,
                    },
                )
                created.append(WarrantyCreatedItem(id_garantia=warranty_code, producto=item.producto, sku=item.sku, parent_warranty_code=parent_code, parent_item_index=idx))
                ids.append(warranty_code)
        else:
            for item, d in zip(data.items, derived):
                code_source = _code_source_for_tipo(d["sucursal_carga"], item.deposito, item.tipo_ingreso)
                warranty_code = next_warranty_code(conn, code_source)
                guarantee_id = insert_guarantee(
                    conn, warranty_code, item, user,
                    sucursal_carga=d["sucursal_carga"],
                    sucursal_carga_branch_id=d["sucursal_carga_bid"],
                    sucursal_responsable_override=d["sucursal_responsable"],
                    sucursal_responsable_id_override=d["sucursal_responsable_id"],
                    company_id_override=d["company_id"],
                )
                insert_item(conn, guarantee_id, item, item_index=1)
                add_history(conn, guarantee_id, warranty_code, user, "created", new_status=DEFAULT_STATUSES[0], note="Garantía creada", details={"items": 1, "grouped": False, "fecha_ingreso": date_input_from_iso(ingreso_at_from_input(item.fecha_ingreso)), "tipo_ingreso": item.tipo_ingreso})
                created.append(WarrantyCreatedItem(id_garantia=warranty_code, producto=item.producto, sku=item.sku))
                ids.append(warranty_code)
        conn.commit()
    audit("warranties.create", user=user, resource_type="warranty", resource_id=",".join(unique_keep_order(ids)), details={"count": len(created), "ids": unique_keep_order(ids), "source": "database"})
    return WarrantyCreateResponse(ok=True, count=len(created), ids=ids, items=created)

@router.get("/list", response_model=WarrantyListResponse)
def list_warranties(
    user: Annotated[Any, Depends(require_permission("warranties.view"))],
    q: str = "",
    sucursal: str = "",
    estado: str = "",
    review_status: str = "",
    deposito: str = "",
    marca: str = "",
    proveedor: str = "",
    tipo_ingreso: str = "",
    origen_ingreso: str = "",
    transit_status: str = "",
    ubicacion_actual: str = "",
    estado_retiro_proveedor: str = "",
    demora_min: int = Query(default=0, ge=0, le=3650),
    fecha_desde: str = "",
    fecha_hasta: str = "",
    limit: int = Query(default=200, ge=1, le=1000),
    sucursal_logistics: bool = Query(default=False),
):
    # ── Auto-filtro operativo para sucursal ────────────────────────────────────
    # Para usuarios sin permisos de gestión, el listado NO es un historial global:
    # es la bandeja de garantías que todavía están físicamente en su sucursal.
    # Cuando la garantía sale por remito hacia Chiclana, deja de aparecer acá.
    # EXCEPCIÓN: sucursal_logistics=True (usado por WarrantySucursalPage) incluye
    # garantías en tránsito y en depósito para que el encargado pueda ver su historial.
    user_perms = set(getattr(user, "permissions", []) or [])
    roles = _user_role_keys(user)
    # Fase 38: no depender solo de user.permissions para decidir alcance.
    # En algunas sesiones el objeto usuario pasa require_permission() mediante
    # user.has(), pero permissions viene incompleto; entonces Gestión quedaba
    # filtrada por la unidad principal del admin/gestor y no mostraba garantías
    # recién aprobadas de otras sucursales.
    can_manage = (
        "*" in user_perms
        or "warranties.manage" in user_perms
        or "warranties.manage_provider" in user_perms
        or bool(roles & WARRANTY_PRIVILEGED_ROLES)
        or bool(getattr(user, "has", lambda _p: False)("warranties.manage"))
        or bool(getattr(user, "has", lambda _p: False)("warranties.manage_provider"))
    )
    user_sucursal = str(getattr(user, "sucursal", "") or "").strip()
    user_branch_id = str(getattr(user, "branch_id", "") or "").strip()
    user_branch_type = str(getattr(user, "branch_type", "") or "").strip().lower()
    is_branch_operator = not can_manage and user_branch_type not in {"deposit", "admin"}
    # Fallback visual para usuarios legacy. La seguridad real se controla abajo
    # con branch_id / sucursal_responsable_id cuando existen.
    if is_branch_operator and user_sucursal and not sucursal.strip() and not user_branch_id:
        sucursal = user_sucursal

    with db_connect() as conn:
        ensure_warranty_tables(conn)
        rows = conn.execute("SELECT * FROM guarantees ORDER BY ingreso_at DESC, id DESC").fetchall()
        all_items = conn.execute("SELECT * FROM guarantee_items ORDER BY id").fetchall()
    by_gid: dict[int, list[sqlite3.Row]] = {}
    for item in all_items:
        by_gid.setdefault(int(item["guarantee_id"]), []).append(item)

    summaries = [row_to_summary(row, by_gid.get(int(row["id"]), [])) for row in rows]
    q_tokens = normalize_text(q).split()
    suc_key = normalize_text(sucursal)
    dep_key = normalize_text(deposito)
    est_key = normalize_text(estado)
    marca_key = normalize_text(marca)
    proveedor_key = normalize_text(proveedor)
    tipo_ingreso_key = normalize_text(tipo_ingreso)
    origen_ingreso_key = str(origen_ingreso or "").strip()
    review_key = str(review_status or "").strip()
    date_from = parse_date_filter(fecha_desde)
    date_to = parse_date_filter(fecha_hasta)

    def match(item: WarrantySummary) -> bool:
        if not can_manage:
            if user_branch_id:
                if item.branch_id != user_branch_id and item.sucursal_responsable_id != user_branch_id:
                    return False
            elif user_sucursal and normalize_text(item.sucursal) != normalize_text(user_sucursal):
                return False

            # Para sucursal/vendedor, Listado funciona como "garantías en mi local".
            # Si ya tienen remito activo, están en tránsito, llegaron a depósito,
            # fueron al proveedor o se finalizaron, deben verse en Gestión/Remitos,
            # no en este listado operativo de sucursal.
            # EXCEPCIÓN: sucursal_logistics=True (WarrantySucursalPage) necesita ver
            # todas las garantías de la sucursal sin importar su estado de tránsito.
            if is_branch_operator and not sucursal_logistics:
                ubicacion = str(item.ubicacion_actual or "").strip().lower()
                transit = str(item.transit_status or "").strip().lower()
                status_norm = normalize_status(item.estado)
                has_active_remito = bool(str(item.remito_interno or "").strip()) and transit != "cancelado"
                if status_norm in {"9 - ANULADA", "10 - FINALIZADO"}:
                    return False
                if ubicacion:
                    # La ubicación real de una garantía cargada en local debe ser
                    # el nombre de la sucursal (ej. Caseros). Se conserva
                    # 'sucursal' como fallback para datos legacy.
                    allowed_locations = {"sucursal"}
                    if item.sucursal:
                        allowed_locations.add(str(item.sucursal or "").strip().lower())
                    if user_sucursal:
                        allowed_locations.add(str(user_sucursal or "").strip().lower())
                    if ubicacion not in allowed_locations:
                        return False
                else:
                    # Compatibilidad con datos previos: si no hay ubicación explícita,
                    # sólo se considera en sucursal si no tiene remito ni tránsito.
                    if has_active_remito or transit in {"en_transito", "en_deposito", "llegado"}:
                        return False
                if has_active_remito or transit in {"en_transito", "en_deposito", "llegado"}:
                    return False
        if suc_key:
            # Fase 35: el filtro "Sucursal" debe servir tanto para la sucursal
            # de carga como para la sucursal responsable/comercial.
            # Caso clave: cliente deja el producto en depósito, pero compró en Caseros.
            # La garantía debe aparecer en revisión/gestión al filtrar Caseros aunque
            # la unidad de carga sea Depósito Chiclana.
            sucursal_values = {
                normalize_text(item.sucursal),
                normalize_text(item.sucursal_responsable),
            }
            if suc_key not in sucursal_values:
                return False
        if dep_key and normalize_text(item.deposito) != dep_key and normalize_text(item.lugar_llegada) != dep_key:
            return False
        if est_key and not status_matches(item.estado, estado):
            return False
        if tipo_ingreso_key and normalize_text(item.tipo_ingreso) != tipo_ingreso_key:
            return False
        if origen_ingreso_key and item.origen_ingreso != origen_ingreso_key:
            return False
        if marca_key and marca_key not in normalize_text(" ".join(item.productos)) and marca_key not in normalize_text(item.producto_principal):
            # Se busca marca real contra items porque el resumen principal prioriza producto.
            item_rows = by_gid.get(next((int(r["id"]) for r in rows if str(r["warranty_code"] or "") == item.id_garantia), -1), [])
            if not any(normalize_text(x["marca"]) == marca_key for x in item_rows):
                return False
        if proveedor_key and normalize_text(item.provider_name) != proveedor_key:
            return False
        if demora_min and (item.dias_sin_respuesta is None or item.dias_sin_respuesta < demora_min):
            return False
        if review_key and not review_status_matches(item.review_status, review_status):
            return False
        transit_key = str(transit_status or "").strip().lower()
        if transit_key and item.transit_status != transit_key:
            return False
        ubicacion_key = str(ubicacion_actual or "").strip().lower()
        if ubicacion_key and item.ubicacion_actual != ubicacion_key:
            return False
        retiro_key = str(estado_retiro_proveedor or "").strip().lower()
        if retiro_key and item.estado_retiro_proveedor != retiro_key:
            return False
        ingreso_date = parse_date_filter(item.ingreso)
        if date_from and ingreso_date and ingreso_date < date_from:
            return False
        if date_to and ingreso_date and ingreso_date > date_to:
            return False
        if q_tokens:
            haystack = normalize_text(" ".join([
                item.id_garantia, item.producto_principal, " ".join(item.productos), item.sku,
                item.serie, item.falla, item.responsable, item.sucursal, item.estado, item.deposito,
                item.provider_name, item.id_de_caso,
            ]))
            return all(token in haystack for token in q_tokens)
        return True

    filtered = [item for item in summaries if match(item)]
    total = len(filtered)
    return WarrantyListResponse(items=filtered[:limit], total=total, limit=limit)



# ── Reset producción / limpieza de datos de prueba ───────────────────────────

RESET_PRESERVED_ITEMS = [
    "usuarios, roles y permisos",
    "empleados",
    "empresas",
    "sucursales y depósitos",
    "configuración operativa",
    "configuración de Google Sheets",
    "productos y proveedores",
]
RESET_TABLES_IN_DELETE_ORDER = [
    "guarantee_history",
    "guarantee_items",
    "guarantee_exports",
    "guarantee_sync_logs",
    "guarantee_counters",
    "warranty_remitos",
    "guarantees",
]


def _is_reset_admin(user: Any) -> bool:
    roles = _user_role_keys(user)
    perms = set(getattr(user, "permissions", []) or [])
    return "*" in perms or "warranties.reset_data" in perms or bool(roles & {"SUPERADMIN", "ADMIN", "ADMINISTRADOR", "GERENTE"})


def _require_reset_admin(user: Any) -> None:
    if not _is_reset_admin(user):
        raise HTTPException(status_code=403, detail="Solo un administrador puede resetear datos de prueba de garantías.")


def _table_count(conn: sqlite3.Connection, table: str) -> int:
    if not _table_exists(conn, table):
        return 0
    return int(conn.execute(f"SELECT COUNT(*) AS c FROM {table}").fetchone()["c"] or 0)


def _generated_export_files_count() -> int:
    path = warranty_exports_dir()
    if not path.exists():
        return 0
    return sum(1 for p in path.glob("*.xlsx") if p.is_file())


def _reset_summary(conn: sqlite3.Connection) -> WarrantyResetSummary:
    return WarrantyResetSummary(
        guarantees=_table_count(conn, "guarantees"),
        guarantee_items=_table_count(conn, "guarantee_items"),
        guarantee_history=_table_count(conn, "guarantee_history"),
        remitos=_table_count(conn, "warranty_remitos"),
        exports=_table_count(conn, "guarantee_exports"),
        sync_logs=_table_count(conn, "guarantee_sync_logs"),
        counters=_table_count(conn, "guarantee_counters"),
        generated_export_files=_generated_export_files_count(),
    )


def _fetch_table_rows(conn: sqlite3.Connection, table: str) -> list[dict[str, Any]]:
    if not _table_exists(conn, table):
        return []
    rows = conn.execute(f"SELECT * FROM {table}").fetchall()
    return [dict(row) for row in rows]


def _create_warranty_reset_backup(conn: sqlite3.Connection, user: Any) -> Path:
    settings = get_settings()
    backup_dir = settings.outputs_dir / "warranties" / "reset_backups"
    backup_dir.mkdir(parents=True, exist_ok=True)
    stamp = now_ar().strftime("%Y%m%d-%H%M%S")
    filename = f"backup-garantias-pre-reset-{stamp}.json"
    path = backup_dir / filename
    payload = {
        "created_at": utc_now_iso(),
        "created_by": getattr(user, "username", "") or "",
        "created_by_name": getattr(user, "display_name", "") or "",
        "summary": _reset_summary(conn).model_dump(),
        "preserved": RESET_PRESERVED_ITEMS,
        "tables": {table: _fetch_table_rows(conn, table) for table in RESET_TABLES_IN_DELETE_ORDER},
    }
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
    return path


def _delete_generated_warranty_export_files() -> int:
    path = warranty_exports_dir()
    if not path.exists():
        return 0
    deleted = 0
    for file_path in path.glob("*.xlsx"):
        try:
            if file_path.is_file():
                file_path.unlink()
                deleted += 1
        except Exception:
            # No bloquear el reset por un archivo abierto. Quedará registrado en el backup/auditoría.
            continue
    return deleted


@router.get("/production-reset/preview", response_model=WarrantyResetPreviewResponse)
def preview_warranty_production_reset(user: Annotated[Any, Depends(require_current_user)]):
    _require_reset_admin(user)
    with db_connect() as conn:
        ensure_warranty_tables(conn)
        summary = _reset_summary(conn)
    return WarrantyResetPreviewResponse(
        generated_at=utc_now_iso(),
        summary=summary,
        preserved=RESET_PRESERVED_ITEMS,
        warning="Esta acción limpia datos operativos de garantías de prueba y reinicia correlativos. No borra usuarios, empresas, sucursales, depósitos ni configuración.",
        confirmation_phrase=RESET_CONFIRMATION_PHRASE,
    )


@router.post("/production-reset/backup")
def create_warranty_production_reset_backup(user: Annotated[Any, Depends(require_current_user)]):
    _require_reset_admin(user)
    with db_connect() as conn:
        ensure_warranty_tables(conn)
        backup_path = _create_warranty_reset_backup(conn, user)
    audit("warranties.production_reset.backup", user=user, resource_type="warranty_reset", resource_id=backup_path.name)
    return FileResponse(path=backup_path, media_type="application/json", filename=backup_path.name)


@router.post("/production-reset/execute", response_model=WarrantyResetResponse)
def execute_warranty_production_reset(data: WarrantyResetRequest, user: Annotated[Any, Depends(require_current_user)]):
    _require_reset_admin(user)
    if data.confirmation.strip().upper() != RESET_CONFIRMATION_PHRASE:
        raise HTTPException(status_code=400, detail=f"Confirmación inválida. Escribí exactamente: {RESET_CONFIRMATION_PHRASE}")

    reset_at = utc_now_iso()
    with db_connect() as conn:
        ensure_warranty_tables(conn)
        summary_before = _reset_summary(conn)
        backup_path = _create_warranty_reset_backup(conn, user)
        for table in RESET_TABLES_IN_DELETE_ORDER:
            if _table_exists(conn, table):
                conn.execute(f"DELETE FROM {table}")
        # Reiniciar AUTOINCREMENT de tablas limpiadas. Los códigos visibles GAR/REM/ENV
        # también quedan en cero porque se eliminan counters/exports/remitos.
        if _table_exists(conn, "sqlite_sequence"):
            placeholders = ",".join("?" for _ in RESET_TABLES_IN_DELETE_ORDER)
            conn.execute(f"DELETE FROM sqlite_sequence WHERE name IN ({placeholders})", RESET_TABLES_IN_DELETE_ORDER)
        conn.commit()

    deleted_files = _delete_generated_warranty_export_files() if data.reset_generated_files else 0
    audit(
        "warranties.production_reset.executed",
        user=user,
        resource_type="warranty_reset",
        resource_id=backup_path.name,
        details={"summary_before": summary_before.model_dump(), "deleted_generated_files": deleted_files},
    )
    return WarrantyResetResponse(
        ok=True,
        reset_at=reset_at,
        summary_before=summary_before,
        backup_file=backup_path.name,
        deleted_generated_files=deleted_files,
        message="Datos operativos de garantías limpiados. Usuarios, empresas, sucursales, depósitos, roles, permisos y configuración se conservaron.",
    )


@router.get("/counters", response_model=WarrantyCountersResponse)
def get_warranty_counters(_user: Annotated[Any, Depends(require_permission("warranties.manage"))]):
    with db_connect() as conn:
        ensure_warranty_tables(conn)
        rows = conn.execute("SELECT year, sucursal_code, last_number FROM guarantee_counters ORDER BY year, sucursal_code").fetchall()
    return WarrantyCountersResponse(counters=[WarrantyCounterInfo(year=int(row["year"]), sucursal=str(row["sucursal_code"]), last_number=int(row["last_number"])) for row in rows])


@router.post("/counters/resync", response_model=WarrantyCountersResponse)
def resync_warranty_counters(user: Annotated[Any, Depends(require_permission("warranties.manage"))]):
    counters: dict[tuple[int, str], int] = {}
    with db_connect() as conn:
        ensure_warranty_tables(conn)
        rows = conn.execute("SELECT warranty_code FROM guarantees").fetchall()
        for row in rows:
            m = re.fullmatch(r"GAR-(\d{4})-([A-Z0-9]+)-(\d+)", str(row["warranty_code"] or "").upper())
            if not m:
                continue
            key = (int(m.group(1)), m.group(2))
            counters[key] = max(counters.get(key, 0), int(m.group(3)))
        conn.execute("DELETE FROM guarantee_counters")
        now = utc_now_iso()
        for (year, code), last in counters.items():
            conn.execute("INSERT INTO guarantee_counters (year, sucursal_code, last_number, updated_at) VALUES (?, ?, ?, ?)", (year, code, last, now))
        conn.commit()
    audit("warranties.counters.resync", user=user, resource_type="warranty_counter", details={"counters": {f"{year}-{code}": last for (year, code), last in counters.items()}, "source": "database"})
    return get_warranty_counters(user)



@router.get("/review-queue", response_model=WarrantyListResponse)
def review_queue(
    _user: Annotated[Any, Depends(require_permission("warranties.review"))],
    q: str = "",
    sucursal: str = "",
    deposito: str = "",
    limit: int = Query(default=300, ge=1, le=1000),
):
    deny_plain_deposit_operator(_user, "ver bandeja de revisión")

    # Fase 37: Revisión debe leer la cola real de ingreso/revisión, no depender de
    # filtros pensados para el listado operativo de sucursal ni de campos legacy.
    # Caso clave: una garantía cargada en Caseros con ubicación_actual=Caseros debe
    # aparecer en Revisión aunque también esté disponible para remito. El remito
    # mueve lo físico; revisión controla la calidad de datos.
    q_tokens = normalize_text(q).split()
    suc_key = normalize_text(sucursal)
    dep_key = normalize_text(deposito)
    if dep_key in {"TODOS", "ALL"}:
        dep_key = ""
    if suc_key in {"TODOS", "ALL"}:
        suc_key = ""

    with db_connect() as conn:
        ensure_warranty_tables(conn)
        rows = conn.execute("SELECT * FROM guarantees ORDER BY ingreso_at DESC, id DESC").fetchall()
        all_items = conn.execute("SELECT * FROM guarantee_items ORDER BY id").fetchall()

    by_gid: dict[int, list[sqlite3.Row]] = {}
    for item_row in all_items:
        by_gid.setdefault(int(item_row["guarantee_id"]), []).append(item_row)

    summaries = [row_to_summary(row, by_gid.get(int(row["id"]), [])) for row in rows]

    def _branch_label_from_id(branch_id: str) -> str:
        if not branch_id:
            return ""
        try:
            info = _fetch_branch_info(branch_id)
            return str((info or {}).get("name") or "")
        except Exception:
            return ""

    def _matches_sucursal(item: WarrantySummary) -> bool:
        if not suc_key:
            return True
        values = {
            normalize_text(item.sucursal),
            normalize_text(item.sucursal_responsable),
            normalize_text(item.ubicacion_actual),
            normalize_text(_branch_label_from_id(item.branch_id)),
            normalize_text(_branch_label_from_id(item.sucursal_responsable_id)),
        }
        values.discard("")
        return suc_key in values

    def _matches_lugar(item: WarrantySummary) -> bool:
        if not dep_key:
            return True
        values = {
            normalize_text(item.deposito),
            normalize_text(item.lugar_llegada),
            normalize_text(item.ubicacion_actual),
            normalize_text(item.ubicacion_actual_label),
        }
        values.discard("")
        return dep_key in values

    def _matches_search(item: WarrantySummary) -> bool:
        if not q_tokens:
            return True
        haystack = normalize_text(" ".join([
            item.id_garantia,
            item.parent_warranty_code,
            item.producto_principal,
            " ".join(item.productos),
            item.sku,
            item.serie,
            item.falla,
            item.responsable,
            item.sucursal,
            item.sucursal_responsable,
            item.deposito,
            item.ubicacion_actual,
            item.provider_name,
        ]))
        return all(token in haystack for token in q_tokens)

    def _is_in_review_queue(item: WarrantySummary) -> bool:
        if item.cancelled:
            return False
        if review_status_matches(item.review_status, REVIEW_APPROVED):
            return False
        # Pendiente / en revisión / requiere corrección entran siempre aunque el
        # estado operativo venga escrito con alias viejos. También entran los
        # INGRESO nuevos aunque review_status esté vacío por compatibilidad.
        return (
            status_matches(item.estado, DEFAULT_STATUSES[0])
            or review_status_matches(item.review_status, REVIEW_PENDING)
            or review_status_matches(item.review_status, REVIEW_IN_PROGRESS)
            or review_status_matches(item.review_status, REVIEW_INCOMPLETE)
        )

    filtered = [
        item for item in summaries
        if _is_in_review_queue(item)
        and _matches_sucursal(item)
        and _matches_lugar(item)
        and _matches_search(item)
    ]
    return WarrantyListResponse(items=filtered[:limit], total=len(filtered), limit=limit)


@router.post("/{warranty_id}/take-review", response_model=WarrantyDetailResponse)
def take_warranty_into_review(warranty_id: str, data: WarrantyReviewRequest, user: Annotated[Any, Depends(require_permission("warranties.review"))]):
    deny_plain_deposit_operator(user, "tomar garantías en revisión")
    """Marca la garantía como 'en revisión interna' (el depósito la está revisando activamente)."""
    note = (data.note or "").strip()
    if not (user.has("warranties.create") or user.has("warranties.manage") or user.has("warranties.manage_provider")):
        raise HTTPException(status_code=403, detail="No tenés permiso para editar la base de ingreso")

    with db_connect() as conn:
        result = fetch_guarantee_with_items(conn, warranty_id)
        if not result:
            raise HTTPException(status_code=404, detail="Garantía no encontrada")
        row, _items = result
        try:
            cancelled_val = row["cancelled"]
        except (KeyError, IndexError):
            cancelled_val = 0
        if str(cancelled_val or 0) == "1":
            raise HTTPException(status_code=400, detail="La garantía está anulada")
        current_rs = str(row["review_status"] or REVIEW_PENDING)
        if review_status_matches(current_rs, REVIEW_APPROVED):
            raise HTTPException(status_code=400, detail="La garantía ya fue revisada y aprobada")
        if review_status_matches(current_rs, REVIEW_IN_PROGRESS):
            raise HTTPException(status_code=400, detail="La garantía ya está en revisión interna")
        current_status = str(row["status"] or "")
        new_status = DEFAULT_STATUSES[0]
        now = utc_now_iso()
        actor_username = getattr(user, "username", "") or ""
        actor_name = getattr(user, "display_name", "") or ""
        conn.execute(
            """
            UPDATE guarantees
            SET review_status = ?, status = ?, review_started_at = ?, review_started_by = ?,
                updated_at = ?, updated_by = ?, updated_by_name = ?, synced_to_google_sheet = 0
            WHERE id = ?
            """,
            (REVIEW_IN_PROGRESS, new_status, now, actor_username, now, actor_username, actor_name, int(row["id"])),
        )
        add_history(
            conn,
            int(row["id"]),
            str(row["warranty_code"]),
            user,
            "review_started",
            old_status=current_status,
            new_status=new_status,
            note=note or "Tomada en revisión interna",
            details={"review_status": REVIEW_IN_PROGRESS, "review_started_at": now, "status_normalizado": new_status},
        )
        conn.commit()
    audit("warranties.review.take", user=user, resource_type="warranty", resource_id=warranty_id, details={"note": note})
    return get_warranty_detail(warranty_id, user)


@router.post("/{warranty_id}/mark-incomplete", response_model=WarrantyDetailResponse)
def mark_warranty_incomplete(warranty_id: str, data: WarrantyReviewRequest, user: Annotated[Any, Depends(require_permission("warranties.mark_incomplete"))]):
    deny_plain_deposit_operator(user, "pedir corrección de garantías")
    note = (data.note or "").strip()
    if not note:
        raise HTTPException(status_code=400, detail="Indicá qué debe corregir la sucursal antes de devolver la garantía")
    with db_connect() as conn:
        result = fetch_guarantee_with_items(conn, warranty_id)
        if not result:
            raise HTTPException(status_code=404, detail="Garantía no encontrada")
        row, _items = result
        try:
            cancelled_val = row["cancelled"]
        except (KeyError, IndexError):
            cancelled_val = 0
        if str(cancelled_val or 0) == "1":
            raise HTTPException(status_code=400, detail="La garantía está anulada")
        if review_status_matches(str(row["review_status"] or REVIEW_PENDING), REVIEW_APPROVED):
            raise HTTPException(status_code=400, detail="La garantía ya fue aprobada; corregila desde gestión/admin si corresponde")
        current_status = str(row["status"] or "")
        new_status = DEFAULT_STATUSES[0]
        now = utc_now_iso()
        actor_username = getattr(user, "username", "") or ""
        actor_name = getattr(user, "display_name", "") or ""
        conn.execute(
            """
            UPDATE guarantees
            SET review_status = ?, status = ?, review_note = ?, correction_requested_at = ?, correction_requested_by = ?,
                updated_at = ?, updated_by = ?, updated_by_name = ?, synced_to_google_sheet = 0
            WHERE id = ?
            """,
            (REVIEW_INCOMPLETE, new_status, note, now, actor_username, now, actor_username, actor_name, int(row["id"])),
        )
        add_history(
            conn,
            int(row["id"]),
            str(row["warranty_code"]),
            user,
            "review_correction_requested",
            old_status=current_status,
            new_status=new_status,
            note=note,
            details={"review_status": REVIEW_INCOMPLETE, "correction_requested_at": now, "status_normalizado": new_status},
        )
        conn.commit()
    audit("warranties.review.incomplete", user=user, resource_type="warranty", resource_id=warranty_id, details={"note": note})
    return get_warranty_detail(warranty_id, user)


@router.post("/{warranty_id}/approve-review", response_model=WarrantyDetailResponse)
def approve_warranty_review(warranty_id: str, data: WarrantyReviewRequest, user: Annotated[Any, Depends(require_permission("warranties.approve_review"))]):
    deny_plain_deposit_operator(user, "aprobar garantías")
    note = (data.note or "").strip()
    with db_connect() as conn:
        result = fetch_guarantee_with_items(conn, warranty_id)
        if not result:
            raise HTTPException(status_code=404, detail="Garantía no encontrada")
        row, _items = result
        try:
            cancelled_val = row["cancelled"]
        except (KeyError, IndexError):
            cancelled_val = 0
        if str(cancelled_val or 0) == "1":
            raise HTTPException(status_code=400, detail="La garantía está anulada")
        if review_status_matches(str(row["review_status"] or REVIEW_PENDING), REVIEW_APPROVED):
            raise HTTPException(status_code=400, detail="La garantía ya fue revisada y aprobada")
        current_status = str(row["status"] or "")
        new_status = DEFAULT_STATUSES[1]   # configurable — actualmente "2 - PENDIENTE"
        now = utc_now_iso()
        actor_username = getattr(user, "username", "") or ""
        actor_name = getattr(user, "display_name", "") or ""
        conn.execute(
            """
            UPDATE guarantees
            SET review_status = ?, reviewed_by = ?, reviewed_by_name = ?, reviewed_at = ?, review_note = ?,
                status = ?, correction_requested_at = '', correction_requested_by = '',
                updated_at = ?, updated_by = ?, updated_by_name = ?, synced_to_google_sheet = 0
            WHERE id = ?
            """,
            (
                REVIEW_APPROVED,
                actor_username,
                actor_name,
                now,
                note,
                new_status,
                now,
                actor_username,
                actor_name,
                int(row["id"]),
            ),
        )
        add_history(
            conn,
            int(row["id"]),
            str(row["warranty_code"]),
            user,
            "review_approved",
            old_status=current_status,
            new_status=new_status,
            note=note or "Revisión aprobada",
            details={"review_status": REVIEW_APPROVED},
        )
        conn.commit()
    audit("warranties.review.approve", user=user, resource_type="warranty", resource_id=warranty_id, details={"status": new_status})
    return get_warranty_detail(warranty_id, user)




@router.get("/management", response_model=WarrantyListResponse)
def management_warranties(
    _user: Annotated[Any, Depends(require_permission("warranties.manage_provider"))],
    q: str = "",
    marca: str = "",
    proveedor: str = "",
    sucursal: str = "",
    deposito: str = "",
    estado: str = "",
    review_status: str = "revisada",
    include_pending: bool = False,
    demora_min: int = Query(default=0, ge=0, le=3650),
    limit: int = Query(default=300, ge=1, le=1000),
):
    deny_plain_deposit_operator(_user, "gestionar proveedor")
    # Las garantías "2 - PENDIENTE" pasan por Exportación antes de llegar a Gestión.
    # Por defecto las ocultamos aquí para forzar el flujo correcto.
    # Mostramos TODAS las garantías con review_status="revisada", incluyendo
    # las "2 - PENDIENTE" que acaban de ser aprobadas en revisión y aún no tienen
    # lote de exportación. El frontend las distingue visualmente.
    return list_warranties(
        _user,
        q=q,
        marca=marca,
        proveedor=proveedor,
        sucursal=sucursal,
        deposito=deposito,
        estado=estado,
        review_status=review_status,
        demora_min=demora_min,
        limit=limit,
    )


@router.get("/delayed", response_model=WarrantyListResponse)
def delayed_warranties(
    _user: Annotated[Any, Depends(require_permission("warranties.manage_provider"))],
    q: str = "",
    marca: str = "",
    proveedor: str = "",
    sucursal: str = "",
    deposito: str = "",
    limit: int = Query(default=300, ge=1, le=1000),
):
    deny_plain_deposit_operator(_user, "ver gestión proveedor")
    return list_warranties(
        _user,
        q=q,
        marca=marca,
        proveedor=proveedor,
        sucursal=sucursal,
        deposito=deposito,
        demora_min=7,
        limit=limit,
    )


@router.post("/{warranty_id}/confirm-shipment", response_model=WarrantyDetailResponse)
def confirm_warranty_shipment(warranty_id: str, data: ConfirmShipmentRequest, user: Annotated[Any, Depends(require_permission("warranties.manage_provider"))]):
    deny_plain_deposit_operator(user, "confirmar ENV/mail al proveedor")
    """Confirma que el lote fue enviado al proveedor. Valida el código ENV y pasa a ENVIADO AL PROVEEDOR."""
    code_input = data.shipment_code.strip().upper()
    provider = (data.provider_name or "").strip()
    nota = (data.nota or "").strip()
    with db_connect() as conn:
        result = fetch_guarantee_with_items(conn, warranty_id)
        if not result:
            raise HTTPException(status_code=404, detail="Garantía no encontrada")
        row, _items = result
        # ENV/mail es un aviso administrativo al proveedor: puede confirmarse
        # aunque el producto todavía esté viajando a Chiclana. La entrega física
        # se controla más adelante con solicitud/retiro proveedor.
        stored_code = str(row["shipment_code"] or "").strip().upper()
        if not stored_code:
            raise HTTPException(status_code=400, detail="Esta garantía no tiene un lote de exportación asignado.")
        if code_input != stored_code:
            raise HTTPException(status_code=400, detail=f"El código ingresado no coincide con el lote asignado ({stored_code}).")
        old_status = str(row["status"] or "")
        new_status = "4 - ENVIADO AL PROVEEDOR"
        now = utc_now_iso()
        updates: dict[str, Any] = {
            "status": new_status,
            "sent_to_provider_at": str(row["sent_to_provider_at"] or "") or now,
            "fecha_ultimo_mail_proveedor": now,
            "synced_to_google_sheet": 0,
        }
        if provider:
            updates["provider_name"] = provider
        update_guarantee_provider_fields(
            conn,
            row=row,
            user=user,
            updates=updates,
            action="shipment_confirmed",
            old_status=old_status,
            new_status=new_status,
            note=nota or f"Envío al proveedor confirmado. Lote: {stored_code}",
            details={"shipment_code": stored_code, "provider_name": provider},
        )
        conn.commit()
    audit("warranties.shipment.confirmed", user=user, resource_type="warranty", resource_id=warranty_id, details={"shipment_code": stored_code})
    return get_warranty_detail(warranty_id, user)


@router.post("/{warranty_id}/send-provider", response_model=WarrantyDetailResponse)
def send_warranty_to_provider(warranty_id: str, data: WarrantyProviderSendRequest, user: Annotated[Any, Depends(require_permission("warranties.manage_provider"))]):
    deny_plain_deposit_operator(user, "enviar garantías al proveedor")
    provider = data.provider_name.strip()
    case_id = (data.provider_case_id or "").strip()
    note = (data.note or "").strip()
    with db_connect() as conn:
        result = fetch_guarantee_with_items(conn, warranty_id)
        if not result:
            raise HTTPException(status_code=404, detail="Garantía no encontrada")
        row, _items = result
        assert_internal_logistics_ready_for_provider(row)
        old_status = str(row["status"] or "")
        new_status = "4 - ENVIADO AL PROVEEDOR"
        now = utc_now_iso()
        updates = {
            "provider_name": provider,
            "provider_case_id": case_id,
            "sent_to_provider_at": str(row["sent_to_provider_at"] or "") or now,
            "status": new_status,
        }
        update_guarantee_provider_fields(
            conn,
            row=row,
            user=user,
            updates=updates,
            action="provider_sent",
            old_status=old_status,
            new_status=new_status,
            note=note or f"Enviada al proveedor {provider}",
            details={"provider_name": provider, "provider_case_id": case_id},
        )
        conn.commit()
    audit("warranties.provider.sent", user=user, resource_type="warranty", resource_id=warranty_id, details={"provider_name": provider, "provider_case_id": case_id})
    return get_warranty_detail(warranty_id, user)


@router.post("/{warranty_id}/provider-pickup-request", response_model=WarrantyDetailResponse)
def register_provider_pickup_request(warranty_id: str, data: ProviderPickupRequest, user: Annotated[Any, Depends(require_permission("warranties.register_provider_response"))]):
    deny_plain_deposit_operator(user, "registrar retiro solicitado por proveedor")
    """Registra que el proveedor aceptó/solicitó retiro.

    No mueve físicamente la garantía ni cambia a EN EL PROVEEDOR.
    Si todavía no está en Chiclana, queda como retiro_solicitado para disparar urgencia logística.
    Si ya está en depósito, queda listo_para_retiro.
    """
    note = (data.note or "").strip()
    with db_connect() as conn:
        result = fetch_guarantee_with_items(conn, warranty_id)
        if not result:
            raise HTTPException(status_code=404, detail="Garantía no encontrada")
        row, _items = result
        if not provider_flow_started(row):
            raise HTTPException(status_code=400, detail="Primero confirmá el envío del ENV/mail al proveedor.")
        status_key = canonical_status_key(str(row["status"] or ""))
        if status_key in {"RESUELTO", "RECHAZADO", "ANULADA", "FINALIZADO"}:
            raise HTTPException(status_code=400, detail="La garantía ya está cerrada o resuelta; no corresponde solicitar retiro.")
        old_status = str(row["status"] or "")
        now = utc_now_iso()
        pickup_status = "listo_para_retiro" if internal_logistics_ready_for_provider(row) else "retiro_solicitado"
        updates: dict[str, Any] = {
            "estado_retiro_proveedor": pickup_status,
            "fecha_solicitud_retiro_proveedor": now,
            "last_provider_response_at": now,
            "nota_retiro_proveedor": note,
        }
        if data.provider_case_id is not None and data.provider_case_id.strip():
            updates["provider_case_id"] = data.provider_case_id.strip()
        update_guarantee_provider_fields(
            conn,
            row=row,
            user=user,
            updates=updates,
            action="provider_pickup_requested",
            old_status=old_status,
            new_status=old_status,
            note=note or ("Proveedor solicitó retiro. Producto listo en Chiclana." if pickup_status == "listo_para_retiro" else "Proveedor solicitó retiro. Traer urgente a Depósito Chiclana."),
            details={"pickup_status": pickup_status, "fecha_retiro_acordada": data.fecha_retiro_acordada or ""},
        )
        conn.commit()

    # Alerta al Gestor de Garantías cuando el producto aún NO está en el depósito
    if pickup_status == "retiro_solicitado":
        _notify_gestor_garantias_pickup(
            "⚠️ Retiro urgente pendiente",
            f"El proveedor solicitó retiro de la garantía {warranty_id} pero el producto no está en depósito. Coordinar traslado urgente.",
        )

    audit("warranties.provider.pickup_requested", user=user, resource_type="warranty", resource_id=warranty_id, details={"pickup_status": pickup_status})
    return get_warranty_detail(warranty_id, user)


@router.post("/{warranty_id}/provider-response", response_model=WarrantyDetailResponse)
def register_provider_response(warranty_id: str, data: WarrantyProviderResponseRequest, user: Annotated[Any, Depends(require_permission("warranties.register_provider_response"))]):
    deny_plain_deposit_operator(user, "registrar respuestas del proveedor")
    note = (data.note or "").strip()
    case_id = (data.provider_case_id or "").strip()
    requested_status = (data.estado or "").strip()
    with db_connect() as conn:
        result = fetch_guarantee_with_items(conn, warranty_id)
        if not result:
            raise HTTPException(status_code=404, detail="Garantía no encontrada")
        row, _items = result
        if not provider_flow_started(row):
            raise HTTPException(status_code=400, detail="Todavía no se confirmó el envío al proveedor. Primero confirmá el lote ENV/mail enviado.")
        old_status = str(row["status"] or "")
        new_status = validate_status_or_400(requested_status) if requested_status else "6 - RESPONDIDO POR PROVEEDOR"
        target_key = canonical_status_key(new_status)
        if target_key in {"EN EL PROVEEDOR", "RESUELTO"}:
            assert_provider_has_physical_product(row)
        updates = {
            "last_provider_response_at": utc_now_iso(),
            "status": new_status,
        }
        if case_id:
            updates["provider_case_id"] = case_id
        if is_resolved_status(new_status) and not row["fecha_resolucion"]:
            updates["fecha_resolucion"] = utc_now_iso()
        update_guarantee_provider_fields(
            conn,
            row=row,
            user=user,
            updates=updates,
            action="provider_response",
            old_status=old_status,
            new_status=new_status,
            note=note or "Respuesta del proveedor registrada",
            details={"provider_case_id": case_id, "last_provider_response_at": updates["last_provider_response_at"]},
        )
        conn.commit()
    audit("warranties.provider.response", user=user, resource_type="warranty", resource_id=warranty_id, details={"status": new_status})
    return get_warranty_detail(warranty_id, user)


@router.post("/{warranty_id}/resend-provider-mail", response_model=WarrantyDetailResponse)
def resend_provider_mail(warranty_id: str, data: WarrantyResendMailRequest, user: Annotated[Any, Depends(require_permission("warranties.register_claim"))]):
    deny_plain_deposit_operator(user, "reenviar mails/reclamos al proveedor")
    """Registra mail/reclamo reenviado al proveedor y reinicia el contador de días sin respuesta."""
    note = (data.note or "").strip() or "Mail vuelto a enviar al proveedor"
    with db_connect() as conn:
        result = fetch_guarantee_with_items(conn, warranty_id)
        if not result:
            raise HTTPException(status_code=404, detail="Garantía no encontrada")
        row, _items = result
        if not provider_flow_started(row):
            raise HTTPException(status_code=400, detail="Todavía no se confirmó el envío al proveedor. No corresponde reenviar mail.")
        status_key = canonical_status_key(str(row["status"] or ""))
        if status_key in {"RESUELTO", "RECHAZADO", "ANULADA", "FINALIZADO"}:
            raise HTTPException(status_code=400, detail="La garantía ya está cerrada o resuelta; no corresponde reenviar mail.")
        now = utc_now_iso()
        update_guarantee_provider_fields(
            conn,
            row=row,
            user=user,
            updates={
                "fecha_ultimo_mail_proveedor": now,
                "last_claim_at": now,
            },
            action="provider_mail_resent",
            old_status=str(row["status"] or ""),
            new_status=str(row["status"] or ""),
            note=note,
            details={"fecha_ultimo_mail_proveedor": now},
        )
        conn.commit()
    audit("warranties.provider.mail_resent", user=user, resource_type="warranty", resource_id=warranty_id, details={"note": note})
    return get_warranty_detail(warranty_id, user)


@router.post("/{warranty_id}/claim", response_model=WarrantyDetailResponse)
def register_warranty_claim(warranty_id: str, data: WarrantyClaimRequest, user: Annotated[Any, Depends(require_permission("warranties.register_claim"))]):
    deny_plain_deposit_operator(user, "registrar reclamos al proveedor")
    note = data.note.strip()
    with db_connect() as conn:
        result = fetch_guarantee_with_items(conn, warranty_id)
        if not result:
            raise HTTPException(status_code=404, detail="Garantía no encontrada")
        row, _items = result
        if not provider_flow_started(row):
            raise HTTPException(status_code=400, detail="Todavía no se confirmó el envío al proveedor. No corresponde registrar reclamos.")
        update_guarantee_provider_fields(
            conn,
            row=row,
            user=user,
            updates={"last_claim_at": utc_now_iso()},
            action="provider_claim",
            old_status=str(row["status"] or ""),
            new_status=str(row["status"] or ""),
            note=note,
            details={"last_claim_at": utc_now_iso()},
        )
        conn.commit()
    audit("warranties.provider.claim", user=user, resource_type="warranty", resource_id=warranty_id, details={"note": note})
    return get_warranty_detail(warranty_id, user)


@router.post("/{warranty_id}/status", response_model=WarrantyDetailResponse)
def change_warranty_status(warranty_id: str, data: WarrantyStatusChangeRequest, user: Annotated[Any, Depends(require_permission("warranties.change_status"))]):
    deny_plain_deposit_operator(user, "cambiar estados operativos")
    new_status = validate_status_or_400(data.estado)
    note = (data.note or "").strip()
    resultado = normalize_resolution_result(data.resultado_resolucion)

    is_resuelto = "RESUELTO" in canonical_status_key(new_status)
    is_finalizado = "FINALIZADO" in canonical_status_key(new_status)

    if is_resuelto:
        if not resultado:
            raise HTTPException(status_code=400, detail="Indicá cómo se resolvió la garantía: nota de crédito, reparación o cambio de equipo.")
        if resultado not in RESOLUTION_OPTIONS:
            raise HTTPException(status_code=400, detail=f"Resultado de resolución inválido: {resultado}")

    def clean(value: str | None) -> str:
        return (value or "").strip()

    with db_connect() as conn:
        result = fetch_guarantee_with_items(conn, warranty_id)
        if not result:
            raise HTTPException(status_code=404, detail="Garantía no encontrada")
        row, _items = result
        target_key = canonical_status_key(new_status)
        # Solo el retiro físico/EN EL PROVEEDOR exige que la garantía esté en Chiclana/depósito.
        # Las respuestas administrativas del proveedor pueden registrarse antes para disparar urgencias.
        if target_key == "EN EL PROVEEDOR":
            assert_internal_logistics_ready_for_provider(row)
        if target_key == "RESUELTO":
            assert_provider_has_physical_product(row)
        if target_key in {"EN EL PROVEEDOR", "RESPONDIDO POR PROVEEDOR", "RESUELTO", "RECHAZADO", "ANULADA"} and not provider_flow_started(row):
            raise HTTPException(status_code=400, detail="No se puede avanzar proveedor sin confirmar antes el ENV/mail enviado.")
        old_status = str(row["status"] or "")
        updates: dict[str, Any] = {"status": new_status}
        if target_key == "EN EL PROVEEDOR":
            now_pickup = utc_now_iso()
            updates["estado_retiro_proveedor"] = "retirado"
            updates["fecha_retiro_proveedor"] = str(row["fecha_retiro_proveedor"] or "") if "fecha_retiro_proveedor" in row.keys() and row["fecha_retiro_proveedor"] else now_pickup
            updates["fecha_retiro"] = str(row["fecha_retiro"] or "") if "fecha_retiro" in row.keys() and row["fecha_retiro"] else now_pickup
            updates["ubicacion_actual"] = "proveedor"

        if is_resolved_status(new_status) and not row["fecha_resolucion"]:
            updates["fecha_resolucion"] = utc_now_iso()
        if is_finalizado:
            if "fecha_finalizacion" in row.keys() and not row["fecha_finalizacion"]:
                updates["fecha_finalizacion"] = utc_now_iso()
            if data.finalizacion is not None:
                updates["finalizacion"] = clean(data.finalizacion)
            elif note:
                updates["finalizacion"] = note

        # Fase 12: resolución normalizada. Mantiene resolution_reference/resolution_note
        # por compatibilidad visual, pero guarda campos específicos para reportes.
        if resultado:
            updates["resultado_resolucion"] = resultado
            if resultado == "nota_credito":
                nc_number = clean(data.numero_nota_credito) or clean(data.resolution_reference)
                nc_amount = clean(data.importe_nota_credito)
                nc_date = clean(data.fecha_nota_credito)
                if nc_number:
                    updates["numero_nota_credito"] = nc_number
                    updates["resolution_reference"] = nc_number
                if nc_amount:
                    updates["importe_nota_credito"] = nc_amount
                if nc_date:
                    updates["fecha_nota_credito"] = nc_date
                if data.resolution_note is not None:
                    updates["resolution_note"] = clean(data.resolution_note)
            elif resultado == "reparacion":
                detail = clean(data.detalle_reparacion) or clean(data.resolution_note)
                repair_date = clean(data.fecha_reparacion)
                if detail:
                    updates["detalle_reparacion"] = detail
                    updates["resolution_note"] = detail
                if repair_date:
                    updates["fecha_reparacion"] = repair_date
                if data.resolution_reference is not None:
                    updates["resolution_reference"] = clean(data.resolution_reference)
            elif resultado == "cambio_equipo":
                replacement_product = clean(data.producto_reemplazo) or clean(data.resolution_note)
                replacement_sku = clean(data.sku_reemplazo)
                replacement_serial = clean(data.serie_reemplazo)
                replacement_received = clean(data.fecha_recepcion_reemplazo)
                if replacement_product:
                    updates["producto_reemplazo"] = replacement_product
                    updates["resolution_note"] = replacement_product
                if replacement_sku:
                    updates["sku_reemplazo"] = replacement_sku
                if replacement_serial:
                    updates["serie_reemplazo"] = replacement_serial
                    updates["resolution_reference"] = replacement_serial
                elif data.resolution_reference is not None:
                    updates["resolution_reference"] = clean(data.resolution_reference)
                if replacement_received:
                    updates["fecha_recepcion_reemplazo"] = replacement_received

        if data.resolution_note is not None and not resultado:
            updates["resolution_note"] = clean(data.resolution_note)
        if data.resolution_reference is not None and not resultado:
            updates["resolution_reference"] = clean(data.resolution_reference)

        history_note = note or f"Estado actualizado a {new_status}"
        if resultado:
            history_note = f"{history_note} | Resolución: {RESOLUTION_OPTIONS[resultado]}"
        if updates.get("numero_nota_credito"):
            history_note = f"{history_note} | NC: {updates['numero_nota_credito']}"
        if updates.get("importe_nota_credito"):
            history_note = f"{history_note} | Importe: {updates['importe_nota_credito']}"
        if updates.get("detalle_reparacion"):
            history_note = f"{history_note} | Reparación: {updates['detalle_reparacion']}"
        if updates.get("producto_reemplazo"):
            history_note = f"{history_note} | Cambio: {updates['producto_reemplazo']}"

        action = "status_changed"
        if is_resuelto:
            action = "resolution_registered"
        elif is_finalizado:
            action = "warranty_finalized"

        update_guarantee_provider_fields(
            conn,
            row=row,
            user=user,
            updates=updates,
            action=action,
            old_status=old_status,
            new_status=new_status,
            note=history_note,
            details={
                "resultado_resolucion": resultado,
                "updates": {k: v for k, v in updates.items() if k not in {"updated_at", "updated_by", "updated_by_name", "synced_to_google_sheet"}},
            },
        )
        conn.commit()
    audit("warranties.status.change", user=user, resource_type="warranty", resource_id=warranty_id, details={"old_status": old_status, "new_status": new_status, "resultado_resolucion": resultado})
    return get_warranty_detail(warranty_id, user)


EXPORT_ELIGIBLE_STATUS = DEFAULT_STATUSES[1]  # 2 - PENDIENTE
EXPORT_READY_STATUS = DEFAULT_STATUSES[2]     # 3 - LISTO PARA ENVIAR


def _normalize_warranty_code_list(values: list[str]) -> list[str]:
    seen: set[str] = set()
    output: list[str] = []
    for raw in values:
        code = str(raw or "").strip().upper()
        if not code or code in seen:
            continue
        seen.add(code)
        output.append(code)
    return output


def _export_validation_error(row: sqlite3.Row | None, code: str) -> str:
    if row is None:
        return f"{code}: no existe."
    if int(row["cancelled"] or 0):
        return f"{code}: está anulada/cancelada."
    if not status_matches(str(row["status"] or ""), EXPORT_ELIGIBLE_STATUS):
        return f"{code}: debe estar en {EXPORT_ELIGIBLE_STATUS}."
    if not review_status_matches(str(row["review_status"] or REVIEW_PENDING), REVIEW_APPROVED):
        return f"{code}: debe estar revisada por depósito."
    if str(row["shipment_code"] or "").strip():
        return f"{code}: ya pertenece al lote {row['shipment_code']}."
    # El ENV es aviso administrativo/mail al proveedor, no entrega física.
    # Por eso NO exigimos que ya esté en Chiclana para crear el lote.
    return ""


def validate_export_selection(conn: sqlite3.Connection, warranty_codes: list[str]) -> list[str]:
    codes = _normalize_warranty_code_list(warranty_codes)
    if not codes:
        raise HTTPException(status_code=400, detail="Seleccioná al menos una garantía para crear el lote ENV.")
    placeholders = ",".join("?" * len(codes))
    rows = conn.execute(f"SELECT * FROM guarantees WHERE warranty_code IN ({placeholders})", codes).fetchall()
    by_code = {str(row["warranty_code"] or "").strip().upper(): row for row in rows}
    errors: list[str] = []
    for code in codes:
        msg = _export_validation_error(by_code.get(code), code)
        if msg:
            errors.append(msg)
    if errors:
        preview = " | ".join(errors[:8])
        extra = "" if len(errors) <= 8 else f" | +{len(errors) - 8} más"
        raise HTTPException(status_code=400, detail=f"No se puede crear el ENV: {preview}{extra}")
    return codes


def export_ready_warranty_codes(conn: sqlite3.Connection, warranty_codes: list[str]) -> set[str]:
    codes = _normalize_warranty_code_list(warranty_codes)
    if not codes:
        return set()
    placeholders = ",".join("?" * len(codes))
    rows = conn.execute(f"SELECT * FROM guarantees WHERE warranty_code IN ({placeholders})", codes).fetchall()
    ready: set[str] = set()
    for row in rows:
        code = str(row["warranty_code"] or "").strip().upper()
        if not _export_validation_error(row, code):
            ready.add(code)
    return ready


@router.get("/export/provider-suggestions")
def export_provider_suggestions(
    _user: Annotated[Any, Depends(require_permission("warranties.export"))],
    q: str = "",
    limit: int = Query(default=25, ge=1, le=100),
):
    """Sugerencias para proveedor del ENV. El campo sigue permitiendo escritura manual."""
    deny_plain_deposit_operator(_user, "ver proveedores para exportación")
    needle = normalize_text(q)
    suggestions: list[str] = []
    seen: set[str] = set()

    def add(value: Any) -> None:
        text = str(value or "").strip()
        if not text:
            return
        key = normalize_text(text)
        if not key or key in seen:
            return
        if needle and needle not in key:
            return
        seen.add(key)
        suggestions.append(text)

    with db_connect() as conn:
        ensure_warranty_tables(conn)
        for row in conn.execute("SELECT DISTINCT provider_name FROM guarantees WHERE TRIM(COALESCE(provider_name, '')) <> '' ORDER BY provider_name LIMIT 200").fetchall():
            add(row["provider_name"])
        for row in conn.execute("SELECT DISTINCT provider_name FROM guarantee_exports WHERE TRIM(COALESCE(provider_name, '')) <> '' ORDER BY provider_name LIMIT 200").fetchall():
            add(row["provider_name"])
        for row in conn.execute("SELECT DISTINCT marca FROM guarantee_items WHERE TRIM(COALESCE(marca, '')) <> '' ORDER BY marca LIMIT 200").fetchall():
            add(row["marca"])
    return {"items": suggestions[:limit]}


@router.get("/export/eligible", response_model=WarrantyListResponse)
def export_eligible_warranties(
    _user: Annotated[Any, Depends(require_permission("warranties.export"))],
    q: str = "",
    marca: str = "",
    proveedor: str = "",
    sucursal: str = "",
    deposito: str = "",
    limit: int = Query(default=500, ge=1, le=1000),
):
    """Devuelve garantías revisadas, pendientes y sin ENV. No exige ubicación física en Chiclana."""
    result = list_warranties(
        _user,
        q=q,
        marca=marca,
        proveedor=proveedor,
        sucursal=sucursal,
        deposito=deposito,
        estado=EXPORT_ELIGIBLE_STATUS,
        review_status=REVIEW_APPROVED,
        demora_min=0,
        limit=limit,
    )
    with db_connect() as conn:
        ensure_warranty_tables(conn)
        ready_codes = export_ready_warranty_codes(conn, [i.id_garantia for i in result.items])
    elegibles = [i for i in result.items if i.id_garantia.strip().upper() in ready_codes]
    return WarrantyListResponse(items=elegibles, total=len(elegibles), limit=result.limit)


@router.post("/export/batch", response_model=WarrantyExportInfo)
def export_warranty_batch(data: WarrantyBatchExportRequest, user: Annotated[Any, Depends(require_permission("warranties.export"))]):
    deny_plain_deposit_operator(user, "exportar ENV")
    """Crea un lote ENV y Excel. No confirma mail/proveedor; deja las garantías en LISTO PARA ENVIAR."""
    proveedor = (data.proveedor or "").strip()
    nota = (data.nota or "").strip()
    formato = normalize_export_format(data.formato)
    logo_brand = normalize_export_logo(data.logo_brand)
    with db_connect() as conn:
        ensure_warranty_tables(conn)
        warranty_codes = validate_export_selection(conn, list(data.warranty_ids))
        shipment_code = next_shipment_code(conn)
        rows = collect_export_rows_by_ids(conn, warranty_codes)
        if not rows:
            raise HTTPException(status_code=400, detail="No se encontraron garantías para los IDs indicados.")
        stamp = now_ar().strftime("%Y%m%d-%H%M%S")
        provider_part = safe_filename_part(proveedor or "lote")
        extension = "pdf" if formato == "pdf" else "xlsx"
        file_name = f"garantias-{provider_part}-{shipment_code}-{stamp}.{extension}"
        file_path = warranty_exports_dir() / file_name
        if formato == "pdf":
            build_provider_pdf(rows, file_path, provider_name=proveedor, shipment_code=shipment_code, logo_brand=logo_brand)
        else:
            build_provider_excel(rows, file_path, provider_name=proveedor, shipment_code=shipment_code, logo_brand=logo_brand)
        now = utc_now_iso()
        cur = conn.execute(
            """
            INSERT INTO guarantee_exports
                (created_at, created_by, created_by_name, provider_name, marca, filters_json, file_path, file_name, row_count, shipment_code, warranty_ids_json, file_format, logo_brand)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                now,
                getattr(user, "username", "") or "",
                getattr(user, "display_name", "") or "",
                proveedor,
                "",
                json.dumps({"warranty_ids": list(data.warranty_ids), "proveedor": proveedor}, ensure_ascii=False),
                str(file_path),
                file_name,
                len(rows),
                shipment_code,
                json.dumps(list(data.warranty_ids), ensure_ascii=False),
                formato,
                logo_brand,
            ),
        )
        export_id = int(cur.lastrowid)
        touched: set[int] = set()
        for row in rows:
            gid = int(row.get("guarantee_id") or 0)
            wcode = str(row.get("warranty_code") or "")
            if not gid or gid in touched:
                continue
            touched.add(gid)
            current_status = str(row.get("status") or "")
            # Pasamos a LISTO PARA ENVIAR. El estado ENVIADO AL PROVEEDOR
            # se registra recién cuando Gestión confirma el mail/ENV enviado.
            new_status = EXPORT_READY_STATUS
            updates: dict[str, Any] = {
                "status": new_status,
                "shipment_code": shipment_code,
                "shipment_file_name": file_name,
                "synced_to_google_sheet": 0,
            }
            if proveedor:
                updates["provider_name"] = proveedor
            set_clause = ", ".join(f"{k} = ?" for k in updates)
            conn.execute(
                f"UPDATE guarantees SET {set_clause}, updated_at = ?, updated_by = ?, updated_by_name = ? WHERE id = ?",
                (*updates.values(), now, getattr(user, "username", "") or "", getattr(user, "display_name", "") or "", gid),
            )
            add_history(
                conn,
                gid,
                wcode,
                user,
                "batch_exported",
                old_status=current_status,
                new_status=new_status,
                note=nota or f"Excel generado para lote {shipment_code}. Pendiente de confirmación de envío.",
                details={"export_id": export_id, "shipment_code": shipment_code, "file_name": file_name, "file_format": formato, "logo_brand": logo_brand},
            )
        conn.commit()
        export_row = conn.execute("SELECT * FROM guarantee_exports WHERE id = ?", (export_id,)).fetchone()
    audit("warranties.export.batch", user=user, resource_type="warranty_export", resource_id=str(export_id), details={"row_count": len(rows), "shipment_code": shipment_code, "file_name": file_name, "file_format": formato, "logo_brand": logo_brand, "warranty_ids": warranty_codes})
    return export_info_from_row(export_row)


@router.post("/export/provider", response_model=WarrantyExportInfo)
def export_warranties_for_provider(data: WarrantyExportRequest, user: Annotated[Any, Depends(require_permission("warranties.export"))]):
    deny_plain_deposit_operator(user, "exportar garantías al proveedor")
    filters = data.model_dump()
    provider_part = safe_filename_part(data.proveedor or data.marca or "proveedor")
    stamp = now_ar().strftime("%Y%m%d-%H%M%S")
    file_name = f"garantias-{provider_part}-{stamp}-{uuid4().hex[:6]}.xlsx"
    file_path = warranty_exports_dir() / file_name
    with db_connect() as conn:
        ensure_warranty_tables(conn)
        rows = collect_export_rows(conn, data)
        build_provider_excel(rows, file_path, provider_name=(data.proveedor or "").strip(), logo_brand="gv_electro")
        now = utc_now_iso()
        cur = conn.execute(
            """
            INSERT INTO guarantee_exports (created_at, created_by, created_by_name, provider_name, marca, filters_json, file_path, file_name, row_count)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                now,
                getattr(user, "username", "") or "",
                getattr(user, "display_name", "") or "",
                (data.proveedor or "").strip(),
                (data.marca or "").strip(),
                json.dumps(filters, ensure_ascii=False),
                str(file_path),
                file_name,
                len(rows),
            ),
        )
        export_id = int(cur.lastrowid)
        touched: set[int] = set()
        for row in rows:
            gid = int(row.get("guarantee_id") or 0)
            if not gid or gid in touched:
                continue
            touched.add(gid)
            add_history(
                conn,
                gid,
                str(row.get("warranty_code") or ""),
                user,
                "excel_exported",
                note="Excel para proveedor generado",
                details={"export_id": export_id, "filters": filters, "file_name": file_name},
            )
        conn.commit()
        export_row = conn.execute("SELECT * FROM guarantee_exports WHERE id = ?", (export_id,)).fetchone()
    audit("warranties.export", user=user, resource_type="warranty_export", resource_id=str(export_id), details={"row_count": len(rows), "filters": filters, "file_name": file_name})
    return export_info_from_row(export_row)


@router.get("/exports", response_model=WarrantyExportListResponse)
def list_warranty_exports(_user: Annotated[Any, Depends(require_permission("warranties.export"))], limit: int = Query(default=50, ge=1, le=200)):
    deny_plain_deposit_operator(_user, "ver exportaciones ENV")
    with db_connect() as conn:
        ensure_warranty_tables(conn)
        rows = conn.execute("SELECT * FROM guarantee_exports ORDER BY id DESC LIMIT ?", (limit,)).fetchall()
    return WarrantyExportListResponse(items=[export_info_from_row(row) for row in rows])


@router.get("/exports/{export_id}/download")
def download_warranty_export(export_id: int, _user: Annotated[Any, Depends(require_permission("warranties.export"))]):
    deny_plain_deposit_operator(_user, "descargar exportaciones ENV")
    with db_connect() as conn:
        ensure_warranty_tables(conn)
        row = conn.execute("SELECT * FROM guarantee_exports WHERE id = ?", (export_id,)).fetchone()
    if not row:
        raise HTTPException(status_code=404, detail="Exportación no encontrada")
    file_path = Path(str(row["file_path"] or ""))
    exports_dir = warranty_exports_dir().resolve()
    try:
        resolved = file_path.resolve()
    except Exception:
        raise HTTPException(status_code=404, detail="Archivo de exportación no disponible")
    if not str(resolved).startswith(str(exports_dir)) or not resolved.exists():
        raise HTTPException(status_code=404, detail="Archivo de exportación no disponible")
    suffix = resolved.suffix.lower()
    media_type = "application/pdf" if suffix == ".pdf" else "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    return FileResponse(
        path=resolved,
        filename=str(row["file_name"] or resolved.name),
        media_type=media_type,
    )



@router.get("/sync/status", response_model=WarrantySyncStatus)
def warranty_sync_status(_user: Annotated[Any, Depends(require_permission("warranties.sync_logs"))]):
    with db_connect() as conn:
        ensure_warranty_tables(conn)
        total = int(conn.execute("SELECT COUNT(*) AS c FROM guarantees").fetchone()["c"] or 0)
        pending = int(conn.execute("SELECT COUNT(*) AS c FROM guarantees WHERE COALESCE(synced_to_google_sheet, 0) = 0").fetchone()["c"] or 0)
        last = conn.execute("SELECT * FROM guarantee_sync_logs ORDER BY id DESC LIMIT 1").fetchone()
        error_rows = conn.execute("SELECT errors_json FROM guarantee_sync_logs WHERE status IN ('failed', 'partial') ORDER BY id DESC LIMIT 3").fetchall()
    errors: list[str] = []
    for row in error_rows:
        try:
            parsed = json.loads(row["errors_json"] or "[]")
            if isinstance(parsed, list):
                errors.extend(str(x) for x in parsed[:5])
        except Exception:
            pass
    return WarrantySyncStatus(
        last_sync_at=format_datetime_ar(parse_iso_datetime(last["finished_at"] or last["started_at"])) if last else "",
        last_sync_type=str(last["sync_type"] or "") if last else "",
        last_sync_status=str(last["status"] or "") if last else "",
        last_sync_user=str(last["actor_name"] or last["actor_username"] or "") if last else "",
        pending_to_sheet=pending,
        total_guarantees=total,
        errors=errors[:10],
    )


@router.get("/sync/logs", response_model=WarrantySyncLogsResponse)
def warranty_sync_logs(_user: Annotated[Any, Depends(require_permission("warranties.sync_logs"))], limit: int = Query(default=30, ge=1, le=200)):
    with db_connect() as conn:
        ensure_warranty_tables(conn)
        rows = conn.execute("SELECT * FROM guarantee_sync_logs ORDER BY id DESC LIMIT ?", (limit,)).fetchall()
    return WarrantySyncLogsResponse(items=[sync_log_info(row) for row in rows])


@router.post("/sync/setup-sheet")
def setup_warranty_sheet(user: Annotated[Any, Depends(require_permission("warranties.sync_to_sheet"))]):
    """Crea/verifica la planilla espejo de garantías.

    La app sigue siendo la fuente principal. Google Sheets queda como espejo/reporting.
    Se mantiene 00_RAW_GARANTIAS por compatibilidad histórica y se agregan pestañas
    normalizadas para garantías, remitos, lotes ENV y eventos.
    """
    spreadsheet_id = require_spreadsheet_id()
    created = ensure_mirror_sheets()
    created_names = [name for name, was_created in created.items() if was_created]
    audit("warranties.sheet.setup", user=user, resource_type="warranty", details={"spreadsheet_id": spreadsheet_id, "sheets": list(created.keys()), "created": created_names})
    return {
        "ok": True,
        "spreadsheet_id": spreadsheet_id,
        "sheet": sheet_raw_name(),
        "tab_created": bool(created_names),
        "headers_count": sum(len(DEFAULT_RAW_HEADERS) if name == sheet_raw_name() else len(MIRROR_SHEETS.get(name, [])) for name in created.keys()),
        "message": f"Planilla espejo verificada. Pestañas listas: {', '.join(created.keys())}." + (f" Creadas: {', '.join(created_names)}." if created_names else ""),
    }


@router.post("/sync/push-to-sheet", response_model=WarrantySyncResult)
def push_warranties_to_sheet(user: Annotated[Any, Depends(require_permission("warranties.sync_to_sheet"))]):
    started = utc_now_iso()
    errors: list[str] = []
    rows_processed = rows_updated = rows_skipped = 0
    sheet_counts: dict[str, int] = {}
    try:
        service = sheets_service()
        spreadsheet_id = require_spreadsheet_id()
        ensure_mirror_sheets()
        with db_connect() as conn:
            ensure_warranty_tables(conn)
            raw_values, row_ranges = warranty_rows_for_sheet(conn)
            # 00_RAW_GARANTIAS se mantiene por compatibilidad. Las pestañas nuevas son espejo normalizado.
            write_sheet_values(service, spreadsheet_id, sheet_raw_name(), DEFAULT_RAW_HEADERS, raw_values)
            rows_processed += len(raw_values)
            sheet_counts[sheet_raw_name()] = len(raw_values)
            for sheet_name, headers in MIRROR_SHEETS.items():
                values = mirror_rows_for_sheet(conn, sheet_name)
                write_sheet_values(service, spreadsheet_id, sheet_name, headers, values)
                rows_processed += len(values)
                sheet_counts[sheet_name] = len(values)

        now = utc_now_iso()
        with db_connect() as conn:
            ensure_warranty_tables(conn)
            # Se marca cada garantía como sincronizada contra el espejo completo.
            guarantee_ids = [int(row["id"]) for row in conn.execute("SELECT id FROM guarantees").fetchall()]
            for gid in guarantee_ids:
                conn.execute(
                    """
                    UPDATE guarantees
                    SET synced_to_google_sheet = 1, last_google_sync_at = ?, sync_error = ''
                    WHERE id = ?
                    """,
                    (now, gid),
                )
            finished = utc_now_iso()
            status_value = "success" if not errors else "partial"
            insert_sync_log(
                conn,
                sync_type="push_to_sheet",
                status_value=status_value,
                started_at=started,
                finished_at=finished,
                user=user,
                rows_processed=rows_processed,
                rows_created=0,
                rows_updated=len(guarantee_ids),
                rows_skipped=rows_skipped,
                errors=errors,
            )
            conn.commit()
        audit("warranties.sync.push", user=user, resource_type="warranty", details={"rows_processed": rows_processed, "sheets": sheet_counts})
        info = [f"{name}: {count} filas" for name, count in sheet_counts.items()]
        return WarrantySyncResult(ok=not errors, sync_type="push_to_sheet", status=status_value, started_at=format_datetime_ar(parse_iso_datetime(started) or now_ar()), finished_at=format_datetime_ar(parse_iso_datetime(finished) or now_ar()), rows_processed=rows_processed, rows_created=0, rows_updated=len(guarantee_ids), rows_skipped=rows_skipped, errors=errors + info)
    except Exception as exc:
        errors.append(str(exc))
        finished = utc_now_iso()
        with db_connect() as conn:
            ensure_warranty_tables(conn)
            insert_sync_log(conn, sync_type="push_to_sheet", status_value="failed", started_at=started, finished_at=finished, user=user, rows_processed=rows_processed, rows_created=0, rows_updated=rows_updated, rows_skipped=rows_skipped, errors=errors)
            conn.commit()
        audit("warranties.sync.push.failed", user=user, resource_type="warranty", details={"errors": errors})
        return WarrantySyncResult(ok=False, sync_type="push_to_sheet", status="failed", started_at=format_datetime_ar(parse_iso_datetime(started) or now_ar()), finished_at=format_datetime_ar(parse_iso_datetime(finished) or now_ar()), rows_processed=rows_processed, rows_created=0, rows_updated=rows_updated, rows_skipped=rows_skipped, errors=errors)


@router.post("/sync/pull-from-sheet", response_model=WarrantySyncResult)
def pull_warranties_from_sheet(user: Annotated[Any, Depends(require_permission("warranties.sync_from_sheet"))]):
    started = utc_now_iso()
    errors: list[str] = []
    created = updated = skipped = processed = 0
    try:
        headers, raw_rows = read_raw_sheet_rows()
        index = make_header_index(headers)
        grouped: dict[str, list[dict[str, str]]] = {}
        for raw in raw_rows:
            code = get_sheet_cell(raw, index, "ID GARANTIA")
            if not code:
                continue
            record = {header: get_sheet_cell(raw, index, header) for header in DEFAULT_RAW_HEADERS}
            grouped.setdefault(code, []).append(record)
        with db_connect() as conn:
            ensure_warranty_tables(conn)
            for code, rows in grouped.items():
                processed += len(rows)
                action, message = import_sheet_group(conn, code, rows, user)
                if action == "created":
                    created += 1
                elif action == "skipped":
                    skipped += 1
                elif action == "conflict":
                    skipped += 1
                    errors.append(message)
                elif action == "updated":
                    updated += 1
            finished = utc_now_iso()
            status_value = "success" if not errors else "partial"
            insert_sync_log(conn, sync_type="pull_from_sheet", status_value=status_value, started_at=started, finished_at=finished, user=user, rows_processed=processed, rows_created=created, rows_updated=updated, rows_skipped=skipped, errors=errors)
            conn.commit()
        audit("warranties.sync.pull", user=user, resource_type="warranty", details={"rows_processed": processed, "created": created, "skipped": skipped, "errors": errors[:20]})
        return WarrantySyncResult(ok=not errors, sync_type="pull_from_sheet", status=status_value, started_at=format_datetime_ar(parse_iso_datetime(started) or now_ar()), finished_at=format_datetime_ar(parse_iso_datetime(finished) or now_ar()), rows_processed=processed, rows_created=created, rows_updated=updated, rows_skipped=skipped, errors=errors[:50])
    except Exception as exc:
        errors.append(str(exc))
        finished = utc_now_iso()
        with db_connect() as conn:
            ensure_warranty_tables(conn)
            insert_sync_log(conn, sync_type="pull_from_sheet", status_value="failed", started_at=started, finished_at=finished, user=user, rows_processed=processed, rows_created=created, rows_updated=updated, rows_skipped=skipped, errors=errors)
            conn.commit()
        audit("warranties.sync.pull.failed", user=user, resource_type="warranty", details={"errors": errors})
        return WarrantySyncResult(ok=False, sync_type="pull_from_sheet", status="failed", started_at=format_datetime_ar(parse_iso_datetime(started) or now_ar()), finished_at=format_datetime_ar(parse_iso_datetime(finished) or now_ar()), rows_processed=processed, rows_created=created, rows_updated=updated, rows_skipped=skipped, errors=errors)



# =========================================================
# Dashboard y métricas
# =========================================================

FINAL_STATUS_LABELS = [
    "10 - FINALIZADO",
]


def status_equals(value: Any, expected: str) -> bool:
    return status_matches(value, expected)


def is_rejected_status(value: Any) -> bool:
    return "RECHAZADO" in normalize_text(value)


def is_final_status(value: Any) -> bool:
    return any(status_matches(value, status_value) for status_value in FINAL_STATUS_LABELS)


def dashboard_date_key(iso_value: str) -> str:
    dt = parse_iso_datetime(iso_value)
    if not dt:
        return "Sin fecha"
    local = dt.astimezone(ZoneInfo("America/Argentina/Buenos_Aires"))
    return local.strftime("%Y-%m")


def avg(values: list[int | float]) -> float:
    clean = [float(v) for v in values if v is not None]
    if not clean:
        return 0
    return round(sum(clean) / len(clean), 1)


def ordered_points(counter: dict[str, int | float], *, limit: int | None = None, preferred_order: list[str] | None = None) -> list[WarrantyDashboardPoint]:
    if preferred_order:
        ordered: list[tuple[str, int | float]] = [(label, counter.get(label, 0)) for label in preferred_order if counter.get(label, 0)]
        rest = [(label, value) for label, value in counter.items() if label not in preferred_order and value]
        rest.sort(key=lambda pair: float(pair[1]), reverse=True)
        data = ordered + rest
    else:
        data = sorted(counter.items(), key=lambda pair: float(pair[1]), reverse=True)
    if limit is not None:
        data = data[:limit]
    return [WarrantyDashboardPoint(label=str(label), value=float(value)) for label, value in data]


def delay_range_label(days: int) -> str:
    if days <= 3:
        return "0 a 3 días"
    if days <= 7:
        return "4 a 7 días"
    if days <= 14:
        return "8 a 14 días"
    if days <= 30:
        return "15 a 30 días"
    return "Más de 30 días"


def dashboard_matches(row: sqlite3.Row, items: list[sqlite3.Row], filters: dict[str, Any]) -> bool:
    date_from = parse_date_filter(filters.get("fecha_desde"))
    date_to = parse_date_filter(filters.get("fecha_hasta"))
    ingreso_dt = parse_iso_datetime(row["ingreso_at"] or row["created_at"])
    ingreso_date = ingreso_dt.date() if ingreso_dt else None
    if date_from and ingreso_date and ingreso_date < date_from:
        return False
    if date_to and ingreso_date and ingreso_date > date_to:
        return False
    if filters.get("estado") and not status_matches(row["status"], filters.get("estado")):
        return False
    if filters.get("sucursal") and normalize_text(row["sucursal"]) != normalize_text(filters.get("sucursal")):
        return False
    if filters.get("deposito"):
        wanted = normalize_text(filters.get("deposito"))
        if wanted not in {normalize_text(row["deposito"]), normalize_text(row["lugar_llegada"])}:
            return False
    if filters.get("proveedor") and normalize_text(row["provider_name"]) != normalize_text(filters.get("proveedor")):
        return False
    if filters.get("marca"):
        wanted = normalize_text(filters.get("marca"))
        if not any(normalize_text(item["marca"]) == wanted for item in items):
            return False
    return True


@router.get("/config", response_model=WarrantyConfigResponse)
def get_warranty_config(_user: Annotated[Any, Depends(require_permission("warranties.config"))]):
    values = warranty_config_values()
    with db_connect() as conn:
        ensure_warranty_tables(conn)
        providers_count = int((conn.execute("SELECT COUNT(*) AS total FROM providers WHERE is_active = 1").fetchone() or {"total": 0})["total"]) if _table_exists(conn, "providers") else 0
        brands_count = int((conn.execute("SELECT COUNT(*) AS total FROM product_brands WHERE is_active = 1").fetchone() or {"total": 0})["total"]) if _table_exists(conn, "product_brands") else 0
        mapped_brands_count = int((conn.execute("SELECT COUNT(DISTINCT brand_id) AS total FROM brand_providers").fetchone() or {"total": 0})["total"]) if _table_exists(conn, "brand_providers") else 0
        pending_review_count = int(conn.execute("SELECT COUNT(*) AS total FROM guarantees WHERE review_status != ? AND cancelled = 0", (REVIEW_APPROVED,)).fetchone()["total"])
        active_count = int(conn.execute("SELECT COUNT(*) AS total FROM guarantees WHERE cancelled = 0").fetchone()["total"])
    return WarrantyConfigResponse(
        config=WarrantyConfigCatalog(
            statuses=values["statuses"],
            final_statuses=values["final_statuses"],
            sucursales=values["sucursales"],
            depositos=values["depositos"],
            delay_ranges=values["delay_ranges"],
            required_review_fields=values["required_review_fields"],
            sheet_raw=values["raw_sheet"],
            spreadsheet_url=values["spreadsheet_url"],
        ),
        providers_count=providers_count,
        brands_count=brands_count,
        mapped_brands_count=mapped_brands_count,
        unmapped_brands_count=max(0, brands_count - mapped_brands_count),
        pending_review_count=pending_review_count,
        active_count=active_count,
    )


@router.patch("/config", response_model=WarrantyConfigResponse)
def save_warranty_config(data: WarrantyConfigSaveRequest, user: Annotated[Any, Depends(require_permission("warranties.config"))]):
    root = load_operational_config()
    warranties_cfg = root.get("warranties", {}) if isinstance(root.get("warranties"), dict) else {}
    if data.statuses is not None:
        clean = [x.strip() for x in data.statuses if str(x).strip()]
        warranties_cfg["statuses"] = unique_keep_order(clean or DEFAULT_STATUSES)
        warranties_cfg["estados"] = warranties_cfg["statuses"]
    if data.final_statuses is not None:
        warranties_cfg["final_statuses"] = unique_keep_order([x.strip() for x in data.final_statuses if str(x).strip()] or DEFAULT_FINAL_STATUSES)
    if data.sucursales is not None:
        warranties_cfg["sucursales"] = unique_keep_order([x.strip() for x in data.sucursales if str(x).strip()] or DEFAULT_SUCURSALES)
    if data.depositos is not None:
        warranties_cfg["depositos"] = unique_keep_order([x.strip() for x in data.depositos if str(x).strip()] or DEFAULT_DEPOSITOS)
    if data.delay_ranges is not None:
        values = sorted({int(x) for x in data.delay_ranges if int(x) > 0})
        warranties_cfg["delay_ranges"] = values or DEFAULT_DELAY_RANGES
    if data.required_review_fields is not None:
        warranties_cfg["required_review_fields"] = unique_keep_order([x.strip() for x in data.required_review_fields if str(x).strip()] or DEFAULT_REQUIRED_REVIEW_FIELDS)
    if data.raw_sheet is not None:
        warranties_cfg["raw_sheet"] = data.raw_sheet.strip() or "00_RAW_GARANTIAS"
    if data.spreadsheet_url is not None:
        warranties_cfg["spreadsheet_url"] = data.spreadsheet_url.strip()
    root["warranties"] = warranties_cfg
    save_operational_config(root, updated_by=getattr(user, "username", "system") or "system")
    audit("warranties.config.save", user=user, resource_type="warranty_config", resource_id="system", details={"section": "warranties"})
    return get_warranty_config(user)


def _table_exists(conn: sqlite3.Connection, name: str) -> bool:
    row = conn.execute("SELECT name FROM sqlite_master WHERE type='table' AND name=?", (name,)).fetchone()
    return bool(row)




def _count_value(conn: sqlite3.Connection, sql: str, params: tuple[Any, ...] = ()) -> int:
    row = conn.execute(sql, params).fetchone()
    if not row:
        return 0
    return int(row[0] or 0)


def _diagnostic_item(key: str, label: str, status_value: str, detail: str, count: int = 0) -> WarrantyDiagnosticItem:
    return WarrantyDiagnosticItem(key=key, label=label, status=status_value, detail=detail, count=count)


@router.get("/diagnostics", response_model=WarrantyDiagnosticsResponse)
def warranty_diagnostics(actor: Annotated[Any, Depends(require_permission("warranties.dashboard"))]) -> WarrantyDiagnosticsResponse:
    """Resumen de cierre operativo del módulo Garantías.

    No modifica datos. Sirve para detectar puntos pendientes antes de usar el flujo completo
    en producción: catálogo, proveedores, revisión, sincronización y configuración.
    """
    cfg = runtime_warranty_config()
    product_cfg = runtime_product_catalog_config()
    items: list[WarrantyDiagnosticItem] = []
    next_actions: list[str] = []
    with db_connect() as conn:
        ensure_product_catalog_tables(conn)
        guarantee_rows = conn.execute("SELECT * FROM guarantees WHERE cancelled = 0").fetchall()
        active_guarantees = len(guarantee_rows)
        pending_review = sum(1 for row in guarantee_rows if status_matches(row["status"], DEFAULT_STATUSES[0]) or review_status_matches(row["review_status"], REVIEW_PENDING))
        needs_correction = sum(1 for row in guarantee_rows if review_status_matches(row["review_status"], REVIEW_INCOMPLETE))
        pending_provider = sum(1 for row in guarantee_rows if status_matches(row["status"], "2 - PENDIENTE"))
        sent_without_case = sum(
            1
            for row in guarantee_rows
            if (status_matches(row["status"], "4 - ENVIADO AL PROVEEDOR") or status_matches(row["status"], "5 - EN EL PROVEEDOR"))
            and not str(row["provider_case_id"] or "").strip()
        )
        pending_sync = sum(1 for row in guarantee_rows if int(row["synced_to_google_sheet"] or 0) == 0)
        products_count = _count_value(conn, "SELECT COUNT(*) FROM products WHERE is_active = 1")
        providers_count = _count_value(conn, "SELECT COUNT(*) FROM providers WHERE is_active = 1")
        brands_count = _count_value(conn, "SELECT COUNT(*) FROM product_brands WHERE is_active = 1")
        unmapped_brands = _count_value(conn, """
            SELECT COUNT(*)
            FROM product_brands b
            WHERE b.is_active = 1
              AND NOT EXISTS (
                SELECT 1
                FROM brand_providers bp
                JOIN providers p ON p.id = bp.provider_id AND p.is_active = 1
                WHERE bp.brand_id = b.id
              )
        """)
        deposit_branches = conn.execute(
            "SELECT id, name, code FROM branches WHERE is_active = 1 AND type = 'deposit' ORDER BY name COLLATE NOCASE"
        ).fetchall() if _table_exists(conn, "branches") else []
        physical_branches_count = _count_value(conn, "SELECT COUNT(*) FROM branches WHERE is_active = 1 AND type = 'physical'") if _table_exists(conn, "branches") else 0
        missing_org_fields = _count_value(conn, """
            SELECT COUNT(*) FROM guarantees
            WHERE cancelled = 0 AND (
                COALESCE(branch_id, '') = ''
                OR COALESCE(company_id, '') = ''
                OR COALESCE(tipo_ingreso, '') = ''
                OR COALESCE(origen_ingreso, '') = ''
                OR COALESCE(ubicacion_actual, '') = ''
            )
        """)
        deposit_without_responsible = _count_value(conn, """
            SELECT COUNT(*) FROM guarantees
            WHERE cancelled = 0
              AND tipo_ingreso = 'cliente_deposito'
              AND COALESCE(sucursal_responsable_id, '') = ''
              AND COALESCE(sucursal_responsable, '') = ''
        """)
        env_with_remito = _count_value(conn, """
            SELECT COUNT(*) FROM guarantees
            WHERE cancelled = 0
              AND COALESCE(shipment_code, '') <> ''
              AND COALESCE(remito_interno, '') <> ''
        """)
        deposito_disponible_remito_risk = _count_value(conn, """
            SELECT COUNT(*) FROM guarantees
            WHERE cancelled = 0
              AND origen_ingreso = 'deposito'
              AND COALESCE(remito_interno, '') <> ''
        """)
    products_ok = products_count > 0
    providers_ok = providers_count > 0
    sheet_ok = bool(cfg.get("spreadsheet_id") or cfg.get("spreadsheet_url"))
    product_sheet_ok = bool(product_cfg.get("spreadsheet_id") or product_cfg.get("spreadsheet_url"))
    items.append(_diagnostic_item(
        "product_catalog",
        "Catálogo de productos",
        "ok" if products_ok else "warning",
        f"{products_count} productos activos disponibles para búsquedas y autocompletado.",
        products_count,
    ))
    if not products_ok:
        next_actions.append("Actualizar el catálogo desde la Planilla Madre de Ventas.")
    items.append(_diagnostic_item(
        "product_source",
        "Fuente de productos",
        "ok" if product_sheet_ok else "warning",
        f"Hoja configurada: {product_cfg.get('sheet_name') or 'sin configurar'}.",
        products_count,
    ))
    if not product_sheet_ok:
        next_actions.append("Configurar la Planilla Madre de Ventas en Configuración operativa > Productos.")
    items.append(_diagnostic_item(
        "providers",
        "Proveedores",
        "ok" if providers_ok else "warning",
        f"{providers_count} proveedores activos cargados.",
        providers_count,
    ))
    if not providers_ok:
        next_actions.append("Cargar proveedores y vincularlos con marcas principales.")
    items.append(_diagnostic_item(
        "brand_mapping",
        "Marcas vinculadas a proveedor",
        "ok" if unmapped_brands == 0 and brands_count > 0 else "warning",
        f"{unmapped_brands} marcas activas sin proveedor asignado sobre {brands_count} marcas detectadas.",
        unmapped_brands,
    ))
    if unmapped_brands:
        next_actions.append("Completar la relación Marca → Proveedor en Productos y proveedores.")
    items.append(_diagnostic_item(
        "review_queue",
        "Revisión interna",
        "ok" if pending_review == 0 and needs_correction == 0 else "warning",
        f"{pending_review} en ingreso y {needs_correction} requieren corrección.",
        pending_review + needs_correction,
    ))
    if pending_review or needs_correction:
        next_actions.append("Revisar la bandeja de garantías antes de enviarlas a proveedor.")
    items.append(_diagnostic_item(
        "provider_management",
        "Gestión con proveedor",
        "ok" if sent_without_case == 0 else "warning",
        f"{pending_provider} listas para proveedor. {sent_without_case} enviadas/en revisión sin ID de caso.",
        pending_provider + sent_without_case,
    ))
    if sent_without_case:
        next_actions.append("Completar ID de caso en garantías ya enviadas o en revisión.")
    items.append(_diagnostic_item(
        "sheet_sync",
        "Sincronización Google Sheet",
        "ok" if pending_sync == 0 else "warning",
        f"{pending_sync} garantías activas pendientes de sincronizar.",
        pending_sync,
    ))
    if pending_sync:
        next_actions.append("Actualizar Google Sheet desde el panel de sincronización cuando corresponda.")
    items.append(_diagnostic_item(
        "sheet_config",
        "Google Sheet de garantías",
        "ok" if sheet_ok else "warning",
        f"Hoja raw: {cfg.get('raw_sheet') or 'sin hoja configurada'}.",
        active_guarantees,
    ))
    if not sheet_ok:
        next_actions.append("Configurar la Google Sheet de garantías en Configuración operativa > Garantías.")

    deposit_names = ", ".join(str(r["name"] or r["code"] or "") for r in deposit_branches) or "sin depósitos activos"
    deposits_ok = len(deposit_branches) >= 3
    items.append(_diagnostic_item(
        "org_deposits",
        "Depósitos como branches",
        "ok" if deposits_ok else "warning",
        f"{len(deposit_branches)} depósitos activos detectados: {deposit_names}.",
        len(deposit_branches),
    ))
    if not deposits_ok:
        next_actions.append("Crear/asignar Chiclana, Corrales y Cachi como branches type=deposit.")

    items.append(_diagnostic_item(
        "org_physical_branches",
        "Sucursales físicas activas",
        "ok" if physical_branches_count > 0 else "warning",
        f"{physical_branches_count} sucursales físicas activas disponibles para sucursal_responsable.",
        physical_branches_count,
    ))
    if not physical_branches_count:
        next_actions.append("Revisar la configuración de sucursales físicas en Organización.")

    items.append(_diagnostic_item(
        "org_truth_fields",
        "Fuente de verdad organizativa",
        "ok" if missing_org_fields == 0 else "warning",
        f"{missing_org_fields} garantías activas tienen incompletos branch_id/company_id/tipo/origen/ubicación.",
        missing_org_fields,
    ))
    if missing_org_fields:
        next_actions.append("Revisar/backfillear garantías heredadas con campos organizativos nuevos.")

    items.append(_diagnostic_item(
        "org_deposit_responsible",
        "Cliente en depósito con sucursal responsable",
        "ok" if deposit_without_responsible == 0 else "warning",
        f"{deposit_without_responsible} garantías cliente_deposito no tienen sucursal responsable.",
        deposit_without_responsible,
    ))
    if deposit_without_responsible:
        next_actions.append("Completar sucursal responsable en garantías cargadas desde depósito por cliente.")

    items.append(_diagnostic_item(
        "flow_env_remito_overlap",
        "ENV y remito interno superpuestos",
        "ok" if env_with_remito == 0 else "warning",
        f"{env_with_remito} garantías tienen a la vez shipment_code y remito_interno. Puede ser válido, pero revisar que no se estén confundiendo flujos.",
        env_with_remito,
    ))
    if env_with_remito:
        next_actions.append("Auditar que REM sea traslado interno y ENV sea lote proveedor, sin usar uno como sustituto del otro.")

    items.append(_diagnostic_item(
        "flow_deposit_remito_risk",
        "Remitos en ingresos de depósito",
        "ok" if deposito_disponible_remito_risk == 0 else "warning",
        f"{deposito_disponible_remito_risk} garantías con origen depósito tienen remito interno asignado.",
        deposito_disponible_remito_risk,
    ))
    if deposito_disponible_remito_risk:
        next_actions.append("Revisar remitos asociados a garantías que ya ingresaron en depósito.")

    # Fase 29 — auditoría pre-producción del flujo real de garantías.
    # No modifica datos: detecta registros que podrían romper el flujo antes de salir a producción.
    canonical_statuses = {canonical_status_key(label) for label in DEFAULT_STATUSES}
    canonical_review_statuses = {
        canonical_status_key(REVIEW_PENDING),
        canonical_status_key(REVIEW_IN_PROGRESS),
        canonical_status_key(REVIEW_INCOMPLETE),
        canonical_status_key(REVIEW_APPROVED),
    }
    invalid_status = sum(
        1 for row in guarantee_rows
        if canonical_status_key(normalize_status(row["status"])) not in canonical_statuses
    )
    invalid_review_status = sum(
        1 for row in guarantee_rows
        if canonical_status_key(row["review_status"] or REVIEW_PENDING) not in canonical_review_statuses
    )
    pendiente_sin_revision = sum(
        1 for row in guarantee_rows
        if status_matches(row["status"], "2 - PENDIENTE")
        and not review_status_matches(row["review_status"], REVIEW_APPROVED)
    )
    revisada_en_ingreso = sum(
        1 for row in guarantee_rows
        if status_matches(row["status"], "1 - INGRESO")
        and review_status_matches(row["review_status"], REVIEW_APPROVED)
    )
    listo_sin_env = sum(
        1 for row in guarantee_rows
        if status_matches(row["status"], "3 - LISTO PARA ENVIAR")
        and not str(row["shipment_code"] or "").strip()
    )
    enviado_sin_mail = sum(
        1 for row in guarantee_rows
        if status_matches(row["status"], "4 - ENVIADO AL PROVEEDOR")
        and not (str(row["fecha_ultimo_mail_proveedor"] or "").strip() or str(row["sent_to_provider_at"] or "").strip())
    )
    retiro_solicitado_urgente = sum(
        1 for row in guarantee_rows
        if normalize_text(row["estado_retiro_proveedor"] or "") in {"RETIRO_SOLICITADO", "SOLICITADO"}
        and not internal_logistics_ready_for_provider(row)
    )
    retiro_solicitado_listo = sum(
        1 for row in guarantee_rows
        if normalize_text(row["estado_retiro_proveedor"] or "") in {"RETIRO_SOLICITADO", "SOLICITADO", "LISTO_PARA_RETIRO"}
        and internal_logistics_ready_for_provider(row)
        and not status_matches(row["status"], "5 - EN EL PROVEEDOR")
    )
    en_proveedor_sin_fecha = sum(
        1 for row in guarantee_rows
        if status_matches(row["status"], "5 - EN EL PROVEEDOR")
        and not (str(row["fecha_retiro_proveedor"] or "").strip() or str(row["fecha_retiro"] or "").strip())
    )
    resuelta_sin_resultado = sum(
        1 for row in guarantee_rows
        if status_matches(row["status"], "7 - RESUELTO")
        and not str(row["resultado_resolucion"] or "").strip()
    )
    finalizada_sin_cierre = sum(
        1 for row in guarantee_rows
        if status_matches(row["status"], "10 - FINALIZADO")
        and not (str(row["fecha_finalizacion"] or "").strip() or str(row["finalizacion"] or "").strip())
    )
    remitos_en_transito = 0
    remitos_con_env = 0
    if _table_exists(conn, "warranty_remitos"):
        remitos_en_transito = _count_value(conn, """
            SELECT COUNT(*) FROM warranty_remitos
            WHERE status IN ('en_transito', 'despachado')
        """)
        remitos_con_env = _count_value(conn, """
            SELECT COUNT(*) FROM warranty_remitos
            WHERE COALESCE(shipment_code, '') <> ''
        """)

    items.append(_diagnostic_item(
        "flow_canonical_statuses",
        "Estados canónicos",
        "ok" if invalid_status == 0 and invalid_review_status == 0 else "error",
        f"{invalid_status} garantías con estado no canónico y {invalid_review_status} con review_status no canónico.",
        invalid_status + invalid_review_status,
    ))
    if invalid_status or invalid_review_status:
        next_actions.append("Normalizar estados/review_status heredados antes de cargar datos reales.")

    state_logic_issues = pendiente_sin_revision + revisada_en_ingreso + listo_sin_env + enviado_sin_mail
    items.append(_diagnostic_item(
        "flow_state_coherence",
        "Coherencia estado → acción",
        "ok" if state_logic_issues == 0 else "warning",
        f"{pendiente_sin_revision} pendientes sin revisión aprobada, {revisada_en_ingreso} revisadas en ingreso, {listo_sin_env} listas sin ENV y {enviado_sin_mail} enviadas sin fecha de mail.",
        state_logic_issues,
    ))
    if state_logic_issues:
        next_actions.append("Revisar garantías que no respetan la secuencia INGRESO → PENDIENTE → ENV → MAIL.")

    provider_flow_issues = retiro_solicitado_urgente + en_proveedor_sin_fecha + resuelta_sin_resultado + finalizada_sin_cierre
    items.append(_diagnostic_item(
        "flow_provider_followup",
        "Seguimiento proveedor",
        "ok" if provider_flow_issues == 0 else "warning",
        f"{retiro_solicitado_urgente} retiros solicitados sin producto listo, {retiro_solicitado_listo} listos para retiro, {en_proveedor_sin_fecha} en proveedor sin fecha, {resuelta_sin_resultado} resueltas sin resolución y {finalizada_sin_cierre} finalizadas sin cierre.",
        provider_flow_issues + retiro_solicitado_listo,
    ))
    if retiro_solicitado_urgente:
        next_actions.append("Priorizar remitos urgentes hacia Chiclana para retiros solicitados por proveedor.")
    if retiro_solicitado_listo:
        next_actions.append("Coordinar retiro proveedor para garantías listas físicamente en Chiclana.")
    if resuelta_sin_resultado:
        next_actions.append("Completar si la resolución fue Nota de crédito, Reparación o Cambio de equipo.")

    items.append(_diagnostic_item(
        "flow_internal_remitos",
        "Remitos internos",
        "ok" if remitos_con_env == 0 else "error",
        f"{remitos_en_transito} remitos en tránsito y {remitos_con_env} remitos con shipment_code/ENV asociado.",
        remitos_en_transito + remitos_con_env,
    ))
    if remitos_con_env:
        next_actions.append("Corregir remitos con shipment_code: REM y ENV deben seguir separados.")

    status_value = "ok"
    if any(item.status == "error" for item in items):
        status_value = "error"
    elif any(item.status == "warning" for item in items):
        status_value = "warning"
    return WarrantyDiagnosticsResponse(status=status_value, generated_at=format_datetime_ar(), items=items, next_actions=next_actions[:8])


@router.get("/dashboard", response_model=WarrantyDashboardResponse)
def warranty_dashboard(
    _user: Annotated[Any, Depends(require_permission("warranties.dashboard"))],
    fecha_desde: str = "",
    fecha_hasta: str = "",
    marca: str = "",
    proveedor: str = "",
    sucursal: str = "",
    deposito: str = "",
    estado: str = "",
):
    filters = {
        "fecha_desde": fecha_desde,
        "fecha_hasta": fecha_hasta,
        "marca": marca,
        "proveedor": proveedor,
        "sucursal": sucursal,
        "deposito": deposito,
        "estado": estado,
    }
    with db_connect() as conn:
        ensure_warranty_tables(conn)
        rows = conn.execute("SELECT * FROM guarantees WHERE COALESCE(cancelled, 0) = 0 ORDER BY ingreso_at DESC, id DESC").fetchall()
        all_items = conn.execute("SELECT * FROM guarantee_items ORDER BY id").fetchall()
    by_gid: dict[int, list[sqlite3.Row]] = {}
    for item in all_items:
        by_gid.setdefault(int(item["guarantee_id"]), []).append(item)
    selected: list[tuple[sqlite3.Row, list[sqlite3.Row]]] = []
    for row in rows:
        items = by_gid.get(int(row["id"]), [])
        if dashboard_matches(row, items, filters):
            selected.append((row, items))

    by_status: dict[str, int] = {}
    by_brand: dict[str, int] = {}
    by_provider: dict[str, int] = {}
    by_branch: dict[str, int] = {}
    by_deposit: dict[str, int] = {}
    by_delay_range: dict[str, int] = {"0 a 3 días": 0, "4 a 7 días": 0, "8 a 14 días": 0, "15 a 30 días": 0, "Más de 30 días": 0}
    monthly: dict[str, int] = {}
    final_resolutions: dict[str, int] = {}
    resolution_by_provider_values: dict[str, list[int]] = {}

    pending_days_values: list[int] = []
    resolution_days_values: list[int] = []
    no_response_values: list[int] = []
    critical_candidates: list[tuple[int, WarrantySummary]] = []

    ingreso_count = review_pending = pending_provider = sent_provider = in_review = resolved = rejected = delayed_7 = delayed_15 = 0

    for row, items in selected:
        status_value = str(row["status"] or "")
        by_status[status_value or "Sin estado"] = by_status.get(status_value or "Sin estado", 0) + 1
        by_branch[str(row["sucursal"] or "Sin sucursal")] = by_branch.get(str(row["sucursal"] or "Sin sucursal"), 0) + 1
        by_deposit[str(row["deposito"] or "Sin depósito")] = by_deposit.get(str(row["deposito"] or "Sin depósito"), 0) + 1
        provider_label = str(row["provider_name"] or "Sin proveedor")
        by_provider[provider_label] = by_provider.get(provider_label, 0) + 1
        monthly_key = dashboard_date_key(str(row["ingreso_at"] or row["created_at"] or ""))
        monthly[monthly_key] = monthly.get(monthly_key, 0) + 1
        for item in items:
            brand = str(item["marca"] or "Sin marca")
            by_brand[brand] = by_brand.get(brand, 0) + 1

        pending_days = compute_pending_days(row)
        pending_days_values.append(pending_days)
        if not is_final_status(status_value):
            by_delay_range[delay_range_label(pending_days)] = by_delay_range.get(delay_range_label(pending_days), 0) + 1
        no_response = compute_no_response_days(row)
        if no_response is not None:
            no_response_values.append(no_response)
        if status_equals(status_value, "1 - INGRESO"):
            ingreso_count += 1
        if str(row["review_status"] or REVIEW_PENDING) != REVIEW_APPROVED:
            review_pending += 1
        if status_equals(status_value, "2 - PENDIENTE"):
            pending_provider += 1
        if status_equals(status_value, "4 - ENVIADO AL PROVEEDOR"):
            sent_provider += 1
        if status_equals(status_value, "5 - EN EL PROVEEDOR"):
            in_review += 1
        if is_final_status(status_value):
            final_resolutions[status_value] = final_resolutions.get(status_value, 0) + 1
            if is_rejected_status(status_value):
                rejected += 1
            else:
                resolved += 1
            resolution_days = compute_pending_days(row)
            resolution_days_values.append(resolution_days)
            resolution_by_provider_values.setdefault(provider_label, []).append(resolution_days)
        if not is_final_status(status_value) and pending_days >= 7:
            delayed_7 += 1
        if not is_final_status(status_value) and pending_days >= 15:
            delayed_15 += 1
        urgency_score = max(pending_days, int(no_response or 0))
        if urgency_score >= 7 and not is_final_status(status_value):
            critical_candidates.append((urgency_score, row_to_summary(row, items)))

    resolution_provider_avg = {label: avg(values) for label, values in resolution_by_provider_values.items() if values}
    critical_candidates.sort(key=lambda pair: pair[0], reverse=True)
    monthly_points = [WarrantyDashboardPoint(label=label, value=float(monthly[label])) for label in sorted(monthly.keys())]
    return WarrantyDashboardResponse(
        metrics=WarrantyDashboardMetrics(
            total=len(selected),
            ingreso=ingreso_count,
            pendientes_revision=review_pending,
            pendientes_proveedor=pending_provider,
            enviadas_proveedor=sent_provider,
            en_revision=in_review,
            resueltas=resolved,
            rechazadas=rejected,
            demoradas_7=delayed_7,
            demoradas_15=delayed_15,
            promedio_dias_pendiente=avg(pending_days_values),
            promedio_resolucion=avg(resolution_days_values),
            promedio_dias_sin_respuesta=avg(no_response_values),
        ),
        by_status=ordered_points(by_status, preferred_order=DEFAULT_STATUSES),
        by_brand=ordered_points(by_brand, limit=10),
        by_provider=ordered_points(by_provider, limit=10),
        by_branch=ordered_points(by_branch),
        by_deposit=ordered_points(by_deposit),
        by_delay_range=ordered_points(by_delay_range, preferred_order=["0 a 3 días", "4 a 7 días", "8 a 14 días", "15 a 30 días", "Más de 30 días"]),
        monthly_entries=monthly_points[-12:],
        avg_resolution_by_provider=ordered_points(resolution_provider_avg, limit=10),
        final_resolutions=ordered_points(final_resolutions, preferred_order=FINAL_STATUS_LABELS),
        critical=[summary for _score, summary in critical_candidates[:15]],
        filters={key: value for key, value in filters.items() if value},
    )


@router.post("/{warranty_id}/cancel", response_model=WarrantyDetailResponse)
def cancel_warranty(warranty_id: str, data: WarrantyCancelRequest, user: Annotated[Any, Depends(require_permission("warranties.cancel"))]):
    reason = data.reason.strip()
    with db_connect() as conn:
        result = fetch_guarantee_with_items(conn, warranty_id)
        if not result:
            raise HTTPException(status_code=404, detail="Garantía no encontrada")
        row, _items = result
        if int(row["cancelled"] or 0):
            raise HTTPException(status_code=400, detail="La garantía ya se encuentra anulada.")
        old_status = str(row["status"] or "")
        now = utc_now_iso()
        conn.execute(
            """
            UPDATE guarantees
            SET cancelled = 1, cancel_reason = ?, cancelled_by = ?, cancelled_at = ?, status = ?,
                updated_at = ?, updated_by = ?, updated_by_name = ?, synced_to_google_sheet = 0
            WHERE id = ?
            """,
            (
                reason,
                getattr(user, "username", "") or "",
                now,
                CANCELLED_STATUS,
                now,
                getattr(user, "username", "") or "",
                getattr(user, "display_name", "") or "",
                int(row["id"]),
            ),
        )
        add_history(conn, int(row["id"]), str(row["warranty_code"]), user, "cancelled", old_status=old_status, new_status=CANCELLED_STATUS, note=reason)
        conn.commit()
    audit("warranties.cancel", user=user, resource_type="warranty", resource_id=warranty_id, details={"reason": reason})
    return get_warranty_detail(warranty_id, user)


@router.delete("/{warranty_id}")
def delete_warranty(warranty_id: str, user: Annotated[Any, Depends(require_permission("warranties.delete"))]):
    """Eliminación definitiva para correcciones de carga/pruebas.

    La anulación sigue siendo la acción recomendada para casos reales.
    Esta acción queda registrada en auditoría global antes de borrar los datos del módulo.
    """
    with db_connect() as conn:
        result = fetch_guarantee_with_items(conn, warranty_id)
        if not result:
            raise HTTPException(status_code=404, detail="Garantía no encontrada")
        row, items = result
        guarantee_id = int(row["id"])
        snapshot = {
            "warranty_code": str(row["warranty_code"] or warranty_id),
            "status": str(row["status"] or ""),
            "review_status": str(row["review_status"] or ""),
            "items": len(items),
            "responsible": str(row["responsible_name"] or row["responsible_username"] or ""),
        }
        conn.execute("DELETE FROM guarantee_items WHERE guarantee_id = ?", (guarantee_id,))
        conn.execute("DELETE FROM guarantee_history WHERE guarantee_id = ?", (guarantee_id,))
        conn.execute("DELETE FROM guarantees WHERE id = ?", (guarantee_id,))
        conn.commit()
    audit("warranties.delete", user=user, resource_type="warranty", resource_id=warranty_id, details=snapshot)
    return {"ok": True, "deleted": warranty_id}




def _entry_base_edit_allowed(row: sqlite3.Row, user: Any) -> bool:
    """Permite corrección de base solo antes de que el caso avance."""
    if user.has("warranties.manage") or user.has("warranties.manage_provider"):
        return True
    if int(row["cancelled"] or 0):
        return False
    status_ok = status_matches(str(row["status"] or ""), DEFAULT_STATUSES[0])
    review_ok = review_status_matches(str(row["review_status"] or REVIEW_PENDING), REVIEW_PENDING) or review_status_matches(str(row["review_status"] or REVIEW_PENDING), REVIEW_INCOMPLETE)
    if not (status_ok and review_ok):
        return False
    username = str(getattr(user, "username", "") or "")
    user_branch_id = str(getattr(user, "branch_id", "") or "")
    user_sucursal = normalize_text(getattr(user, "sucursal", "") or getattr(user, "branch_name", "") or "")
    if username and username == str(row["created_by"] or ""):
        return True
    if user_branch_id and user_branch_id in {str(row["branch_id"] or ""), str(row["sucursal_responsable_id"] or "")}:
        return True
    if user_sucursal and user_sucursal == normalize_text(row["sucursal"] or ""):
        return True
    return False


@router.patch("/{warranty_id}/entry-base", response_model=WarrantyDetailResponse)
def update_warranty_entry_base(warranty_id: str, data: WarrantyEntryBaseUpdateRequest, user: Annotated[Any, Depends(require_current_user)]):
    """Edita la base de una garantía recién ingresada.

    Alcance Fase 3: fecha de ingreso, datos de cliente, observaciones/fotos, proveedor sugerido
    y productos asociados. No modifica estado, revisión, remitos ni ENV.
    """
    with db_connect() as conn:
        result = fetch_guarantee_with_items(conn, warranty_id)
        if not result:
            raise HTTPException(status_code=404, detail="Garantía no encontrada")
        row, _items = result
        if not _entry_base_edit_allowed(row, user):
            raise HTTPException(status_code=403, detail="Esta garantía ya no puede editarse desde ingreso. Pedile a un gestor/admin que la corrija.")

        updates: dict[str, str] = {}
        notes: list[str] = []
        details: dict[str, Any] = {"updated_fields": [], "item_changes": []}

        if data.fecha_ingreso is not None:
            new_ingreso = ingreso_at_from_input(data.fecha_ingreso, fallback_now=False)
            old_date = date_input_from_iso(row["ingreso_at"])
            new_date = date_input_from_iso(new_ingreso)
            if new_date and new_date != old_date:
                updates["ingreso_at"] = new_ingreso
                notes.append(f"Fecha de ingreso: {old_date or '-'} → {new_date}")

        simple_fields = {
            "observaciones": "observations",
            "photos_reference": "photos_reference",
            "proveedor": "provider_name",
            "cliente_nombre": "cliente_nombre",
            "cliente_telefono": "cliente_telefono",
            "cliente_email": "cliente_email",
            "numero_factura": "numero_factura",
            "fecha_compra": "fecha_compra",
        }
        for input_name, column_name in simple_fields.items():
            value = getattr(data, input_name)
            if value is None:
                continue
            clean_value = str(value or "").strip()
            if clean_value != str(row[column_name] or ""):
                updates[column_name] = clean_value
                notes.append(f"{input_name}: actualizado")

        item_changes: list[dict[str, Any]] = []
        if data.items:
            existing_items = {int(item["id"]): item for item in conn.execute("SELECT * FROM guarantee_items WHERE guarantee_id = ?", (int(row["id"]),)).fetchall()}
            allowed_item_fields = ["producto", "sku", "marca", "tipo", "serie", "falla", "observaciones"]
            for incoming in data.items:
                item_row = existing_items.get(int(incoming.row_number))
                if not item_row:
                    continue
                item_updates: dict[str, str] = {}
                for field in allowed_item_fields:
                    value = getattr(incoming, field)
                    if value is None:
                        continue
                    clean_value = str(value or "").strip()
                    if clean_value != str(item_row[field] or ""):
                        item_updates[field] = clean_value
                        item_changes.append({"item_id": incoming.row_number, "field": field, "old": str(item_row[field] or ""), "new": clean_value})
                if item_updates:
                    item_updates["updated_at"] = utc_now_iso()
                    assignments_item = ", ".join([f"{key} = ?" for key in item_updates])
                    conn.execute(f"UPDATE guarantee_items SET {assignments_item} WHERE id = ? AND guarantee_id = ?", [*item_updates.values(), int(incoming.row_number), int(row["id"])])

        if not updates and not item_changes:
            return get_warranty_detail(warranty_id, user)

        updates["updated_at"] = utc_now_iso()
        updates["updated_by"] = getattr(user, "username", "") or ""
        updates["updated_by_name"] = getattr(user, "display_name", "") or ""
        updates["synced_to_google_sheet"] = "0"
        assignments = ", ".join([f"{key} = ?" for key in updates])
        conn.execute(f"UPDATE guarantees SET {assignments} WHERE id = ?", [*updates.values(), int(row["id"])])

        details["updated_fields"] = list(updates.keys())
        details["item_changes"] = item_changes
        add_history(
            conn,
            int(row["id"]),
            str(row["warranty_code"]),
            user,
            "entry_corrected" if was_correction else "entry_base_updated",
            old_status=str(row["status"] or ""),
            new_status=updates.get("status", str(row["status"] or "")),
            note="; ".join(notes) or ("Productos actualizados" if item_changes else "Ingreso actualizado"),
            details=details,
        )
        conn.commit()
    audit("warranties.entry_base_update", user=user, resource_type="warranty", resource_id=warranty_id, details={"fields": list(updates.keys()), "item_changes": len(item_changes), "was_correction": bool(locals().get("was_correction", False))})
    return get_warranty_detail(warranty_id, user)


@router.get("/{warranty_id}", response_model=WarrantyDetailResponse)
def get_warranty_detail(warranty_id: str, _user: Annotated[Any, Depends(require_permission("warranties.view"))]):
    with db_connect() as conn:
        result = fetch_guarantee_with_items(conn, warranty_id)
        if not result:
            raise HTTPException(status_code=404, detail="Garantía no encontrada")
        row, items = result
        summary = row_to_summary(row, items)
        rows = [item_to_row(row, item, index) for index, item in enumerate(items, start=1)]
        history = history_for_guarantee(conn, int(row["id"]))
    return WarrantyDetailResponse(summary=summary, rows=rows, history=history)


@router.get("/{warranty_id}/history")
def get_warranty_history(warranty_id: str, _user: Annotated[Any, Depends(require_permission("warranties.view"))]):
    with db_connect() as conn:
        result = fetch_guarantee_with_items(conn, warranty_id)
        if not result:
            raise HTTPException(status_code=404, detail="Garantía no encontrada")
        row, _items = result
        return history_for_guarantee(conn, int(row["id"]))


@router.patch("/{warranty_id}", response_model=WarrantyDetailResponse)
def update_warranty(warranty_id: str, data: WarrantyUpdateRequest, user: Annotated[Any, Depends(require_permission("warranties.manage"))]):
    with db_connect() as conn:
        result = fetch_guarantee_with_items(conn, warranty_id)
        if not result:
            raise HTTPException(status_code=404, detail="Garantía no encontrada")
        row, _items = result
        current_status = str(row["status"] or "")
        current_review_status = str(row["review_status"] or REVIEW_PENDING)
        updates: dict[str, str] = {}
        notes: list[str] = []
        if data.estado is not None and data.estado.strip() and data.estado.strip() != current_status:
            updates["status"] = data.estado.strip()
            notes.append(f"Estado: {current_status or '-'} → {data.estado.strip()}")
        if data.sucursal is not None and data.sucursal.strip() != str(row["sucursal"] or ""):
            updates["sucursal"] = data.sucursal.strip()
            updates["sucursal_code"] = sucursal_code(data.sucursal.strip())
            notes.append(f"Sucursal: {row['sucursal'] or '-'} → {data.sucursal.strip() or '-'}")
        if data.deposito is not None and data.deposito.strip() != str(row["deposito"] or ""):
            updates["deposito"] = data.deposito.strip()
            notes.append(f"Depósito: {row['deposito'] or '-'} → {data.deposito.strip() or '-'}")
        if data.lugar_llegada is not None and data.lugar_llegada.strip() != str(row["lugar_llegada"] or ""):
            updates["lugar_llegada"] = data.lugar_llegada.strip()
        if data.ubicacion_actual is not None and data.ubicacion_actual.strip() != str(row["ubicacion_actual"] or ""):
            old_ub = str(row["ubicacion_actual"] or "-")
            new_ub = data.ubicacion_actual.strip()
            updates["ubicacion_actual"] = new_ub
            notes.append(f"Ubicación: {old_ub} → {new_ub or '-'}")
        if data.sucursal_responsable is not None and data.sucursal_responsable.strip() != str(row["sucursal_responsable"] or ""):
            old_sr = str(row["sucursal_responsable"] or "-")
            new_sr = data.sucursal_responsable.strip()
            updates["sucursal_responsable"] = new_sr
            notes.append(f"Suc. responsable: {old_sr} → {new_sr or '-'}")
        if data.observaciones is not None:
            updates["observations"] = data.observaciones.strip()
        if data.photos_reference is not None:
            updates["photos_reference"] = data.photos_reference.strip()
        if data.append_observation and data.append_observation.strip():
            stamp = format_datetime_ar()
            note = f"[{stamp} - {getattr(user, 'display_name', '')}] {data.append_observation.strip()}"
            current = str(row["observations"] or "").strip()
            updates["observations"] = f"{current}\n{note}".strip() if current else note
            notes.append("Se agregó una observación")
        item_changes: list[dict[str, Any]] = []
        if data.items:
            existing_items = {int(item["id"]): item for item in conn.execute("SELECT * FROM guarantee_items WHERE guarantee_id = ?", (int(row["id"]),)).fetchall()}
            allowed_item_fields = ["producto", "sku", "marca", "tipo", "serie", "falla", "observaciones"]
            for incoming in data.items:
                item_row = existing_items.get(int(incoming.row_number))
                if not item_row:
                    continue
                item_updates: dict[str, str] = {}
                for field in allowed_item_fields:
                    value = getattr(incoming, field)
                    if value is None:
                        continue
                    clean_value = str(value or "").strip()
                    if clean_value != str(item_row[field] or ""):
                        item_updates[field] = clean_value
                        item_changes.append({"item_id": incoming.row_number, "field": field, "old": str(item_row[field] or ""), "new": clean_value})
                if item_updates:
                    item_updates["updated_at"] = utc_now_iso()
                    assignments_item = ", ".join([f"{key} = ?" for key in item_updates])
                    conn.execute(f"UPDATE guarantee_items SET {assignments_item} WHERE id = ? AND guarantee_id = ?", [*item_updates.values(), int(incoming.row_number), int(row["id"])])
        if is_resolved_status(updates.get("status", current_status)) and not row["fecha_resolucion"]:
            updates["fecha_resolucion"] = utc_now_iso()
        # Si la garantía fue marcada como requiere_correccion y el usuario la corrige,
        # vuelve a INGRESO + pendiente_revision para entrar nuevamente a la cola.
        was_correction = review_status_matches(current_review_status, REVIEW_INCOMPLETE)
        if was_correction and (updates or item_changes):
            updates["status"] = DEFAULT_STATUSES[0]
            updates["review_status"] = REVIEW_PENDING
            updates["review_note"] = ""
            updates["correction_resubmitted_at"] = utc_now_iso()
            updates["correction_resubmitted_by"] = getattr(user, "username", "") or ""
            notes.append("Corregida y enviada nuevamente a revisión")
        if not updates and not item_changes:
            return get_warranty_detail(warranty_id, user)
        updates["updated_at"] = utc_now_iso()
        updates["updated_by"] = getattr(user, "username", "") or ""
        updates["updated_by_name"] = getattr(user, "display_name", "") or ""
        updates["synced_to_google_sheet"] = "0"
        assignments = ", ".join([f"{key} = ?" for key in updates])
        conn.execute(f"UPDATE guarantees SET {assignments} WHERE id = ?", [*updates.values(), int(row["id"])])
        add_history(
            conn,
            int(row["id"]),
            str(row["warranty_code"]),
            user,
            "updated",
            old_status=current_status,
            new_status=updates.get("status", current_status),
            note="; ".join(notes) or ("Productos actualizados" if item_changes else "Garantía actualizada"),
            details={"updated_fields": list(updates.keys()), "item_changes": item_changes},
        )
        conn.commit()
    audit("warranties.update", user=user, resource_type="warranty", resource_id=warranty_id, details={"source": "database", "fields": list(updates.keys())})
    return get_warranty_detail(warranty_id, user)
