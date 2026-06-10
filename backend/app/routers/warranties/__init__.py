from __future__ import annotations

import json
import re
import time
from datetime import date, datetime, timezone
from pathlib import Path
from uuid import uuid4
from typing import Annotated, Any
from zoneinfo import ZoneInfo

from fastapi import APIRouter, Depends, HTTPException, Query, status
from fastapi.responses import FileResponse
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.units import mm
from reportlab.pdfgen import canvas
from reportlab.platypus import Table, TableStyle, Paragraph
from reportlab.lib.styles import ParagraphStyle
from pydantic import BaseModel, Field, model_validator
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table as XLTable, TableStyleInfo

# Imports relativos: subimos 1 nivel extra porque ahora vivimos en
# routers/warranties/__init__.py en vez de routers/warranties.py.
from ...access import assigned_deposit_names, ensure_active_user, user_has, user_role_keys, users_with_permission
from ...audit import audit
from ...auth import require_current_user, require_permission
from ...brand_assets import brand_logo_path
from ...config import get_settings
from ...google_sheets import quote_sheet_name, sheets_service
from ...operational_config import runtime_warranty_config, load_operational_config, save_operational_config
from ...product_catalog import search_products as search_local_products, get_provider_for_brand, ensure_product_catalog_tables, runtime_product_catalog_config
from ...warranty_helpers import (
    AR_TZ,
    CANCELLED_STATUS,
    REVIEW_APPROVED,
    REVIEW_IN_PROGRESS,
    REVIEW_INCOMPLETE,
    REVIEW_LABELS,
    REVIEW_PENDING,
    format_date_ar,
    format_datetime_ar,
    header_key,
    normalize_text,
    now_ar,
    parse_date_filter,
    parse_iso_datetime,
    utc_now_iso,
)
from ...warranties_db import (
    RESET_TABLES_PG,
    pg_add_history,
    pg_cancel_guarantee,
    pg_collect_export_rows,
    pg_collect_export_rows_by_codes,
    pg_create_export,
    pg_delete_guarantee,
    pg_export_table_rows,
    pg_fetch_guarantee_rows_by_codes,
    pg_fetch_all_guarantee_rows,
    pg_fetch_guarantee_with_items,
    pg_get_export,
    pg_history_for_guarantee,
    pg_insert_guarantee,
    pg_insert_item,
    pg_insert_sync_log,
    pg_list_exports,
    pg_list_sync_logs,
    pg_mark_all_synced,
    pg_next_shipment_code,
    pg_next_warranty_code,
    pg_provider_suggestions,
    pg_reset_summary,
    pg_reset_warranty_tables,
    pg_sync_status,
    pg_clear_item_correction_notes,
    pg_update_item_fields,
    pg_update_guarantee_fields,
)
from ..notifications import notify_many

router = APIRouter(prefix="/api/warranties", tags=["warranties"])

# La app pasa a operar Garantías desde DB propia.
# Google Sheet queda como fuente auxiliar para productos/opciones y espejo futuro.

DEFAULT_STATUSES = [
    "1 - INGRESO",
    "2 - PENDIENTE",
    "3 - LISTO PARA ENVIAR",
    "4 - ENVIADO AL PROVEEDOR",
    "5 - EN EL PROVEEDOR",
    "6 - RESPONDIDO POR PROVEEDOR",
    "7 - RESUELTO",
    "8 - RECHAZADO",
    "9 - ANULADA",
    "10 - FINALIZADO",
]
DEFAULT_SUCURSALES = ["1 - CASEROS", "2 - LANUS", "3 - CANNING", "4 - NORCENTER"]
DEFAULT_DEPOSITOS = ["6 - CHICLANA", "7 - CORRALES", "8 - CACHI"]
DEFAULT_FINAL_STATUSES = ["10 - FINALIZADO"]

# Sub-tipos de resolución cuando estado = "7 - RESUELTO".
# Fase 12: solo se consideran resoluciones finales reales del proveedor.
# RESUELTO no equivale a FINALIZADO; el cierre se hace aparte con estado 10.
RESOLUTION_OPTIONS = {
    "nota_credito": "Nota de crédito",
    "reparacion": "Reparación",
    "cambio_equipo": "Cambio de equipo",
}
RESOLUTION_ALIASES = {
    "nc": "nota_credito",
    "nota de credito": "nota_credito",
    "nota de crédito": "nota_credito",
    "cambio": "cambio_equipo",
    "cambio_aprobado": "cambio_equipo",
    "cambio aprobado": "cambio_equipo",
    "reparado": "reparacion",
    "reparación": "reparacion",
}

def normalize_resolution_result(value: Any) -> str:
    raw = str(value or "").strip().lower()
    if not raw:
        return ""
    key = raw.replace("-", "_").replace(" ", "_")
    if key in RESOLUTION_OPTIONS:
        return key
    text_key = raw.replace("_", " ")
    return RESOLUTION_ALIASES.get(key) or RESOLUTION_ALIASES.get(text_key) or key

# Estado logístico del retiro del proveedor.
# No reemplaza al estado principal: indica si el proveedor ya pidió retirar y
# si la mercadería está lista físicamente para entregar.
PROVIDER_PICKUP_STATUSES = {
    "sin_solicitud": "Sin solicitud",
    "retiro_solicitado": "Retiro solicitado",
    "listo_para_retiro": "Listo para retiro",
    "retirado": "Retirado por proveedor",
}

# Fase 26 — roles/acciones: un usuario operativo de depósito NO es gestor.
# Aunque el catálogo de permisos haya quedado viejo o se le hayan colado permisos,
# si su único rol operativo es DEPOSITO no debe poder revisar, exportar, gestionar proveedor
# ni resolver garantías. Debe poder cargar cliente en depósito, recibir remitos y mover depósito→depósito.
#
# Fase 0 roles/permisos: el "escape" de este modo restringido se decide por
# PERMISOS, no por nombres de rol. Si el usuario suma cualquier capacidad de
# gestion (por otro rol o por permiso extra), deja de ser "deposito puro".
# Antes se chequeaba contra WARRANTY_PRIVILEGED_ROLES (nombres hardcodeados),
# lo que bloqueaba combos multi-rol validos (ej. DEPOSITO + un rol nuevo con
# permisos de export que no estuviera en la lista).
_DEPOSIT_ESCAPE_PERMISSIONS = (
    "warranties.manage",
    "warranties.manage_provider",
    "warranties.review",
    "warranties.gestor.panel",
    "warranties.dashboard",
    "warranties.export",
    "branches.cross_select",
)

def _user_role_keys(user: Any) -> set[str]:
    return user_role_keys(user)

def is_plain_deposit_operator(user: Any) -> bool:
    roles = _user_role_keys(user)
    if "DEPOSITO" not in roles:
        return False
    return not any(user_has(user, p) for p in _DEPOSIT_ESCAPE_PERMISSIONS)

def deny_plain_deposit_operator(user: Any, action: str = "esta acción") -> None:
    if is_plain_deposit_operator(user):
        raise HTTPException(status_code=403, detail=f"El rol Depósito operativo no puede realizar {action}.")

def normalize_provider_pickup_status(value: Any) -> str:
    raw = str(value or "").strip().lower().replace(" ", "_").replace("-", "_")
    return raw if raw in PROVIDER_PICKUP_STATUSES else "sin_solicitud"

# Tipo de respuesta del proveedor (Fase B — flujo claro de posventa).
# Cuando el proveedor responde (estado "6 - RESPONDIDO POR PROVEEDOR") puede pedir
# una de tres cosas. Es una dimensión paralela al estado principal, igual que
# review_status o estado_retiro_proveedor: NO reemplaza al estado, lo califica.
#   · retiro     → el proveedor pasa a retirar el equipo (dispara logística de retiro).
#   · revision   → hay que enviarle el equipo al proveedor para que lo revise.
#   · correccion → el proveedor pide corregir datos/serie antes de continuar.
PROVIDER_RESPONSE_TYPES = {
    "retiro": "Solicitó retiro",
    "revision": "Solicitó revisión",
    "correccion": "Pidió corrección",
}
PROVIDER_RESPONSE_ALIASES = {
    "retiro_solicitado": "retiro",
    "solicito_retiro": "retiro",
    "solicita_retiro": "retiro",
    "pickup": "retiro",
    "revision_solicitada": "revision",
    "solicito_revision": "revision",
    "solicita_revision": "revision",
    "envio_a_proveedor": "revision",
    "correccion_solicitada": "correccion",
    "solicito_correccion": "correccion",
    "pidio_correccion": "correccion",
    "correction": "correccion",
}

def normalize_provider_response_type(value: Any) -> str:
    raw = str(value or "").strip().lower().replace(" ", "_").replace("-", "_")
    if not raw:
        return ""
    if raw in PROVIDER_RESPONSE_TYPES:
        return raw
    return PROVIDER_RESPONSE_ALIASES.get(raw, "")

DEFAULT_DELAY_RANGES = [3, 7, 14, 30]
DEFAULT_REQUIRED_REVIEW_FIELDS = ["producto", "sku", "marca", "serie", "falla", "sucursal", "deposito"]
# CANCELLED_STATUS / REVIEW_* / REVIEW_LABELS viven en app.warranty_helpers (re-exportados arriba).

# ── Tipos de ingreso ──────────────────────────────────────────────────────────
# Cómo llegó la garantía al sistema. Determina origen_ingreso y ubicacion_actual.
VALID_TIPO_INGRESO = [
    "cliente_sucursal",           # cliente dejó el producto en una sucursal
    "cliente_deposito",           # cliente trajo el producto directamente al depósito
    "falla_recepcion_mercaderia", # falla detectada al recibir mercadería en depósito
    "stock_interno",              # problema detectado en stock propio
    "otro",                       # cualquier otro caso
]
# Tipos permitidos para un usuario VENDEDOR (rol sin warranties.manage).
# Todo ingreso desde sucursal debe ser cliente_sucursal.
TIPOS_INGRESO_VENDEDOR = {"cliente_sucursal"}
TIPO_INGRESO_LABELS: dict[str, str] = {
    "cliente_sucursal":           "Cliente en sucursal",
    "cliente_deposito":           "Cliente en depósito",
    "falla_recepcion_mercaderia": "Falla al recibir mercadería",
    "stock_interno":              "Stock interno",
    "otro":                       "Otro",
}
UBICACION_LABELS: dict[str, str] = {
    "sucursal":              "En sucursal",
    "en_transito":           "En tránsito",
    "deposito":              "En depósito",
    "en_transito_proveedor": "En tránsito al proveedor",
    "proveedor":             "En el proveedor",
    "devuelto":              "Devuelto",
    "entregado":             "Entregado al cliente",
    "desconocida":           "Ubicación desconocida",
}

DEFAULT_RAW_HEADERS = [
    "ID GARANTIA",
    "RESPONSABLE",
    "INGRESO",
    "PRODUCTO",
    "SKU",
    "MARCA",
    "SERIE",
    "FALLA",
    "SUCURSAL",
    "DEPOSITO",
    "ESTADO",
    "DIAS PENDIENTE",
    "FECHA DE INICIO DE GESTION",
    "ID DE CASO",
    "DIAS SIN RESPUESTA",
    "FECHA DE RETIRO",
    "FECHA DE RESOLUCION",
    "OBSERVACIONES",
    "VUELVE A",
    "FINALIZACION",
    "TIPO",
    "LUGAR LLEGADA",
    "USUARIO",
    "FECHA ULTIMA ACTUALIZACION",
    "ACTUALIZADO POR",
]

# Fase 27 — Espejo Google Sheets.
# La app sigue siendo fuente principal; estas pestañas son espejo/reporting.
# 00_RAW_GARANTIAS se mantiene por compatibilidad histórica/importación.
MIRROR_SHEETS: dict[str, list[str]] = {
    "GARANTIAS": [
        "ID GARANTIA", "FECHA INGRESO", "EMPRESA", "SUCURSAL CARGA", "SUCURSAL RESPONSABLE",
        "ORIGEN INGRESO", "TIPO INGRESO", "UBICACION ACTUAL", "DEPOSITO DESTINO",
        "ESTADO", "REVISION", "REMITO INTERNO", "ENV", "PROVEEDOR",
        "FECHA ENVIO PROVEEDOR", "FECHA ULTIMO MAIL", "DIAS SIN RESPUESTA",
        "RETIRO PROVEEDOR", "FECHA RETIRO PROVEEDOR", "RESPUESTA PROVEEDOR",
        "RESULTADO RESOLUCION", "FECHA RESOLUCION", "FECHA FINALIZACION",
        "CLIENTE", "TELEFONO", "EMAIL", "FACTURA", "FECHA COMPRA",
        "RESPONSABLE", "CREADO POR", "ACTUALIZADO", "ACTUALIZADO POR", "OBSERVACIONES",
    ],
    "GARANTIA_ITEMS": [
        "ID GARANTIA", "ITEM", "PRODUCTO", "SKU", "MARCA", "TIPO", "SERIE", "FALLA",
        "PROVEEDOR", "OBSERVACIONES", "UPDATED_AT",
    ],
    "REMITOS": [
        "CODIGO REMITO", "TIPO REMITO", "EMPRESA", "ORIGEN", "DESTINO", "ESTADO",
        "FECHA CREACION", "FECHA DESPACHO", "FECHA LLEGADA", "CREADO POR", "DESPACHADO POR",
        "RECIBIDO POR", "CANTIDAD ITEMS", "PDF", "NOTA",
    ],
    "REMITO_ITEMS": [
        "CODIGO REMITO", "ID GARANTIA", "PRODUCTO", "SKU", "SERIE", "SUCURSAL RESPONSABLE",
        "ORIGEN", "DESTINO", "ESTADO REMITO", "UPDATED_AT",
    ],
    "LOTES_ENV": [
        "ENV", "PROVEEDOR", "MARCA", "ESTADO LOTE", "FECHA CREACION", "CREADO POR",
        "ARCHIVO EXCEL", "CANTIDAD ITEMS", "FECHA MAIL", "FECHA ULTIMO MAIL", "OBSERVACIONES",
    ],
    "LOTE_ITEMS": [
        "ENV", "ID GARANTIA", "PRODUCTO", "SKU", "SERIE", "PROVEEDOR", "ESTADO GARANTIA",
        "RESPUESTA PROVEEDOR", "RESULTADO RESOLUCION", "UPDATED_AT",
    ],
    "EVENTOS": [
        "FECHA", "ID GARANTIA", "USUARIO", "NOMBRE", "ACCION", "ESTADO ANTERIOR", "ESTADO NUEVO",
        "REVISION ANTERIOR", "REVISION NUEVA", "DETALLE", "METADATA",
    ],
}

PRODUCT_CACHE: dict[str, Any] = {"loaded_at": 0.0, "items": []}


class WarrantyItemIn(BaseModel):
    # ── Datos del artículo ─────────────────────────────────────────────────────
    producto: str = Field(min_length=1)
    sku: str | None = None
    marca: str | None = None
    tipo: str | None = None
    serie: str | None = None
    falla: str = Field(min_length=1)
    observaciones: str | None = None
    # ── Origen y ubicación ─────────────────────────────────────────────────────
    tipo_ingreso: str = Field(min_length=1)     # obligatorio — determina origen y ubicación inicial
    sucursal: str = ""                          # obligatorio solo si tipo_ingreso = "cliente_sucursal"
    deposito: str = Field(min_length=1)
    lugar_llegada: str | None = None
    # ── Sucursal comercialmente responsable (depósito/gestor solo) ───────────
    # sucursal_responsable_id: branch_id real del sistema (preferido).
    # sucursal_responsable: texto de display (se deriva del ID si se conoce; fallback legado).
    # Los VENDEDORES no envían ninguno — el backend los deriva de su asignación.
    sucursal_responsable_id: str = ""   # branch_id de la sucursal responsable
    sucursal_responsable: str = ""      # nombre de display (derivado del ID o texto libre)
    # ── Proveedor (sugerido por catálogo, opcional) ────────────────────────────
    proveedor: str | None = None
    # ── Datos del cliente (opcionales) ────────────────────────────────────────
    cliente_nombre: str | None = None
    cliente_telefono: str | None = None
    cliente_email: str | None = None
    numero_factura: str | None = None
    fecha_compra: str | None = None
    # Fecha real en la que ingresó físicamente/operativamente la garantía.
    # Si no viene, se usa la fecha/hora actual.
    fecha_ingreso: str | None = None

    @model_validator(mode="after")
    def _validate_tipo_and_sucursal(self) -> "WarrantyItemIn":
        tipo = (self.tipo_ingreso or "").strip()
        if tipo not in VALID_TIPO_INGRESO:
            raise ValueError(
                f"tipo_ingreso inválido: '{tipo}'. "
                f"Valores permitidos: {', '.join(VALID_TIPO_INGRESO)}"
            )
        if tipo == "cliente_sucursal" and not (self.sucursal or "").strip():
            raise ValueError(
                "sucursal es obligatoria cuando tipo_ingreso es 'cliente_sucursal'"
            )
        return self


class WarrantyCreateRequest(BaseModel):
    items: list[WarrantyItemIn] = Field(min_length=1, max_length=100)
    group_under_one_id: bool = False


class WarrantyCreatedItem(BaseModel):
    id_garantia: str
    producto: str
    sku: str | None = None
    parent_warranty_code: str = ""
    parent_item_index: int | None = None


class WarrantyCreateResponse(BaseModel):
    ok: bool
    count: int
    ids: list[str]
    items: list[WarrantyCreatedItem]


class WarrantyRow(BaseModel):
    row_number: int
    id_garantia: str
    responsable: str = ""
    usuario: str = ""
    ingreso: str = ""
    producto: str = ""
    sku: str = ""
    marca: str = ""
    tipo: str = ""
    serie: str = ""
    falla: str = ""
    sucursal: str = ""
    deposito: str = ""
    lugar_llegada: str = ""
    estado: str = ""
    observaciones: str = ""
    correction_note: str = ""
    actualizado_por: str = ""
    fecha_ultima_actualizacion: str = ""
    cancelled: bool = False
    cancel_reason: str = ""
    cancelled_by: str = ""
    cancelled_at: str = ""


class WarrantySummary(BaseModel):
    id_garantia: str
    parent_warranty_code: str = ""
    parent_item_index: int | None = None
    grouped_item_label: str = ""
    ingreso: str = ""
    ingreso_iso: str = ""
    responsible_username: str = ""
    responsable: str = ""
    usuario: str = ""
    producto_principal: str = ""
    productos: list[str] = []
    cantidad_items: int = 0
    marca: str = ""
    sku: str = ""
    serie: str = ""
    falla: str = ""
    sucursal: str = ""
    sucursal_code: str = ""
    branch_id: str = ""
    company_id: str = ""
    sucursal_responsable: str = ""
    sucursal_responsable_id: str = ""
    deposito: str = ""
    lugar_llegada: str = ""
    estado: str = ""
    review_status: str = "pendiente_revision"
    review_status_label: str = "Pendiente de revisión"
    reviewed_by: str = ""
    reviewed_by_name: str = ""
    reviewed_at: str = ""
    review_note: str = ""
    observaciones: str = ""
    photos_reference: str = ""
    # ── Origen / tipo / ubicación física (Fase 1) ──────────────────────────────
    tipo_ingreso: str = ""
    tipo_ingreso_label: str = ""
    origen_ingreso: str = ""
    ubicacion_actual: str = ""
    ubicacion_actual_label: str = ""
    # ── Datos del cliente (Fase 1) ─────────────────────────────────────────────
    cliente_nombre: str = ""
    cliente_telefono: str = ""
    cliente_email: str = ""
    numero_factura: str = ""
    fecha_compra: str = ""
    # ── Proveedor / gestión ────────────────────────────────────────────────────
    provider_name: str = ""
    id_de_caso: str = ""
    fecha_envio_proveedor: str = ""
    fecha_ultima_respuesta: str = ""
    fecha_ultimo_reclamo: str = ""
    estado_retiro_proveedor: str = "sin_solicitud"
    estado_retiro_proveedor_label: str = "Sin solicitud"
    # Qué pidió el proveedor cuando respondió (retiro | revision | correccion).
    provider_response_type: str = ""
    provider_response_type_label: str = ""
    provider_correction_note: str = ""
    fecha_solicitud_retiro_proveedor: str = ""
    fecha_retiro_proveedor: str = ""
    dias_pendiente: int = 0
    dias_sin_respuesta: int | None = None
    shipment_code: str = ""
    shipment_file_name: str = ""
    resolution_note: str = ""
    resolution_reference: str = ""
    resultado_resolucion: str = ""
    resultado_resolucion_label: str = ""
    numero_nota_credito: str = ""
    importe_nota_credito: str = ""
    fecha_nota_credito: str = ""
    detalle_reparacion: str = ""
    fecha_reparacion: str = ""
    producto_reemplazo: str = ""
    sku_reemplazo: str = ""
    serie_reemplazo: str = ""
    fecha_recepcion_reemplazo: str = ""
    fecha_finalizacion: str = ""
    finalizacion: str = ""
    remito_interno: str = ""
    remito_proveedor: str = ""
    transit_status: str = ""  # '' | 'en_transito' | 'en_deposito'
    synced_to_google_sheet: bool = False
    fecha_ultima_sincronizacion: str = ""
    actualizado_por: str = ""
    fecha_ultima_actualizacion: str = ""
    cancelled: bool = False
    cancel_reason: str = ""
    cancelled_by: str = ""
    cancelled_at: str = ""


class WarrantyListResponse(BaseModel):
    items: list[WarrantySummary]
    total: int
    limit: int


class WarrantyDetailResponse(BaseModel):
    summary: WarrantySummary
    rows: list[WarrantyRow]
    history: list[dict[str, Any]]


class WarrantyItemUpdateRequest(BaseModel):
    row_number: int
    producto: str | None = None
    sku: str | None = None
    marca: str | None = None
    tipo: str | None = None
    serie: str | None = None
    falla: str | None = None
    observaciones: str | None = None


class WarrantyUpdateRequest(BaseModel):
    estado: str | None = None
    sucursal: str | None = None
    deposito: str | None = None
    lugar_llegada: str | None = None
    ubicacion_actual: str | None = None
    sucursal_responsable: str | None = None
    observaciones: str | None = None
    photos_reference: str | None = None
    append_observation: str | None = None
    items: list[WarrantyItemUpdateRequest] | None = None


class WarrantyEntryBaseUpdateRequest(BaseModel):
    """Edición controlada de la base del ingreso.

    Pensado para corregir datos recién cargados antes de que avancen a revisión/gestión.
    No permite cambiar estado, remitos, ENV ni proveedor operativo.
    """
    fecha_ingreso: str | None = None
    observaciones: str | None = None
    photos_reference: str | None = None
    proveedor: str | None = None
    cliente_nombre: str | None = None
    cliente_telefono: str | None = None
    cliente_email: str | None = None
    numero_factura: str | None = None
    fecha_compra: str | None = None
    items: list[WarrantyItemUpdateRequest] | None = None


class WarrantyReviewRequest(BaseModel):
    note: str | None = None


class WarrantyProviderSendRequest(BaseModel):
    provider_name: str = Field(min_length=1)
    provider_case_id: str | None = None
    note: str | None = None


class WarrantyItemCorrection(BaseModel):
    row_number: int            # id del ítem (guarantee_items.id) — coincide con WarrantyRow.row_number
    note: str = ""             # qué corregir en ese ítem ("" = sin corrección para ese ítem)


class WarrantyProviderResponseRequest(BaseModel):
    note: str | None = None
    provider_case_id: str | None = None
    estado: str | None = None
    # Tipo de respuesta del proveedor: retiro | revision | correccion.
    response_type: str | None = None
    # Detalle general de qué corregir (cuando no se especifica por ítem).
    correction_note: str | None = None
    # Corrección por ítem/serie (cuando response_type = correccion).
    item_corrections: list[WarrantyItemCorrection] | None = None
    # Fecha acordada con el proveedor para el retiro (solo cuando response_type='retiro').
    # ISO date o ISO datetime. Si se manda, se registra junto con la solicitud.
    fecha_retiro_acordada: str | None = None


class WarrantyProviderCorrectionResolveRequest(BaseModel):
    note: str | None = None


class WarrantyClaimRequest(BaseModel):
    note: str = Field(min_length=1)


class WarrantyResendMailRequest(BaseModel):
    note: str | None = None


class WarrantyStatusChangeRequest(BaseModel):
    estado: str = Field(min_length=1)
    note: str | None = None
    resolution_note: str | None = None       # Detalle de la resolución (motivo rechazo, descripción reparación, etc.)
    resolution_reference: str | None = None  # Referencia numérica (N° NC, remito cambio, etc.)
    resultado_resolucion: str | None = None  # nota_credito | reparacion | cambio_equipo (requerido si estado = "7 - RESUELTO")
    numero_nota_credito: str | None = None
    importe_nota_credito: str | None = None
    fecha_nota_credito: str | None = None
    detalle_reparacion: str | None = None
    fecha_reparacion: str | None = None
    producto_reemplazo: str | None = None
    sku_reemplazo: str | None = None
    serie_reemplazo: str | None = None
    fecha_recepcion_reemplazo: str | None = None
    finalizacion: str | None = None


class WarrantyExportRequest(BaseModel):
    marca: str | None = None
    proveedor: str | None = None
    estado: str | None = None
    sucursal: str | None = None
    deposito: str | None = None
    fecha_desde: str | None = None
    fecha_hasta: str | None = None


class WarrantyExportInfo(BaseModel):
    id: int
    created_at: str
    created_by: str = ""
    provider_name: str = ""
    marca: str = ""
    filters: dict[str, Any] = {}
    file_name: str
    row_count: int
    download_url: str
    shipment_code: str = ""
    file_format: str = "excel"
    logo_brand: str = "gv_electro"


class WarrantyExportListResponse(BaseModel):
    items: list[WarrantyExportInfo]


class WarrantyBatchExportRequest(BaseModel):
    warranty_ids: list[str] = Field(min_length=1)
    proveedor: str | None = None
    nota: str | None = None
    formato: str | None = "excel"
    logo_brand: str | None = "gv_electro"


class WarrantyExportRegenerateRequest(BaseModel):
    proveedor: str | None = None
    nota: str | None = None
    formato: str | None = None
    logo_brand: str | None = None


def normalize_export_format(value: Any) -> str:
    raw = str(value or "excel").strip().lower()
    return "pdf" if raw == "pdf" else "excel"


def normalize_export_logo(value: Any) -> str:
    raw = str(value or "gv_electro").strip().lower().replace("-", "_")
    if raw in {"abc", "abc_electro", "electro_abc"}:
        return "abc_electro"
    return "gv_electro"


class ConfirmShipmentRequest(BaseModel):
    shipment_code: str = Field(min_length=1)
    provider_name: str | None = None
    nota: str | None = None

class ProviderPickupRequest(BaseModel):
    note: str | None = None
    provider_case_id: str | None = None
    fecha_retiro_acordada: str | None = None


class WarrantySyncStatus(BaseModel):
    last_sync_at: str = ""
    last_sync_type: str = ""
    last_sync_status: str = ""
    last_sync_user: str = ""
    pending_to_sheet: int = 0
    total_guarantees: int = 0
    errors: list[str] = []


class WarrantySyncResult(BaseModel):
    ok: bool
    sync_type: str
    status: str
    started_at: str
    finished_at: str
    rows_processed: int = 0
    rows_created: int = 0
    rows_updated: int = 0
    rows_skipped: int = 0
    errors: list[str] = []


class WarrantySyncLogInfo(BaseModel):
    id: int
    sync_type: str
    status: str
    started_at: str
    finished_at: str
    actor_username: str = ""
    actor_name: str = ""
    rows_processed: int = 0
    rows_created: int = 0
    rows_updated: int = 0
    rows_skipped: int = 0
    errors: list[str] = []


class WarrantySyncLogsResponse(BaseModel):
    items: list[WarrantySyncLogInfo]




class WarrantyConfigCatalog(BaseModel):
    statuses: list[str] = []
    final_statuses: list[str] = []
    sucursales: list[str] = []
    depositos: list[str] = []
    delay_ranges: list[int] = []
    required_review_fields: list[str] = []
    sheet_raw: str = ""
    spreadsheet_url: str = ""
    products_source_label: str = "Catálogo local"


class WarrantyConfigResponse(BaseModel):
    config: WarrantyConfigCatalog
    providers_count: int = 0
    brands_count: int = 0
    mapped_brands_count: int = 0
    unmapped_brands_count: int = 0
    pending_review_count: int = 0
    active_count: int = 0


class WarrantyConfigSaveRequest(BaseModel):
    statuses: list[str] | None = None
    final_statuses: list[str] | None = None
    sucursales: list[str] | None = None
    depositos: list[str] | None = None
    delay_ranges: list[int] | None = None
    required_review_fields: list[str] | None = None
    raw_sheet: str | None = None
    spreadsheet_url: str | None = None


class WarrantyCancelRequest(BaseModel):
    reason: str = Field(min_length=3)


class WarrantyDashboardPoint(BaseModel):
    label: str
    value: float
    extra: dict[str, Any] = {}


class WarrantyDashboardMetrics(BaseModel):
    total: int = 0
    ingreso: int = 0
    pendientes_revision: int = 0
    pendientes_proveedor: int = 0
    enviadas_proveedor: int = 0
    en_revision: int = 0
    resueltas: int = 0
    rechazadas: int = 0
    demoradas_7: int = 0
    demoradas_15: int = 0
    promedio_dias_pendiente: float = 0
    promedio_resolucion: float = 0
    promedio_dias_sin_respuesta: float = 0


class WarrantyDashboardResponse(BaseModel):
    metrics: WarrantyDashboardMetrics
    by_status: list[WarrantyDashboardPoint] = []
    by_brand: list[WarrantyDashboardPoint] = []
    by_provider: list[WarrantyDashboardPoint] = []
    by_branch: list[WarrantyDashboardPoint] = []
    by_deposit: list[WarrantyDashboardPoint] = []
    by_delay_range: list[WarrantyDashboardPoint] = []
    monthly_entries: list[WarrantyDashboardPoint] = []
    avg_resolution_by_provider: list[WarrantyDashboardPoint] = []
    final_resolutions: list[WarrantyDashboardPoint] = []
    critical: list[WarrantySummary] = []
    filters: dict[str, Any] = {}




class WarrantyDiagnosticItem(BaseModel):
    key: str
    label: str
    status: str = "ok"
    detail: str = ""
    count: int = 0


class WarrantyDiagnosticsResponse(BaseModel):
    status: str = "ok"
    generated_at: str
    items: list[WarrantyDiagnosticItem] = []
    next_actions: list[str] = []

class WarrantyCounterInfo(BaseModel):
    year: int
    sucursal: str
    last_number: int


class WarrantyCountersResponse(BaseModel):
    counters: list[WarrantyCounterInfo]


class WarrantyResetSummary(BaseModel):
    guarantees: int = 0
    guarantee_items: int = 0
    guarantee_history: int = 0
    remitos: int = 0
    exports: int = 0
    sync_logs: int = 0
    counters: int = 0
    generated_export_files: int = 0


class WarrantyResetPreviewResponse(BaseModel):
    ok: bool = True
    generated_at: str
    summary: WarrantyResetSummary
    preserved: list[str]
    warning: str
    confirmation_phrase: str


class WarrantyResetRequest(BaseModel):
    confirmation: str = Field(min_length=1)
    reset_generated_files: bool = True


class WarrantyResetResponse(BaseModel):
    ok: bool
    reset_at: str
    summary_before: WarrantyResetSummary
    backup_file: str
    deleted_generated_files: int = 0
    message: str


RESET_CONFIRMATION_PHRASE = "RESET GARANTIAS PRODUCCION"


# =========================================================
# Utilidades generales
# =========================================================

# Fase 2.5h.3: este router ya opera contra Postgres mediante app.db.db_session y helpers pg_*.


# utc_now_iso / now_ar / normalize_text viven en app.warranty_helpers (re-exportados arriba).


def canonical_status_key(value: Any) -> str:
    """Clave estable para comparar estados aunque vengan con número o texto.

    La normalización separa revisión interna (review_status) de estados operativos.
    Compatibiliza estados viejos sin volver a usarlos como fuente de verdad.
    """
    original = normalize_text(value)
    if not original:
        return ""
    # Casos numerados antiguos: antes el proveedor usaba "5 - EN REVISION".
    if re.match(r"^5\s+EN\s+REVISION$", original):
        return "EN EL PROVEEDOR"
    text = re.sub(r"^\d+\s+", "", original).strip()
    aliases = {
        "INGRESADO": "INGRESO",
        "PENDIENTE REVISION": "INGRESO",
        "PENDIENTE DE REVISION": "INGRESO",
        # Estados internos escritos en status por versiones anteriores. Ahora viven en review_status.
        "EN REVISION": "INGRESO",
        "EN REVISION INTERNA": "INGRESO",
        "CORRECCION PENDIENTE": "INGRESO",
        # Variantes del flujo proveedor.
        "EN REVISION PROVEEDOR": "EN EL PROVEEDOR",
        "EN REVISION DEL PROVEEDOR": "EN EL PROVEEDOR",
        "PROVEEDOR": "EN EL PROVEEDOR",
        "EN PROVEEDOR": "EN EL PROVEEDOR",
        "RESPONDIDO": "RESPONDIDO POR PROVEEDOR",
        "RESPUESTA PROVEEDOR": "RESPONDIDO POR PROVEEDOR",
        "RESPONDIDO PROVEEDOR": "RESPONDIDO POR PROVEEDOR",
        "LISTO RETIRO": "LISTO PARA RETIRO",
        "NC": "NOTA DE CREDITO",
        "NOTA CREDITO": "NOTA DE CREDITO",
        "ANULADO": "ANULADA",
        "CANCELADO": "ANULADA",
    }
    return aliases.get(text, text)


def normalize_status(value: Any) -> str:
    """Devuelve la etiqueta canónica visible de un estado operativo.

    Fase 16 hotfix: algunos filtros del listado usan normalize_status(), pero
    versiones anteriores sólo dejaron canonical_status_key(). Esta función mantiene
    compatibilidad con estados viejos y devuelve siempre una etiqueta oficial
    cuando puede mapearla.
    """
    key = canonical_status_key(value)
    if not key:
        return ""
    for label in DEFAULT_STATUSES:
        if canonical_status_key(label) == key:
            return label
    # Compatibilidad con valores viejos que no existen en DEFAULT_STATUSES.
    if key == canonical_status_key(CANCELLED_STATUS):
        return "9 - ANULADA"
    return str(value or "").strip()


def is_deposit_operator_user(user: Any) -> bool:
    branch_type_key = normalize_text(getattr(user, "branch_type", "") or "")
    branch_name_key = normalize_text(getattr(user, "branch_name", "") or getattr(user, "sucursal", "") or "")
    roles_keys = {normalize_text(r) for r in user_role_keys(user)}
    return (
        branch_type_key in {"DEPOSIT", "DEPOSITO"}
        or "DEPOSITO" in roles_keys
        or branch_name_key.startswith("DEPOSITO ")
        or branch_name_key == "DEPOSITO"
        or bool(assigned_deposit_names(user))
    )


def ensure_warranty_intake_access(user: Any) -> None:
    ensure_active_user(user)
    if user_has(user, "warranties.view") or user_has(user, "warranties.create") or is_deposit_operator_user(user):
        return
    raise HTTPException(status_code=403, detail="No tenes permiso para realizar esta accion")


def status_matches(value: Any, expected: Any) -> bool:
    return canonical_status_key(value) == canonical_status_key(expected)


def _canonical_review_status(value: Any) -> str:
    """Normaliza un valor de review_status a su constante canónica.

    NO usa canonical_status_key() porque esa función mapea 'EN REVISION' y
    'PENDIENTE REVISION' a 'INGRESO' (para compatibilidad con estados operativos
    viejos), lo que hace que pendiente_revision y en_revision sean indistinguibles.
    """
    s = normalize_text(value)
    if not s:
        return REVIEW_PENDING
    # Quitar prefijo numérico si existiera
    s = re.sub(r"^\d+\s+", "", s).strip()
    aliases: dict[str, str] = {
        # pendiente_revision
        "PENDIENTE REVISION":          REVIEW_PENDING,
        "PENDIENTE DE REVISION":       REVIEW_PENDING,
        "PENDING":                     REVIEW_PENDING,
        # en_revision
        "EN REVISION":                 REVIEW_IN_PROGRESS,
        "EN REVISION INTERNA":         REVIEW_IN_PROGRESS,
        # requiere_correccion
        "REQUIERE CORRECCION":         REVIEW_INCOMPLETE,
        "CORRECCION PENDIENTE":        REVIEW_INCOMPLETE,
        "INCOMPLETA":                  REVIEW_INCOMPLETE,
        # revisada
        "REVISADA":                    REVIEW_APPROVED,
    }
    return aliases.get(s, s)


def review_status_matches(value: Any, expected: str) -> bool:
    """Compara review_status con normalización propia (sin canonical_status_key)."""
    return _canonical_review_status(value) == _canonical_review_status(expected)


# header_key / format_date_ar / format_datetime_ar / parse_iso_datetime / parse_date_filter
# viven en app.warranty_helpers (re-exportados arriba).


def ingreso_at_from_input(value: Any, *, fallback_now: bool = True) -> str:
    """Normaliza una fecha de ingreso a ISO.

    Acepta YYYY-MM-DD, DD/MM/YYYY o un ISO existente. Para fechas sin hora,
    guardamos mediodía de Argentina para evitar corrimientos visuales por zona horaria.
    """
    text = str(value or "").strip()
    if not text:
        return utc_now_iso() if fallback_now else ""
    dt = parse_iso_datetime(text)
    if dt:
        return dt.isoformat()
    d = parse_date_filter(text)
    if d:
        return datetime(d.year, d.month, d.day, 12, 0, tzinfo=ZoneInfo("America/Argentina/Buenos_Aires")).isoformat()
    raise HTTPException(status_code=400, detail="Fecha de ingreso inválida. Usá formato AAAA-MM-DD o DD/MM/AAAA.")


def date_input_from_iso(value: Any) -> str:
    dt = parse_iso_datetime(value)
    if not dt:
        return ""
    return dt.astimezone(ZoneInfo("America/Argentina/Buenos_Aires")).date().isoformat()


def clean_option_value(value: Any) -> str:
    text = str(value or "").strip()
    return re.sub(r"\s+", " ", text)


def strip_numeric_prefix(value: Any) -> str:
    return re.sub(r"^\s*\d+\s*[-.)]\s*", "", str(value or "").strip()).strip()


def canonical_deposit_display(value: Any) -> str:
    clean = strip_numeric_prefix(value)
    key = normalize_text(clean)
    if key in {"CHICLANA", "DEPOSITO CHICLANA"}:
        return "Depósito Chiclana"
    if key in {"CORRALES", "DEPOSITO CORRALES"}:
        return "Depósito Corrales"
    if key in {"CACHI", "DEPOSITO CACHI"}:
        return "Depósito Cachi"
    return clean


def clean_select_options(values: list[Any], *, deposit: bool = False) -> list[str]:
    cleaned: list[str] = []
    for value in values or []:
        text = canonical_deposit_display(value) if deposit else strip_numeric_prefix(value)
        text = clean_option_value(text)
        if text:
            cleaned.append(text)
    return unique_keep_order(cleaned)


def unique_keep_order(values: list[str]) -> list[str]:
    seen: set[str] = set()
    out: list[str] = []
    for value in values:
        clean = clean_option_value(value)
        if not clean:
            continue
        key = header_key(clean)
        if key in seen:
            continue
        seen.add(key)
        out.append(clean)
    return out


def sucursal_code(value: Any) -> str:
    clean = normalize_text(strip_numeric_prefix(value))
    compact = clean.replace(" ", "")
    mapping = {
        "CASEROS": "CAS",
        "LANUS": "LAN",
        "LANUSOESTE": "LAN",
        "CANNING": "CAN",
        "NORCENTER": "NOR",
        "NORTE": "NOR",
        "NORTH": "NOR",
        "SUR": "SUR",
    }
    if compact in mapping:
        return mapping[compact]
    for key, code in mapping.items():
        if key in compact:
            return code
    compact = re.sub(r"[^A-Z0-9]+", "", compact)
    return (compact[:3] or "GEN").upper()


# ──────────────────────────────────────────────────────────────────────────
# Resolvers de texto → IDs canónicos (import histórico desde Excel/Sheet).
# El Excel viejo trae textos legibles ("Sur", "Norte", "Depósito Chiclana");
# acá los convertimos a branch_id / company_id / user_id de Postgres.
# Reusan branches_db.get_branch_by_name_or_code para match directo y un dict
# de aliases para casos que no matchean (ej. "Sur" → branch_id "sur" cuyo
# name oficial es "Lanús").
# ──────────────────────────────────────────────────────────────────────────

_BRANCH_TEXT_ALIASES: dict[str, str] = {
    "SUR": "sur",
    "NORTE": "norte",
    "CHICLANA": "deposito_chiclana",
    "DEPOSITO CHICLANA": "deposito_chiclana",
    "CORRALES": "deposito_corrales",
    "DEPOSITO CORRALES": "deposito_corrales",
    "CACHI": "deposito_cachi",
    "DEPOSITO CACHI": "deposito_cachi",
}

_COMPANY_TEXT_ALIASES: dict[str, str] = {
    "ELECTRO GV": "electro_gv",
    "ELECTROGV": "electro_gv",
    "ELECTRO ABC SRL": "electro_abc_srl",
    "ELECTRO ABC": "electro_abc_srl",
    "ABC ELECTRO": "electro_abc_srl",
    "ABCELECTRO": "electro_abc_srl",
}


def resolve_branch_id_from_text(value: Any) -> str | None:
    """Resuelve texto de sucursal/depósito (Excel viejo) a branch_id de Postgres.

    Prueba en orden:
      1. Match directo contra `branches.name` o `branches.code` (case-insensitive).
      2. Diccionario de alias (Sur→sur, Norte→norte, etc.).
    Retorna None si no encuentra.
    """
    text = str(value or "").strip()
    if not text:
        return None
    from ...branches_db import get_branch_by_name_or_code as _by_nc

    # El Excel viejo a veces trae "1 - CASEROS", "2 - LANUS" — limpiar antes del match.
    clean = strip_numeric_prefix(text)
    found = _by_nc(clean) or _by_nc(text)
    if found:
        return str(found["id"])
    key = normalize_text(clean)
    return _BRANCH_TEXT_ALIASES.get(key)


def resolve_company_id_from_text(value: Any) -> str | None:
    """Resuelve texto de empresa a company_id de Postgres.

    Primero prueba match directo contra el slug (los IDs ya son `electro_gv`,
    `electro_abc_srl`), después contra alias por nombre legible.
    """
    text = str(value or "").strip()
    if not text:
        return None
    from sqlalchemy import select as _select

    from ...db import db_session as _db_session
    from ...models.org import Company as _Company

    with _db_session() as session:
        c = session.get(_Company, text)
        if c and c.is_active:
            return str(c.id)
        # Match por name (case-insensitive)
        from sqlalchemy import func as _func

        c2 = session.scalar(
            _select(_Company).where(
                _Company.is_active.is_(True),
                _func.lower(_Company.name) == text.lower(),
            )
        )
        if c2:
            return str(c2.id)
    key = normalize_text(text)
    return _COMPANY_TEXT_ALIASES.get(key)


def resolve_user_id_from_username(session: Any, username: Any) -> int | None:
    """Busca user.id por username (case-insensitive). None si no existe."""
    u = str(username or "").strip()
    if not u:
        return None
    from sqlalchemy import func as _func, select as _select

    from ...models.auth import User as _User

    return session.scalar(
        _select(_User.id).where(_func.lower(_User.username) == u.lower())
    )


def _origen_from_tipo(tipo: str) -> str:
    """Deriva origen_ingreso a partir del tipo_ingreso."""
    return "sucursal" if tipo == "cliente_sucursal" else "deposito"


def _ubicacion_from_tipo(tipo: str) -> str:
    """Deriva ubicación genérica legacy a partir del tipo_ingreso."""
    if tipo == "cliente_sucursal":
        return "sucursal"
    if tipo in ("cliente_deposito", "falla_recepcion_mercaderia", "stock_interno"):
        return "deposito"
    return "desconocida"


def _initial_ubicacion_actual(tipo: str, sucursal_carga: str = "", deposito_carga: str = "") -> str:
    """Ubicación física real al crear la garantía.

    Regla operativa fase 36:
    - Cliente en sucursal: el equipo queda físicamente en esa sucursal
      (Caseros, Canning, etc.), no en el depósito destino.
    - Cliente en depósito / falla recepción / stock interno: el equipo queda
      físicamente en el depósito de carga (Chiclana, Corrales, Cachi).

    Los valores genéricos 'sucursal'/'deposito' quedan solo como fallback legacy.
    """
    clean_tipo = (tipo or "").strip()
    if clean_tipo == "cliente_sucursal":
        return (sucursal_carga or "").strip() or "sucursal"
    if clean_tipo in ("cliente_deposito", "falla_recepcion_mercaderia", "stock_interno"):
        return (deposito_carga or "").strip() or "deposito"
    return "desconocida"


def _location_is_deposit_value(value: str) -> bool:
    key = normalize_text(value)
    return key == "DEPOSITO" or key.startswith("DEPOSITO ")


def _code_source_for_tipo(item_sucursal: str, item_deposito: str, tipo: str) -> str:
    """Devuelve la fuente de código (sucursal o depósito) para next_warranty_code."""
    if tipo == "cliente_sucursal" and item_sucursal.strip():
        return item_sucursal
    return item_deposito


def _fetch_branch_info(branch_id: str) -> dict[str, str] | None:
    """Busca en la tabla branches el nombre, empresa y tipo de una branch por su ID.

    Fase 2.5h.2b: porteado a Postgres vía SQLAlchemy.
    """
    bid = (branch_id or "").strip()
    if not bid:
        return None
    try:
        from ...db import db_session as _db_session
        from ...models.org import Branch as _Branch, Company as _Company
        from sqlalchemy import select as _select

        with _db_session() as session:
            b = session.get(_Branch, bid)
            if not b:
                return None
            company_name = ""
            if b.company_id:
                company_name = session.scalar(_select(_Company.name).where(_Company.id == b.company_id)) or ""
            return {
                "id": b.id or "",
                "name": b.name or "",
                "code": b.code or "",
                "type": b.type or "",
                "company_id": b.company_id or "",
                "company_name": company_name,
            }
    except Exception:
        return None


def _branch_key(value: Any) -> str:
    return normalize_text(strip_numeric_prefix(value))


def _branch_by_name(branches: list[dict[str, str]], name: str, branch_type: str | None = None) -> dict[str, str] | None:
    key = _branch_key(name)
    if not key:
        return None
    for branch in branches:
        if branch_type and str(branch.get("type") or "") != branch_type:
            continue
        if _branch_key(branch.get("name", "")) == key or _branch_key(branch.get("code", "")) == key:
            return branch
    return None


def _fetch_branches_operativas() -> list[dict[str, str]]:
    """Devuelve las branches físicas y de depósito activas del sistema (physical + deposit).
    Usadas en el selector de sucursal_responsable y para validar IDs del frontend.

    Fase 2.5h.2b: porteado a Postgres vía SQLAlchemy.
    """
    try:
        from ...db import db_session as _db_session
        from ...models.org import Branch as _Branch, Company as _Company
        from sqlalchemy import case as _case, select as _select

        type_order = _case(
            (_Branch.type == "physical", 1),
            (_Branch.type == "deposit", 2),
            else_=3,
        )
        with _db_session() as session:
            rows = session.execute(
                _select(_Branch, _Company.name.label("company_name"))
                .join(_Company, _Company.id == _Branch.company_id, isouter=True)
                .where(_Branch.is_active.is_(True), _Branch.type.in_(["physical", "deposit", "admin"]))
                .order_by(type_order, _Branch.name)
            ).all()
            return [
                {
                    "id": b.id or "",
                    "name": b.name or "",
                    "code": b.code or "",
                    "type": b.type or "",
                    "company_id": b.company_id or "",
                    "company_name": company_name or "",
                }
                for b, company_name in rows
            ]
    except Exception:
        return []

def _warranty_central_deposit_from_branches(branches: list[dict[str, str]]) -> dict[str, str] | None:
    """Depósito operativo principal de Garantías.

    Regla actual del negocio: Chiclana es el depósito destino principal.
    Corrales y Cachi son depósitos de guarda, por lo que nunca deben ser
    fallback automático para remitos/ingresos de sucursal.
    """
    deposits = [b for b in branches if str(b.get("type") or "") == "deposit"]
    for b in deposits:
        key = _branch_key(f"{b.get('code','')} {b.get('name','')}")
        if "chiclana" in key:
            return b
    return None

def _warranty_central_deposit_name(branches: list[dict[str, str]] | None = None) -> str:
    if branches is None:
        branches = _fetch_branches_operativas()
    branch = _warranty_central_deposit_from_branches(branches)
    if branch and str(branch.get("name") or "").strip():
        return str(branch.get("name") or "").strip()
    for value in DEFAULT_DEPOSITOS:
        if "chiclana" in normalize_text(value):
            return strip_numeric_prefix(value)
    return "Depósito Chiclana"



# ──────────────────────────────────────────────────────────────────────────────
# Fuente de verdad organizativa para Garantías
# ──────────────────────────────────────────────────────────────────────────────
# Estos campos son los que deben conducir la lógica nueva:
#   company_id                 -> empresa imputada/relacionada
#   branch_id                  -> unidad de carga real (sucursal o depósito asignado)
#   sucursal_responsable_id    -> sucursal comercial responsable, cuando aplica
#   origen_ingreso             -> sucursal | deposito
#   tipo_ingreso               -> cliente_sucursal | cliente_deposito | falla_recepcion_mercaderia | ...
#   ubicacion_actual           -> sucursal | deposito | en_transito | proveedor | ...
# Los campos texto heredados (sucursal, deposito, lugar_llegada, transit_status)
# se mantienen para compatibilidad/display, pero no deberían ser la única fuente
# de permisos ni de decisiones nuevas.
def _derive_sucursal_fields(item: "WarrantyItemIn", user: Any, is_vendedor_sucursal: bool) -> tuple[str, str]:
    """Devuelve (sucursal_carga, sucursal_responsable).

    sucursal_carga       → origen físico del ingreso (se guarda en guarantees.sucursal).
    sucursal_responsable → rama comercialmente responsable (nueva columna).

    Reglas:
    · VENDEDOR (is_vendedor_sucursal=True): ambos campos = sucursal del usuario.
      El tipo de ingreso siempre es cliente_sucursal; no puede escalar ni sobre-escribir.
    · GESTOR/DEPOSITO — cliente_sucursal:
        carga = item.sucursal; responsable = item.sucursal_responsable o item.sucursal.
    · GESTOR/DEPOSITO — cliente_deposito:
        carga = sucursal del usuario (depósito); responsable = item.sucursal_responsable (obligatorio).
    · GESTOR/DEPOSITO — otros (falla_recepcion_mercaderia, stock_interno, otro):
        carga = sucursal del usuario (depósito); responsable = item.sucursal_responsable o user_branch.
    """
    tipo = (item.tipo_ingreso or "").strip()
    user_branch = str(getattr(user, "sucursal", "") or getattr(user, "branch_name", "") or "").strip()

    if is_vendedor_sucursal:
        return user_branch, user_branch

    suc_item = (item.sucursal or "").strip()
    suc_resp = (item.sucursal_responsable or "").strip()

    if tipo == "cliente_sucursal":
        return suc_item, suc_resp or suc_item

    if tipo == "cliente_deposito":
        # suc_resp ya fue validado como obligatorio en create_warranty_entries
        return user_branch, suc_resp

    # falla_recepcion_mercaderia, stock_interno, otro → depósito es carga y responsable
    return user_branch, suc_resp or user_branch


def _notify_gestor_garantias_pickup(title: str, message: str) -> None:
    """Notifica a usuarios con warranties.remitos.provider_delivery cuando se necesita acción urgente de logística."""
    try:
        targets = users_with_permission("warranties.remitos.provider_delivery")
        if targets:
            notify_many(targets, title, message, type_="warning")
    except Exception:
        pass  # notificaciones no son criticas


def is_provider_waiting_closed_status(status_value: str) -> bool:
    """Estados que cortan el contador de días sin respuesta del proveedor."""
    s = canonical_status_key(status_value)
    return any(token in s for token in [
        "RESPONDIDO POR PROVEEDOR", "RESUELTO", "RECHAZADO", "ANULADA", "FINALIZADO",
        # Tokens antiguos (mantener para historial)
        "LISTO PARA RETIRO", "APROBADO CAMBIO", "NOTA DE CREDITO", "REPARADO",
    ])


def is_resolved_status(status_value: str) -> bool:
    """Estados donde el proveedor ya definió o el caso quedó cerrado administrativamente."""
    s = canonical_status_key(status_value)
    return any(token in s for token in [
        "RESUELTO", "RECHAZADO", "ANULADA", "FINALIZADO",
        # Tokens antiguos (mantener para historial)
        "LISTO PARA RETIRO", "APROBADO CAMBIO", "NOTA DE CREDITO", "REPARADO",
    ])



def internal_logistics_ready_for_provider(row: dict[str, Any]) -> bool:
    """Indica si la garantía puede avanzar a flujo proveedor.

    Para garantías nacidas en sucursal, primero deben llegar físicamente a
    Chiclana/depósito mediante remito interno. Esto evita estados incoherentes
    como "ENVIADO AL PROVEEDOR" mientras el producto sigue en tránsito interno.
    Las garantías nacidas directamente en depósito se consideran listas.
    """
    keys = set(row.keys())
    origin = str(row["origen_ingreso"] or "").strip().lower() if "origen_ingreso" in keys else ""
    transit = str(row["transit_status"] or "").strip().lower() if "transit_status" in keys else ""
    location = str(row["ubicacion_actual"] or "").strip().lower() if "ubicacion_actual" in keys else ""
    if origin == "sucursal":
        return transit == "en_deposito" or location == "deposito" or _location_is_deposit_value(location)
    return True


def assert_internal_logistics_ready_for_provider(row: dict[str, Any]) -> None:
    if internal_logistics_ready_for_provider(row):
        return
    remito = str(row["remito_interno"] or "").strip() if "remito_interno" in row.keys() else ""
    if remito:
        raise HTTPException(
            status_code=400,
            detail=f"La garantía todavía no llegó a Depósito Chiclana. Confirmá la llegada del remito {remito} antes de avanzar con proveedor.",
        )
    raise HTTPException(
        status_code=400,
        detail="La garantía todavía está en sucursal. Generá/confirmá el remito interno a Depósito Chiclana antes de avanzar con proveedor.",
    )


def provider_flow_started(row: dict[str, Any]) -> bool:
    status_key = canonical_status_key(str(row["status"] or ""))
    return bool(str(row["sent_to_provider_at"] or "").strip()) or status_key in {
        "ENVIADO AL PROVEEDOR",
        "EN EL PROVEEDOR",
        "RESPONDIDO POR PROVEEDOR",
        "RESUELTO",
        "RECHAZADO",
        "ANULADA",
        "FINALIZADO",
    }


def provider_has_physical_product(row: dict[str, Any]) -> bool:
    """True cuando ya se registró el retiro físico del proveedor."""
    keys = set(row.keys())
    status_key = canonical_status_key(str(row["status"] or ""))
    location = str(row["ubicacion_actual"] or "").strip().lower() if "ubicacion_actual" in keys else ""
    pickup = normalize_provider_pickup_status(row["estado_retiro_proveedor"]) if "estado_retiro_proveedor" in keys else "sin_solicitud"
    return status_key == "EN EL PROVEEDOR" or location == "proveedor" or pickup == "retirado"


def assert_provider_has_physical_product(row: dict[str, Any]) -> None:
    if provider_has_physical_product(row):
        return
    raise HTTPException(
        status_code=400,
        detail="El proveedor todavía no tiene físicamente el producto. Primero registrá el retiro desde Depósito Chiclana.",
    )

def days_between(start_iso: str, end_iso: str | None = None) -> int:
    start = parse_iso_datetime(start_iso)
    if not start:
        return 0
    end = parse_iso_datetime(end_iso or "") or datetime.now(timezone.utc)
    return max(0, (end.date() - start.date()).days)


def compute_pending_days(row: dict[str, Any]) -> int:
    resolution = row["fecha_resolucion"] or row["cancelled_at"] or ""
    return days_between(row["ingreso_at"] or row["created_at"], resolution or None)


def compute_no_response_days(row: dict[str, Any]) -> int | None:
    if not row["sent_to_provider_at"]:
        return None
    if is_provider_waiting_closed_status(row["status"] or ""):
        return 0
    keys = set(row.keys())
    # El contador debe arrancar desde el último mail enviado/re-enviado al proveedor.
    # Las respuestas cierran la espera mediante is_provider_waiting_closed_status.
    last_mail = str(row["fecha_ultimo_mail_proveedor"] or "").strip() if "fecha_ultimo_mail_proveedor" in keys else ""
    base = last_mail or row["sent_to_provider_at"]
    return days_between(base)


# =========================================================
# Google Sheets: opciones/productos auxiliares
# =========================================================

def require_spreadsheet_id() -> str:
    spreadsheet_id = runtime_warranty_config().get("spreadsheet_id")
    if not spreadsheet_id:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Falta configurar la planilla de Garantías en Configuración operativa.",
        )
    return str(spreadsheet_id)


def get_values(sheet_name: str, a1: str) -> list[list[Any]]:
    service = sheets_service()
    spreadsheet_id = require_spreadsheet_id()
    result = service.spreadsheets().values().get(
        spreadsheetId=spreadsheet_id,
        range=f"{quote_sheet_name(sheet_name)}!{a1}",
    ).execute()
    return result.get("values", [])


def find_column(headers: list[str], candidates: list[str], fallback_index: int | None = None) -> int | None:
    keys = {header_key(c) for c in candidates}
    for index, header in enumerate(headers):
        if header_key(header) in keys:
            return index
    if fallback_index is not None and fallback_index < len(headers):
        return fallback_index
    return None


def read_options_from_sheet(sheet_name: str, max_rows: int = 1000) -> dict[str, list[str]]:
    try:
        values = get_values(sheet_name, f"A1:ZZ{max_rows}")
    except Exception:
        return {"sucursales": [], "depositos": [], "estados": []}
    if not values:
        return {"sucursales": [], "depositos": [], "estados": []}
    headers = [str(x).strip() for x in values[0]]
    sucursal_col = find_column(headers, ["SUCURSAL", "LOCAL", "ORIGEN", "SUCURSAL ORIGEN"])
    deposito_col = find_column(headers, ["DEPOSITO", "DEPÓSITO", "DESTINO", "DEPOSITO DESTINO", "LUGAR LLEGADA", "LUGAR DONDE LLEGA"])
    estado_col = find_column(headers, ["ESTADO", "STATUS", "ESTADO GARANTIA", "ESTADO GARANTÍA"])

    def collect(col: int | None) -> list[str]:
        if col is None:
            return []
        out: list[str] = []
        for row in values[1:]:
            if col >= len(row):
                continue
            value = clean_option_value(row[col])
            if not value:
                continue
            if header_key(value) in {"SUCURSAL", "DEPOSITO", "ESTADO", "STATUS"}:
                continue
            out.append(value)
        return unique_keep_order(out)

    return {"sucursales": collect(sucursal_col), "depositos": collect(deposito_col), "estados": collect(estado_col)}


def merge_options(*groups: list[str]) -> list[str]:
    merged: list[str] = []
    for group in groups:
        merged.extend(group or [])
    return unique_keep_order(merged)


def warranty_config_values() -> dict[str, Any]:
    cfg = runtime_warranty_config()
    statuses = list(cfg.get("statuses") or cfg.get("estados") or [])
    if not statuses:
        statuses = list(DEFAULT_STATUSES)
    # Estados canónicos. La config vieja puede traer estados desactualizados; se dejan
    # fuera para que el flujo visual y operativo sea único.
    statuses = unique_keep_order(DEFAULT_STATUSES)
    sucursales = list(cfg.get("sucursales") or DEFAULT_SUCURSALES)
    depositos = list(cfg.get("depositos") or DEFAULT_DEPOSITOS)
    # Finalizado es el cierre real. RESUELTO/RECHAZADO no significan cierre automático.
    final_statuses = list(DEFAULT_FINAL_STATUSES)
    delay_ranges_raw = cfg.get("delay_ranges") or DEFAULT_DELAY_RANGES
    delay_ranges: list[int] = []
    for value in delay_ranges_raw:
        try:
            number = int(value)
            if number > 0 and number not in delay_ranges:
                delay_ranges.append(number)
        except Exception:
            continue
    delay_ranges = sorted(delay_ranges or DEFAULT_DELAY_RANGES)
    required_review_fields = list(cfg.get("required_review_fields") or DEFAULT_REQUIRED_REVIEW_FIELDS)
    return {
        "statuses": statuses,
        "sucursales": sucursales,
        "depositos": depositos,
        "final_statuses": final_statuses,
        "delay_ranges": delay_ranges,
        "required_review_fields": required_review_fields,
        "raw_sheet": str(cfg.get("raw_sheet") or "00_RAW_GARANTIAS"),
        "spreadsheet_url": str(cfg.get("spreadsheet_url") or ""),
    }


def runtime_warranty_options() -> dict[str, Any]:
    cfg = runtime_warranty_config()
    values = warranty_config_values()

    # ── Branches reales del sistema (fuente de verdad) ────────────────────────
    branches_operativas = _fetch_branches_operativas()
    branches_physical = [b for b in branches_operativas if b["type"] == "physical"]
    branches_deposit  = [b for b in branches_operativas if b["type"] == "deposit"]

    # Sucursales: derivadas de branches físicas del sistema.
    # Si no hay branches configuradas todavía, caer a la config operativa como fallback.
    if branches_physical:
        sucursales = clean_select_options([b["name"] for b in branches_physical])
    else:
        sucursales = clean_select_options(list(values["sucursales"] or DEFAULT_SUCURSALES))

    # Depósitos: derivados de branches de tipo deposit del sistema.
    if branches_deposit:
        depositos = clean_select_options([b["name"] for b in branches_deposit], deposit=True)
    else:
        depositos = clean_select_options(list(values["depositos"] or DEFAULT_DEPOSITOS), deposit=True)

    warranty_central_deposit = _warranty_central_deposit_from_branches(branches_operativas)
    warranty_central_deposit_name = _warranty_central_deposit_name(branches_operativas)

    # Estados: forzar DEFAULT_STATUSES como única fuente de verdad visual/operativa.
    estados = unique_keep_order(DEFAULT_STATUSES)

    return {
        "sucursales": sucursales,
        "depositos": depositos,
        "warranty_central_deposit": warranty_central_deposit or {"id": "", "name": warranty_central_deposit_name, "code": "CHICLANA", "type": "deposit", "company_id": "", "company_name": ""},
        "estados": estados,
        "estado_default": str(cfg.get("estado_default") or DEFAULT_STATUSES[0]),
        "review_statuses": [{"value": key, "label": value} for key, value in REVIEW_LABELS.items()],
        "tipos_ingreso": [{"value": k, "label": v} for k, v in TIPO_INGRESO_LABELS.items()],
        "ubicacion_labels": UBICACION_LABELS,
        "resolution_options": [{"value": k, "label": v} for k, v in RESOLUTION_OPTIONS.items()],
        "final_statuses": values["final_statuses"],
        "delay_ranges": values["delay_ranges"],
        "required_review_fields": values["required_review_fields"],
        # Branches con IDs reales — el frontend las usa para selectores con ID.
        "branches_operativas": branches_operativas,
        "source": {"raw_sheet": values["raw_sheet"], "product_sheet": "Catálogo local", "mode": "database_primary"},
    }


def load_product_catalog() -> list[dict[str, str]]:
    cfg = runtime_warranty_config()
    now = time.time()
    cache_seconds = int(cfg.get("product_cache_seconds", 300) or 300)
    if PRODUCT_CACHE["items"] and now - float(PRODUCT_CACHE["loaded_at"]) < cache_seconds:
        return PRODUCT_CACHE["items"]
    try:
        values = get_values(str(cfg.get("product_sheet") or "Productos PVP"), "A:Z")
    except Exception:
        PRODUCT_CACHE["loaded_at"] = now
        PRODUCT_CACHE["items"] = []
        return []
    if not values:
        PRODUCT_CACHE["loaded_at"] = now
        PRODUCT_CACHE["items"] = []
        return []
    headers = [str(x).strip() for x in values[0]]
    producto_col = find_column(headers, ["PRODUCTO", "DESCRIPCION", "DESCRIPCIÓN", "ARTICULO", "ARTÍCULO", "NOMBRE"], fallback_index=2)
    sku_col = find_column(headers, ["SKU", "CODIGO", "CÓDIGO"], fallback_index=3)
    marca_col = find_column(headers, ["MARCA"], fallback_index=0)
    tipo_col = find_column(headers, ["TIPO", "RUBRO", "TIPO PRODUCTO"], fallback_index=1)
    items: list[dict[str, str]] = []
    for raw in values[1:]:
        def get(col: int | None) -> str:
            if col is None or col >= len(raw):
                return ""
            return str(raw[col]).strip()
        producto = get(producto_col)
        sku = get(sku_col)
        marca = get(marca_col)
        tipo = get(tipo_col)
        if not producto and not sku:
            continue
        label = " — ".join(part for part in [producto, sku] if part)
        items.append({"producto": producto, "sku": sku, "marca": marca, "tipo": tipo, "label": label, "search": normalize_text(" ".join([producto, sku, marca, tipo]))})
    PRODUCT_CACHE["loaded_at"] = now
    PRODUCT_CACHE["items"] = items
    return items


# =========================================================
# DB mappers
# =========================================================

def row_to_summary(row: dict[str, Any], items: list[dict[str, Any]]) -> WarrantySummary:
    products = [str(item["producto"] or "") for item in items if str(item["producto"] or "").strip()]
    first = items[0] if items else None
    ingreso = format_date_ar(parse_iso_datetime(row["ingreso_at"]) or now_ar())
    updated = format_datetime_ar(parse_iso_datetime(row["updated_at"]) or now_ar())
    return WarrantySummary(
        id_garantia=str(row["warranty_code"] or ""),
        parent_warranty_code=str(row["parent_warranty_code"] or "") if "parent_warranty_code" in row.keys() else "",
        parent_item_index=int(row["parent_item_index"] or 0) if "parent_item_index" in row.keys() and row["parent_item_index"] else None,
        grouped_item_label=(f"Ítem {int(row['parent_item_index']):02d} de {row['parent_warranty_code']}" if "parent_warranty_code" in row.keys() and row["parent_warranty_code"] and "parent_item_index" in row.keys() and row["parent_item_index"] else ""),
        ingreso=ingreso,
        ingreso_iso=date_input_from_iso(row["ingreso_at"]),
        responsible_username=str(row["responsible_username"] or ""),
        responsable=str(row["responsible_name"] or ""),
        usuario=str(row["created_by"] or ""),
        producto_principal=(products[0] if products else (str(first["producto"] or "") if first else "")),
        productos=products[:12],
        cantidad_items=len(items),
        marca=str(first["marca"] or "") if first else "",
        sku=str(first["sku"] or "") if first else "",
        serie=str(first["serie"] or "") if first else "",
        falla=str(first["falla"] or "") if first else "",
        sucursal=str(row["sucursal"] or ""),
        sucursal_code=str(row["sucursal_code"] or ""),
        branch_id=str(row["branch_id"] or "") if "branch_id" in row.keys() else "",
        company_id=str(row["company_id"] or "") if "company_id" in row.keys() else "",
        sucursal_responsable=str(row["sucursal_responsable"] or "") if "sucursal_responsable" in row.keys() else "",
        sucursal_responsable_id=str(row["sucursal_responsable_id"] or "") if "sucursal_responsable_id" in row.keys() else "",
        deposito=str(row["deposito"] or ""),
        lugar_llegada=str(row["lugar_llegada"] or row["deposito"] or ""),
        estado=str(row["status"] or ""),
        tipo_ingreso=str(row["tipo_ingreso"] or "") if "tipo_ingreso" in row.keys() else "",
        tipo_ingreso_label=TIPO_INGRESO_LABELS.get(str(row["tipo_ingreso"] or ""), str(row["tipo_ingreso"] or "")) if "tipo_ingreso" in row.keys() else "",
        origen_ingreso=str(row["origen_ingreso"] or "") if "origen_ingreso" in row.keys() else "",
        ubicacion_actual=str(row["ubicacion_actual"] or "") if "ubicacion_actual" in row.keys() else "",
        ubicacion_actual_label=UBICACION_LABELS.get(str(row["ubicacion_actual"] or ""), str(row["ubicacion_actual"] or "")) if "ubicacion_actual" in row.keys() else "",
        cliente_nombre=str(row["cliente_nombre"] or "") if "cliente_nombre" in row.keys() else "",
        cliente_telefono=str(row["cliente_telefono"] or "") if "cliente_telefono" in row.keys() else "",
        cliente_email=str(row["cliente_email"] or "") if "cliente_email" in row.keys() else "",
        numero_factura=str(row["numero_factura"] or "") if "numero_factura" in row.keys() else "",
        fecha_compra=str(row["fecha_compra"] or "") if "fecha_compra" in row.keys() else "",
        review_status=str(row["review_status"] or REVIEW_PENDING),
        review_status_label=REVIEW_LABELS.get(str(row["review_status"] or REVIEW_PENDING), str(row["review_status"] or REVIEW_PENDING)),
        reviewed_by=str(row["reviewed_by"] or ""),
        reviewed_by_name=str(row["reviewed_by_name"] or ""),
        reviewed_at=format_datetime_ar(parse_iso_datetime(row["reviewed_at"])) if row["reviewed_at"] else "",
        review_note=str(row["review_note"] or ""),
        observaciones=str(row["observations"] or ""),
        photos_reference=str(row["photos_reference"] or ""),
        provider_name=str(row["provider_name"] or ""),
        id_de_caso=str(row["provider_case_id"] or ""),
        fecha_envio_proveedor=format_datetime_ar(parse_iso_datetime(row["sent_to_provider_at"])) if row["sent_to_provider_at"] else "",
        fecha_ultima_respuesta=format_datetime_ar(parse_iso_datetime(row["last_provider_response_at"])) if row["last_provider_response_at"] else "",
        fecha_ultimo_reclamo=format_datetime_ar(parse_iso_datetime(row["last_claim_at"])) if "last_claim_at" in row.keys() and row["last_claim_at"] else "",
        estado_retiro_proveedor=normalize_provider_pickup_status(row["estado_retiro_proveedor"]) if "estado_retiro_proveedor" in row.keys() else "sin_solicitud",
        estado_retiro_proveedor_label=PROVIDER_PICKUP_STATUSES.get(normalize_provider_pickup_status(row["estado_retiro_proveedor"]) if "estado_retiro_proveedor" in row.keys() else "sin_solicitud", "Sin solicitud"),
        provider_response_type=normalize_provider_response_type(row["provider_response_type"]) if "provider_response_type" in row.keys() else "",
        provider_response_type_label=PROVIDER_RESPONSE_TYPES.get(normalize_provider_response_type(row["provider_response_type"]) if "provider_response_type" in row.keys() else "", ""),
        provider_correction_note=str(row["provider_correction_note"] or "") if "provider_correction_note" in row.keys() else "",
        fecha_solicitud_retiro_proveedor=format_datetime_ar(parse_iso_datetime(row["fecha_solicitud_retiro_proveedor"])) if "fecha_solicitud_retiro_proveedor" in row.keys() and row["fecha_solicitud_retiro_proveedor"] else "",
        fecha_retiro_proveedor=format_datetime_ar(parse_iso_datetime(row["fecha_retiro_proveedor"])) if "fecha_retiro_proveedor" in row.keys() and row["fecha_retiro_proveedor"] else (format_datetime_ar(parse_iso_datetime(row["fecha_retiro"])) if "fecha_retiro" in row.keys() and row["fecha_retiro"] else ""),
        dias_pendiente=compute_pending_days(row),
        dias_sin_respuesta=compute_no_response_days(row),
        shipment_code=str(row["shipment_code"] or "") if "shipment_code" in row.keys() else "",
        shipment_file_name=str(row["shipment_file_name"] or "") if "shipment_file_name" in row.keys() else "",
        resolution_note=str(row["resolution_note"] or "") if "resolution_note" in row.keys() else "",
        resolution_reference=str(row["resolution_reference"] or "") if "resolution_reference" in row.keys() else "",
        resultado_resolucion=normalize_resolution_result(row["resultado_resolucion"]) if "resultado_resolucion" in row.keys() else "",
        resultado_resolucion_label=RESOLUTION_OPTIONS.get(normalize_resolution_result(row["resultado_resolucion"]) if "resultado_resolucion" in row.keys() else "", ""),
        numero_nota_credito=str(row["numero_nota_credito"] or "") if "numero_nota_credito" in row.keys() else "",
        importe_nota_credito=str(row["importe_nota_credito"] or "") if "importe_nota_credito" in row.keys() else "",
        fecha_nota_credito=str(row["fecha_nota_credito"] or "") if "fecha_nota_credito" in row.keys() else "",
        detalle_reparacion=str(row["detalle_reparacion"] or "") if "detalle_reparacion" in row.keys() else "",
        fecha_reparacion=str(row["fecha_reparacion"] or "") if "fecha_reparacion" in row.keys() else "",
        producto_reemplazo=str(row["producto_reemplazo"] or "") if "producto_reemplazo" in row.keys() else "",
        sku_reemplazo=str(row["sku_reemplazo"] or "") if "sku_reemplazo" in row.keys() else "",
        serie_reemplazo=str(row["serie_reemplazo"] or "") if "serie_reemplazo" in row.keys() else "",
        fecha_recepcion_reemplazo=str(row["fecha_recepcion_reemplazo"] or "") if "fecha_recepcion_reemplazo" in row.keys() else "",
        fecha_finalizacion=format_datetime_ar(parse_iso_datetime(row["fecha_finalizacion"])) if "fecha_finalizacion" in row.keys() and row["fecha_finalizacion"] else "",
        finalizacion=str(row["finalizacion"] or "") if "finalizacion" in row.keys() else "",
        remito_interno=str(row["remito_interno"] or "") if "remito_interno" in row.keys() else "",
        remito_proveedor=str(row["remito_proveedor"] or "") if "remito_proveedor" in row.keys() else "",
        transit_status=str(row["transit_status"] or "") if "transit_status" in row.keys() else "",
        synced_to_google_sheet=bool(row["synced_to_google_sheet"]),
        fecha_ultima_sincronizacion=format_datetime_ar(parse_iso_datetime(row["last_google_sync_at"])) if row["last_google_sync_at"] else "",
        actualizado_por=str(row["updated_by_name"] or row["updated_by"] or ""),
        fecha_ultima_actualizacion=updated,
        cancelled=bool(row["cancelled"]),
        cancel_reason=str(row["cancel_reason"] or ""),
        cancelled_by=str(row["cancelled_by"] or ""),
        cancelled_at=format_datetime_ar(parse_iso_datetime(row["cancelled_at"])) if row["cancelled_at"] else "",
    )


def item_to_row(warranty_row: dict[str, Any], item: dict[str, Any], index: int) -> WarrantyRow:
    return WarrantyRow(
        row_number=int(item["id"] or index),
        id_garantia=str(warranty_row["warranty_code"] or ""),
        responsable=str(warranty_row["responsible_name"] or ""),
        usuario=str(warranty_row["created_by"] or ""),
        ingreso=format_date_ar(parse_iso_datetime(warranty_row["ingreso_at"]) or now_ar()),
        producto=str(item["producto"] or ""),
        sku=str(item["sku"] or ""),
        marca=str(item["marca"] or ""),
        tipo=str(item["tipo"] or ""),
        serie=str(item["serie"] or ""),
        falla=str(item["falla"] or ""),
        sucursal=str(warranty_row["sucursal"] or ""),
        deposito=str(warranty_row["deposito"] or ""),
        lugar_llegada=str(warranty_row["lugar_llegada"] or warranty_row["deposito"] or ""),
        estado=str(warranty_row["status"] or ""),
        observaciones=str(item["observaciones"] or warranty_row["observations"] or ""),
        correction_note=str(item["correction_note"] or "") if "correction_note" in item.keys() else "",
        actualizado_por=str(warranty_row["updated_by_name"] or warranty_row["updated_by"] or ""),
        fecha_ultima_actualizacion=format_datetime_ar(parse_iso_datetime(warranty_row["updated_at"]) or now_ar()),
    )


def fetch_all_guarantee_summaries() -> list[WarrantySummary]:
    rows, all_items = pg_fetch_all_guarantee_rows()
    by_gid: dict[int, list[dict[str, Any]]] = {}
    for item in all_items:
        by_gid.setdefault(int(item["guarantee_id"]), []).append(item)
    return [row_to_summary(row, by_gid.get(int(row["id"]), [])) for row in rows]


def validate_status_or_400(estado: str) -> str:
    clean = str(estado or "").strip()
    allowed = {canonical_status_key(x): x for x in DEFAULT_STATUSES}
    key = canonical_status_key(clean)
    if key not in allowed:
        raise HTTPException(status_code=400, detail=f"Estado inválido para garantías: {clean}")
    return allowed[key]


def warranty_exports_dir() -> Path:
    settings = get_settings()
    settings.ensure_dirs()
    path = settings.outputs_dir / "warranties" / "exports"
    path.mkdir(parents=True, exist_ok=True)
    return path


def safe_filename_part(value: Any, fallback: str = "general") -> str:
    text = normalize_text(value).lower().replace(" ", "-")
    text = re.sub(r"[^a-z0-9-]+", "", text).strip("-")
    return text or fallback


def export_row_matches(row: dict[str, Any], filters: WarrantyExportRequest) -> bool:
    if filters.marca and normalize_text(row.get("marca")) != normalize_text(filters.marca):
        return False
    if filters.proveedor and normalize_text(row.get("provider_name")) != normalize_text(filters.proveedor):
        return False
    if filters.estado and normalize_text(row.get("status")) != normalize_text(filters.estado):
        return False
    if filters.sucursal and normalize_text(row.get("sucursal")) != normalize_text(filters.sucursal):
        return False
    if filters.deposito:
        dep = normalize_text(row.get("deposito"))
        llegada = normalize_text(row.get("lugar_llegada"))
        wanted = normalize_text(filters.deposito)
        if wanted not in {dep, llegada}:
            return False
    date_from = parse_date_filter(filters.fecha_desde)
    date_to = parse_date_filter(filters.fecha_hasta)
    ingreso_dt = parse_iso_datetime(row.get("ingreso_at"))
    ingreso_date = ingreso_dt.date() if ingreso_dt else None
    if date_from and ingreso_date and ingreso_date < date_from:
        return False
    if date_to and ingreso_date and ingreso_date > date_to:
        return False
    return True


def collect_export_rows(filters: WarrantyExportRequest) -> list[dict[str, Any]]:
    output: list[dict[str, Any]] = []
    for item in pg_collect_export_rows():
        if export_row_matches(item, filters):
            output.append(item)
    return output


def export_info_from_row(row: dict[str, Any]) -> WarrantyExportInfo:
    filters: dict[str, Any] = {}
    try:
        parsed = json.loads(row["filters_json"] or "{}")
        if isinstance(parsed, dict):
            filters = parsed
    except Exception:
        filters = {}
    return WarrantyExportInfo(
        id=int(row["id"]),
        created_at=format_datetime_ar(parse_iso_datetime(row["created_at"]) or now_ar()),
        created_by=str(row["created_by_name"] or row["created_by"] or ""),
        provider_name=str(row["provider_name"] or ""),
        marca=str(row["marca"] or ""),
        filters=filters,
        file_name=str(row["file_name"] or ""),
        row_count=int(row["row_count"] or 0),
        download_url=f"/api/warranties/exports/{int(row['id'])}/download",
        shipment_code=str(row["shipment_code"] or "") if "shipment_code" in row.keys() else "",
        file_format=str(row["file_format"] or "excel") if "file_format" in row.keys() else "excel",
        logo_brand=str(row["logo_brand"] or "gv_electro") if "logo_brand" in row.keys() else "gv_electro",
    )


def _export_brand_info(logo_brand: str) -> dict[str, str]:
    brand = normalize_export_logo(logo_brand)
    if brand == "abc_electro":
        return {"key": "abc_electro", "label": "ABC Electro", "logo_file": "abc_electro.png", "accent": "2563EB"}
    return {"key": "gv_electro", "label": "GV Electro", "logo_file": "gv_electro.png", "accent": "1E293B"}


def _export_logo_path(logo_brand: str) -> Path | None:
    return brand_logo_path(logo_brand)


def build_provider_excel(rows: list[dict[str, Any]], file_path: Path, *, provider_name: str = "", shipment_code: str = "", logo_brand: str = "gv_electro") -> None:
    """Excel externo profesional para proveedor.

    La cabecera puede incluir logo/título, pero la tabla mantiene solo los campos
    que el proveedor necesita: ID, producto, SKU, serie y falla.
    """
    brand = _export_brand_info(logo_brand)
    accent = brand["accent"]
    dark = "0F172A"
    muted = "64748B"
    soft = "F8FAFC"
    row_alt = "F1F5F9"

    wb = Workbook()
    ws = wb.active
    ws.title = "Garantías"
    ws.sheet_view.showGridLines = False

    # Configuración de impresión: que el archivo se vea presentable al abrirlo
    # y también al exportarlo/imprimirlo desde Excel.
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins.left = 0.35
    ws.page_margins.right = 0.35
    ws.page_margins.top = 0.45
    ws.page_margins.bottom = 0.45

    # Anchos pensados para lectura real, no solo para que entren los datos.
    widths = [24, 64, 20, 22, 56]
    for index, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(index)].width = width

    # Bloque de encabezado profesional.
    for row_num in range(1, 7):
        ws.row_dimensions[row_num].height = 24
        for col_num in range(1, 6):
            ws.cell(row=row_num, column=col_num).fill = PatternFill("solid", fgColor=soft)

    ws.row_dimensions[1].height = 32
    ws.row_dimensions[2].height = 24
    ws.row_dimensions[3].height = 22
    ws.row_dimensions[4].height = 22
    ws.row_dimensions[5].height = 18
    ws.row_dimensions[6].height = 8

    logo_path = _export_logo_path(logo_brand)
    if logo_path:
        try:
            logo = XLImage(str(logo_path))
            logo.width = 86
            logo.height = 86
            ws.add_image(logo, "A1")
        except Exception:
            pass

    # Encabezado textual. Se deja espacio al logo en A1:A4.
    ws.merge_cells("B1:E1")
    ws["B1"] = "LOTE DE GARANTÍAS PARA PROVEEDOR"
    ws["B1"].font = Font(bold=True, size=18, color=dark)
    ws["B1"].alignment = Alignment(horizontal="left", vertical="center")

    ws.merge_cells("B2:E2")
    ws["B2"] = f"N° {shipment_code or 'ENV'}"
    ws["B2"].font = Font(bold=True, size=11, color=accent)
    ws["B2"].alignment = Alignment(horizontal="left", vertical="center")

    ws.merge_cells("B3:E3")
    ws["B3"] = f"Proveedor: {provider_name or '—'}"
    ws["B3"].font = Font(bold=True, size=11, color=dark)
    ws["B3"].alignment = Alignment(horizontal="left", vertical="center")

    ws.merge_cells("B4:E4")
    ws["B4"] = f"Fecha de emisión: {format_datetime_ar(now_ar())} · Cantidad: {len(rows)} garantía(s)"
    ws["B4"].font = Font(size=9, color=muted)
    ws["B4"].alignment = Alignment(horizontal="left", vertical="center")

    ws.merge_cells("A6:E6")
    ws["A6"] = ""
    ws["A6"].fill = PatternFill("solid", fgColor=accent)

    headers = ["ID GARANTÍA", "PRODUCTO", "SKU", "N° SERIE", "FALLA"]
    header_row = 8
    for col, value in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col, value=value)
        cell.fill = PatternFill("solid", fgColor=dark)
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[header_row].height = 26

    thin = Side(style="thin", color="CBD5E1")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for row_idx, item in enumerate(rows, start=header_row + 1):
        values = [
            item.get("warranty_code") or "",
            item.get("producto") or "",
            item.get("sku") or "",
            item.get("serie") or "",
            item.get("falla") or "",
        ]
        fill = PatternFill("solid", fgColor=row_alt if (row_idx - header_row) % 2 == 0 else "FFFFFF")
        # Altura dinámica simple para que fallas o productos largos no queden aplastados.
        product_len = len(str(values[1] or ""))
        falla_len = len(str(values[4] or ""))
        ws.row_dimensions[row_idx].height = max(28, min(64, 18 + 10 * max(product_len // 55, falla_len // 45)))
        for col, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.fill = fill
            cell.border = border
            cell.font = Font(size=10, color="111827")
            if col == 1:
                cell.font = Font(size=10, bold=True, color=dark)
            if col in (3, 4):
                # Texto explícito para evitar que Excel convierta series/SKU largos en números.
                cell.number_format = "@"
            cell.alignment = Alignment(vertical="top", wrap_text=True, horizontal="left")

    # Bordes de encabezado al final para que queden por encima de los estilos.
    for cell in ws[header_row]:
        cell.border = border

    last_row = header_row + max(len(rows), 1)
    if not rows:
        ws.cell(row=header_row + 1, column=1, value="Sin garantías seleccionadas")
        ws.merge_cells(start_row=header_row + 1, start_column=1, end_row=header_row + 1, end_column=5)
        ws.cell(row=header_row + 1, column=1).alignment = Alignment(horizontal="center")
        ws.cell(row=header_row + 1, column=1).font = Font(italic=True, color=muted)
        last_row = header_row + 1

    try:
        table_ref = f"A{header_row}:E{last_row}"
        tab = XLTable(displayName="TablaGarantiasProveedor", ref=table_ref)
        style = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False, showRowStripes=False, showColumnStripes=False)
        tab.tableStyleInfo = style
        ws.add_table(tab)
    except Exception:
        ws.auto_filter.ref = f"A{header_row}:E{last_row}"

    ws.freeze_panes = f"A{header_row + 1}"
    ws.print_title_rows = f"{header_row}:{header_row}"

    # Sin leyendas internas ni datos de empresa: el logo identifica la marca emisora.

    file_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(file_path)

def build_provider_pdf(rows: list[dict[str, Any]], file_path: Path, *, provider_name: str = "", shipment_code: str = "", logo_brand: str = "gv_electro") -> None:
    """PDF externo profesional para proveedor con tabla mínima.

    Fase 44: se genera en horizontal y con celdas envueltas para evitar
    superposición en productos/SKU/series largos.
    """
    brand = _export_brand_info(logo_brand)
    file_path.parent.mkdir(parents=True, exist_ok=True)

    page_size = landscape(A4)
    c = canvas.Canvas(str(file_path), pagesize=page_size)
    width, height = page_size
    margin_x = 12 * mm
    top_y = height - 12 * mm
    logo_path = _export_logo_path(logo_brand)

    # Encabezado horizontal: más aire y más ancho útil para la tabla.
    if logo_path:
        try:
            c.drawImage(str(logo_path), margin_x, top_y - 24 * mm, width=24 * mm, height=24 * mm, mask="auto")
        except Exception:
            pass

    c.setFillColor(colors.HexColor("#111827"))
    c.setFont("Helvetica-Bold", 18)
    c.drawRightString(width - margin_x, top_y - 6 * mm, "LOTE DE GARANTÍAS")
    c.setFont("Helvetica-Bold", 11)
    c.drawRightString(width - margin_x, top_y - 13 * mm, "PARA PROVEEDOR")
    c.setFont("Helvetica-Bold", 10)
    c.drawRightString(width - margin_x, top_y - 20 * mm, f"N° {shipment_code or 'ENV'}")
    c.setFont("Helvetica", 8)
    c.drawRightString(width - margin_x, top_y - 25 * mm, format_datetime_ar(now_ar()))

    y = top_y - 32 * mm
    c.setStrokeColor(colors.HexColor("#334155"))
    c.setLineWidth(0.8)
    c.line(margin_x, y, width - margin_x, y)

    y -= 8 * mm
    c.setFont("Helvetica-Bold", 10)
    c.setFillColor(colors.HexColor("#0F172A"))
    c.drawString(margin_x, y, f"Proveedor: {provider_name or '—'}")
    y -= 9 * mm

    header_style = ParagraphStyle(
        "ProviderHeaderCell",
        fontName="Helvetica-Bold",
        fontSize=7.5,
        leading=9,
        textColor=colors.white,
        alignment=0,
    )
    cell_style = ParagraphStyle(
        "ProviderBodyCell",
        fontName="Helvetica",
        fontSize=7.4,
        leading=9.2,
        textColor=colors.HexColor("#111827"),
        wordWrap="CJK",
        splitLongWords=True,
    )

    def pcell(value: Any, style: ParagraphStyle = cell_style) -> Paragraph:
        text = str(value or "")
        text = (
            text.replace("&", "&amp;")
            .replace("<", "&lt;")
            .replace(">", "&gt;")
            .replace("\n", "<br/>")
        )
        return Paragraph(text, style)

    headers = ["ID GARANTIA", "PRODUCTO", "SKU", "N° SERIE", "FALLA"]
    data: list[list[Any]] = [[pcell(h, header_style) for h in headers]]
    for item in rows:
        data.append([
            pcell(item.get("warranty_code") or ""),
            pcell(item.get("producto") or ""),
            pcell(item.get("sku") or ""),
            pcell(item.get("serie") or ""),
            pcell(item.get("falla") or ""),
        ])

    # Anchos para A4 apaisado. El producto y falla tienen prioridad; SKU/serie
    # también envuelven para no invadir columnas vecinas.
    col_widths = [38 * mm, 96 * mm, 42 * mm, 38 * mm, 58 * mm]
    table = Table(data, colWidths=col_widths, repeatRows=1, hAlign="LEFT")
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#" + brand["accent"])),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#CBD5E1")),
        ("LEFTPADDING", (0, 0), (-1, -1), 4),
        ("RIGHTPADDING", (0, 0), (-1, -1), 4),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8FAFC")]),
    ]))

    # Permitir múltiples páginas si hay muchos ítems.
    available_height = y - 14 * mm
    parts = table.splitOn(c, width - 2 * margin_x, available_height) or [table]
    for idx, part in enumerate(parts):
        if idx > 0:
            c.showPage()
            y = height - 16 * mm
            available_height = y - 14 * mm
        tw, th = part.wrapOn(c, width - 2 * margin_x, available_height)
        part.drawOn(c, margin_x, y - th)

    c.setFont("Helvetica", 6.5)
    c.setFillColor(colors.HexColor("#94A3B8"))
    c.drawCentredString(width / 2, 8 * mm, f"{shipment_code or ''}")
    c.save()


# =========================================================
# Sincronización controlada con Google Sheets
# =========================================================

def sheet_raw_name() -> str:
    cfg = runtime_warranty_config()
    return str(cfg.get("raw_sheet") or "00_RAW_GARANTIAS")


def _spreadsheet_titles(service: Any, spreadsheet_id: str) -> set[str]:
    meta = service.spreadsheets().get(spreadsheetId=spreadsheet_id, fields="sheets(properties(title))").execute()
    return {str(sheet.get("properties", {}).get("title", "")) for sheet in meta.get("sheets", [])}


def ensure_sheet_with_headers(service: Any, spreadsheet_id: str, sheet_name: str, headers: list[str], existing_titles: set[str] | None = None) -> bool:
    """Crea/verifica una pestaña y escribe headers. Devuelve True si la creó."""
    created = False
    titles = existing_titles if existing_titles is not None else _spreadsheet_titles(service, spreadsheet_id)
    if sheet_name not in titles:
        service.spreadsheets().batchUpdate(
            spreadsheetId=spreadsheet_id,
            body={"requests": [{"addSheet": {"properties": {"title": sheet_name}}}]},
        ).execute()
        titles.add(sheet_name)
        created = True
    service.spreadsheets().values().update(
        spreadsheetId=spreadsheet_id,
        range=f"{quote_sheet_name(sheet_name)}!A1:{get_column_letter(len(headers))}1",
        valueInputOption="USER_ENTERED",
        body={"values": [headers]},
    ).execute()
    return created


def ensure_mirror_sheets() -> dict[str, bool]:
    service = sheets_service()
    spreadsheet_id = require_spreadsheet_id()
    existing = _spreadsheet_titles(service, spreadsheet_id)
    created: dict[str, bool] = {}
    # Raw histórico + pestañas espejo nuevas.
    created[sheet_raw_name()] = ensure_sheet_with_headers(service, spreadsheet_id, sheet_raw_name(), DEFAULT_RAW_HEADERS, existing)
    for name, headers in MIRROR_SHEETS.items():
        created[name] = ensure_sheet_with_headers(service, spreadsheet_id, name, headers, existing)
    return created


def _fmt_sheet_dt(value: Any) -> str:
    if not value:
        return ""
    return format_datetime_ar(parse_iso_datetime(value)) if parse_iso_datetime(value) else str(value)


def _pg_item_summary(items_list: list[dict[str, Any]]) -> tuple[int, str, str, str, str]:
    """Resumen de items de una garantía (cantidad + datos del primero)."""
    if not items_list:
        return 0, "", "", "", ""
    sorted_items = sorted(items_list, key=lambda it: int(it.get("id") or 0))
    first = sorted_items[0]
    return (
        len(sorted_items),
        str(first.get("producto") or ""),
        str(first.get("sku") or ""),
        str(first.get("serie") or ""),
        str(first.get("falla") or ""),
    )


def mirror_rows_for_sheet(sheet_name: str) -> list[list[Any]]:
    """Fase 2.5h.2e · espejo Google Sheets desde Postgres.

    Reemplaza al helper legacy de espejo: ahora lee
    de Postgres usando ``pg_fetch_all_guarantee_rows`` + ``pg_list_remitos`` +
    ``pg_list_exports`` + ``pg_all_history``. La estructura visible (orden y
    columnas por sheet) se mantiene idéntica para no romper consumidores.
    """
    from ...remitos_db import pg_list_remitos
    sheet = sheet_name.upper().strip()

    if sheet == "GARANTIAS":
        all_rows, all_items = pg_fetch_all_guarantee_rows()
        rows = sorted(all_rows, key=lambda r: (r.get("ingreso_at") or "", r.get("warranty_code") or "", int(r.get("id") or 0)))
        items_by_gid: dict[int, list[dict[str, Any]]] = {}
        for it in all_items:
            items_by_gid.setdefault(int(it["guarantee_id"]), []).append(it)
        values: list[list[Any]] = []
        for row in rows:
            _count, producto, sku, serie, falla = _pg_item_summary(items_by_gid.get(int(row["id"]), []))
            days = compute_no_response_days(row)
            values.append([
                row.get("warranty_code"),
                _fmt_sheet_dt(row.get("ingreso_at")),
                row.get("company_id") or "",
                row.get("sucursal") or "",
                row.get("sucursal_responsable") or row.get("sucursal") or "",
                row.get("origen_ingreso") or "",
                row.get("tipo_ingreso") or "",
                row.get("ubicacion_actual") or row.get("lugar_llegada") or "",
                row.get("deposito") or "",
                normalize_status(row.get("status")),
                row.get("review_status") or "",
                row.get("remito_interno") or "",
                row.get("shipment_code") or "",
                row.get("provider_name") or "",
                _fmt_sheet_dt(row.get("sent_to_provider_at")),
                _fmt_sheet_dt(row.get("fecha_ultimo_mail_proveedor")),
                "" if days is None else days,
                row.get("estado_retiro_proveedor") or "",
                _fmt_sheet_dt(row.get("fecha_retiro")),
                row.get("provider_response_type") or "",
                row.get("resultado_resolucion") or "",
                _fmt_sheet_dt(row.get("fecha_resolucion")),
                _fmt_sheet_dt(row.get("fecha_finalizacion")),
                row.get("cliente_nombre") or "",
                row.get("cliente_telefono") or "",
                row.get("cliente_email") or "",
                row.get("numero_factura") or "",
                row.get("fecha_compra") or "",
                row.get("responsible_name") or row.get("responsible_username") or "",
                row.get("created_by_name") or row.get("created_by") or "",
                _fmt_sheet_dt(row.get("updated_at")),
                row.get("updated_by_name") or row.get("updated_by") or "",
                row.get("observations") or "",
            ])
        return values

    if sheet == "GARANTIA_ITEMS":
        all_rows, all_items = pg_fetch_all_guarantee_rows()
        by_gid = {int(g["id"]): g for g in all_rows}
        sorted_items = sorted(
            all_items,
            key=lambda it: (
                by_gid.get(int(it["guarantee_id"]), {}).get("ingreso_at") or "",
                by_gid.get(int(it["guarantee_id"]), {}).get("warranty_code") or "",
                int(it.get("id") or 0),
            ),
        )
        return [[
            by_gid.get(int(it["guarantee_id"]), {}).get("warranty_code") or "",
            it.get("item_index") or it.get("id"),
            it.get("producto") or "", it.get("sku") or "", it.get("marca") or "",
            it.get("tipo") or "", it.get("serie") or "", it.get("falla") or "",
            "",  # proveedor (columna histórica del item, ya no existe en el modelo)
            it.get("observaciones") or "",
            _fmt_sheet_dt(by_gid.get(int(it["guarantee_id"]), {}).get("updated_at")),
        ] for it in sorted_items]

    if sheet in {"REMITOS", "REMITO_ITEMS"}:
        remitos = pg_list_remitos(limit=100000)
        remitos.sort(key=lambda r: (r.get("created_at") or "", r.get("remito_code") or ""))
        if sheet == "REMITOS":
            return [[
                r.get("remito_code"),
                r.get("tipo_remito") or "sucursal_a_deposito",
                r.get("company_brand") or "",
                r.get("origen_sucursal") or "",
                r.get("destino_deposito") or "",
                r.get("status") or "",
                _fmt_sheet_dt(r.get("created_at")),
                _fmt_sheet_dt(r.get("fecha_despacho")),
                _fmt_sheet_dt(r.get("fecha_llegada")),
                r.get("created_by_name") or r.get("created_by") or "",
                r.get("despachado_por_name") or r.get("despachado_por") or "",
                r.get("recibido_por_name") or r.get("recibido_por") or "",
                len(r.get("warranty_ids") or []),
                f"/api/warranties/remitos/pdf/{r.get('remito_code')}",
                r.get("nota") or "",
            ] for r in remitos]
        # REMITO_ITEMS — desnormalizado: 1 fila por warranty_code en cada remito.
        all_rows, all_items = pg_fetch_all_guarantee_rows()
        by_code = {str(g.get("warranty_code") or ""): g for g in all_rows}
        items_by_gid: dict[int, list[dict[str, Any]]] = {}
        for it in all_items:
            items_by_gid.setdefault(int(it["guarantee_id"]), []).append(it)
        values: list[list[Any]] = []
        for r in remitos:
            for code in (r.get("warranty_ids") or []):
                g = by_code.get(str(code))
                if g:
                    _count, producto, sku, serie, _falla = _pg_item_summary(items_by_gid.get(int(g["id"]), []))
                else:
                    producto = sku = serie = ""
                values.append([
                    r.get("remito_code"), code, producto, sku, serie,
                    (g or {}).get("sucursal_responsable") or (g or {}).get("sucursal") or "",
                    r.get("origen_sucursal") or "", r.get("destino_deposito") or "", r.get("status") or "",
                    _fmt_sheet_dt((g or {}).get("updated_at") or r.get("created_at")),
                ])
        return values

    if sheet in {"LOTES_ENV", "LOTE_ITEMS"}:
        exports = pg_list_exports(limit=100000)
        exports.sort(key=lambda e: (e.get("created_at") or "", e.get("shipment_code") or ""))
        if sheet == "LOTES_ENV":
            return [[
                e.get("shipment_code") or "", e.get("provider_name") or "", e.get("marca") or "",
                "excel_generado",  # estado simplificado post-port (el legacy hacía un sub-SELECT por status)
                _fmt_sheet_dt(e.get("created_at")),
                e.get("created_by_name") or e.get("created_by") or "",
                e.get("file_name") or "",
                e.get("row_count") or 0,
                "", "",
                json.dumps(e.get("filters") or {}, ensure_ascii=False),
            ] for e in exports]
        # LOTE_ITEMS
        all_rows, all_items = pg_fetch_all_guarantee_rows()
        by_code = {str(g.get("warranty_code") or ""): g for g in all_rows}
        items_by_gid: dict[int, list[dict[str, Any]]] = {}
        for it in all_items:
            items_by_gid.setdefault(int(it["guarantee_id"]), []).append(it)
        values: list[list[Any]] = []
        for e in exports:
            for code in (e.get("warranty_ids") or []):
                g = by_code.get(str(code))
                if g:
                    _count, producto, sku, serie, _falla = _pg_item_summary(items_by_gid.get(int(g["id"]), []))
                else:
                    producto = sku = serie = ""
                values.append([
                    e.get("shipment_code") or "", code, producto, sku, serie,
                    (g or {}).get("provider_name") or e.get("provider_name") or "",
                    normalize_status((g or {}).get("status") or ""),
                    (g or {}).get("provider_response_type") or "",
                    (g or {}).get("resultado_resolucion") or "",
                    _fmt_sheet_dt((g or {}).get("updated_at") or e.get("created_at")),
                ])
        return values

    if sheet == "EVENTOS":
        from ...warranties_db import pg_all_history as _pg_all_history
        events = _pg_all_history()
        values: list[list[Any]] = []
        for e in events:
            values.append([
                _fmt_sheet_dt(e.get("created_at")),
                e.get("warranty_code") or "",
                e.get("actor_username") or "",
                e.get("actor_name") or "",
                e.get("action") or "",
                e.get("old_status") or "",
                e.get("new_status") or "",
                "",  # previous_review_status — el legacy lo dejaba vacío en general
                "",  # new_review_status — idem
                e.get("note") or "",
                e.get("details_json") or "",
            ])
        return values

    return []


def write_sheet_values(service: Any, spreadsheet_id: str, sheet_name: str, headers: list[str], values: list[list[Any]]) -> None:
    ensure_sheet_with_headers(service, spreadsheet_id, sheet_name, headers)
    last_col = get_column_letter(len(headers))
    service.spreadsheets().values().clear(
        spreadsheetId=spreadsheet_id,
        range=f"{quote_sheet_name(sheet_name)}!A2:{last_col}50000",
        body={},
    ).execute()
    if values:
        service.spreadsheets().values().update(
            spreadsheetId=spreadsheet_id,
            range=f"{quote_sheet_name(sheet_name)}!A2:{last_col}{len(values) + 1}",
            valueInputOption="USER_ENTERED",
            body={"values": values},
        ).execute()


def warranty_sheet_row(guarantee_row: dict[str, Any], item_row: dict[str, Any]) -> list[Any]:
    def g(key: str, default: Any = "") -> Any:
        try:
            if hasattr(guarantee_row, "keys") and key in guarantee_row.keys():
                return guarantee_row[key]
            if isinstance(guarantee_row, dict):
                return guarantee_row.get(key, default)
        except Exception:
            pass
        return default

    def i(key: str, default: Any = "") -> Any:
        try:
            if hasattr(item_row, "keys") and key in item_row.keys():
                return item_row[key]
            if isinstance(item_row, dict):
                return item_row.get(key, default)
        except Exception:
            pass
        return default

    pending_days = compute_pending_days(guarantee_row)
    no_response_days = compute_no_response_days(guarantee_row)
    fecha_inicio_gestion = g("reviewed_at") or g("sent_to_provider_at") or ""
    observations = str(i("observaciones") or g("observations") or "")
    return [
        g("warranty_code"),
        g("responsible_name"),
        format_date_ar(parse_iso_datetime(g("ingreso_at")) or now_ar()),
        i("producto"),
        i("sku"),
        i("marca"),
        i("serie"),
        i("falla"),
        g("sucursal"),
        g("deposito"),
        g("status"),
        pending_days,
        format_datetime_ar(parse_iso_datetime(fecha_inicio_gestion)) if fecha_inicio_gestion else "",
        g("provider_case_id"),
        "" if no_response_days is None else no_response_days,
        format_datetime_ar(parse_iso_datetime(g("fecha_retiro"))) if g("fecha_retiro") else "",
        format_datetime_ar(parse_iso_datetime(g("fecha_resolucion"))) if g("fecha_resolucion") else "",
        observations,
        g("vuelve_a"),
        g("finalizacion"),
        i("tipo"),
        g("lugar_llegada"),
        g("created_by"),
        format_datetime_ar(parse_iso_datetime(g("updated_at"))) if g("updated_at") else "",
        g("updated_by_name") or g("updated_by"),
    ]


def warranty_rows_for_sheet() -> tuple[list[list[Any]], dict[int, tuple[int, int]]]:
    """Fase 2.5h.2e · raw sheet (00_RAW_GARANTIAS) desde Postgres."""
    all_rows, all_items = pg_fetch_all_guarantee_rows()
    rows = [r for r in all_rows if not (r.get("cancelled") or 0)]
    rows.sort(key=lambda r: (r.get("ingreso_at") or "", r.get("warranty_code") or "", int(r.get("id") or 0)))
    items_by_gid: dict[int, list[dict[str, Any]]] = {}
    for it in all_items:
        items_by_gid.setdefault(int(it["guarantee_id"]), []).append(it)
    values: list[list[Any]] = []
    row_ranges: dict[int, tuple[int, int]] = {}
    next_row = 2
    for guarantee in rows:
        items = sorted(items_by_gid.get(int(guarantee["id"]), []), key=lambda it: int(it.get("id") or 0))
        if not items:
            continue
        start = next_row
        for item in items:
            values.append(warranty_sheet_row(guarantee, item))
            next_row += 1
        row_ranges[int(guarantee["id"])] = (start, next_row - 1)
    return values, row_ranges


def sync_log_info(row: dict[str, Any]) -> WarrantySyncLogInfo:
    errors: list[str] = []
    try:
        parsed = json.loads(row["errors_json"] or "[]")
        if isinstance(parsed, list):
            errors = [str(x) for x in parsed]
    except Exception:
        errors = []
    return WarrantySyncLogInfo(
        id=int(row["id"]),
        sync_type=str(row["sync_type"] or ""),
        status=str(row["status"] or ""),
        started_at=format_datetime_ar(parse_iso_datetime(row["started_at"]) or now_ar()),
        finished_at=format_datetime_ar(parse_iso_datetime(row["finished_at"])) if row["finished_at"] else "",
        actor_username=str(row["actor_username"] or ""),
        actor_name=str(row["actor_name"] or ""),
        rows_processed=int(row["rows_processed"] or 0),
        rows_created=int(row["rows_created"] or 0),
        rows_updated=int(row["rows_updated"] or 0),
        rows_skipped=int(row["rows_skipped"] or 0),
        errors=errors,
    )


# =========================================================
# Endpoints
# =========================================================

# Ingreso de garantias se movio a `warranties/intake.py`.
# Se registra antes de list/review/export/lifecycle para mantener rutas especificas.
from . import intake as _intake_module  # noqa: E402
router.include_router(_intake_module.router)






# Listing se movio a `warranties/listing.py`. Se registra aca, antes del
# catch-all `/{warranty_id}` que queda en este archivo hasta mover lifecycle.py.
from . import listing as _listing_module  # noqa: E402
list_warranties = _listing_module.list_warranties
router.include_router(_listing_module.router)




# ── Reset producción / limpieza de datos de prueba ───────────────────────────

RESET_PRESERVED_ITEMS = [
    "usuarios, roles y permisos",
    "empleados",
    "empresas",
    "sucursales y depósitos",
    "configuración operativa",
    "configuración de Google Sheets",
    "productos y proveedores",
]
RESET_TABLES_IN_DELETE_ORDER = [
    "guarantee_history",
    "guarantee_items",
    "guarantee_exports",
    "guarantee_sync_logs",
    "guarantee_counters",
    "warranty_remitos",
    "guarantees",
]


# Los 3 endpoints de production-reset + sus helpers privados (_is_reset_admin,
# _require_reset_admin, _generated_export_files_count, _reset_summary_pg,
# _create_warranty_reset_backup_pg, _delete_generated_warranty_export_files)
# se movieron a `warranties/reset.py` (Fase B.1).
# El sub-router se registra al final de este archivo con include_router.


# Revision se movio a `warranties/review.py`. Se registra antes de
# provider/lifecycle para conservar el orden de rutas especificas.
from . import review as _review_module  # noqa: E402
router.include_router(_review_module.router)















# Flujo proveedor se movio a `warranties/provider.py`. Se registra antes
# de exports/config/lifecycle para mantener rutas especificas primero.
from . import provider as _provider_module  # noqa: E402
router.include_router(_provider_module.router)

















EXPORT_ELIGIBLE_STATUS = DEFAULT_STATUSES[1]  # 2 - PENDIENTE
EXPORT_READY_STATUS = DEFAULT_STATUSES[2]     # 3 - LISTO PARA ENVIAR


# Exportaciones ENV se movieron a `warranties/exports.py`.
# Se registra aca, antes del catch-all `/{warranty_id}` que queda en este
# archivo hasta mover lifecycle.py.
from . import exports as _exports_module  # noqa: E402
router.include_router(_exports_module.router)























# Los 4 endpoints /sync/* (status, logs, setup-sheet, push-to-sheet) se
# movieron a warranties/sync.py (Fase B.1).
# /sync/pull-from-sheet fue eliminado en Fase 2.5h.2e (PG es fuente única).



# =========================================================
# Dashboard y métricas
# =========================================================

# Configuracion, diagnosticos y dashboard se movieron a `warranties/config.py`.
# Se registra antes de lifecycle para evitar que el catch-all tape /config.
from . import config as _config_module  # noqa: E402
router.include_router(_config_module.router)



















# Fase 2.5h.2d · GET /config porteado a Postgres.






# Fase 2.5h.2d · /diagnostics porteado a Postgres. Las funciones helper que
# inspeccionan filas ya aceptan dicts de pg_fetch_all_guarantee_rows.


# Fase 2.5h.2d · /dashboard porteado a Postgres. Las agregaciones siguen
# haciéndose en memoria sobre los dicts que devuelve pg_fetch_all_guarantee_rows
# (cantidades chicas; rendimiento idéntico al SQLite legacy).

















# ── Sub-routers (Fase B.1) ──────────────────────────────────────────────────
# Se importan al FINAL para evitar ciclos: los sub-routers importan símbolos
# (constantes, modelos, helpers compartidos) de este __init__.py.
from . import reset as _reset_module  # noqa: E402
from . import sync as _sync_module  # noqa: E402
from . import lifecycle as _lifecycle_module  # noqa: E402
router.include_router(_reset_module.router)
router.include_router(_sync_module.router)
# Lifecycle va ultimo porque contiene GET /{warranty_id}.
get_warranty_detail = _lifecycle_module.get_warranty_detail
router.include_router(_lifecycle_module.router)
