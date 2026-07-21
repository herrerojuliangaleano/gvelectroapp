"""Tool interno: Stock valorizado por sucursal.

Toma el Excel crudo de "stock valorizado" que exporta el ERP, deja solo las
columnas útiles (Código, Descripción, Modelo, Dispon, Costo, Valuación),
elimina las filas con Dispon < 0, y arma un workbook con 2 hojas:
- **Detalle**: las filas filtradas.
- **Resumen**: ítems, cantidad total (unidades) y valuación total.

Luego lo sube a Google Drive (carpeta "Valorizado"), dentro de una subcarpeta
del mes según la fecha del día, convertido a Google Sheet, con el nombre según
la sucursal. Cada sucursal sube su propio archivo del día.
"""
from __future__ import annotations

import io
import re
import unicodedata
from datetime import date
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font

from .google_sheets import drive_service

# Carpeta raíz "Valorizado" en Drive.
STOCK_FOLDER_ID = "18OHMbSj3kLJoB9ADC0l-E2DgFImuVc-G"

_MESES = [
    "", "01-Enero", "02-Febrero", "03-Marzo", "04-Abril", "05-Mayo", "06-Junio",
    "07-Julio", "08-Agosto", "09-Septiembre", "10-Octubre", "11-Noviembre", "12-Diciembre",
]

_XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
_SHEET_MIME = "application/vnd.google-apps.spreadsheet"
_FOLDER_MIME = "application/vnd.google-apps.folder"

# Columnas de salida (orden pedido) y su nombre normalizado de origen.
_COLUMNAS = [
    ("codigo", "Código"),
    ("descripcion", "Descripción"),
    ("modelo", "Modelo"),
    ("dispon", "Dispon"),
    ("costo", "Costo"),
    ("valuacion", "Valuación"),
]

_TOTAL_ROW_PREFIXES = ("total", "totales")
_TOTAL_ROW_EXACT = {
    "total cantidad",
    "total cantidades",
    "total valorizado",
    "total valuacion",
    "total valuación",
    "total general",
}

_FILENAME_DATE_PATTERNS = (
    re.compile(r"(?<!\d)(?P<day>\d{1,2})[\s._-](?P<month>\d{1,2})[\s._-](?P<year>\d{2,4})(?!\d)"),
    re.compile(r"(?<!\d)(?P<year>\d{4})[\s._-](?P<month>\d{1,2})[\s._-](?P<day>\d{1,2})(?!\d)"),
)

_SUCURSAL_ALIASES = (
    ("Caseros", ("caseros", "electrogv", "gv")),
    ("Canning", ("canning",)),
    ("Lanus", ("lanus", "sur")),
    ("Norcenter", ("norcenter", "norte", "northcenter", "nortecenter")),
)


def _norm(s: Any) -> str:
    return "".join(
        c for c in unicodedata.normalize("NFKD", str(s)) if not unicodedata.combining(c)
    ).strip().lower()


def _norm_filename(s: str) -> str:
    normalized = _norm(Path(s or "").stem)
    return re.sub(r"[^a-z0-9]+", " ", normalized).strip()


def fecha_desde_nombre_archivo(filename: str) -> date | None:
    """Extrae fechas tipo 11-07-2026 o 2026-07-11 desde el nombre del archivo."""
    text = _norm_filename(filename)
    for pattern in _FILENAME_DATE_PATTERNS:
        match = pattern.search(text)
        if not match:
            continue
        day = int(match.group("day"))
        month = int(match.group("month"))
        year = int(match.group("year"))
        if year < 100:
            year += 2000
        try:
            return date(year, month, day)
        except ValueError:
            continue
    return None


def sucursal_desde_nombre_archivo(filename: str) -> str | None:
    """Infiere la sucursal desde nombres como `stock valorizado norte 11-07-2026.xlsx`."""
    text = f" {_norm_filename(filename)} "
    for sucursal, aliases in _SUCURSAL_ALIASES:
        if any(f" {alias} " in text for alias in aliases):
            return sucursal
    return None


def _is_total_row(row: pd.Series) -> bool:
    """Detecta filas resumen del export ERP, por ejemplo `TOTAL CANTIDAD`."""
    text_fields = (_COLUMNAS[0][1], _COLUMNAS[1][1], _COLUMNAS[2][1])
    for field in text_fields:
        value = _norm(row.get(field, ""))
        if not value or value == "nan":
            continue
        if value in _TOTAL_ROW_EXACT:
            return True
        if any(value.startswith(prefix + " ") for prefix in _TOTAL_ROW_PREFIXES):
            return True
    return False


def procesar_stock(xlsx_bytes: bytes) -> tuple[bytes, dict[str, Any]]:
    """Procesa el Excel crudo. Devuelve (xlsx_procesado, resumen)."""
    df = pd.read_excel(io.BytesIO(xlsx_bytes), sheet_name=0, dtype=object)
    cols = {_norm(c): c for c in df.columns}
    faltan = [norm for norm, _ in _COLUMNAS if norm not in cols]
    if faltan:
        raise ValueError(
            "El archivo no tiene las columnas esperadas. Faltan: "
            + ", ".join(faltan)
            + ". ¿Es el export de 'stock valorizado' del ERP?"
        )

    det = df[[cols[norm] for norm, _ in _COLUMNAS]].copy()
    det.columns = [display for _, display in _COLUMNAS]

    total_original = len(det)
    filas_total_mask = det.apply(_is_total_row, axis=1)
    filas_total_eliminadas = int(filas_total_mask.sum())
    if filas_total_eliminadas:
        det = det[~filas_total_mask].copy()

    for c in ("Dispon", "Costo", "Valuación"):
        det[c] = pd.to_numeric(det[c], errors="coerce")

    det = det[det["Dispon"].fillna(0) >= 0].reset_index(drop=True)
    eliminados = total_original - filas_total_eliminadas - len(det)

    cantidad_total = int(det["Dispon"].sum())
    valuacion_total = float(det["Valuación"].sum())

    resumen_df = pd.DataFrame(
        {
            "Métrica": ["Ítems (productos)", "Cantidad total (unidades)", "Valuación total"],
            "Valor": [len(det), cantidad_total, valuacion_total],
        }
    )

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        det.to_excel(writer, sheet_name="Detalle", index=False)
        resumen_df.to_excel(writer, sheet_name="Resumen", index=False)

    buf.seek(0)
    wb = load_workbook(buf)
    for name in ("Detalle", "Resumen"):
        ws = wb[name]
        for cell in ws[1]:
            cell.font = Font(bold=True)
        for col in ws.columns:
            letter = col[0].column_letter
            largo = max((len(str(c.value)) if c.value is not None else 0) for c in col)
            ws.column_dimensions[letter].width = min(max(largo + 2, 10), 60)
    ws = wb["Detalle"]
    for r in range(2, ws.max_row + 1):
        for letter in ("D", "E", "F"):  # Dispon, Costo, Valuación
            ws[f"{letter}{r}"].number_format = "#,##0"
    ws = wb["Resumen"]
    ws["B3"].number_format = "#,##0"
    ws["B4"].number_format = "#,##0"

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue(), {
        "items": len(det),
        "cantidad_total": cantidad_total,
        "valuacion_total": valuacion_total,
        "eliminados": eliminados,
        "filas_total_eliminadas": filas_total_eliminadas,
    }


def _nombre_carpeta_mes(fecha: date) -> str:
    return f"{_MESES[fecha.month]} {fecha.year}"


def _obtener_o_crear_carpeta(drive, parent_id: str, nombre: str) -> str:
    safe = nombre.replace("\\", "\\\\").replace("'", "\\'")
    query = (
        f"'{parent_id}' in parents and name = '{safe}' "
        f"and mimeType = '{_FOLDER_MIME}' and trashed = false"
    )
    res = drive.files().list(
        q=query, fields="files(id,name)", spaces="drive",
        supportsAllDrives=True, includeItemsFromAllDrives=True,
    ).execute()
    encontrados = res.get("files", [])
    if encontrados:
        return encontrados[0]["id"]
    creada = drive.files().create(
        body={"name": nombre, "mimeType": _FOLDER_MIME, "parents": [parent_id]},
        fields="id", supportsAllDrives=True,
    ).execute()
    return creada["id"]


def subir_a_drive(
    xlsx_bytes: bytes, sucursal: str, fecha: date, resumen: dict[str, Any] | None = None
) -> dict[str, Any]:
    """Sube el xlsx procesado a Drive como Google Sheet, en carpeta del mes → sucursal.

    Guarda los totales como `appProperties` del archivo, para poder armar después
    el mensaje de WhatsApp sin reabrir cada planilla.
    """
    from googleapiclient.http import MediaIoBaseUpload

    drive = drive_service()
    sucursal = sucursal.strip()
    mes = _nombre_carpeta_mes(fecha)
    carpeta_mes = _obtener_o_crear_carpeta(drive, STOCK_FOLDER_ID, mes)
    carpeta_suc = _obtener_o_crear_carpeta(drive, carpeta_mes, sucursal)
    nombre = f"Stock valorizado {sucursal.upper()} {fecha.strftime('%d-%m-%Y')}"

    app_props = {"sv_app": "stock-valorizado", "sv_sucursal": sucursal, "sv_fecha": fecha.isoformat()}
    if resumen:
        app_props["sv_valuacion"] = str(resumen.get("valuacion_total", 0))
        app_props["sv_unidades"] = str(resumen.get("cantidad_total", 0))

    media = MediaIoBaseUpload(io.BytesIO(xlsx_bytes), mimetype=_XLSX_MIME, resumable=False)
    creado = drive.files().create(
        body={"name": nombre, "parents": [carpeta_suc], "mimeType": _SHEET_MIME, "appProperties": app_props},
        media_body=media, fields="id,name,webViewLink", supportsAllDrives=True,
    ).execute()
    return {
        "sheet_id": creado["id"],
        "sheet_name": creado.get("name", nombre),
        "sheet_url": creado.get("webViewLink"),
        "folder_name": f"{mes} / {sucursal}",
    }


def _fmt_ars(valor: float) -> str:
    """133250332.6 -> '133.250.332,60' (formato argentino)."""
    return f"{valor:,.2f}".replace(",", "|").replace(".", ",").replace("|", ".")


def generar_mensaje_whatsapp(fecha: date) -> dict[str, Any]:
    """Arma el mensaje de WhatsApp con las sucursales subidas para esa fecha."""
    drive = drive_service()
    query = (
        "appProperties has { key='sv_app' and value='stock-valorizado' } and "
        f"appProperties has {{ key='sv_fecha' and value='{fecha.isoformat()}' }} and trashed = false"
    )
    res = drive.files().list(
        q=query, fields="files(id,name,createdTime,appProperties)", spaces="drive",
        orderBy="createdTime desc", pageSize=200,
        supportsAllDrives=True, includeItemsFromAllDrives=True,
    ).execute()

    # Dedupe por sucursal quedándose con la más reciente (orderBy createdTime desc).
    por_suc: dict[str, tuple[float, int]] = {}
    for f in res.get("files", []):
        ap = f.get("appProperties", {})
        suc = ap.get("sv_sucursal") or "?"
        if suc in por_suc:
            continue
        try:
            val = float(ap.get("sv_valuacion") or 0)
            uni = int(float(ap.get("sv_unidades") or 0))
        except (TypeError, ValueError):
            continue
        por_suc[suc] = (val, uni)

    ordenadas = sorted(por_suc.items(), key=lambda kv: kv[1][0], reverse=True)  # por valorizado desc
    bloques = [
        f"*{suc}*\n\n• *Valorizado:* ${_fmt_ars(val)}\n• *Unidades:* {uni}"
        for suc, (val, uni) in ordenadas
    ]
    mensaje = "\n\n--------------------------------\n\n".join(bloques)
    return {
        "fecha": fecha.isoformat(),
        "mensaje": mensaje,
        "sucursales": [
            {"sucursal": suc, "valuacion": val, "unidades": uni}
            for suc, (val, uni) in ordenadas
        ],
    }
