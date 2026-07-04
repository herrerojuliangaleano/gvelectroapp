"""Export Excel del dossier de marca — los datos crudos detrás del informe.

Genera un workbook con una hoja por sección, respetando la métrica elegida
(`units` = solo unidades, `pvp` = solo pesos, `both` = todo). Todas las cifras
son números reales con formato (no texto), listas para tabla dinámica.
"""
from __future__ import annotations

from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

MONEY_FMT = '"$" #,##0'
UNITS_FMT = '#,##0'
PCT_FMT = '0.0"%"'

HEAD_FILL = PatternFill("solid", fgColor="1E3A8A")
HEAD_FONT = Font(bold=True, color="FFFFFF")
BOLD = Font(bold=True)


def _units(metric: str) -> bool:
    return metric in ("units", "both")


def _pvp(metric: str) -> bool:
    return metric in ("pvp", "both")


def _sheet(wb: Workbook, title: str):
    ws = wb.create_sheet(title[:31])
    return ws


def _write_table(ws, start_row: int, headers: list[tuple[str, str | None]], rows: list[list[Any]]) -> int:
    """headers: [(titulo, formato|None)]. Devuelve la fila siguiente al final."""
    for c, (head, _fmt) in enumerate(headers, start=1):
        cell = ws.cell(start_row, c, head)
        cell.fill = HEAD_FILL
        cell.font = HEAD_FONT
        cell.alignment = Alignment(vertical="center")
    for r, row in enumerate(rows, start=start_row + 1):
        for c, value in enumerate(row, start=1):
            cell = ws.cell(r, c, value)
            fmt = headers[c - 1][1]
            if fmt and isinstance(value, (int, float)):
                cell.number_format = fmt
    # anchos según contenido del header + muestra
    for c, (head, _f) in enumerate(headers, start=1):
        muestra = [str(head)] + [str(row[c - 1])[:40] for row in rows[:20] if len(row) >= c]
        ws.column_dimensions[get_column_letter(c)].width = min(46, max(10, max(len(x) for x in muestra) + 2))
    ws.freeze_panes = ws.cell(start_row + 1, 1)
    return start_row + 1 + len(rows)


def build_brand_dossier_xlsx(d: dict[str, Any], metric: str = "both") -> bytes:
    u, p = _units(metric), _pvp(metric)
    wb = Workbook()
    marca = d["marca"]
    f = d["filters"]
    periodo = f"{f['fecha_desde']} al {f['fecha_hasta']}"

    # ── Resumen ──────────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Resumen"
    ws["A1"] = f"Informe de marca — {marca}"
    ws["A1"].font = Font(bold=True, size=16)
    ws["A2"] = f"Período: {periodo} · Fuente: {d.get('source', '')} · Sin costos ni márgenes"
    ws["A2"].font = Font(color="64748B")
    kpis: list[list[Any]] = []
    b, m = d["totals"]["brand"], d["totals"]["market"]
    s = d["share"]
    if u:
        kpis.append(["Unidades de la marca", b["unidades"], UNITS_FMT])
        kpis.append(["Unidades del mercado", m["unidades"], UNITS_FMT])
        kpis.append(["Share en unidades (%)", s["units_pct"], PCT_FMT])
    if p:
        kpis.append(["Facturación de la marca", b["total_vendido"], MONEY_FMT])
        kpis.append(["Facturación del mercado", m["total_vendido"], MONEY_FMT])
        kpis.append(["Share en facturación (%)", s["pvp_pct"], PCT_FMT])
        kpis.append(["PVP promedio marca", b["pvp_promedio"], MONEY_FMT])
        kpis.append(["PVP promedio mercado", m["pvp_promedio"], MONEY_FMT])
    kpis.append(["Ranking por facturación", f"#{s['rank_pvp']} de {s['total_brands']}", None])
    kpis.append(["SKUs vendidos", b["productos"], UNITS_FMT])
    kpis.append(["Índice de precio (100=mercado)", d["price_index_global"], None])
    r = 4
    for label, value, fmt in kpis:
        ws.cell(r, 1, label).font = BOLD
        cell = ws.cell(r, 2, value)
        if fmt and isinstance(value, (int, float)):
            cell.number_format = fmt
        r += 1
    r += 1
    ws.cell(r, 1, "Lecturas principales").font = BOLD
    for h in d.get("highlights", []):
        r += 1
        ws.cell(r, 1, f"• {h}")
    r += 2
    for titulo, key in (("Fortalezas", "fortalezas"), ("Oportunidades", "oportunidades"), ("Próximos pasos", "acciones")):
        ws.cell(r, 1, titulo).font = BOLD
        for item in d.get("conclusions", {}).get(key, []):
            r += 1
            ws.cell(r, 1, f"• {item}")
        r += 2
    ws.column_dimensions["A"].width = 60
    ws.column_dimensions["B"].width = 20

    # ── Evolución mensual ────────────────────────────────────────────────
    ws = _sheet(wb, "Evolución mensual")
    heads: list[tuple[str, str | None]] = [("Mes", None)]
    if u:
        heads += [("Marca u", UNITS_FMT), ("Mercado u", UNITS_FMT), ("Share u %", PCT_FMT)]
    if p:
        heads += [("Marca $", MONEY_FMT), ("Mercado $", MONEY_FMT), ("Share $ %", PCT_FMT)]
    rows = []
    for row in d.get("monthly_series", []):
        out: list[Any] = [row["mes"]]
        if u:
            out += [row["brand_unidades"], row["market_unidades"], row["share_units_pct"]]
        if p:
            out += [row["brand_pvp"], row["market_pvp"], row["share_pvp_pct"]]
        rows.append(out)
    _write_table(ws, 1, heads, rows)

    # ── Evolución diaria ─────────────────────────────────────────────────
    ws = _sheet(wb, "Evolución diaria")
    heads = [("Fecha", None)]
    if u:
        heads += [("Marca u", UNITS_FMT), ("Mercado u", UNITS_FMT)]
    if p:
        heads += [("Marca $", MONEY_FMT), ("Mercado $", MONEY_FMT), ("Share $ %", PCT_FMT)]
    rows = []
    for row in d.get("daily_series", []):
        out = [row["fecha"]]
        if u:
            out += [row["brand_unidades"], row["market_unidades"]]
        if p:
            out += [row["brand_pvp"], row["market_pvp"], row["share_pvp_pct"]]
        rows.append(out)
    _write_table(ws, 1, heads, rows)

    # ── Ranking de marcas ────────────────────────────────────────────────
    ws = _sheet(wb, "Ranking")
    heads = [("#", None), ("Marca", None)]
    if u:
        heads.append(("Unidades", UNITS_FMT))
    if p:
        heads += [("Facturación", MONEY_FMT), ("Share $ %", PCT_FMT), ("PVP promedio", MONEY_FMT)]
    heads.append(("SKUs", UNITS_FMT))
    rows = []
    for i, row in enumerate(d.get("ranking", []), start=1):
        out = [i, row["name"] + (" ◀ (la marca)" if row.get("is_brand") else "")]
        if u:
            out.append(row["unidades"])
        if p:
            out += [row["total_vendido"], row.get("participacion_pct") or 0, row["pvp_promedio"]]
        out.append(row["productos"])
        rows.append(out)
    _write_table(ws, 1, heads, rows)

    # ── Categorías ───────────────────────────────────────────────────────
    ws = _sheet(wb, "Categorías")
    heads = [("Categoría", None)]
    if u:
        heads += [("Marca u", UNITS_FMT), ("Mercado u", UNITS_FMT), ("Share u %", PCT_FMT)]
    if p:
        heads += [("Marca $", MONEY_FMT), ("Mercado $", MONEY_FMT), ("Share $ %", PCT_FMT)]
    heads += [("Mix de la marca %", PCT_FMT), ("Share período ant. %", PCT_FMT), ("Δ share (pts)", "0.0"),
              ("Rank", None), ("Marcas en cat.", UNITS_FMT), ("Líder", None), ("Share líder %", PCT_FMT)]
    if p:
        heads += [("PVP prom. marca", MONEY_FMT), ("PVP prom. mercado", MONEY_FMT), ("Índice precio", "0")]
    rows = []
    for c in d.get("categories", []):
        out = [c["categoria"]]
        if u:
            out += [c["brand_unidades"], c["market_unidades"], c["share_units_pct"]]
        if p:
            out += [c["brand_pvp"], c["market_pvp"], c["share_pvp_pct"]]
        out += [c["brand_mix_pct"], c["share_prev_pct"], c["share_delta_pts"],
                c["rank_in_categoria"], c["marcas_en_categoria"], c["leader_name"], c["leader_share_pct"]]
        if p:
            out += [c["brand_avg_pvp"], c["market_avg_pvp"], c["price_index"]]
        rows.append(out)
    _write_table(ws, 1, heads, rows)

    # ── Tipos ────────────────────────────────────────────────────────────
    ws = _sheet(wb, "Tipos")
    heads = [("Tipo", None)]
    if u:
        heads.append(("Unidades", UNITS_FMT))
    if p:
        heads += [("Facturación", MONEY_FMT), ("Mercado $", MONEY_FMT), ("Share $ %", PCT_FMT)]
    rows = []
    for t in d.get("tipos_top", []):
        out = [t["tipo"]]
        if u:
            out.append(t["unidades"])
        if p:
            out += [t["total_vendido"], t["market_pvp"], t["share_pvp_pct"]]
        rows.append(out)
    _write_table(ws, 1, heads, rows)

    # ── Gamas de precio ──────────────────────────────────────────────────
    if d.get("price_bands"):
        ws = _sheet(wb, "Gamas de precio")
        cortes = d["price_bands"]["cortes"]
        ws["A1"] = f"Cortes (terciles del mercado): Entrada hasta ${cortes['entrada_hasta']:,.0f} · Media hasta ${cortes['media_hasta']:,.0f}".replace(",", ".")
        ws["A1"].font = Font(color="64748B")
        heads = [("Gama", None), ("Desde $", MONEY_FMT), ("Hasta $", MONEY_FMT),
                 ("Marca u", UNITS_FMT), ("Mercado u", UNITS_FMT), ("Share u %", PCT_FMT),
                 ("Mix marca u %", PCT_FMT), ("Mix mercado u %", PCT_FMT)]
        if p:
            heads += [("Marca $", MONEY_FMT), ("Share $ %", PCT_FMT)]
        rows = []
        for band in d["price_bands"]["bands"]:
            out = [band["banda"], band["corte_min"], band["corte_max"],
                   band["brand_unidades"], band["market_unidades"], band["share_units_pct"],
                   band["brand_mix_units_pct"], band["market_mix_units_pct"]]
            if p:
                out += [band["brand_pvp"], band["share_pvp_pct"]]
            rows.append(out)
        _write_table(ws, 3, heads, rows)

    # ── Productos ────────────────────────────────────────────────────────
    ws = _sheet(wb, "Productos")
    heads = [("SKU", None), ("Producto", None), ("Tipo", None)]
    if u:
        heads.append(("Unidades", UNITS_FMT))
    if p:
        heads += [("Facturación", MONEY_FMT), ("PVP promedio", MONEY_FMT), ("% de la marca", PCT_FMT)]
    rows = []
    for prod in d.get("top_products", []):
        out = [prod["sku"], prod["producto"], prod["tipo_producto"]]
        if u:
            out.append(prod["unidades"])
        if p:
            out += [prod["total_vendido"], prod["pvp_promedio"], prod.get("participacion_pct") or 0]
        rows.append(out)
    _write_table(ws, 1, heads, rows)

    # ── Producto × Sucursal ──────────────────────────────────────────────
    matriz = d.get("product_branch_metrics") or []
    if matriz:
        ws = _sheet(wb, "Producto x Sucursal")
        sucursales = [br["sucursal"] for br in d.get("branches", [])]
        heads = [("SKU", None), ("Producto", None)]
        if u:
            heads.append(("Total u", UNITS_FMT))
        if p:
            heads.append(("Total $", MONEY_FMT))
        for suc in sucursales:
            if u:
                heads.append((f"{suc} u", UNITS_FMT))
            if p:
                heads.append((f"{suc} $", MONEY_FMT))
        rows = []
        for row in matriz:
            out: list[Any] = [row["sku"], row["producto"]]
            if u:
                out.append(row["total_unidades"])
            if p:
                out.append(row["total_vendido"])
            for suc in sucursales:
                cell = (row.get("branches") or {}).get(suc) or {}
                if u:
                    out.append(cell.get("unidades") or 0)
                if p:
                    out.append(cell.get("total_vendido") or 0)
            rows.append(out)
        _write_table(ws, 1, heads, rows)

    # ── Sucursales ───────────────────────────────────────────────────────
    ws = _sheet(wb, "Sucursales")
    heads = [("Sucursal", None)]
    if u:
        heads += [("Marca u", UNITS_FMT), ("Total sucursal u", UNITS_FMT), ("Share u %", PCT_FMT)]
    if p:
        heads += [("Marca $", MONEY_FMT), ("Total sucursal $", MONEY_FMT), ("Share $ %", PCT_FMT)]
    heads.append(("Mix de la marca %", PCT_FMT))
    rows = []
    for br in d.get("branches", []):
        out = [br["sucursal"]]
        if u:
            out += [br["brand_unidades"], br.get("market_unidades") or 0, br.get("share_units_in_branch_pct") or 0]
        if p:
            out += [br["brand_pvp"], br["market_pvp"], br["share_in_branch_pct"]]
        out.append(br["brand_mix_pct"])
        rows.append(out)
    _write_table(ws, 1, heads, rows)

    # ── Dinámica de productos (si hay período anterior) ─────────────────
    movers = d.get("product_movers") or {}
    if movers.get("up") or movers.get("down"):
        ws = _sheet(wb, "Alzas y bajas")
        heads = [("Dirección", None), ("SKU", None), ("Producto", None),
                 ("Unidades", UNITS_FMT), ("Unidades período ant.", UNITS_FMT), ("Δ unidades", "+#,##0;-#,##0")]
        rows = []
        for mv in movers.get("up", []):
            rows.append(["ALZA", mv["sku"], mv["producto"], mv["unidades"], mv["unidades_prev"], mv["delta_unidades"]])
        for mv in movers.get("down", []):
            rows.append(["BAJA", mv["sku"], mv["producto"], mv["unidades"], mv["unidades_prev"], mv["delta_unidades"]])
        _write_table(ws, 1, heads, rows)

    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()
