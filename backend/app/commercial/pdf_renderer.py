"""Renderer del PDF del PSI — diseño ejecutivo simplificado (v3).

Layout pedido por el gerente:

  +----+  Reporte PSI                                 Marca: ...
  |LOGO|  ELECTRO GV                                  Período: ...
  +----+                                               Sucursales: Todas
                                                       Responsable: ... (si GERENTE_COMERCIAL)

         [ STOCK TOTAL ]    [ SELL OUT TOTAL ]

  ┌─────┬───────────────┬───────┬────────┬──────────┐
  │ SKU │ DESCRIPCIÓN   │ MARCA │ STOCK  │ SELL-OUT │
  └─────┴───────────────┴───────┴────────┴──────────┘
"""
from __future__ import annotations

from datetime import datetime
from io import BytesIO
from pathlib import Path
from typing import Any, Optional
from zoneinfo import ZoneInfo

from reportlab.lib.colors import HexColor, black, white
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import cm, mm
from reportlab.platypus import (
    Image as RLImage,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)

from ..brand_assets import brand_logo_path
from ..operational_config import load_operational_config


AR_TZ = ZoneInfo("America/Argentina/Buenos_Aires")

# ──────────────────────────────────────────────────────────────────────────
# Paleta
# ──────────────────────────────────────────────────────────────────────────

H_TEXT_1     = "#0F172A"
H_TEXT_2     = "#334155"
H_TEXT_3     = "#64748B"
H_BORDER     = "#E2E8F0"
H_SUBTLE_BG  = "#F8FAFC"
H_ACCENT     = "#2563EB"
H_SUCCESS    = "#10B981"
H_DANGER     = "#EF4444"

C_TEXT_1     = HexColor(H_TEXT_1)
C_TEXT_2     = HexColor(H_TEXT_2)
C_TEXT_3     = HexColor(H_TEXT_3)
C_BORDER     = HexColor(H_BORDER)
C_SUBTLE_BG  = HexColor(H_SUBTLE_BG)
C_ACCENT     = HexColor(H_ACCENT)
C_SUCCESS    = HexColor(H_SUCCESS)
C_DANGER     = HexColor(H_DANGER)


# ──────────────────────────────────────────────────────────────────────────
# Logo
# ──────────────────────────────────────────────────────────────────────────

def _logo_path(logo: str) -> Optional[Path]:
    logo_norm = (logo or "").upper()
    if logo_norm == "NONE":
        return None
    cfg = load_operational_config().get("commercial", {}) or {}
    logos = cfg.get("logos", {}) or {}

    key = "gv_path" if logo_norm == "GV" else "abc_path" if logo_norm == "ABC" else None
    if not key:
        return None
    rel = str(logos.get(key) or "")
    return brand_logo_path(logo_norm, configured_path=rel or None, allow_none=True)


# ──────────────────────────────────────────────────────────────────────────
# Helpers
# ──────────────────────────────────────────────────────────────────────────

def _meta_lines(meta_pairs: list[tuple[str, str]]) -> Paragraph:
    """Bloque metadata derecha — líneas 'Label: value'."""
    lines = []
    for label, value in meta_pairs:
        lines.append(
            f'<font color="{H_TEXT_3}"><b>{label}:</b></font>'
            f' <font color="{H_TEXT_2}">{value}</font>'
        )
    style = ParagraphStyle("Meta", fontName="Helvetica", fontSize=9, alignment=TA_RIGHT, leading=13)
    return Paragraph("<br/>".join(lines), style)


def _kpi_card(label: str, value: str) -> Table:
    return Table(
        [
            [Paragraph(f'<font color="{H_TEXT_3}" size="7.5"><b>{label}</b></font>',
                       ParagraphStyle("KPILabel", fontSize=7.5, alignment=TA_LEFT, leading=9))],
            [Paragraph(f'<font color="{H_TEXT_1}" size="24"><b>{value}</b></font>',
                       ParagraphStyle("KPIValue", fontSize=24, alignment=TA_LEFT, leading=28))],
        ],
        colWidths=[60 * mm],
        style=TableStyle([
            ("BACKGROUND",    (0, 0), (-1, -1), C_SUBTLE_BG),
            ("BOX",           (0, 0), (-1, -1), 0.5, C_BORDER),
            ("LEFTPADDING",   (0, 0), (-1, -1), 10),
            ("RIGHTPADDING",  (0, 0), (-1, -1), 10),
            ("TOPPADDING",    (0, 0), (0, 0), 10),
            ("TOPPADDING",    (0, 1), (-1, -1), 2),
            ("BOTTOMPADDING", (0, 1), (-1, -1), 12),
        ]),
    )


# ──────────────────────────────────────────────────────────────────────────
# Render principal
# ──────────────────────────────────────────────────────────────────────────

# ── Columnas configurables del export PSI ───────────────────────────────────
# El caller elige cuáles incluir (cada proveedor pide distinto). "kind" define
# el formato; width_mm/xl los anchos en PDF (mm) y Excel (unidades openpyxl).
PSI_COLUMNS: dict[str, dict[str, Any]] = {
    "sku":          {"header": "SKU",          "align": "left",  "kind": "text",  "width_mm": 30, "width_xl": 16},
    "descripcion":  {"header": "Descripción",  "align": "left",  "kind": "text",  "width_mm": 74, "width_xl": 48},
    "marca":        {"header": "Marca",        "align": "left",  "kind": "text",  "width_mm": 24, "width_xl": 16},
    "pvp":          {"header": "PVP",          "align": "right", "kind": "money", "width_mm": 26, "width_xl": 14},
    "stock_inicio": {"header": "Stock inicio", "align": "right", "kind": "int",   "width_mm": 24, "width_xl": 13},
    "stock_final":  {"header": "Stock final",  "align": "right", "kind": "int",   "width_mm": 24, "width_xl": 13},
    "sell_out":     {"header": "Sell-out",     "align": "right", "kind": "int",   "width_mm": 24, "width_xl": 13},
}
PSI_DEFAULT_COLUMNS = ["sku", "descripcion", "marca", "pvp", "stock_inicio", "stock_final", "sell_out"]


def _psi_norm_columns(columns: Optional[list[str]]) -> list[str]:
    cols = [c for c in (columns or []) if c in PSI_COLUMNS]
    return cols or list(PSI_DEFAULT_COLUMNS)


def _psi_fmt_val(kind: str, value: Any) -> str:
    if kind == "money":
        if value in (None, ""):
            return "—"
        try:
            return "$ " + f"{float(value):,.0f}".replace(",", ".")
        except (TypeError, ValueError):
            return "—"
    if kind == "int":
        try:
            return f"{int(value):,}".replace(",", ".")
        except (TypeError, ValueError):
            return "0"
    return str(value or "")


def render_psi_pdf(
    *,
    titulo: str,
    items: list[dict[str, Any]],
    totals: dict[str, Any],
    filters_applied: dict[str, Any],
    gfk_files_used: list[dict[str, Any]] | None = None,
    logo: str = "GV",
    responsable: Optional[str] = None,
    responsable_area: str = "Comercial",
    empresa_nombre: Optional[str] = None,
    columns: Optional[list[str]] = None,
) -> bytes:
    """Genera el PDF del PSI. Columnas configurables y título editable.

    - ``columns``: qué columnas incluir (ver PSI_COLUMNS). Cada item ya trae
      pvp / stock_inicio / stock_final calculados por el router.
    - ``titulo``: se imprime tal cual; si viene vacío, no se muestra título.
    - ``responsable``: solo se imprime si != None.
    """
    cols = _psi_norm_columns(columns)
    buffer = BytesIO()
    # A4 apaisado cuando hay muchas columnas; vertical si son pocas.
    page_size = landscape(A4) if len(cols) > 5 else A4
    usable_mm = (page_size[0] / mm) - 30.0  # ancho útil (márgenes 15mm por lado)
    doc = SimpleDocTemplate(
        buffer,
        pagesize=page_size,
        leftMargin=1.5 * cm,
        rightMargin=1.5 * cm,
        topMargin=1.2 * cm,
        bottomMargin=1.2 * cm,
        title=titulo or "Reporte",
    )

    styles = getSampleStyleSheet()
    h_title = ParagraphStyle(
        "HTitle",
        parent=styles["Title"],
        fontName="Helvetica-Bold",
        fontSize=20,
        leading=24,
        alignment=TA_LEFT,
        textColor=C_ACCENT,
        spaceAfter=2,
    )
    h_subtitle = ParagraphStyle(
        "HSub",
        parent=styles["Normal"],
        fontName="Helvetica-Bold",
        fontSize=12,
        leading=15,
        alignment=TA_LEFT,
        textColor=C_TEXT_1,
        spaceAfter=2,
    )
    sub_style = ParagraphStyle(
        "Sub",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=8.5,
        textColor=C_TEXT_3,
        leading=11,
    )
    cell_style = ParagraphStyle(
        "Cell",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=9,
        leading=11,
        textColor=C_TEXT_2,
    )
    cell_bold = ParagraphStyle("CellB", parent=cell_style, fontName="Helvetica-Bold", textColor=C_TEXT_1)

    # Empresa por defecto según logo
    if empresa_nombre is None:
        empresa_nombre = "ELECTRO GV" if logo.upper() == "GV" else "ELECTRO ABC SRL" if logo.upper() == "ABC" else "ELECTRO"

    logo_p = _logo_path(logo)
    if logo_p:
        try:
            logo_img: Any = RLImage(str(logo_p), width=18 * mm, height=18 * mm, kind="proportional")
            logo_img.hAlign = "LEFT"
        except Exception:
            logo_img = Paragraph("", sub_style)
    else:
        logo_img = Paragraph("", sub_style)

    # Bloque izquierdo: logo + titulo (editable, se omite si viene vacío) + empresa.
    titulo_txt = (titulo or "").strip()
    titulo_block = []
    if titulo_txt:
        titulo_block.append(Paragraph(titulo_txt, h_title))
    titulo_block.append(Paragraph(empresa_nombre, h_subtitle))
    left_w = usable_mm - 50.0  # el bloque derecho de metadata usa 50mm
    left_block = Table(
        [[logo_img, titulo_block]],
        colWidths=[22 * mm, (left_w - 22) * mm],
        style=TableStyle([
            ("VALIGN",       (0, 0), (-1, -1), "MIDDLE"),
            ("LEFTPADDING",  (0, 0), (-1, -1), 0),
            ("RIGHTPADDING", (0, 0), (-1, -1), 4),
        ]),
    )

    # ─── HEADER derecho: metadata ────────────────────────────────────────
    filt = filters_applied or {}
    marcas_list = filt.get("marcas") or []
    marca_text = ", ".join(marcas_list) if marcas_list else "Todas"
    pi = filt.get("periodo_inicio") or ""
    pf = filt.get("periodo_fin") or ""

    def _fmt_date(iso: str) -> str:
        try:
            d = datetime.strptime(iso, "%Y-%m-%d")
            return d.strftime("%d/%m/%Y")
        except (TypeError, ValueError):
            return iso

    periodo_text = f"{_fmt_date(pi)} → {_fmt_date(pf)}" if pi and pf else "—"

    meta_pairs: list[tuple[str, str]] = [
        ("Marca",      marca_text),
        ("Período",    periodo_text),
        ("Sucursales", "Todas"),
    ]
    if filt.get("condicion") and filt["condicion"] != "TODO":
        meta_pairs.append(("Condición", filt["condicion"]))
    if responsable:
        full = f"{responsable} · {responsable_area}" if responsable_area else responsable
        meta_pairs.append(("Responsable", full))

    right_block = _meta_lines(meta_pairs)

    header_table = Table(
        [[left_block, right_block]],
        colWidths=[left_w * mm, 50 * mm],
        style=TableStyle([
            ("VALIGN",       (0, 0), (-1, -1), "TOP"),
            ("LEFTPADDING",  (0, 0), (-1, -1), 0),
            ("RIGHTPADDING", (0, 0), (-1, -1), 0),
        ]),
    )

    story: list = [header_table, Spacer(1, 7 * mm)]

    # ─── KPIs: totales de las columnas numéricas que estén seleccionadas ──
    def _tot(key: str) -> int:
        return sum(int(r.get(key) or 0) for r in items)

    kpi_cards = [
        _kpi_card(lbl, f"{_tot(k):,}".replace(",", "."))
        for k, lbl in (("stock_inicio", "STOCK INICIO"), ("stock_final", "STOCK FINAL"), ("sell_out", "SELL OUT"))
        if k in cols
    ]
    if kpi_cards:
        kpi_row = Table(
            [kpi_cards],
            colWidths=[58 * mm] * len(kpi_cards),
            style=TableStyle([
                ("LEFTPADDING",  (0, 0), (-1, -1), 0),
                ("RIGHTPADDING", (0, 0), (-2, 0), 8),
                ("RIGHTPADDING", (-1, 0), (-1, 0), 0),
            ]),
            hAlign="LEFT",
        )
        story.append(kpi_row)
        story.append(Spacer(1, 7 * mm))

    # ─── TABLA SIMPLE ─────────────────────────────────────────────────────
    header_h = ParagraphStyle("H", parent=cell_style, fontName="Helvetica-Bold", textColor=C_TEXT_3)
    header_h_right = ParagraphStyle("HR", parent=header_h, alignment=TA_RIGHT)

    right_cell = ParagraphStyle("RC", parent=cell_style, alignment=TA_RIGHT)
    headers_row = [
        Paragraph(PSI_COLUMNS[c]["header"], header_h_right if PSI_COLUMNS[c]["align"] == "right" else header_h)
        for c in cols
    ]
    data_rows: list[list[Any]] = [headers_row]

    for r in items:
        row_cells: list[Any] = []
        for c in cols:
            spec = PSI_COLUMNS[c]
            txt = _psi_fmt_val(spec["kind"], r.get(c))
            if spec["kind"] in ("int", "money"):
                # Sell-out en verde si vendió "bien" (>= 80% del stock final).
                color = None
                if c == "sell_out":
                    sf = int(r.get("stock_final") or 0)
                    sell = int(r.get("sell_out") or 0)
                    if sf > 0 and sell >= sf * 0.8:
                        color = H_SUCCESS
                    elif sell == 0:
                        color = H_TEXT_3
                elif c in ("stock_inicio", "stock_final") and int(r.get(c) or 0) < 0:
                    color = H_DANGER
                inner = f'<font color="{color}"><b>{txt}</b></font>' if color else f"<b>{txt}</b>"
                row_cells.append(Paragraph(inner, right_cell))
            elif c == "sku":
                row_cells.append(Paragraph(txt, cell_bold))
            else:
                row_cells.append(Paragraph(txt, cell_style))
        data_rows.append(row_cells)

    if len(data_rows) == 1:
        data_rows.append([Paragraph("Sin productos para los filtros aplicados.", cell_style)] + [""] * (len(cols) - 1))

    # Anchos: columnas cortas fijas; la descripción absorbe el resto del ancho útil.
    fixed = sum(PSI_COLUMNS[c]["width_mm"] for c in cols if c != "descripcion")
    col_widths = [
        (max(40.0, usable_mm - fixed) * mm if c == "descripcion" else PSI_COLUMNS[c]["width_mm"] * mm)
        for c in cols
    ]
    main_table = Table(data_rows, colWidths=col_widths, repeatRows=1)
    main_table.setStyle(TableStyle([
        ("BACKGROUND",    (0, 0), (-1, 0), C_SUBTLE_BG),
        ("LINEBELOW",     (0, 0), (-1, 0), 0.5, C_BORDER),
        ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
        ("LINEBELOW",     (0, 1), (-1, -1), 0.25, C_BORDER),
        ("LEFTPADDING",   (0, 0), (-1, -1), 5),
        ("RIGHTPADDING",  (0, 0), (-1, -1), 5),
        ("TOPPADDING",    (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
    ]))
    story.append(main_table)

    # Sin footer, sin leyenda, sin paginación visible (queda mínimo).
    doc.build(story)
    pdf_bytes = buffer.getvalue()
    buffer.close()
    return pdf_bytes


# ──────────────────────────────────────────────────────────────────────────
# Export Excel (.xlsx)
# ──────────────────────────────────────────────────────────────────────────

def render_psi_xlsx(
    *,
    titulo: str,
    items: list[dict[str, Any]],
    totals: dict[str, Any],
    filters_applied: dict[str, Any],
    empresa_nombre: Optional[str] = None,
    responsable: Optional[str] = None,
    responsable_area: str = "Comercial",
    logo: str = "GV",
    columns: Optional[list[str]] = None,
) -> bytes:
    """Genera el reporte PSI como Excel (.xlsx) con columnas configurables.

    Los items ya traen pvp / stock_inicio / stock_final calculados. El título
    es editable (si viene vacío se usa el nombre de la empresa).
    """
    cols = _psi_norm_columns(columns)
    import openpyxl
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reporte"

    if empresa_nombre is None:
        empresa_nombre = "ELECTRO GV" if logo.upper() == "GV" else "ELECTRO ABC SRL" if logo.upper() == "ABC" else "ELECTRO"

    filt = filters_applied or {}
    marca_text = ", ".join(filt.get("marcas") or []) or "Todas"
    pi = filt.get("periodo_inicio") or ""
    pf = filt.get("periodo_fin") or ""

    def _fmt_date(iso: str) -> str:
        try:
            return datetime.strptime(iso, "%Y-%m-%d").strftime("%d/%m/%Y")
        except (TypeError, ValueError):
            return iso

    periodo_text = f"{_fmt_date(pi)} → {_fmt_date(pf)}" if pi and pf else "—"

    # Estilos
    title_font   = Font(name="Calibri", size=18, bold=True, color="2563EB")
    subt_font    = Font(name="Calibri", size=12, bold=True, color="0F172A")
    meta_label   = Font(name="Calibri", size=10, bold=True, color="64748B")
    meta_value   = Font(name="Calibri", size=10, color="334155")
    kpi_label    = Font(name="Calibri", size=9,  bold=True, color="64748B")
    kpi_value    = Font(name="Calibri", size=20, bold=True, color="0F172A")
    header_font  = Font(name="Calibri", size=10, bold=True, color="64748B")
    header_fill  = PatternFill("solid", fgColor="F8FAFC")
    border_thin  = Side(style="thin",  color="E2E8F0")
    grid_border  = Border(left=border_thin, right=border_thin, top=border_thin, bottom=border_thin)

    # ── Header (filas 1-5)
    title_col = "A"
    logo_path = _logo_path(logo)
    if logo_path:
        try:
            logo_img = XLImage(str(logo_path))
            logo_img.width = 64
            logo_img.height = 64
            ws.add_image(logo_img, "A1")
            title_col = "B"
            ws.row_dimensions[1].height = 30
            ws.row_dimensions[2].height = 24
            ws.row_dimensions[3].height = 18
        except Exception:
            title_col = "A"

    ws[f"{title_col}1"] = (titulo or "").strip() or empresa_nombre
    ws[f"{title_col}1"].font = title_font
    ws[f"{title_col}2"] = empresa_nombre
    ws[f"{title_col}2"].font = subt_font

    # Metadata derecha (columnas D-E filas 1-5)
    meta_pairs = [
        ("Marca",      marca_text),
        ("Período",    periodo_text),
        ("Sucursales", "Todas"),
    ]
    if filt.get("condicion") and filt["condicion"] != "TODO":
        meta_pairs.append(("Condición", filt["condicion"]))
    if responsable:
        meta_pairs.append(("Responsable", f"{responsable} · {responsable_area}" if responsable_area else responsable))

    for i, (label, value) in enumerate(meta_pairs, start=1):
        ws.cell(row=i, column=4, value=f"{label}:").font = meta_label
        ws.cell(row=i, column=4).alignment = Alignment(horizontal="right")
        ws.cell(row=i, column=5, value=value).font = meta_value

    # ── KPIs (filas 7-8): totales de las columnas numéricas seleccionadas
    kpi_defs = [(k, lbl) for k, lbl in
                (("stock_inicio", "STOCK INICIO"), ("stock_final", "STOCK FINAL"), ("sell_out", "SELL OUT"))
                if k in cols]
    for i, (k, lbl) in enumerate(kpi_defs):
        col = 1 + i
        ws.cell(row=7, column=col, value=lbl).font = kpi_label
        c8 = ws.cell(row=8, column=col, value=sum(int(r.get(k) or 0) for r in items))
        c8.font = kpi_value
        c8.number_format = "#,##0"

    # ── Tabla dinámica (a partir de la fila 11)
    table_start_row = 11
    for col_idx, c in enumerate(cols, start=1):
        spec = PSI_COLUMNS[c]
        cell = ws.cell(row=table_start_row, column=col_idx, value=spec["header"])
        cell.font = header_font
        cell.fill = header_fill
        cell.border = grid_border
        cell.alignment = Alignment(horizontal="right" if spec["align"] == "right" else "left")

    for offset, r in enumerate(items, start=1):
        row = table_start_row + offset
        for col_idx, c in enumerate(cols, start=1):
            spec = PSI_COLUMNS[c]
            if spec["kind"] == "text":
                cell = ws.cell(row=row, column=col_idx, value=str(r.get(c) or ""))
                if c == "sku":
                    cell.font = Font(name="Calibri", size=10, bold=True)
            else:
                try:
                    num: Any = float(r.get(c)) if r.get(c) not in (None, "") else None
                except (TypeError, ValueError):
                    num = None
                if spec["kind"] == "int" and num is not None:
                    num = int(num)
                cell = ws.cell(row=row, column=col_idx, value=num)
                cell.alignment = Alignment(horizontal="right")
                cell.number_format = '"$" #,##0' if spec["kind"] == "money" else "#,##0"
            cell.border = grid_border

    for col_idx, c in enumerate(cols, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = PSI_COLUMNS[c]["width_xl"]

    # Freeze top-left para que el header de tabla quede fijo al scrollear
    ws.freeze_panes = ws.cell(row=table_start_row + 1, column=1)

    buf = BytesIO()
    wb.save(buf)
    data = buf.getvalue()
    buf.close()
    return data
