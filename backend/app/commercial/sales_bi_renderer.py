from __future__ import annotations

import io
from datetime import datetime
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.graphics.shapes import Circle, Drawing, Line, Rect, String
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import Image as RLImage
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from ..brand_assets import brand_logo_path


NAVY = colors.HexColor("#0F2449")
BLUE = colors.HexColor("#2F7DFF")
CYAN = colors.HexColor("#22C7E8")
GREEN = colors.HexColor("#15B981")
AMBER = colors.HexColor("#F59E0B")
TEXT = colors.HexColor("#111827")
MUTED = colors.HexColor("#64748B")
SOFT = colors.HexColor("#F4F7FB")


def _money(value: Any) -> str:
    try:
        n = float(value or 0)
    except Exception:
        n = 0.0
    return "$ " + f"{n:,.0f}".replace(",", ".")


def _num(value: Any) -> str:
    try:
        n = float(value or 0)
    except Exception:
        n = 0.0
    return f"{n:,.0f}".replace(",", ".")


def _period(report: dict[str, Any]) -> str:
    filters = report.get("filters", {}) or {}
    return f"{filters.get('fecha_desde') or ''} al {filters.get('fecha_hasta') or ''}"


def _logo_path(logo: str) -> Path | None:
    return brand_logo_path(logo or "GV", allow_none=True)


def _bar_chart(items: list[dict[str, Any]], *, label_key: str, value_key: str, width: int = 250, height: int = 120, color=BLUE) -> Drawing:
    drawing = Drawing(width, height)
    top_items = items[:6]
    max_value = max([float(i.get(value_key) or 0) for i in top_items] or [1])
    y = height - 18
    for item in top_items:
        label = str(item.get(label_key) or "")[:22]
        value = float(item.get(value_key) or 0)
        drawing.add(String(0, y + 2, label, fontSize=7, fillColor=TEXT))
        drawing.add(String(width - 56, y + 2, _money(value), fontSize=7, fillColor=MUTED))
        drawing.add(Rect(0, y - 8, width - 70, 6, fillColor=colors.HexColor("#E5EAF3"), strokeColor=None))
        drawing.add(Rect(0, y - 8, max(2, (width - 70) * value / max_value), 6, fillColor=color, strokeColor=None))
        y -= 18
    return drawing


def _line_chart(series: list[dict[str, Any]], *, width: int = 250, height: int = 110) -> Drawing:
    drawing = Drawing(width, height)
    values = [float(p.get("total_cobrado") or 0) for p in series]
    if not values:
        drawing.add(String(20, height / 2, "Sin datos para el periodo", fontSize=9, fillColor=MUTED))
        return drawing
    max_value = max(values) or 1
    min_value = min(values)
    span = max(max_value - min_value, 1)
    pad_x = 12
    pad_y = 16
    drawing.add(Line(pad_x, pad_y, width - pad_x, pad_y, strokeColor=colors.HexColor("#CBD5E1")))
    drawing.add(Line(pad_x, pad_y, pad_x, height - pad_y, strokeColor=colors.HexColor("#CBD5E1")))
    points = []
    for idx, value in enumerate(values):
        x = pad_x + (width - pad_x * 2) * (idx / max(1, len(values) - 1))
        y = pad_y + (height - pad_y * 2) * ((value - min_value) / span)
        points.append((x, y))
    for (x1, y1), (x2, y2) in zip(points, points[1:]):
        drawing.add(Line(x1, y1, x2, y2, strokeColor=BLUE, strokeWidth=2))
    for x, y in points:
        drawing.add(Circle(x, y, 2.2, fillColor=BLUE, strokeColor=None))
    drawing.add(String(pad_x, height - 10, _money(max_value), fontSize=7, fillColor=MUTED))
    return drawing


def render_sellers_pdf(report: dict[str, Any], *, compare: dict[str, Any] | None = None, logo: str = "GV", title: str = "Informe de vendedores") -> bytes:
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        rightMargin=12 * mm,
        leftMargin=12 * mm,
        topMargin=10 * mm,
        bottomMargin=10 * mm,
    )
    styles = getSampleStyleSheet()
    h1 = ParagraphStyle("H1", parent=styles["Heading1"], fontSize=22, leading=26, textColor=colors.white, spaceAfter=4)
    small = ParagraphStyle("Small", parent=styles["Normal"], fontSize=8, leading=10, textColor=colors.white)
    section = ParagraphStyle("Section", parent=styles["Heading2"], fontSize=13, leading=16, textColor=NAVY, spaceBefore=8, spaceAfter=6)

    logo_p = _logo_path(logo)
    logo_flow = RLImage(str(logo_p), width=22 * mm, height=22 * mm, kind="proportional") if logo_p else Paragraph("", small)
    header = Table(
        [[
            [Paragraph(title, h1), Paragraph(f"Periodo: {_period(report)} · Generado {datetime.now().strftime('%d/%m/%Y %H:%M')}", small)],
            logo_flow,
        ]],
        colWidths=[235 * mm, 28 * mm],
        style=TableStyle([
            ("BACKGROUND", (0, 0), (-1, -1), NAVY),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("LEFTPADDING", (0, 0), (-1, -1), 12),
            ("RIGHTPADDING", (0, 0), (-1, -1), 12),
            ("TOPPADDING", (0, 0), (-1, -1), 12),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 12),
        ]),
    )

    totals = report.get("totals", {}) or {}
    kpis = [
        ("Cobrado", _money(totals.get("total_cobrado"))),
        ("Vendido", _money(totals.get("total_vendido"))),
        ("Unidades", _num(totals.get("unidades"))),
        ("Tickets", _num(totals.get("tickets"))),
        ("Ticket prom.", _money(totals.get("ticket_promedio"))),
        ("Saldo", _money(totals.get("saldo"))),
    ]
    if "margen_porcentaje" in totals:
        kpis.append(("Margen", f"{float(totals.get('margen_porcentaje') or 0):.1f}%"))
    kpi_table = Table(
        [[Paragraph(f"<b>{label}</b><br/><font size='14'>{value}</font>", styles["Normal"]) for label, value in kpis]],
        colWidths=[260 / len(kpis) * mm] * len(kpis),
        style=TableStyle([
            ("BACKGROUND", (0, 0), (-1, -1), SOFT),
            ("TEXTCOLOR", (0, 0), (-1, -1), TEXT),
            ("BOX", (0, 0), (-1, -1), 0.25, colors.HexColor("#D8E1F0")),
            ("INNERGRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#D8E1F0")),
            ("TOPPADDING", (0, 0), (-1, -1), 8),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
        ]),
    )

    story: list[Any] = [header, Spacer(1, 8), kpi_table, Spacer(1, 8)]
    story.append(Paragraph("Metricas y graficos", section))
    story.append(Table(
        [[
            _line_chart(report.get("daily_series", []), width=250, height=120),
            _bar_chart(report.get("sellers", []), label_key="vendedor", value_key="total_cobrado", width=250, height=120, color=BLUE),
            _bar_chart(report.get("brand_mix", []), label_key="name", value_key="total_cobrado", width=250, height=120, color=GREEN),
        ]],
        colWidths=[88 * mm, 88 * mm, 88 * mm],
    ))

    sellers = report.get("sellers", [])[:12]
    rows = [["Vendedor", "Sucursal", "Cobrado", "Unid.", "Tickets", "Ticket prom.", "Part. suc.", "Part. emp."]]
    for s in sellers:
        rows.append([
            s.get("vendedor", ""),
            s.get("sucursal", ""),
            _money(s.get("total_cobrado")),
            _num(s.get("unidades")),
            _num(s.get("tickets")),
            _money(s.get("ticket_promedio")),
            f"{float(s.get('sucursal_participacion_pct') or s.get('participacion_pct') or 0):.1f}%",
            f"{float(s.get('empresa_participacion_pct') or s.get('participacion_pct') or 0):.1f}%",
        ])
    story.append(Paragraph("Ranking de vendedores", section))
    story.append(Table(rows, repeatRows=1, colWidths=[56 * mm, 28 * mm, 32 * mm, 20 * mm, 20 * mm, 30 * mm, 22 * mm, 22 * mm], style=TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), NAVY),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#D8E1F0")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, SOFT]),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
    ])))

    if compare:
        story.append(Paragraph("Comparacion de periodo", section))
        delta = compare.get("delta", {}) or {}
        rows = [["Metrica", "Actual", "Comparado", "Diferencia", "%"]]
        labels = {
            "total_cobrado": "Cobrado",
            "unidades": "Unidades",
            "tickets": "Tickets",
            "ticket_promedio": "Ticket prom.",
            "saldo": "Saldo",
        }
        for key, label in labels.items():
            d = delta.get(key, {}) or {}
            is_money = key in {"total_cobrado", "ticket_promedio", "saldo"}
            fmt = _money if is_money else _num
            rows.append([label, fmt(d.get("actual")), fmt(d.get("comparado")), fmt(d.get("delta")), "-" if d.get("delta_pct") is None else f"{d.get('delta_pct'):.1f}%"])
        story.append(Table(rows, repeatRows=1, colWidths=[48 * mm, 38 * mm, 38 * mm, 38 * mm, 24 * mm], style=TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), NAVY),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#D8E1F0")),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
        ])))

    doc.build(story)
    return buffer.getvalue()


def _write_rows(ws, rows: list[list[Any]]) -> None:
    for row in rows:
        ws.append(row)
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top")
    for col_idx in range(1, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = min(42, max(12, len(str(ws.cell(1, col_idx).value or "")) + 4))
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="0F2449")


def render_sellers_xlsx(report: dict[str, Any], *, compare: dict[str, Any] | None = None, logo: str = "GV", title: str = "Informe de vendedores") -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Resumen"
    ws["A1"] = title
    ws["A1"].font = Font(size=18, bold=True, color="0F2449")
    ws["A2"] = f"Periodo: {_period(report)}"
    logo_p = _logo_path(logo)
    if logo_p:
        try:
            img = XLImage(str(logo_p))
            img.width = 64
            img.height = 64
            ws.add_image(img, "F1")
        except Exception:
            pass
    totals = report.get("totals", {}) or {}
    rows = [["Metrica", "Valor"]]
    for key in ("total_cobrado", "total_vendido", "saldo", "unidades", "tickets", "ticket_promedio", "participacion_pct", "diferencia", "margen_porcentaje"):
        if key in totals:
            rows.append([key, totals.get(key)])
    _write_rows(ws, rows)

    ws_sellers = wb.create_sheet("Vendedores")
    _write_rows(ws_sellers, [[
        "Vendedor", "Sucursal principal", "Cobrado", "Vendido", "Saldo", "Unidades", "Tickets", "Ticket promedio", "Part. sucursal %", "Part. empresa %", "Ranking sucursal", "Ranking empresa", "Margen %"
    ]] + [[
        s.get("vendedor"), s.get("sucursal"), s.get("total_cobrado"), s.get("total_vendido"), s.get("saldo"), s.get("unidades"),
        s.get("tickets"), s.get("ticket_promedio"),
        s.get("sucursal_participacion_pct", s.get("participacion_pct")),
        s.get("empresa_participacion_pct", s.get("participacion_pct")),
        s.get("rank_sucursal", ""),
        s.get("rank_empresa", ""),
        s.get("margen_porcentaje", ""),
    ] for s in report.get("sellers", [])])

    ws_daily = wb.create_sheet("Evolucion")
    _write_rows(ws_daily, [["Fecha", "Cobrado", "Vendido", "Unidades", "Tickets"]] + [[
        d.get("fecha"), d.get("total_cobrado"), d.get("total_vendido"), d.get("unidades"), d.get("tickets")
    ] for d in report.get("daily_series", [])])

    ws_products = wb.create_sheet("Top productos")
    _write_rows(ws_products, [["SKU", "Producto", "Marca", "Cobrado", "Unidades", "Tickets"]] + [[
        p.get("sku"), p.get("producto"), p.get("marca"), p.get("total_cobrado"), p.get("unidades"), p.get("tickets")
    ] for p in report.get("top_products", [])])

    ws_detail = wb.create_sheet("Detalle")
    detail_headers = ["fecha", "sucursal", "tipo", "remito", "vendedor", "sku", "producto", "marca", "categoria", "cantidad", "pvp", "costo", "total_cobrado", "saldo", "margen_porcentaje", "product_match_status"]
    _write_rows(ws_detail, [detail_headers] + [[row.get(h, "") for h in detail_headers] for row in report.get("detail", [])])

    if compare:
        ws_compare = wb.create_sheet("Comparacion")
        rows = [["Metrica", "Actual", "Comparado", "Delta", "Delta %"]]
        for key, value in (compare.get("delta", {}) or {}).items():
            rows.append([key, value.get("actual"), value.get("comparado"), value.get("delta"), value.get("delta_pct")])
        _write_rows(ws_compare, rows)

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()
