"""Importación histórica de garantías desde Excel (.xlsx).

Lee un workbook generado por el push-to-sheet de la app y crea filas en
`guarantees`, `guarantee_items`, `remitos`, `remito_items`, `guarantee_history`
y `guarantee_exports`. Diseñado para traer datos del sistema viejo a Postgres
una sola vez. Es idempotente: si un `warranty_code` (o `remito_code`,
`shipment_code`) ya existe, lo skipea y reporta.

Pestañas leídas:
  GARANTIAS, GARANTIA_ITEMS  → guarantees + guarantee_items
  REMITOS, REMITO_ITEMS      → remitos + remito_items
  EVENTOS                    → guarantee_history
  LOTES_ENV                  → guarantee_exports

Resolvers de IDs:
  EMPRESA texto       → company_id (electro_gv / electro_abc_srl)
  SUCURSAL texto      → branch_id  (caseros / sur / norte / canning / deposito_*)
  USUARIO texto       → user_id    (por username o display_name; NULL si no existe)
"""
from __future__ import annotations

import json
from datetime import date, datetime, time, timezone
from typing import Any
from zoneinfo import ZoneInfo

import openpyxl
from sqlalchemy import func, select

from .db import db_session
from .models.auth import User
from .models.remitos import Remito, RemitoItem
from .models.warranties import (
    Guarantee,
    GuaranteeExport,
    GuaranteeHistory,
    GuaranteeItem,
)
from .routers.warranties import (
    resolve_branch_id_from_text,
    resolve_company_id_from_text,
    resolve_user_id_from_username,
    sucursal_code as compute_sucursal_code,
)

AR_TZ = ZoneInfo("America/Argentina/Buenos_Aires")


# ──────────────────────────────────────────────────────────────────────────
# Helpers de parseo
# ──────────────────────────────────────────────────────────────────────────

def _clean(value: Any) -> str:
    """Stringifica y limpia. Floats enteros se imprimen sin '.0' (teléfonos)."""
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _parse_date(value: Any) -> date | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    s = str(value).strip()
    if not s:
        return None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(s[:10], fmt).date()
        except Exception:
            pass
    return None


def _parse_dt(value: Any) -> datetime | None:
    """Convierte a datetime UTC. Asume zona AR si viene naive."""
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        if value.tzinfo is None:
            return value.replace(tzinfo=AR_TZ).astimezone(timezone.utc)
        return value.astimezone(timezone.utc)
    if isinstance(value, date):
        return datetime.combine(value, time.min, tzinfo=AR_TZ).astimezone(timezone.utc)
    return None


def _parse_json(value: Any) -> dict:
    if value is None or value == "":
        return {}
    if isinstance(value, dict):
        return value
    try:
        return json.loads(str(value))
    except Exception:
        return {}


def _read_sheet(wb: openpyxl.Workbook, sheet_name: str) -> list[dict[str, Any]]:
    """Lee una hoja como lista de dicts (header → value). Saltea filas vacías."""
    if sheet_name not in wb.sheetnames:
        return []
    ws = wb[sheet_name]
    headers: list[str] = []
    out: list[dict[str, Any]] = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            headers = [str(h or "").strip() for h in row]
            continue
        if not any(c is not None and c != "" for c in row):
            continue
        out.append({h: v for h, v in zip(headers, row) if h})
    return out


def _resolve_user_by_display_or_username(session, text: str) -> int | None:
    """Busca user.id primero por username, después por display_name (case-insensitive)."""
    t = (text or "").strip()
    if not t:
        return None
    uid = resolve_user_id_from_username(session, t)
    if uid:
        return int(uid)
    return session.scalar(
        select(User.id).where(func.lower(User.display_name) == t.lower())
    )


# ──────────────────────────────────────────────────────────────────────────
# Importadores por tabla
# ──────────────────────────────────────────────────────────────────────────

def _import_guarantees(wb, session, actor_user_id: int | None, stats: dict, code_to_id: dict[str, int]) -> None:
    g_rows = _read_sheet(wb, "GARANTIAS")
    i_rows = _read_sheet(wb, "GARANTIA_ITEMS")

    items_by_code: dict[str, list[dict]] = {}
    for it in i_rows:
        code = _clean(it.get("ID GARANTIA"))
        if code:
            items_by_code.setdefault(code, []).append(it)

    for g in g_rows:
        code = _clean(g.get("ID GARANTIA"))
        if not code:
            stats["warranties_skipped_empty"] += 1
            continue

        existing = session.scalar(select(Guarantee.id).where(Guarantee.warranty_code == code))
        if existing:
            code_to_id[code] = int(existing)
            stats["warranties_skipped_existing"] += 1
            continue

        empresa_text = _clean(g.get("EMPRESA"))
        branch_carga_text = _clean(g.get("SUCURSAL CARGA"))
        branch_resp_text = _clean(g.get("SUCURSAL RESPONSABLE"))
        branch_id = resolve_branch_id_from_text(branch_carga_text)
        sucursal_responsable_id = resolve_branch_id_from_text(branch_resp_text) or branch_id

        company_id = resolve_company_id_from_text(empresa_text)
        if not company_id and branch_id:
            from .branches_db import get_branch as _get_branch
            b = _get_branch(branch_id)
            if b:
                company_id = b.get("company_id")
        if not company_id:
            company_id = "electro_gv"
            stats["warnings"].append(f"{code}: empresa indeterminada, usando electro_gv")

        ingreso_at = _parse_dt(g.get("FECHA INGRESO")) or datetime.now(timezone.utc)
        actualizado_at = _parse_dt(g.get("ACTUALIZADO")) or ingreso_at
        suc_text = branch_carga_text or branch_resp_text or ""

        guarantee = Guarantee(
            warranty_code=code,
            parent_item_index=0,
            company_id=company_id,
            branch_id=branch_id,
            sucursal_responsable_id=sucursal_responsable_id,
            sucursal=suc_text,
            sucursal_code=compute_sucursal_code(suc_text),
            sucursal_responsable=branch_resp_text,
            deposito=_clean(g.get("DEPOSITO DESTINO")),
            lugar_llegada=_clean(g.get("DEPOSITO DESTINO")),
            status=_clean(g.get("ESTADO")) or "1 - INGRESO",
            review_status=_clean(g.get("REVISION")) or "pendiente_revision",
            tipo_ingreso=_clean(g.get("TIPO INGRESO")),
            origen_ingreso=_clean(g.get("ORIGEN INGRESO")),
            ubicacion_actual=_clean(g.get("UBICACION ACTUAL")) or "deposito",
            transit_status="",
            review_note="",
            provider_name=_clean(g.get("PROVEEDOR")),
            provider_case_id="",
            sent_to_provider_at=_parse_dt(g.get("FECHA ENVIO PROVEEDOR")),
            fecha_ultimo_mail_proveedor=_parse_dt(g.get("FECHA ULTIMO MAIL")),
            provider_response_type=_clean(g.get("RESPUESTA PROVEEDOR")),
            provider_correction_note="",
            estado_retiro_proveedor=_clean(g.get("RETIRO PROVEEDOR")) or "sin_solicitud",
            fecha_retiro_proveedor=_parse_dt(g.get("FECHA RETIRO PROVEEDOR")),
            nota_retiro_proveedor="",
            remito_proveedor="",
            remito_interno=_clean(g.get("REMITO INTERNO")),
            lugar_salida_transito="",
            fecha_resolucion=_parse_dt(g.get("FECHA RESOLUCION")),
            resultado_resolucion=_clean(g.get("RESULTADO RESOLUCION")),
            resolution_note="",
            resolution_reference="",
            numero_nota_credito="",
            detalle_reparacion="",
            producto_reemplazo="",
            sku_reemplazo="",
            serie_reemplazo="",
            fecha_finalizacion=_parse_dt(g.get("FECHA FINALIZACION")),
            finalizacion="",
            vuelve_a="",
            shipment_code=_clean(g.get("ENV")),
            shipment_file_name="",
            observations=_clean(g.get("OBSERVACIONES")),
            photos_reference="",
            cancelled=False,
            cancel_reason="",
            cliente_nombre=_clean(g.get("CLIENTE")),
            cliente_telefono=_clean(g.get("TELEFONO")),
            cliente_email=_clean(g.get("EMAIL")),
            numero_factura=_clean(g.get("FACTURA")),
            fecha_compra=_parse_date(g.get("FECHA COMPRA")),
            ingreso_at=ingreso_at,
            created_at=ingreso_at,
            updated_at=actualizado_at,
            responsible_user_id=actor_user_id,
            created_by_user_id=actor_user_id,
            updated_by_user_id=actor_user_id,
            synced_to_google_sheet=False,
            google_sheet_row_id="",
            sync_error="",
        )
        session.add(guarantee)
        session.flush()
        code_to_id[code] = int(guarantee.id)
        stats["warranties_created"] += 1

        items = items_by_code.get(code) or [{}]
        for idx, it in enumerate(items, 1):
            item_index_raw = _clean(it.get("ITEM"))
            try:
                item_index = int(float(item_index_raw)) if item_index_raw else idx
            except ValueError:
                item_index = idx
            session.add(GuaranteeItem(
                guarantee_id=guarantee.id,
                item_index=item_index,
                producto=_clean(it.get("PRODUCTO")),
                sku=_clean(it.get("SKU")),
                marca=_clean(it.get("MARCA")),
                tipo=_clean(it.get("TIPO")),
                serie=_clean(it.get("SERIE")),
                falla=_clean(it.get("FALLA")),
                observaciones=_clean(it.get("OBSERVACIONES")),
                correction_note="",
            ))
            stats["items_created"] += 1


def _import_remitos(wb, session, actor_user_id: int | None, stats: dict, code_to_id: dict[str, int]) -> None:
    r_rows = _read_sheet(wb, "REMITOS")
    ri_rows = _read_sheet(wb, "REMITO_ITEMS")

    items_by_remito: dict[str, list[dict]] = {}
    for it in ri_rows:
        rcode = _clean(it.get("CODIGO REMITO"))
        if rcode:
            items_by_remito.setdefault(rcode, []).append(it)

    for r in r_rows:
        rcode = _clean(r.get("CODIGO REMITO"))
        if not rcode:
            stats["remitos_skipped_empty"] += 1
            continue

        existing = session.scalar(select(Remito.id).where(Remito.remito_code == rcode))
        if existing:
            stats["remitos_skipped_existing"] += 1
            continue

        origen_text = _clean(r.get("ORIGEN"))
        destino_text = _clean(r.get("DESTINO"))
        remito = Remito(
            remito_code=rcode,
            shipment_code="",
            tipo_remito=_clean(r.get("TIPO REMITO")) or "sucursal_a_deposito",
            company_brand=_clean(r.get("EMPRESA")) or "gv_electro",
            origen_branch_id=resolve_branch_id_from_text(origen_text),
            destino_branch_id=resolve_branch_id_from_text(destino_text),
            origen_sucursal=origen_text,
            destino_deposito=destino_text,
            proveedor="",
            status=_clean(r.get("ESTADO")) or "pendiente",
            nota=_clean(r.get("NOTA")),
            pdf_path=_clean(r.get("PDF")),
            created_by_user_id=_resolve_user_by_display_or_username(session, _clean(r.get("CREADO POR"))) or actor_user_id,
            despachado_por_user_id=_resolve_user_by_display_or_username(session, _clean(r.get("DESPACHADO POR"))),
            recibido_por_user_id=_resolve_user_by_display_or_username(session, _clean(r.get("RECIBIDO POR"))),
            created_at=_parse_dt(r.get("FECHA CREACION")) or datetime.now(timezone.utc),
            fecha_despacho=_parse_dt(r.get("FECHA DESPACHO")),
            fecha_llegada=_parse_dt(r.get("FECHA LLEGADA")),
        )
        session.add(remito)
        session.flush()
        stats["remitos_created"] += 1

        for it in items_by_remito.get(rcode, []):
            warranty_code = _clean(it.get("ID GARANTIA"))
            gid = code_to_id.get(warranty_code)
            if not gid:
                stats["warnings"].append(f"Remito {rcode}: garantía {warranty_code} no encontrada (skip item).")
                continue
            session.add(RemitoItem(
                remito_id=remito.id,
                guarantee_id=gid,
                created_at=_parse_dt(it.get("UPDATED_AT")) or remito.created_at,
            ))
            stats["remito_items_created"] += 1


def _import_events(wb, session, actor_user_id: int | None, stats: dict, code_to_id: dict[str, int]) -> None:
    rows = _read_sheet(wb, "EVENTOS")
    for ev in rows:
        warranty_code = _clean(ev.get("ID GARANTIA"))
        gid = code_to_id.get(warranty_code)
        if not gid:
            stats["events_skipped_no_guarantee"] += 1
            continue
        username = _clean(ev.get("USUARIO"))
        nombre = _clean(ev.get("NOMBRE"))
        actor_id = (
            resolve_user_id_from_username(session, username)
            or _resolve_user_by_display_or_username(session, nombre)
        )
        session.add(GuaranteeHistory(
            guarantee_id=gid,
            warranty_code=warranty_code,
            action=_clean(ev.get("ACCION")) or "imported",
            old_status=_clean(ev.get("ESTADO ANTERIOR")),
            new_status=_clean(ev.get("ESTADO NUEVO")),
            field_name="",
            old_value=_clean(ev.get("REVISION ANTERIOR")),
            new_value=_clean(ev.get("REVISION NUEVA")),
            note=_clean(ev.get("DETALLE")),
            details=_parse_json(ev.get("METADATA")),
            actor_user_id=actor_id,
            created_at=_parse_dt(ev.get("FECHA")) or datetime.now(timezone.utc),
        ))
        stats["events_created"] += 1


def _import_exports(wb, session, actor_user_id: int | None, stats: dict) -> None:
    """LOTES_ENV → guarantee_exports. Lista de warranty_ids viene en OBSERVACIONES (JSON)."""
    rows = _read_sheet(wb, "LOTES_ENV")
    for lote in rows:
        env_code = _clean(lote.get("ENV"))
        if not env_code:
            continue
        # Skip si ya existe por file_name
        archivo = _clean(lote.get("ARCHIVO EXCEL"))
        existing = session.scalar(
            select(GuaranteeExport.id).where(GuaranteeExport.file_name == archivo)
        ) if archivo else None
        if existing:
            stats["exports_skipped_existing"] += 1
            continue
        metadata = _parse_json(lote.get("OBSERVACIONES"))
        warranty_ids = metadata.get("warranty_ids") if isinstance(metadata, dict) else []
        created_by_text = _clean(lote.get("CREADO POR"))
        session.add(GuaranteeExport(
            created_by_user_id=_resolve_user_by_display_or_username(session, created_by_text) or actor_user_id,
            provider_name=_clean(lote.get("PROVEEDOR")),
            marca=_clean(lote.get("MARCA")),
            warranty_ids=warranty_ids or [],
            filters={"env": env_code, "estado_lote": _clean(lote.get("ESTADO LOTE"))},
            file_path=archivo or env_code,
            file_name=archivo or env_code,
            file_format="excel",
            logo_brand="gv_electro",
            row_count=int(float(_clean(lote.get("CANTIDAD ITEMS")) or 0) or 0),
        ))
        stats["exports_created"] += 1


# ──────────────────────────────────────────────────────────────────────────
# Orquestador público
# ──────────────────────────────────────────────────────────────────────────

def import_from_xlsx(file_bytes: bytes, actor: Any) -> dict:
    """Importa el contenido del workbook a Postgres en una sola transacción.

    Args:
        file_bytes: contenido binario del .xlsx (UploadFile.file.read())
        actor: usuario autenticado que dispara el import. Se usa como
            actor_user_id en created_by/updated_by/responsible_user_id de las
            garantías cargadas (los usuarios originales del Excel se intentan
            mapear por username; si no existen, queda el actor).

    Returns:
        dict con stats: created/skipped por tabla, lista de warnings, errores.
    """
    import io

    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    stats: dict = {
        "warranties_created": 0,
        "warranties_skipped_empty": 0,
        "warranties_skipped_existing": 0,
        "items_created": 0,
        "remitos_created": 0,
        "remitos_skipped_empty": 0,
        "remitos_skipped_existing": 0,
        "remito_items_created": 0,
        "events_created": 0,
        "events_skipped_no_guarantee": 0,
        "exports_created": 0,
        "exports_skipped_existing": 0,
        "warnings": [],
        "errors": [],
    }
    code_to_id: dict[str, int] = {}

    with db_session() as session:
        actor_user_id = resolve_user_id_from_username(session, getattr(actor, "username", ""))
        try:
            _import_guarantees(wb, session, actor_user_id, stats, code_to_id)
            _import_remitos(wb, session, actor_user_id, stats, code_to_id)
            _import_events(wb, session, actor_user_id, stats, code_to_id)
            _import_exports(wb, session, actor_user_id, stats)
            session.commit()
            stats["ok"] = True
        except Exception as exc:
            session.rollback()
            stats["ok"] = False
            stats["errors"].append(str(exc))

    return stats
